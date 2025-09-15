from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse, Http404
from django.db.models import Q, Count
from django.core.paginator import Paginator
from django.utils import timezone
from datetime import timedelta, datetime
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
import json
from calendar import monthrange
from django.views.decorators.http import require_POST

from .models import PRFRequest, EmergencyLoan
from .forms import PRFRequestForm, PRFFilterForm, PRFActionForm, EmergencyLoanForm
from notification.models import Notification
from finance.models import Loan, LoanType

def get_chart_data(request, period):
    """Generate chart data based on period and filters"""
    now = timezone.now()
    today = now.date()

    if period == 'week':
        # Start of current week (Monday)
        start_of_week = today - timedelta(days=today.weekday())
        end_of_week = start_of_week + timedelta(days=6)
        title = 'This Week'
        date_range = 7
        prfs = PRFRequest.objects.filter(status__in=['pending', 'approved', 'disapproved'], created_at__date__gte=start_of_week, created_at__date__lte=end_of_week)
    elif period == 'quarter':
        # Current calendar quarter
        current_month = today.month
        quarter = (current_month - 1) // 3 + 1
        start_month = 3 * (quarter - 1) + 1
        end_month = start_month + 2
        start_of_quarter = today.replace(month=start_month, day=1)
        last_day = calendar.monthrange(today.year, end_month)[1]
        end_of_quarter = today.replace(month=end_month, day=last_day)
        title = 'This Quarter'
        # 3 months, show by week
        prfs = PRFRequest.objects.filter(status__in=['pending', 'approved', 'disapproved'], created_at__date__gte=start_of_quarter, created_at__date__lte=end_of_quarter)
        # For quarter: show 12-14 weeks depending on the quarter
        week_count = ((end_of_quarter - start_of_quarter).days // 7) + 1
        chart_data = []
        categories = ['government', 'banking', 'hr_payroll']
        for week in range(week_count):
            week_start = start_of_quarter + timedelta(weeks=week)
            week_end = week_start + timedelta(days=6)
            month_name = week_start.strftime('%b')
            week_label = f"{month_name} W{week % 4 + 1}"
            week_data = {'date': week_label}
            total_count = 0
            for category in categories:
                count = prfs.filter(
                    prf_category=category,
                    created_at__date__gte=week_start,
                    created_at__date__lte=week_end
                ).count()
                week_data[category] = count
                total_count += count
            week_data['total'] = total_count
            chart_data.append(week_data)
        return {'data': chart_data, 'title': title}
    else:
        # This Month: from the 1st to the last day of the current month
        start_date = today.replace(day=1)
        last_day = calendar.monthrange(today.year, today.month)[1]
        end_date = today.replace(day=last_day)
        title = 'This Month'
        date_range = (end_date - start_date).days + 1
        prfs = PRFRequest.objects.filter(status__in=['pending', 'approved', 'disapproved'], created_at__date__gte=start_date, created_at__date__lte=end_date)

    chart_data = []
    categories = ['government', 'banking', 'hr_payroll']
    for i in range(date_range):
        date = (start_of_week if period == 'week' else start_date) + timedelta(days=i)
        date_label = date.strftime('%a' if period == 'week' else '%b %d')
        day_data = {'date': date_label}
        total_count = 0
        for category in categories:
            count = prfs.filter(
                prf_category=category,
                created_at__date=date
            ).count()
            day_data[category] = count
            total_count += count
        day_data['total'] = total_count
        chart_data.append(day_data)
    return {'data': chart_data, 'title': title}

@login_required
def employee_prf_view(request):
    if request.user.hr_admin:
        return redirect('admin_prf')
    
    user_prfs = PRFRequest.objects.filter(employee=request.user)
    
    context = {
        'prfs': user_prfs,
        'form': PRFRequestForm(),
    }
    return render(request, 'prf/user-prf.html', context)

@login_required
def submit_prf_request(request):
    if request.method == 'POST':
        form = PRFRequestForm(request.POST)
        if form.is_valid():
            prf_request = form.save(commit=False)
            prf_request.employee = request.user
            
            # Emergency Loan specific validation - check for existing loan but don't create PRF yet
            if prf_request.prf_type == 'emergency_loan':
                # Check if user has existing Emergency Loan before proceeding
                try:
                    emergency_loan_type, created = LoanType.objects.get_or_create(
                        loan_type="Emergency Loan",
                        defaults={'description': 'Emergency loan for employees'}
                    )
                    
                    existing_loan = Loan.objects.filter(
                        employee=request.user,
                        loan_type=emergency_loan_type,
                        is_active=True
                    ).first()
                    
                    # Check if existing loan balance is more than one monthly deduction
                    if existing_loan and existing_loan.current_balance > existing_loan.monthly_deduction:
                        return JsonResponse({
                            'success': False,
                            'message': 'existing_emergency_loan',
                            'current_balance': float(existing_loan.current_balance),
                            'show_existing_loan_error': True
                        })
                
                except Exception as e:
                    return JsonResponse({
                        'success': False,
                        'message': f'Error checking existing loans: {str(e)}'
                    })
                
                # No existing loan found - show Emergency Loan modal without creating PRF yet
                return JsonResponse({
                    'success': True,
                    'message': 'Please complete Emergency Loan details.',
                    'show_emergency_loan_modal': True,
                    'prf_data': {
                        'prf_category': prf_request.prf_category,
                        'prf_type': prf_request.prf_type,
                        'purpose': prf_request.purpose
                    }
                })
            
            # For other loan types, require control number
            loan_types = ['pagibig_loan', 'sss_loan', 'medical_loan', 'educational_loan', 'coop_loan']
            if prf_request.prf_type in loan_types and not prf_request.control_number:
                return JsonResponse({
                    'success': False,
                    'message': 'Control number is required for this PRF type.'
                })
            
            prf_request.save()
            return JsonResponse({
                'success': True,
                'message': 'PRF request submitted successfully!'
            })
        else:
            return JsonResponse({
                'success': False,
                'message': 'Please fill in all required fields correctly.'
            })
    
    return JsonResponse({'success': False, 'message': 'Invalid request method.'})

@login_required
def get_prf_types(request):
    category = request.GET.get('category')
    if not category:
        return JsonResponse({'types': []})
    
    category_types = {
        'government': [
            ('pagibig_loan', 'PAG-IBIG Loan'),
            ('pagibig_cert_payment', 'PAG-IBIG Certificate of Payment'),
            ('pagibig_cert_contribution', 'PAG-IBIG Certificate of Contribution'),
            ('philhealth_form', 'PHILHEALTH Form'),
            ('sss_loan', 'SSS Loan'),
            ('sss_maternity', 'SSS Maternity Benefits'),
            ('sss_sickness', 'SSS Sickness Benefits'),
            ('bir_form', 'BIR Form (2316/1902)'),
        ],
        'banking': [
            ('rcbc_maintenance', 'RCBC Maintenance Form'),
            ('bank_deposit', 'Bank Deposit'),
        ],
        'hr_payroll': [
            ('payroll_adjustment', 'Payroll Adjustment'),
            ('id_replacement', 'ID Replacement'),
            ('pcoe_compensation', 'PCOE with Compensation'),
            ('certificate_employment', 'Certificate of Employment'),
            ('clearance_form', 'Clearance Form'),
            ('emergency_loan', 'Emergency Loan'),
            ('medical_loan', 'Medical Assistance Loan'),
            ('educational_loan', 'Educational Assistance Loan'),
            ('coop_loan', 'Coop Loan'),
            ('uniform_ppe', 'Uniform / Caps / PPE / T-shirt'),
            ('others', 'Others'),
        ]
    }
    
    types = category_types.get(category, [])
    return JsonResponse({'types': types})

@login_required
def get_prf_detail(request, prf_id):
    prf = get_object_or_404(PRFRequest, id=prf_id, employee=request.user)
    
    data = {
        'id': prf.id,
        'prf_category': prf.get_prf_category_display(),
        'prf_type': prf.get_prf_type_display(),
        'purpose': prf.purpose,
        'control_number': prf.control_number or 'N/A',
        'status': prf.get_status_display(),
        'admin_remarks': prf.admin_remarks or 'No remarks',
        'created_at': prf.created_at.strftime('%B %d, %Y at %I:%M %p'),
        'updated_at': prf.updated_at.strftime('%B %d, %Y at %I:%M %p'),
    }
    
    # Add Emergency Loan details if this is an Emergency Loan PRF
    if prf.prf_type == 'emergency_loan' and hasattr(prf, 'emergency_loan'):
        emergency_loan = prf.emergency_loan
        data['emergency_loan'] = {
            'amount': emergency_loan.amount,
            'amount_display': f"₱{emergency_loan.amount:,}",
            'number_of_cutoff': emergency_loan.number_of_cutoff,
            'cutoff_display': f"{emergency_loan.number_of_cutoff} Cut-off{'s' if emergency_loan.number_of_cutoff > 1 else ''} ({emergency_loan.number_of_cutoff * 0.5} month{'s' if emergency_loan.number_of_cutoff > 1 else ''})",
            'starting_date': emergency_loan.starting_date.isoformat(),
            'starting_date_display': emergency_loan.formatted_starting_date,
            'employee_full_name': emergency_loan.employee_full_name,
            'deduction_per_cutoff': emergency_loan.deduction_per_cutoff,
            'deduction_per_cutoff_display': f"₱{emergency_loan.deduction_per_cutoff:,.2f}",
            'created_at': emergency_loan.created_at.strftime('%B %d, %Y at %I:%M %p'),
        }
    
    return JsonResponse(data)

@login_required
def admin_dashboard(request):
    if not request.user.hr_admin:
        return redirect('user_prf')

    period = request.GET.get('period', 'month')
    chart_result = get_chart_data(request, period)
    chart_data = chart_result['data']
    chart_title = chart_result['title']

    # Calculate month start and end for display
    month_start = month_end = None
    if period == 'month':
        now = timezone.now()
        month_start = now.replace(day=1)
        last_day = monthrange(now.year, now.month)[1]
        month_end = now.replace(day=last_day)

    prfs = PRFRequest.objects.all().select_related('employee', 'processed_by')

    search = request.GET.get('search', '')
    if search:
        prfs = prfs.filter(
            Q(employee__username__icontains=search) |
            Q(employee__firstname__icontains=search) |
            Q(employee__lastname__icontains=search) |
            Q(employee__idnumber__icontains=search) |
            Q(prf_type__icontains=search) |
            Q(prf_control_number__icontains=search) |
            Q(control_number__icontains=search) |
            Q(purpose__icontains=search) |
            Q(status__icontains=search)
        )

    filter_form = PRFFilterForm(request.GET)
    if filter_form.is_valid():
        if filter_form.cleaned_data['prf_type']:
            prfs = prfs.filter(prf_type=filter_form.cleaned_data['prf_type'])
        if filter_form.cleaned_data['start_date']:
            prfs = prfs.filter(created_at__date__gte=filter_form.cleaned_data['start_date'])
        if filter_form.cleaned_data['end_date']:
            prfs = prfs.filter(created_at__date__lte=filter_form.cleaned_data['end_date'])
    
    # Order by status (pending first) and then by creation date (newest first)
    from django.db.models import Case, When, IntegerField
    prfs = prfs.annotate(
        status_order=Case(
            When(status='pending', then=1),
            When(status='approved', then=2),
            When(status='disapproved', then=3),
            When(status='cancelled', then=4),
            default=5,
            output_field=IntegerField()
        )
    ).order_by('status_order', '-created_at')
    
    paginator = Paginator(prfs, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    now = timezone.now()
    start_of_this_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    start_of_prev_month = (start_of_this_month - timedelta(days=1)).replace(day=1)

    total_this = PRFRequest.objects.filter(created_at__gte=start_of_this_month).count()
    total_prev = PRFRequest.objects.filter(created_at__gte=start_of_prev_month, created_at__lt=start_of_this_month).count()
    pending_this = PRFRequest.objects.filter(status='pending', created_at__gte=start_of_this_month).count()
    pending_prev = PRFRequest.objects.filter(status='pending', created_at__gte=start_of_prev_month, created_at__lt=start_of_this_month).count()
    approved_this = PRFRequest.objects.filter(status='approved', created_at__gte=start_of_this_month).count()
    approved_prev = PRFRequest.objects.filter(status='approved', created_at__gte=start_of_prev_month, created_at__lt=start_of_this_month).count()
    disapproved_this = PRFRequest.objects.filter(status='disapproved', created_at__gte=start_of_this_month).count()
    disapproved_prev = PRFRequest.objects.filter(status='disapproved', created_at__gte=start_of_prev_month, created_at__lt=start_of_this_month).count()

    def calc_percent(this, prev):
        if prev == 0:
            return 100 if this > 0 else 0
        return round(((this - prev) / prev) * 100)

    context = {
        'chart_data': json.dumps(chart_data),
        'prfs': page_obj,
        'chart_title': chart_title,
        'page_obj': page_obj,
        'filter_form': filter_form,
        'search': search,
        'total_requests': PRFRequest.objects.count(),
        'pending_requests': PRFRequest.objects.filter(status='pending').count(),
        'approved_requests': PRFRequest.objects.filter(status='approved').count(),
        'disapproved_requests': PRFRequest.objects.filter(status='disapproved').count(),
        'total_requests_percent': abs(calc_percent(total_this, total_prev)),
        'total_requests_positive': total_this >= total_prev,
        'pending_requests_percent': abs(calc_percent(pending_this, pending_prev)),
        'pending_requests_positive': pending_this >= pending_prev,
        'approved_requests_percent': abs(calc_percent(approved_this, approved_prev)),
        'approved_requests_positive': approved_this >= approved_prev,
        'disapproved_requests_percent': abs(calc_percent(disapproved_this, disapproved_prev)),
        'disapproved_requests_positive': disapproved_this >= disapproved_prev,
        'month_start': month_start,
        'month_end': month_end,
    }
    return render(request, 'prf/admin-prf.html', context)

@login_required
def get_chart_data_ajax(request):
    if not request.user.hr_admin:
        return JsonResponse({'success': False, 'message': 'Unauthorized'})

    period = request.GET.get('period', 'month')
    chart_result = get_chart_data(request, period)

    return JsonResponse({
        'success': True,
        'data': chart_result['data'],
        'title': chart_result['title']
    })

@login_required
def get_table_data_ajax(request):
    if not request.user.hr_admin:
        return JsonResponse({'success': False, 'message': 'Unauthorized'})

    prfs = PRFRequest.objects.all().select_related('employee', 'processed_by')

    # Apply search filter
    search = request.GET.get('search', '')
    if search:
        prfs = prfs.filter(
            Q(employee__username__icontains=search) |
            Q(employee__firstname__icontains=search) |
            Q(employee__lastname__icontains=search) |
            Q(employee__idnumber__icontains=search) |
            Q(prf_type__icontains=search) |
            Q(prf_control_number__icontains=search) |
            Q(control_number__icontains=search) |
            Q(purpose__icontains=search) |
            Q(status__icontains=search)
        )

    # Apply form filters
    filter_form = PRFFilterForm(request.GET)
    if filter_form.is_valid():
        if filter_form.cleaned_data['prf_type']:
            prfs = prfs.filter(prf_type=filter_form.cleaned_data['prf_type'])
        if filter_form.cleaned_data['start_date']:
            prfs = prfs.filter(created_at__date__gte=filter_form.cleaned_data['start_date'])
        if filter_form.cleaned_data['end_date']:
            prfs = prfs.filter(created_at__date__lte=filter_form.cleaned_data['end_date'])

    # Apply popover filters
    filter_field = request.GET.get('filter_field')
    filter_value = request.GET.get('filter_value')

    if filter_field and filter_value:
        if filter_field == 'status':
            if filter_value != 'all':
                status_values = filter_value.split(',')
                prfs = prfs.filter(status__in=status_values)
        elif filter_field == 'category':
            if filter_value != 'all':
                category_values = filter_value.split(',')
                prfs = prfs.filter(prf_category__in=category_values)
        elif filter_field == 'prf_type':
            if filter_value != 'all':
                type_values = filter_value.split(',')
                prfs = prfs.filter(prf_type__in=type_values)
        else:
            # Handle text-based filters
            if ':' in filter_value:
                condition, value = filter_value.split(':', 1)
                if condition == 'is':
                    prfs = prfs.filter(**{f'{filter_field}__iexact': value})
                elif condition == 'is_not':
                    prfs = prfs.exclude(**{f'{filter_field}__iexact': value})
                elif condition == 'contains':
                    prfs = prfs.filter(**{f'{filter_field}__icontains': value})
                elif condition == 'any':
                    prfs = prfs.exclude(**{f'{filter_field}__isnull': True}).exclude(**{f'{filter_field}__exact': ''})

    # Order by status (pending first) and then by creation date (newest first)
    from django.db.models import Case, When, IntegerField
    prfs = prfs.annotate(
        status_order=Case(
            When(status='pending', then=1),
            When(status='approved', then=2),
            When(status='disapproved', then=3),
            When(status='cancelled', then=4),
            default=5,
            output_field=IntegerField()
        )
    ).order_by('status_order', '-created_at')

    # Pagination
    paginator = Paginator(prfs, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # Prepare table data
    table_data = []
    for prf in page_obj:
        table_data.append({
            'id': prf.id,
            'employee_username': prf.employee.username,
            'employee_firstname': prf.employee.firstname,
            'employee_lastname': prf.employee.lastname,
            'prf_type_display': prf.get_prf_type_display(),
            'prf_category_display': prf.get_prf_category_display(),
            'purpose': prf.purpose,
            'status': prf.status,
            'status_display': prf.get_status_display(),
            'created_at': prf.created_at.strftime('%b %d, %Y'),
            'prf_control_number': prf.prf_control_number or 'N/A',
            'can_process': prf.status == 'pending'
        })

    # Pagination info
    pagination_data = {
        'has_previous': page_obj.has_previous(),
        'has_next': page_obj.has_next(),
        'previous_page_number': page_obj.previous_page_number() if page_obj.has_previous() else None,
        'next_page_number': page_obj.next_page_number() if page_obj.has_next() else None,
        'number': page_obj.number,
        'start_index': page_obj.start_index(),
        'end_index': page_obj.end_index(),
        'total_count': paginator.count,
        'page_range': list(paginator.page_range),
        'current_page': page_obj.number
    }

    return JsonResponse({
        'success': True,
        'table_data': table_data,
        'pagination': pagination_data,
        'search': search,
        'has_results': len(table_data) > 0
    })

@require_POST
@login_required
def process_prf_action(request):
    prf_id = request.POST.get('prf_id')
    action = request.POST.get('action')
    remarks = request.POST.get('remarks', '')
    if not prf_id or action not in ['approved', 'disapproved']:
        return JsonResponse({'success': False, 'message': 'Invalid request.'})
    try:
        prf = PRFRequest.objects.select_related('employee').get(pk=prf_id)
        prf.status = action
        prf.processed_by = request.user
        if remarks:
            prf.admin_remarks = remarks
        prf.save()
        
        # If this is an Emergency Loan PRF and it's approved, create a finance.Loan
        if action == 'approved' and prf.prf_type == 'emergency_loan':
            try:
                emergency_loan = prf.emergency_loan
                
                # Get or create Emergency Loan type
                emergency_loan_type, created = LoanType.objects.get_or_create(
                    loan_type="Emergency Loan",
                    defaults={'description': 'Emergency loan for employees'}
                )
                
                # Calculate monthly deduction (loan amount / number of cutoffs)
                monthly_deduction = emergency_loan.amount / emergency_loan.number_of_cutoff
                
                # Create the finance.Loan record
                finance_loan = Loan.objects.create(
                    employee=prf.employee,  # requestor
                    loan_type=emergency_loan_type,  # emergency loan type
                    principal_amount=emergency_loan.amount,  # loan amount
                    current_balance=emergency_loan.amount,  # initially same as principal
                    monthly_deduction=monthly_deduction  # loan amount / number of cutoffs
                )
                
                print(f"Created finance loan for Emergency Loan PRF {prf.id}: Amount={emergency_loan.amount}, Monthly Deduction={monthly_deduction}")
                
            except EmergencyLoan.DoesNotExist:
                print(f"Warning: Emergency Loan details not found for PRF {prf.id}")
            except Exception as e:
                print(f"Error creating finance loan for Emergency Loan PRF {prf.id}: {str(e)}")
        
        # Notification
        Notification.objects.create(
            title=f"PRF Request {action.title()}",
            message=f"Your PRF request ({prf.get_prf_type_display()}) has been {action}.",
            notification_type=action,
            sender=request.user,
            recipient=prf.employee,
            module='prf',
        )
        return JsonResponse({'success': True, 'message': f'PRF request {action}.'})
    except PRFRequest.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'PRF not found.'})

@login_required
def get_admin_prf_detail(request, prf_id):
    if not request.user.hr_admin:
        return JsonResponse({'success': False, 'message': 'Unauthorized'})
    
    prf = get_object_or_404(PRFRequest, id=prf_id)
    
    data = {
        'id': prf.id,
        'employee': prf.employee.username,
        'employee_firstname': prf.employee.firstname,
        'employee_lastname': prf.employee.lastname,
        'employee_idnumber': prf.employee.idnumber,
        'employee_name': f"{prf.employee.firstname} {prf.employee.lastname}",
        'prf_category': prf.get_prf_category_display(),
        'prf_type': prf.get_prf_type_display(),
        'prf_type_value': prf.prf_type,
        'purpose': prf.purpose,
        'control_number': prf.control_number or 'N/A',
        'prf_control_number': prf.prf_control_number or 'N/A',
        'status': prf.status,
        'status_display': prf.get_status_display(),
        'admin_remarks': prf.admin_remarks or '',
        'created_at': prf.created_at.strftime('%B %d, %Y at %I:%M %p'),
        'updated_at': prf.updated_at.strftime('%B %d, %Y at %I:%M %p'),
    }
    
    # Add Emergency Loan details if this is an emergency loan PRF
    if prf.prf_type == 'emergency_loan':
        try:
            emergency_loan = prf.emergency_loan
            data['emergency_loan'] = {
                'amount': str(emergency_loan.amount),
                'number_of_cutoff': emergency_loan.number_of_cutoff,
                'starting_date': emergency_loan.starting_date.strftime('%B %d, %Y') if emergency_loan.starting_date else '',
            }
        except Exception as e:
            print(f"Error fetching emergency loan: {e}")
            data['emergency_loan'] = None
    
    return JsonResponse(data)

@login_required
def export_prfs(request):
    if not request.user.hr_admin:
        return redirect('admin_prf')
    
    prfs = PRFRequest.objects.all().select_related('employee', 'processed_by')
    
    prf_type = request.GET.get('prf_type')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if prf_type:
        prfs = prfs.filter(prf_type=prf_type)
    if start_date:
        prfs = prfs.filter(created_at__date__gte=start_date)
    if end_date:
        prfs = prfs.filter(created_at__date__lte=end_date)
    
    wb = Workbook()
    
    # Sheet 1: PRF Requests
    ws1 = wb.active
    ws1.title = "PRF Requests"
    
    # Define colors for status
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')
    orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    
    # Headers for Sheet 1
    headers = [
        'Date Requested', 'PRF Number', 'ID Number', 'Employee Name', 'PRF Category', 
        'PRF Type', 'Control Number', 'Purpose of Request', 'Status', 'Remarks'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Data for Sheet 1
    for row, prf in enumerate(prfs, 2):
        ws1.cell(row=row, column=1, value=prf.created_at.strftime('%Y-%m-%d'))
        ws1.cell(row=row, column=2, value=prf.prf_control_number or 'N/A')
        ws1.cell(row=row, column=3, value=prf.employee.idnumber or 'N/A')
        ws1.cell(row=row, column=4, value=f"{prf.employee.firstname} {prf.employee.lastname}")
        ws1.cell(row=row, column=5, value=prf.get_prf_category_display())
        ws1.cell(row=row, column=6, value=prf.get_prf_type_display())
        ws1.cell(row=row, column=7, value=prf.control_number or 'N/A')
        ws1.cell(row=row, column=8, value=prf.purpose)
        
        # Status cell with color coding
        status_cell = ws1.cell(row=row, column=9, value=prf.get_status_display())
        if prf.status == 'approved':
            status_cell.fill = green_fill
        elif prf.status == 'disapproved':
            status_cell.fill = red_fill
        elif prf.status == 'pending':
            status_cell.fill = yellow_fill
        elif prf.status == 'cancelled':
            status_cell.fill = orange_fill
        
        ws1.cell(row=row, column=10, value=prf.admin_remarks or 'No remarks')
        
        # Add borders to all cells
        for col in range(1, 11):
            ws1.cell(row=row, column=col).border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # Auto-adjust column widths
    for column in ws1.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws1.column_dimensions[column_letter].width = adjusted_width
    
    # Sheet 2: PRF Overview
    ws2 = wb.create_sheet("PRF Overview")
    
    # Summary statistics
    total_requests = prfs.count()
    approved_count = prfs.filter(status='approved').count()
    disapproved_count = prfs.filter(status='disapproved').count()
    pending_count = prfs.filter(status='pending').count()
    cancelled_count = prfs.filter(status='cancelled').count()
    
    # PRF Type summary
    prf_type_summary = prfs.values('prf_type').annotate(count=Count('id')).order_by('-count')
    
    # PRF Category summary
    prf_category_summary = prfs.values('prf_category').annotate(count=Count('id')).order_by('-count')
    
    # Write summary data
    ws2.cell(row=1, column=1, value="PRF Requests Summary Report").font = Font(bold=True, size=16)
    ws2.cell(row=2, column=1, value=f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Overall statistics
    ws2.cell(row=4, column=1, value="Overall Statistics").font = Font(bold=True, size=14)
    ws2.cell(row=5, column=1, value="Total Requests:").font = Font(bold=True)
    ws2.cell(row=5, column=2, value=total_requests)
    ws2.cell(row=6, column=1, value="Approved:").font = Font(bold=True)
    ws2.cell(row=6, column=2, value=approved_count)
    ws2.cell(row=7, column=1, value="Disapproved:").font = Font(bold=True)
    ws2.cell(row=7, column=2, value=disapproved_count)
    ws2.cell(row=8, column=1, value="Pending:").font = Font(bold=True)
    ws2.cell(row=8, column=2, value=pending_count)
    ws2.cell(row=9, column=1, value="Cancelled:").font = Font(bold=True)
    ws2.cell(row=9, column=2, value=cancelled_count)
    
    # PRF Type breakdown
    ws2.cell(row=11, column=1, value="PRF Type Breakdown").font = Font(bold=True, size=14)
    ws2.cell(row=12, column=1, value="PRF Type").font = Font(bold=True)
    ws2.cell(row=12, column=2, value="Count").font = Font(bold=True)
    
    for i, item in enumerate(prf_type_summary, 13):
        ws2.cell(row=i, column=1, value=dict(PRFRequest.PRF_TYPES)[item['prf_type']])
        ws2.cell(row=i, column=2, value=item['count'])
    
    # PRF Category breakdown
    start_row = 13 + len(prf_type_summary)
    ws2.cell(row=start_row, column=1, value="PRF Category Breakdown").font = Font(bold=True, size=14)
    ws2.cell(row=start_row + 1, column=1, value="Category").font = Font(bold=True)
    ws2.cell(row=start_row + 1, column=2, value="Count").font = Font(bold=True)
    
    for i, item in enumerate(prf_category_summary, start_row + 2):
        ws2.cell(row=i, column=1, value=dict(PRFRequest.PRF_CATEGORIES)[item['prf_category']])
        ws2.cell(row=i, column=2, value=item['count'])
    
    # Create charts
    chart_start_row = start_row + 2 + len(prf_category_summary) + 3
    
    # Status Distribution Pie Chart
    if total_requests > 0:
        ws2.cell(row=chart_start_row, column=1, value="Status Distribution").font = Font(bold=True, size=14)
        
        # Create data for pie chart
        chart_data_start = chart_start_row + 2
        ws2.cell(row=chart_data_start, column=1, value="Status")
        ws2.cell(row=chart_data_start, column=2, value="Count")
        
        status_data = [
            ("Approved", approved_count),
            ("Disapproved", disapproved_count),
            ("Pending", pending_count),
            ("Cancelled", cancelled_count)
        ]
        
        for i, (status, count) in enumerate(status_data, chart_data_start + 1):
            ws2.cell(row=i, column=1, value=status)
            ws2.cell(row=i, column=2, value=count)
        
        # Create pie chart
        pie = PieChart()
        pie.title = "PRF Status Distribution"
        data = Reference(ws2, min_col=2, min_row=chart_data_start, max_row=chart_data_start + len(status_data))
        titles = Reference(ws2, min_col=1, min_row=chart_data_start + 1, max_row=chart_data_start + len(status_data))
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(titles)
        pie.height = 15
        pie.width = 20
        
        ws2.add_chart(pie, f"D{chart_start_row + 2}")
    
    # PRF Type Bar Chart
    if len(prf_type_summary) > 0:
        chart_start_row += 20
        ws2.cell(row=chart_start_row, column=1, value="PRF Type Distribution").font = Font(bold=True, size=14)
        
        # Create data for bar chart
        chart_data_start = chart_start_row + 2
        ws2.cell(row=chart_data_start, column=1, value="PRF Type")
        ws2.cell(row=chart_data_start, column=2, value="Count")
        
        for i, item in enumerate(prf_type_summary, chart_data_start + 1):
            ws2.cell(row=i, column=1, value=dict(PRFRequest.PRF_TYPES)[item['prf_type']])
            ws2.cell(row=i, column=2, value=item['count'])
        
        # Create bar chart
        bar = BarChart()
        bar.title = "PRF Type Distribution"
        bar.style = 10
        bar.x_axis.title = "PRF Type"
        bar.y_axis.title = "Count"
        
        data = Reference(ws2, min_col=2, min_row=chart_data_start, max_row=chart_data_start + len(prf_type_summary))
        cats = Reference(ws2, min_col=1, min_row=chart_data_start + 1, max_row=chart_data_start + len(prf_type_summary))
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.height = 15
        bar.width = 25
        
        ws2.add_chart(bar, f"D{chart_start_row + 2}")
    
    # Auto-adjust column widths for sheet 2
    for column in ws2.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws2.column_dimensions[column_letter].width = adjusted_width

    # Sheet 3: Emergency Loan (if Emergency Loan PRF type is selected)
    if prf_type == 'emergency_loan':
        emergency_prfs = prfs.filter(prf_type='emergency_loan')
        emergency_loans = EmergencyLoan.objects.filter(prf_request__in=emergency_prfs).select_related('prf_request__employee')
        
        if emergency_loans.exists():
            ws3 = wb.create_sheet("Emergency Loan")
            
            # Header
            ws3.merge_cells('A1:L1')
            header_cell = ws3.cell(row=1, column=1, value="RYONAN ELECTRIC PHILIPPINES")
            header_cell.font = Font(bold=True, size=16)
            header_cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Subheader
            ws3.merge_cells('A2:L2')
            date_range = f"Emergency Loan Summary for ({start_date if start_date else 'Start'} to {end_date if end_date else 'End'})"
            subheader_cell = ws3.cell(row=2, column=1, value=date_range)
            subheader_cell.font = Font(italic=True, size=12)
            subheader_cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Headers for Emergency Loan sheet
            emergency_headers = [
                'PRF Number', 'Date', 'Id Number', 'Employee Name', 'Position', 
                'Department', 'Purpose', 'Status', 'EL Control Number', 
                'Emergency Loan Amount', 'Terms of Deduction', 'Start Deduction', 
                'Deduction Per Cut-Off', 'Account Number'
            ]
            
            # Yellow highlight fill for headers
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            
            # Create headers with bold and yellow highlight
            for col, header in enumerate(emergency_headers, 1):
                cell = ws3.cell(row=4, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = yellow_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # Data for Emergency Loan sheet
            for row, loan in enumerate(emergency_loans, 5):
                prf = loan.prf_request
                employee = prf.employee
                
                ws3.cell(row=row, column=1, value=prf.prf_control_number or 'N/A')
                ws3.cell(row=row, column=2, value=prf.created_at.strftime('%Y-%m-%d'))
                ws3.cell(row=row, column=3, value=employee.idnumber or 'N/A')
                ws3.cell(row=row, column=4, value=f"{employee.firstname} {employee.lastname}")
                ws3.cell(row=row, column=5, value=employee.employment_info.position.position or 'N/A')
                ws3.cell(row=row, column=6, value=employee.employment_info.department.department_name or 'N/A')
                ws3.cell(row=row, column=7, value=prf.purpose)
                ws3.cell(row=row, column=8, value=prf.get_status_display())
                ws3.cell(row=row, column=9, value=prf.control_number or 'N/A')
                ws3.cell(row=row, column=10, value=f"₱{loan.amount:,}")
                ws3.cell(row=row, column=11, value=f"{loan.number_of_cutoff} Cut-offs")
                ws3.cell(row=row, column=12, value=loan.starting_date.strftime('%Y-%m-%d') if loan.starting_date else 'N/A')
                ws3.cell(row=row, column=13, value=f"₱{loan.deduction_per_cutoff:,.2f}")
                ws3.cell(row=row, column=14, value=employee.employment_info.bank_account or 'N/A')

                # Add borders and wrap text to all cells
                for col in range(1, 15):
                    cell = ws3.cell(row=row, column=col)
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.alignment = Alignment(wrap_text=True)
            
            # Auto-adjust column widths for emergency loan sheet
            for col_num in range(1, 15):
                column_letter = chr(64 + col_num)  # A, B, C, etc.
                max_length = 0
                for row in ws3.iter_rows(min_col=col_num, max_col=col_num):
                    for cell in row:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                adjusted_width = min(max_length + 2, 30)  # Smaller max width for better fit
                ws3.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=prf_requests_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response

@require_POST
@login_required
def bulk_delete_prfs(request):
    if not request.user.hr_admin:
        return JsonResponse({'success': False, 'message': 'Unauthorized'})
    
    if request.method == 'POST':
        data = json.loads(request.body)
        prf_ids = data.get('prf_ids', [])
        
        if not prf_ids:
            return JsonResponse({'success': False, 'message': 'No PRFs selected'})
        
        deleted_count = PRFRequest.objects.filter(id__in=prf_ids).delete()[0]
        
        return JsonResponse({
            'success': True,
            'message': f'{deleted_count} PRF(s) deleted successfully!'
        })
    
    return JsonResponse({'success': False, 'message': 'Invalid request method.'})

def admin_prf_detail(request, pk):
    if not request.headers.get('x-requested-with') == 'XMLHttpRequest':
        raise Http404
    from .models import PRFRequest
    try:
        prf = PRFRequest.objects.select_related('employee').get(pk=pk)
        data = {
            'id': prf.id,
            'employee_username': prf.employee.username,
            'employee_firstname': prf.employee.firstname,
            'employee_lastname': prf.employee.lastname,
            'employee_idnumber': prf.employee.idnumber,
            'prf_type': prf.get_prf_type_display(),
            'prf_type_value': prf.prf_type,  # Add the raw prf_type value
            'prf_category': prf.get_prf_category_display(),
            'purpose': prf.purpose,
            'status': prf.status,
            'status_display': prf.get_status_display(),
            'created_at': prf.created_at.strftime('%b %d, %Y'),
            'updated_at': prf.updated_at.strftime('%b %d, %Y') if hasattr(prf, 'updated_at') and prf.updated_at else '',
            'control_number': getattr(prf, 'control_number', ''),
            'prf_control_number': getattr(prf, 'prf_control_number', '') or 'N/A',
            'admin_remarks': getattr(prf, 'admin_remarks', '') or '',
        }
        
        # Add Emergency Loan details if this is an emergency loan PRF
        if prf.prf_type == 'emergency_loan':
            try:
                emergency_loan = prf.emergency_loan
                data['emergency_loan'] = {
                    'amount': str(emergency_loan.amount),
                    'number_of_cutoff': emergency_loan.number_of_cutoff,
                    'starting_date': emergency_loan.starting_date.strftime('%B %d, %Y') if emergency_loan.starting_date else '',
                }
            except Exception as e:
                print(f"Error fetching emergency loan: {e}")
                data['emergency_loan'] = None
        
        return JsonResponse(data)
    except PRFRequest.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)

@login_required
@require_POST
def cancel_prf_request(request):
    prf_id = request.POST.get('prf_id')
    if not prf_id:
        return JsonResponse({'success': False, 'message': 'Invalid request.'})
    try:
        prf = PRFRequest.objects.get(pk=prf_id, employee=request.user)
        if prf.status != 'pending':
            return JsonResponse({'success': False, 'message': 'Only pending requests can be cancelled.'})
        prf.status = 'cancelled'
        prf.save()
        # Optionally, create a notification here
        return JsonResponse({'success': True, 'message': 'PRF request cancelled.'})
    except PRFRequest.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'PRF not found.'})


# Emergency Loan Views
@login_required
def get_cutoff_choices(request):
    """Get cut-off choices based on selected amount"""
    amount = request.GET.get('amount')
    if not amount:
        return JsonResponse({'choices': []})
    
    try:
        amount = int(amount)
        choices = EmergencyLoan.get_cutoff_choices(amount)
        choices_list = [{'value': choice[0], 'text': choice[1]} for choice in choices]
        return JsonResponse({'choices': choices_list})
    except ValueError:
        return JsonResponse({'choices': []})

@login_required
def check_existing_emergency_loan(request):
    """Check if user has an existing Emergency Loan"""
    try:
        # Get or create Emergency Loan type
        emergency_loan_type, created = LoanType.objects.get_or_create(
            loan_type="Emergency Loan",
            defaults={'description': 'Emergency loan for employees'}
        )
        
        # Check for existing active Emergency Loan
        existing_loan = Loan.objects.filter(
            employee=request.user,
            loan_type=emergency_loan_type,
            is_active=True
        ).first()
        
        # Check if existing loan balance is more than one monthly deduction
        if existing_loan and existing_loan.current_balance > existing_loan.monthly_deduction:
            return JsonResponse({
                'has_existing_loan': True,
                'current_balance': float(existing_loan.current_balance),
                'principal_amount': float(existing_loan.principal_amount),
                'monthly_deduction': float(existing_loan.monthly_deduction)
            })
        else:
            return JsonResponse({'has_existing_loan': False})
            
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@require_POST
def submit_emergency_loan(request):
    """Submit complete Emergency Loan (PRF + EmergencyLoan + finance.Loan)"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST method required.'})
    
    try:
        # Get PRF data from session or POST
        prf_data = request.POST.get('prf_data')
        if prf_data:
            import json
            prf_data = json.loads(prf_data)
        else:
            # Get from POST if not in session
            prf_data = {
                'prf_category': request.POST.get('prf_category', 'hr_payroll'),
                'prf_type': 'emergency_loan',
                'purpose': request.POST.get('purpose', 'Emergency Loan Request')
            }
        
        # Create Emergency Loan form with POST data
        form = EmergencyLoanForm(request.POST, user=request.user)
        
        if form.is_valid():
            # Double-check for existing Emergency Loan
            emergency_loan_type, created = LoanType.objects.get_or_create(
                loan_type="Emergency Loan",
                defaults={'description': 'Emergency loan for employees'}
            )
            
            existing_loan = Loan.objects.filter(
                employee=request.user,
                loan_type=emergency_loan_type,
                is_active=True
            ).first()
            
            # Check if existing loan balance is more than one monthly deduction
            if existing_loan and existing_loan.current_balance > existing_loan.monthly_deduction:
                return JsonResponse({
                    'success': False,
                    'message': 'existing_loan_error',
                    'current_balance': float(existing_loan.current_balance)
                })
            
            # Generate control number
            control_number = EmergencyLoan.generate_control_number()
            
            # Create PRF Request
            prf_request = PRFRequest.objects.create(
                employee=request.user,
                prf_category=prf_data['prf_category'],
                prf_type=prf_data['prf_type'],
                purpose=prf_data['purpose'],
                control_number=control_number
            )
            
            # Save Emergency Loan details
            emergency_loan = form.save(commit=False)
            emergency_loan.prf_request = prf_request
            emergency_loan.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Emergency Loan request submitted successfully!',
                'control_number': control_number,
                'deduction_per_cutoff': emergency_loan.deduction_per_cutoff,
                'formatted_starting_date': emergency_loan.formatted_starting_date,
                'amount': f"₱{int(request.POST.get('amount')):,}"
            })
        else:
            # Return form errors
            errors = []
            for field, error_list in form.errors.items():
                for error in error_list:
                    errors.append(f"{field}: {error}")
            
            return JsonResponse({
                'success': False,
                'message': 'Form validation errors',
                'errors': errors
            })
            
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'An error occurred: {str(e)}'})

@login_required
def get_emergency_loan_details(request, prf_id):
    try:
        prf_request = PRFRequest.objects.get(id=prf_id, employee=request.user)
        
        if prf_request.prf_type != 'Emergency Loan':
            return JsonResponse({'error': 'Not an Emergency Loan PRF'}, status=400)
        
        if not hasattr(prf_request, 'Emergency Loan'):
            return JsonResponse({'error': 'Emergency Loan details not found'}, status=404)
        
        emergency_loan = prf_request.emergency_loan
        
        return JsonResponse({
            'amount': emergency_loan.amount,
            'number_of_cutoff': emergency_loan.number_of_cutoff,
            'starting_date': emergency_loan.starting_date.isoformat(),
            'deduction_per_cutoff': emergency_loan.deduction_per_cutoff,
            'formatted_starting_date': emergency_loan.formatted_starting_date,
            'employee_full_name': emergency_loan.employee_full_name
        })
        
    except PRFRequest.DoesNotExist:
        return JsonResponse({'error': 'PRF not found'}, status=404)