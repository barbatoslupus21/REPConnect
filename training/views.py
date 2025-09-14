from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.utils import timezone
from django.db.models import Count, Q
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator
from django.template.loader import render_to_string
from django.db import transaction
from datetime import timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from .models import Training, TrainingEvaluation, EvaluationRouting
from .forms import (
    TrainingForm, TrainingEvaluationForm, SupervisorAssessmentForm, EvaluationRoutingForm
)
from userlogin.models import EmployeeLogin
from userprofile.models import EmploymentInformation
from django.db import transaction
from notification.models import Notification


@login_required
def training_dashboard(request):
    if request.user.hr_admin:
        return redirect('admin_training_evaluation')
    else:
        return redirect('user_training_evaluation')
    
@login_required
def training_user_view(request):
    user_trainings = Training.objects.filter(participants=request.user)
    
    trainings_data = []
    total_trainings = user_trainings.count()
    
    completed_trainings = TrainingEvaluation.objects.filter(
        participant=request.user,
        submitted_at__isnull=False,
        training__participants=request.user
    ).count()

    pending_trainings = TrainingEvaluation.objects.filter(
        participant=request.user,
        submitted_at__isnull=True,
        training__participants=request.user
    ).count()

    pending_trainings_list = []
    non_pending_trainings_list = []
    
    for training in user_trainings:
        try:
            evaluation = TrainingEvaluation.objects.get(
                training=training, 
                participant=request.user, 
                submitted_at__isnull=False
            )
            status = evaluation.status
            has_draft = False
            non_pending_trainings_list.append((training, status, has_draft))
        except TrainingEvaluation.DoesNotExist:
            try:
                evaluation = TrainingEvaluation.objects.get(
                    training=training, 
                    participant=request.user, 
                    submitted_at__isnull=True
                )
                status = 'pending'
                has_draft = True
                pending_trainings_list.append((training, status, has_draft))
            except TrainingEvaluation.DoesNotExist:
                status = 'pending'
                has_draft = False
                pending_trainings_list.append((training, status, has_draft))
    
    pending_trainings_list.sort(key=lambda x: x[0].training_date, reverse=True)
    
    non_pending_trainings_list.sort(key=lambda x: x[0].training_date, reverse=True)
    
    all_trainings_ordered = pending_trainings_list + non_pending_trainings_list
    
    for training, status, has_draft in all_trainings_ordered:
        trainings_data.append({
            'training': training,
            'status': status,
            'has_draft': has_draft
        })
    
    pending_supervisor_assessments = []
    pending_manager_reviews = []
    pending_routing_count = 0
    is_approver = False
    user_position_level = None
    approval_tab_label = "My Approvals"
    
    if hasattr(request.user, 'employment_info') and request.user.employment_info.position:
        user_position_level = int(request.user.employment_info.position.level)
        approval_tab_label = "For Evaluation"

        pending_supervisor_assessments = EvaluationRouting.objects.filter(
            approver=request.user
        ).order_by('is_completed', 'id')

            
        pending_routing_count = EvaluationRouting.objects.filter(
            approver=request.user,
            is_completed=False, completed_at__isnull=True
        ).count()

    if not is_approver:
        supervisor_total_count = TrainingEvaluation.objects.filter(
            routing_steps__approver=request.user,
            is_submitted=True
        ).exclude(status='participant_reviewed').count()

        if supervisor_total_count > 0:
            is_approver = True
        
        manager_total_count = EvaluationRouting.objects.filter(
            approver=request.user
        ).count()
        if manager_total_count > 0:
            is_approver = True

    context = {
        'trainings_data': trainings_data,
        'total_trainings': total_trainings,
        'completed_trainings': completed_trainings,
        'pending_trainings': pending_trainings,
        'user_trainings': user_trainings,
        'is_approver': is_approver,
        'user_has_approver': hasattr(request.user, 'employment_info') and request.user.employment_info and request.user.employment_info.approver is not None,
        'user_position_level': user_position_level,
        'approval_tab_label': approval_tab_label,
        'pending_supervisor_assessments': pending_supervisor_assessments,
        'pending_manager_reviews': pending_manager_reviews,
        'pending_routing_count': pending_routing_count,
    }
    
    return render(request, 'training/user_training.html', context)

@login_required
def get_training_evaluation(request, training_id):
    training = get_object_or_404(Training, id=training_id, participants=request.user)
    
    # Get or create evaluation
    evaluation, created = TrainingEvaluation.objects.get_or_create(
        training=training,
        participant=request.user,
        defaults={'status': 'draft'}
    )
    
    form = TrainingEvaluationForm(instance=evaluation)
    
    return JsonResponse({
        'success': True,
        'training': {
            'id': training.id,
            'title': training.title,
            'objective': training.objective,
            'speaker': training.speaker,
            'training_date': training.training_date.strftime('%Y-%m-%d'),
        },
        'evaluation_id': evaluation.id,
        'form_html': render(request, 'training/evaluation_form.html', {
            'form': form,
            'training': training,
            'evaluation': evaluation
        }).content.decode()
    })

@login_required
@require_http_methods(["POST"])
def submit_training_evaluation(request, evaluation_id):
    evaluation = get_object_or_404(
        TrainingEvaluation, 
        id=evaluation_id, 
        participant=request.user
    )
    
    form = TrainingEvaluationForm(request.POST, instance=evaluation)
    
    if form.is_valid():
        with transaction.atomic():
            evaluation = form.save(commit=False)
            evaluation.is_submitted = True
            evaluation.submitted_at = timezone.now()
            evaluation.status = 'supervisor_reviewed'  # Status should be submitted when participant submits
            evaluation.save()
            
            # Create notification for the approver/supervisor
            try:
                if hasattr(request.user, 'employment_info') and request.user.employment_info.approver:
                    approver = request.user.employment_info.approver
                    
                    # Create notification
                    Notification.objects.create(
                        title=f"Training Evaluation Submitted",
                        message=f"{request.user.firstname} {request.user.lastname} has submitted their evaluation for training: {evaluation.training.title}",
                        notification_type='approval',
                        sender=request.user,
                        recipient=approver,
                        module='training'
                    )
                    
                    # Create routing sequence for approval workflow
                    # Step 1: Supervisor approval
                    EvaluationRouting.objects.create(
                        evaluation=evaluation,
                        approver=approver,
                        sequence=1
                    )
                    
                    print(f"Created routing step 1 for supervisor: {approver.firstname} {approver.lastname}")
                    
            except Exception as e:
                # Log the error but don't fail the submission
                print(f"Error creating notification/routing: {e}")
        
        return JsonResponse({
            'success': True,
            'message': 'Training evaluation submitted successfully! Your supervisor will be notified.'
        })
    else:
        return JsonResponse({
            'success': False,
            'errors': form.errors
        })

@login_required
def view_submitted_evaluation(request, training_id):
    training = get_object_or_404(Training, id=training_id, participants=request.user)
    
    try:
        evaluation = TrainingEvaluation.objects.get(
            training=training, 
            participant=request.user,
            submitted_at__isnull=False
        )
    except TrainingEvaluation.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'No submitted evaluation found for this training.'
        })
    
    # Check if this is participant review mode (when status is supervisor_reviewed)
    participant_review_mode = request.GET.get('participant_review') == 'true' and evaluation.status == 'supervisor_reviewed'
    
    supervisor_assessment = None
    # Check if supervisor has filled the assessment fields
    if (evaluation.result_and_impact or 
        evaluation.recommendations or 
        evaluation.overall_assessment):
        supervisor_assessment = {
            'result_and_impact': evaluation.result_and_impact,
            'recommendations': evaluation.recommendations,
            'overall_assessment': evaluation.overall_assessment
        }
    
    from django.template.loader import render_to_string
    form = TrainingEvaluationForm(instance=evaluation)
    
    context_data = {
        'form': form,
        'training': training,
        'evaluation': evaluation,
        'supervisor_assessment': supervisor_assessment,
        'participant_review_mode': participant_review_mode,
        'read_only': True
    }
    
    return JsonResponse({
        'success': True,
        'evaluation_id': evaluation.id,
        'training': {
            'id': training.id,
            'title': training.title,
            'speaker': training.speaker,
            'training_date': training.training_date.strftime('%Y-%m-%d'),
            'objective': training.objective
        },
        'is_submitted': evaluation.is_submitted,
        'submitted_at': evaluation.submitted_at.strftime('%Y-%m-%d %H:%M') if evaluation.submitted_at else None,
        'status': evaluation.status,
        'read_only': True,
        'has_supervisor_assessment': supervisor_assessment is not None,
        'html': render_to_string('training/view_submitted_evaluation.html', context_data, request=request)
    })

@login_required
def get_subordinate_evaluation(request, evaluation_id):
    evaluation = get_object_or_404(
        TrainingEvaluation,
        id=evaluation_id,
        routing_steps__approver=request.user,
        is_submitted=True
    )
    
    is_read_only = evaluation.status != 'supervisor_reviewed'
    
    form = SupervisorAssessmentForm(instance=evaluation)
    
    context_data = {
        'form': form,
        'evaluation': evaluation,
        'assessment': evaluation,
        'is_read_only': is_read_only
    }
    
    return JsonResponse({
        'success': True,
        'html': render_to_string('training/supervisor_assessment_content.html', context_data, request=request)
    })

@login_required
@require_http_methods(["POST"])
def submit_supervisor_assessment(request, assessment_id):
    evaluation = get_object_or_404(
        TrainingEvaluation,
        id=assessment_id,
        participant__employment_info__approver=request.user,
        is_submitted=True
    )
    
    form = SupervisorAssessmentForm(request.POST, instance=evaluation)
    
    if form.is_valid():
        with transaction.atomic():
            evaluation = form.save()
            
            evaluation.status = 'participant_reviewed'
            evaluation.save()
            
            # Create or update supervisor routing step
            supervisor_routing, created = EvaluationRouting.objects.get_or_create(
                evaluation=evaluation,
                approver=request.user,
                sequence=1,
                defaults={
                    'is_completed': True,
                    'completed_at': timezone.now()
                }
            )
            if not created:
                supervisor_routing.is_completed = True
                supervisor_routing.completed_at = timezone.now()
                supervisor_routing.save()
            
            print(f"Supervisor routing step completed for {request.user.firstname} {request.user.lastname}")
        
        try:
            Notification.objects.create(
                title="Training Evaluation - Supervisor Assessment Completed",
                message=f"Your supervisor {request.user.firstname} {request.user.lastname} has completed their assessment for your training: {evaluation.training.title}.",
                notification_type='info',
                sender=request.user,
                recipient=evaluation.participant,
                module='training'
            )
            
            supervisor_manager = None
            if (hasattr(request.user, 'employment_info') and 
                request.user.employment_info and 
                request.user.employment_info.approver):
                supervisor_manager = request.user.employment_info.approver
            
            if supervisor_manager:
                routing_step, created = EvaluationRouting.objects.get_or_create(
                    evaluation=evaluation,
                    approver=supervisor_manager,
                    defaults={
                        'sequence': 2,
                        'is_completed': False
                    }
                )

                Notification.objects.create(
                    title="Training Evaluation - Manager Review Required",
                    message=f"Supervisor {request.user.firstname} {request.user.lastname} has completed their assessment for {evaluation.participant.firstname} {evaluation.participant.lastname}'s training: {evaluation.training.title}. Your review and approval is required.",
                    notification_type='approval',
                    sender=request.user,
                    recipient=supervisor_manager,
                    module='training'
                )

            else:
                evaluation.status = 'approved'
                evaluation.approver_by_manager = True
                evaluation.save()
                
        except Exception as e:
            print(f"Error creating routing: {e}")
        
        return JsonResponse({
            'success': True,
            'message': 'Supervisor assessment submitted successfully!'
        })
    else:
        return JsonResponse({
            'success': False,
            'errors': form.errors
        })

@login_required
def get_manager_review_evaluation(request, evaluation_id):
    # Verify user is a manager (level 3)
    if not (hasattr(request.user, 'employment_info') and 
            hasattr(request.user.employment_info, 'position') and
            request.user.employment_info.position and 
            request.user.employment_info.position.level == 3):
        return JsonResponse({'success': False, 'error': 'Unauthorized'})
    
    evaluation = get_object_or_404(
        TrainingEvaluation,
        id=evaluation_id,
        status='supervisor_reviewed'
    )
    
    # Get the routing step for this manager
    routing_step = get_object_or_404(
        EvaluationRouting,
        evaluation=evaluation,
        approver=request.user,
        is_completed=False
    )
    
    # Create a simple form for manager approval (could be expanded)
    from django import forms
    class ManagerApprovalForm(forms.Form):
        decision = forms.ChoiceField(
            choices=[('approve', 'Approve'), ('reject', 'Reject')],
            widget=forms.RadioSelect()
        )
        comments = forms.CharField(
            required=False,
            widget=forms.Textarea(attrs={'rows': 3, 'placeholder': 'Optional comments'})
        )
    
    form = ManagerApprovalForm()
    
    return JsonResponse({
        'success': True,
        'evaluation': {
            'id': evaluation.id,
            'participant_name': f"{evaluation.participant.firstname} {evaluation.participant.lastname}",
            'training_title': evaluation.training.title,
            'submitted_at': evaluation.submitted_at.strftime('%Y-%m-%d %H:%M'),
            # Include all evaluation fields for display
            'content_related_to_job': evaluation.content_related_to_job,
            'content_explained_clearly': evaluation.content_explained_clearly,
            'content_suitable_for_topic': evaluation.content_suitable_for_topic,
            'program_clear_goals': evaluation.program_clear_goals,
            'program_met_goals': evaluation.program_met_goals,
            'program_easy_to_follow': evaluation.program_easy_to_follow,
            'program_easy_to_understand': evaluation.program_easy_to_understand,
            'speaker_knowledge': evaluation.speaker_knowledge,
            'speaker_clear_communication': evaluation.speaker_clear_communication,
            'speaker_answered_questions': evaluation.speaker_answered_questions,
            'training_organization': evaluation.training_organization,
            'suitable_facilities': evaluation.suitable_facilities,
            'helpful_materials': evaluation.helpful_materials,
            'most_interesting_topic': evaluation.most_interesting_topic,
            'feedback_recommendations': evaluation.feedback_recommendations,
            'future_training_topics': evaluation.future_training_topics,
            'new_things_learned_work': evaluation.new_things_learned_work,
            'how_apply_at_work': evaluation.how_apply_at_work,
            'target_implementation_date': evaluation.target_implementation_date.strftime('%Y-%m-%d') if evaluation.target_implementation_date else '',
            'actual_implementation_date': evaluation.actual_implementation_date.strftime('%Y-%m-%d') if evaluation.actual_implementation_date else '',
            'new_things_learned_personal': evaluation.new_things_learned_personal,
            'how_apply_daily_life': evaluation.how_apply_daily_life,
        },
        'supervisor_assessment': {
            'result_and_impact': evaluation.result_and_impact,
            'recommendations': evaluation.recommendations,
            'overall_assessment': evaluation.overall_assessment,
        },
        'routing_id': routing_step.id,
        'form_html': render(request, 'training/manager_review_form.html', {
            'form': form,
            'evaluation': evaluation,
            'routing': routing_step
        }).content.decode()
    })

@login_required
@require_http_methods(["POST"])
def submit_manager_review(request, routing_id):
    try:
        if routing_id:
            try:
                routing_step = EvaluationRouting.objects.get(
                    id=routing_id,
                    approver=request.user,
                    is_completed=False
                )
                evaluation = routing_step.evaluation
            except EvaluationRouting.DoesNotExist:
                evaluation = get_object_or_404(TrainingEvaluation, id=routing_id)
                routing_step = get_object_or_404(
                    EvaluationRouting,
                    evaluation=evaluation,
                    approver=request.user,
                    is_completed=False
                )
        
        import json
        data = json.loads(request.body) if request.body else request.POST
        decision = data.get('decision', 'reject')
        
        with transaction.atomic():
            if decision == 'approve':
                evaluation.approver_by_manager = True
                evaluation.save()
                
                routing_step.is_completed = True
                routing_step.completed_at = timezone.now()
                routing_step.save()
                
                if routing_step.sequence == 2 or routing_step.sequence == 3:
                    evaluation.status = 'approved'
                    evaluation.save()
                
                Notification.objects.create(
                    recipient=evaluation.participant,
                    sender=request.user,
                    title="Training Evaluation Approved",
                    message=f"Your training evaluation for {evaluation.training.title} has been approved by {request.user}.",
                    notification_type="approved",
                    module="training"
                )
                
            else:
                routing_step.delete()
                
                supervisor_routing = EvaluationRouting.objects.filter(
                    evaluation=evaluation,
                    approver=evaluation.participant.employment_info.approver,
                    is_completed=True
                ).first()
                
                if supervisor_routing:
                    supervisor_routing.is_completed = False
                    supervisor_routing.completed_at = None
                    supervisor_routing.save()
                
                evaluation.status = 'supervisor_reviewed'
                evaluation.save()
                
                if hasattr(evaluation.participant, 'employment_info') and evaluation.participant.employment_info.approver:
                    supervisor = evaluation.participant.employment_info.approver
                    Notification.objects.create(
                        recipient=supervisor,
                        sender=request.user,
                        title="Training Evaluation Disapproved",
                        message=f"Training evaluation for {evaluation.participant.firstname} {evaluation.participant.lastname} - {evaluation.training.title} has been disapproved and requires re-evaluation.",
                        notification_type="disapproved",
                        module="training"
                    )
        
        return JsonResponse({
            'success': True,
            'message': f'Evaluation {decision}d successfully!'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Failed to {decision} evaluation: {str(e)}'
        })

@login_required
def training_admin_view(request):
    search = request.GET.get('search', '')
    trainings = Training.objects.all().order_by('-training_date').prefetch_related('participants')
    
    if search:
        trainings = trainings.filter(
            Q(title__icontains=search) |
            Q(speaker__icontains=search) |
            Q(objective__icontains=search)
        )
    
    # Pagination - 10 trainings per page
    from django.core.paginator import Paginator
    paginator = Paginator(trainings, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    trainings_with_progress = []
    for training in page_obj:
        trainings_with_progress.append({
            'training': training,
            'progress_percentage': training.progress_percentage
        })
    
    # For statistics, use all trainings (not paginated)
    all_trainings = Training.objects.all().order_by('-training_date').prefetch_related('participants')
    total_trainings = all_trainings.count()
    
    all_trainings_with_progress = []
    for training in all_trainings:
        all_trainings_with_progress.append({
            'training': training,
            'progress_percentage': training.progress_percentage
        })
    
    active_trainings = sum(1 for item in all_trainings_with_progress if item['progress_percentage'] < 100)
    
    all_evaluations = TrainingEvaluation.objects.all()
    completed_evaluations = all_evaluations.filter(is_submitted=True).count()
    pending_responses = all_evaluations.filter(is_submitted=False).count()
    
    now = timezone.now()
    first_of_this_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    first_of_last_month = (first_of_this_month - timedelta(days=1)).replace(day=1)
    
    trainings_this_month = Training.objects.filter(created_at__gte=first_of_this_month).count()
    trainings_last_month = Training.objects.filter(created_at__gte=first_of_last_month, created_at__lt=first_of_this_month).count()
    trainings_percent = 0
    trainings_positive = True
    if trainings_last_month > 0:
        trainings_percent = round(((trainings_this_month - trainings_last_month) / trainings_last_month) * 100, 1)
        trainings_positive = trainings_percent >= 0
    elif trainings_this_month > 0:
        trainings_percent = 100.0  # 100% increase when going from 0 to any number
        trainings_positive = True
    
    active_this_month = Training.objects.filter(
        created_at__gte=first_of_this_month
    ).exclude(
        id__in=[t['training'].id for t in trainings_with_progress if t['progress_percentage'] >= 100]
    ).count()
    active_last_month = Training.objects.filter(
        created_at__gte=first_of_last_month, 
        created_at__lt=first_of_this_month
    ).exclude(
        id__in=[]
    ).count()
    active_percent = 0
    active_positive = True
    if active_last_month > 0:
        active_percent = round(((active_this_month - active_last_month) / active_last_month) * 100, 1)
        active_positive = active_percent >= 0
    elif active_this_month > 0:
        active_percent = 100.0  # 100% increase when going from 0 to any number
        active_positive = True
    
    completed_this_month = TrainingEvaluation.objects.filter(
        is_submitted=True, 
        submitted_at__gte=first_of_this_month
    ).count()
    completed_last_month = TrainingEvaluation.objects.filter(
        is_submitted=True, 
        submitted_at__gte=first_of_last_month, 
        submitted_at__lt=first_of_this_month
    ).count()
    completed_percent = 0
    completed_positive = True
    if completed_last_month > 0:
        completed_percent = round(((completed_this_month - completed_last_month) / completed_last_month) * 100, 1)
        completed_positive = completed_percent >= 0
    elif completed_this_month > 0:
        completed_percent = 100.0  # 100% increase when going from 0 to any number
        completed_positive = True
    
    pending_this_month = TrainingEvaluation.objects.filter(
        is_submitted=False, 
        created_at__gte=first_of_this_month
    ).count()
    pending_last_month = TrainingEvaluation.objects.filter(
        is_submitted=False, 
        created_at__gte=first_of_last_month, 
        created_at__lt=first_of_this_month
    ).count()
    pending_percent = 0
    pending_positive = False
    if pending_last_month > 0:
        pending_percent = round(((pending_this_month - pending_last_month) / pending_last_month) * 100, 1)
        pending_positive = pending_percent <= 0
    elif pending_this_month > 0:
        pending_percent = 100.0  # 100% increase when going from 0 to any number
        pending_positive = False  # More pending is negative
    
    responses = TrainingEvaluation.objects.filter(
        is_submitted=True
    ).select_related('participant', 'training').order_by('-submitted_at')[:50]
    
    # Calculate response status distribution for pie chart
    all_trainings = Training.objects.all()
    total_participants = 0
    for training in all_trainings:
        total_participants += training.participants.count()
    
    response_status_counts = {
        'submitted': TrainingEvaluation.objects.filter(status='submitted').count(),
        'supervisor_reviewed': TrainingEvaluation.objects.filter(status='supervisor_reviewed').count(),
        'manager_reviewed': TrainingEvaluation.objects.filter(status='manager_reviewed').count(),
        'approved': TrainingEvaluation.objects.filter(status='approved').count(),
        'disapproved': TrainingEvaluation.objects.filter(status='disapproved').count(),
    }
    
    # Calculate unanswered evaluations (participants who haven't submitted)
    total_evaluations = TrainingEvaluation.objects.count()
    unanswered_evaluations = total_participants - total_evaluations
    if unanswered_evaluations > 0:
        response_status_counts['unanswered'] = unanswered_evaluations
    
    # Calculate fiscal year training creation data (May to April)
    from django.db.models import Count
    from django.db.models.functions import TruncMonth
    import calendar
    
    # Determine current fiscal year
    current_date = now.date()
    if current_date.month >= 5:  # May to December
        fiscal_year_start = current_date.replace(month=5, day=1)
        fiscal_year_end = fiscal_year_start.replace(year=fiscal_year_start.year + 1, month=4, day=30)
    else:  # January to April
        fiscal_year_start = current_date.replace(year=current_date.year - 1, month=5, day=1)
        fiscal_year_end = current_date.replace(month=4, day=30)
    
    # Get monthly training creation data for current fiscal year
    monthly_trainings = Training.objects.filter(
        created_at__date__gte=fiscal_year_start,
        created_at__date__lte=fiscal_year_end
    ).annotate(
        month=TruncMonth('created_at')
    ).values('month').annotate(
        count=Count('id')
    ).order_by('month')
    
    # Create month labels and data arrays
    month_labels = []
    month_data = []
    
    # Generate all months in fiscal year
    current_month = fiscal_year_start
    monthly_data_dict = {item['month'].date().replace(day=1): item['count'] for item in monthly_trainings}
    
    while current_month <= fiscal_year_end:
        month_key = current_month.replace(day=1)
        month_labels.append(calendar.month_abbr[current_month.month])
        month_data.append(monthly_data_dict.get(month_key, 0))
        
        # Move to next month
        if current_month.month == 12:
            current_month = current_month.replace(year=current_month.year + 1, month=1)
        else:
            current_month = current_month.replace(month=current_month.month + 1)
    
    context = {
        'total_trainings': total_trainings,
        'active_trainings': active_trainings,
        'completed_evaluations': completed_evaluations,
        'pending_responses': pending_responses,
        'trainings_with_progress': trainings_with_progress,
        'page_obj': page_obj,
        'search': search,
        'responses': responses,
        # Month-over-month percentages
        'trainings_percent': trainings_percent,
        'trainings_positive': trainings_positive,
        'active_percent': active_percent,
        'active_positive': active_positive,
        'completed_percent': completed_percent,
        'completed_positive': completed_positive,
        'pending_percent': pending_percent,
        'pending_positive': pending_positive,
        # Chart data
        'response_status_counts': response_status_counts,
        'month_labels': month_labels,
        'month_data': month_data,
        'fiscal_year_start': fiscal_year_start,
        'fiscal_year_end': fiscal_year_end,
    }
    
    return render(request, 'training/admin_training.html', context)

@login_required
def ajax_search_trainings(request):
    """AJAX endpoint for real-time training search"""
    search = request.GET.get('search', '')
    page_number = request.GET.get('page', 1)
    
    trainings = Training.objects.all().order_by('-training_date').prefetch_related('participants')
    
    if search:
        trainings = trainings.filter(
            Q(title__icontains=search) |
            Q(speaker__icontains=search) |
            Q(objective__icontains=search)
        )
    
    # Pagination - 10 trainings per page
    paginator = Paginator(trainings, 10)
    page_obj = paginator.get_page(page_number)
    
    trainings_with_progress = []
    for training in page_obj:
        trainings_with_progress.append({
            'training': training,
            'progress_percentage': training.progress_percentage
        })
    
    # Render the table content
    from django.template.loader import render_to_string
    table_html = render_to_string('training/partials/training_table_rows.html', {
        'trainings_with_progress': trainings_with_progress,
        'search': search
    })
    
    # Render pagination
    pagination_html = render_to_string('training/partials/training_pagination.html', {
        'page_obj': page_obj,
        'search': search
    })
    
    return JsonResponse({
        'success': True,
        'table_html': table_html,
        'pagination_html': pagination_html,
        'total_results': paginator.count,
        'current_page': page_obj.number,
        'total_pages': paginator.num_pages
    })

@login_required
def get_participants(request):
    if not getattr(request.user, 'hr_admin', False):
        return JsonResponse({'success': False, 'error': 'Unauthorized'})
    
    participants = EmployeeLogin.objects.filter(
        active=True,
        wire_admin=False,
        clinic_admin=False,
        iad_admin=False,
        accounting_admin=False,
        mis_admin=False
    ).select_related('employment_info__position', 'employment_info__department').order_by('firstname', 'lastname')
    
    participants_data = []
    for participant in participants:
        position_name = "Not Assigned"
        if hasattr(participant, 'employment_info') and participant.employment_info and participant.employment_info.position:
            position_name = participant.employment_info.position.position

        department_name = "Not Assigned"
        if hasattr(participant, 'employment_info') and participant.employment_info and participant.employment_info.department:
            department_name = participant.employment_info.department.department_name

        participants_data.append({
            'id': participant.id,
            'idnumber': participant.idnumber or 'N/A',
            'name': participant.full_name,
            'department': department_name
        })
    
    return JsonResponse({
        'success': True,
        'participants': participants_data
    })

@login_required
@require_http_methods(["POST"])
def create_training(request):
    if not getattr(request.user, 'hr_admin', False):
        return JsonResponse({'success': False, 'error': 'Unauthorized'})
    
    form_data = request.POST.copy()
    form_data.pop('participants', None)
    
    form = TrainingForm(form_data)
    
    if form.is_valid():
        try:
            with transaction.atomic():
                training = form.save()
                
                participant_ids = request.POST.getlist('participants')
                if participant_ids:
                    participants = EmployeeLogin.objects.filter(id__in=participant_ids)
                    training.participants.set(participants)
                    
                    # Create notifications for each participant
                    for participant in participants:
                        Notification.objects.create(
                            sender=request.user,
                            recipient=participant,
                            title=f"New Training: {training.title}",
                            message=f"You have been assigned to a new training session: {training.title}. Training date: {training.training_date}",
                            notification_type='general',
                            module='training'
                        )
                
                return JsonResponse({
                    'success': True,
                    'message': 'Training created successfully! Notifications sent to participants.',
                    'training_id': training.id
                })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error creating training: {str(e)}'
            })
    else:
        return JsonResponse({
            'success': False,
            'errors': form.errors
        })

@login_required
def get_training_details(request, training_id):
    # Check if user is HR admin
    if not getattr(request.user, 'hr_admin', False):
        return JsonResponse({'success': False, 'error': 'Unauthorized'})
    
    training = get_object_or_404(Training, id=training_id)
    
    # Get participant responses
    responses = TrainingEvaluation.objects.filter(
        training=training,
        is_submitted=True
    ).select_related('participant').order_by('-submitted_at')
    
    responses_data = []
    for response in responses:
        responses_data.append({
            'participant_name': f"{response.participant.firstname} {response.participant.lastname}",
            'submitted_at': response.submitted_at.strftime('%Y-%m-%d %H:%M'),
            'status': response.get_status_display(),
            'id': response.id
        })
    
    return JsonResponse({
        'success': True,
        'training': {
            'id': training.id,
            'title': training.title,
            'objective': training.objective,
            'speaker': training.speaker,
            'training_date': training.training_date.strftime('%Y-%m-%d'),
            'progress_percentage': training.progress_percentage,
            'total_participants': training.participants.count(),
        },
        'responses': responses_data
    })

@login_required
@require_http_methods(["POST"])
def delete_training(request, training_id):
    # Check if user is HR admin
    if not getattr(request.user, 'hr_admin', False):
        return JsonResponse({'success': False, 'error': 'Unauthorized'})
    
    training = get_object_or_404(Training, id=training_id)
    training_title = training.title
    training.delete()
    
    return JsonResponse({
        'success': True,
        'message': f'Training "{training_title}" deleted successfully!'
    })

@login_required
def get_evaluation_details(request, evaluation_id):
    # Check if user is HR admin
    if not getattr(request.user, 'hr_admin', False):
        return JsonResponse({'success': False, 'error': 'Unauthorized'})
    
    evaluation = get_object_or_404(TrainingEvaluation, id=evaluation_id, is_submitted=True)
    
    # Render the off-canvas content
    html_content = render_to_string('training/evaluation_details_offcanvas.html', {
        'evaluation': evaluation,
    }, request=request)
    
    return JsonResponse({
        'success': True,
        'html': html_content
    })

@login_required
def export_training_report(request, training_id):
    if not getattr(request.user, 'hr_admin', False):
        return JsonResponse({'success': False, 'error': 'Unauthorized'})
    
    training = get_object_or_404(Training, id=training_id)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Training Evaluation Report"
    
    header_font = Font(bold=True, size=14)
    subheader_font = Font(italic=True, size=12)
    bold_font = Font(bold=True)
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A1:M1')
    ws['A1'] = "RYONAN ELECTRIC PHILIPPINES CORPORATION"
    ws['A1'].font = header_font
    ws['A1'].alignment = center_alignment
    
    ws.merge_cells('A2:M2')
    ws['A2'] = "Training Evaluation Summary"
    ws['A2'].font = subheader_font
    ws['A2'].alignment = center_alignment
    
    current_row = 4
    ws[f'A{current_row}'] = "Training Title:"
    ws[f'A{current_row}'].font = bold_font
    ws[f'B{current_row}'] = training.title
    
    current_row += 1
    ws[f'A{current_row}'] = "Training Objectives:"
    ws[f'A{current_row}'].font = bold_font
    ws[f'B{current_row}'] = training.objective
    
    current_row += 1
    ws[f'A{current_row}'] = "Training Speaker:"
    ws[f'A{current_row}'].font = bold_font
    ws[f'B{current_row}'] = training.speaker
    
    current_row += 1
    ws[f'A{current_row}'] = "Training Date:"
    ws[f'A{current_row}'].font = bold_font
    ws[f'B{current_row}'] = training.training_date.strftime('%B %d, %Y')
    
    evaluations = TrainingEvaluation.objects.filter(training=training, is_submitted=True).select_related('participant__employment_info__department', 'participant__employment_info__position')
    
    current_row += 3
    ws[f'A{current_row}'] = "Training Question Summary"
    ws[f'A{current_row}'].font = bold_font
    
    current_row += 1
    rating_headers = [
        'Question', 'Average Rating', 'Response Count'
    ]
    
    for col, header in enumerate(rating_headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = bold_font
        cell.fill = yellow_fill
        cell.border = border_thin
        cell.alignment = center_alignment
    
    rating_questions = [
        ('Content Related to Job', 'content_related_to_job'),
        ('Content Explained Clearly', 'content_explained_clearly'),
        ('Content Suitable for Topic', 'content_suitable_for_topic'),
        ('Program Clear Goals', 'program_clear_goals'),
        ('Program Met Goals', 'program_met_goals'),
        ('Program Easy to Follow', 'program_easy_to_follow'),
        ('Program Easy to Understand', 'program_easy_to_understand'),
        ('Speaker Knowledge', 'speaker_knowledge'),
        ('Speaker Clear Communication', 'speaker_clear_communication'),
        ('Speaker Answered Questions', 'speaker_answered_questions'),
        ('Training Organization', 'training_organization'),
        ('Suitable Facilities', 'suitable_facilities'),
        ('Helpful Materials', 'helpful_materials'),
    ]
    
    for question_label, field_name in rating_questions:
        current_row += 1
        
        # Calculate average and count
        field_values = [getattr(eval, field_name) for eval in evaluations if getattr(eval, field_name) is not None]
        avg_rating = sum(field_values) / len(field_values) if field_values else 0
        response_count = len(field_values)
        
        ws.cell(row=current_row, column=1, value=question_label).border = border_thin
        ws.cell(row=current_row, column=2, value=f"{avg_rating:.2f}").border = border_thin
        ws.cell(row=current_row, column=3, value=response_count).border = border_thin
    
    # Participant Summary
    current_row += 3
    ws[f'A{current_row}'] = "Participant Summary"
    ws[f'A{current_row}'].font = bold_font
    
    current_row += 1
    # Participant headers
    participant_headers = [
        'Submitted Date', 'ID Number', 'Employee Name', 'Department', 'Position',
        'Content Related to Job', 'Content Explained Clearly', 'Content Suitable for Topic',
        'Program Clear Goals', 'Program Met Goals', 'Program Easy to Follow', 'Program Easy to Understand',
        'Speaker Knowledge', 'Speaker Clear Communication', 'Speaker Answered Questions',
        'Training Organization', 'Suitable Facilities', 'Helpful Materials', 'Supervisor Result & Impact',
        'Supervisor Recommendations', 'Supervisor Overall Assessment'
    ]
    
    for col, header in enumerate(participant_headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = bold_font
        cell.fill = yellow_fill
        cell.border = border_thin
        cell.alignment = center_alignment
    
    # Participant data
    for evaluation in evaluations:
        current_row += 1
        participant_data = [
            evaluation.submitted_at.strftime('%m/%d/%Y %H:%M') if evaluation.submitted_at else '',
            evaluation.participant.idnumber or 'N/A',
            f"{evaluation.participant.firstname} {evaluation.participant.lastname}",
            getattr(getattr(evaluation.participant, 'employment_info', None), 'department', None).department_name if getattr(evaluation.participant, 'employment_info', None) and getattr(evaluation.participant.employment_info, 'department', None) else 'N/A',
            str(getattr(getattr(evaluation.participant, 'employment_info', None), 'position', None).position) if getattr(evaluation.participant, 'employment_info', None) and getattr(evaluation.participant.employment_info, 'position', None) else 'N/A',
            evaluation.content_related_to_job or '',
            evaluation.content_explained_clearly or '',
            evaluation.content_suitable_for_topic or '',
            evaluation.program_clear_goals or '',
            evaluation.program_met_goals or '',
            evaluation.program_easy_to_follow or '',
            evaluation.program_easy_to_understand or '',
            evaluation.speaker_knowledge or '',
            evaluation.speaker_clear_communication or '',
            evaluation.speaker_answered_questions or '',
            evaluation.training_organization or '',
            evaluation.suitable_facilities or '',
            evaluation.helpful_materials or '',
            evaluation.result_and_impact or '',
            evaluation.recommendations or '',
            evaluation.overall_assessment or ''
        ]
        
        for col, value in enumerate(participant_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.border = border_thin
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Training_Evaluation_Report_{training.title}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response

@login_required
def training_details_page(request, training_id):
    try:
        if not getattr(request.user, 'hr_admin', False):
            messages.error(request, 'You do not have permission to access this page.')
            return redirect('training:admin')
        
        training = get_object_or_404(Training, id=training_id)
        
        # Get search query
        search_query = request.GET.get('search', '').strip()
        
        # Get all participants for this training
        participants_queryset = training.participants.all().order_by('firstname', 'lastname')
        
        # Apply search filter if provided
        if search_query:
            participants_queryset = participants_queryset.filter(
                Q(idnumber__icontains=search_query) |
                Q(firstname__icontains=search_query) |
                Q(lastname__icontains=search_query) |
                Q(username__icontains=search_query) |
                Q(employment_info__department__department_name__icontains=search_query) |
                Q(employment_info__position__position__icontains=search_query)
            ).distinct()        # Pagination
        from django.core.paginator import Paginator
        paginator = Paginator(participants_queryset, 10)  # 10 participants per page
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
        
        participants_data = []
        for participant in page_obj:
            try:
                evaluation = TrainingEvaluation.objects.get(training=training, participant=participant)
                status = evaluation.get_status_display()
                is_submitted = evaluation.is_submitted
                submitted_at = evaluation.submitted_at
                evaluation_id = evaluation.id
            except TrainingEvaluation.DoesNotExist:
                status = 'Not Started'
                is_submitted = False
                submitted_at = None
                evaluation_id = None
            
            participants_data.append({
                'id': participant.id,
                'evaluation_id': evaluation_id,
                'idnumber': participant.idnumber or 'N/A',
                'name': f"{participant.firstname} {participant.lastname}",
                'username': participant.username,
                'department': getattr(getattr(participant, 'employment_info', None), 'department', None).department_name if getattr(participant, 'employment_info', None) and getattr(participant.employment_info, 'department', None) else 'N/A',
                'position': str(getattr(getattr(participant, 'employment_info', None), 'position', None).position) if getattr(participant, 'employment_info', None) and getattr(participant.employment_info, 'position', None) else 'N/A',
                'status': status,
                'is_submitted': is_submitted,
                'submitted_at': submitted_at,
            })        # Check if this is an AJAX request
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            # Return only the table and pagination content for AJAX requests
            context = {
                'participants_data': participants_data,
                'page_obj': page_obj,
                'search': search_query,
            }
            html = render_to_string('training/training_details_table.html', context, request=request)
            return JsonResponse({'html': html})
        
        # Calculate totals from all participants (not just current page)
        all_participants = training.participants.all()
        total_participants = all_participants.count()
        submitted_count = TrainingEvaluation.objects.filter(training=training, is_submitted=True).count()
        not_started_count = total_participants - TrainingEvaluation.objects.filter(training=training).count()
        in_progress_count = total_participants - submitted_count - not_started_count
        
        status_counts = {}
        for participant in all_participants:
            try:
                evaluation = TrainingEvaluation.objects.get(training=training, participant=participant)
                status = evaluation.get_status_display()
            except TrainingEvaluation.DoesNotExist:
                status = 'Not Started'
            status_counts[status] = status_counts.get(status, 0) + 1
        
        department_counts = {}
        submitted_evaluations = TrainingEvaluation.objects.filter(training=training, is_submitted=True).select_related('participant__employment_info__department')
        for evaluation in submitted_evaluations:
            dept = getattr(evaluation.participant.employment_info, 'department', 'N/A') if hasattr(evaluation.participant, 'employment_info') else 'N/A'
            department_counts[str(dept)] = department_counts.get(str(dept), 0) + 1
        
        from datetime import datetime, timedelta
        now = datetime.now()
        current_month_start = now.replace(day=1)
        previous_month_start = (current_month_start - timedelta(days=1)).replace(day=1)
        previous_month_end = current_month_start - timedelta(days=1)
        
        # Calculate previous month counts
        # For participants, we'll use the training creation date as reference
        if training.created_at.date() < current_month_start.date():
            previous_participants = total_participants  # Use current as baseline if training existed last month
        else:
            previous_participants = 0  # New training this month
        
        previous_submitted = TrainingEvaluation.objects.filter(
            training=training,
            submitted_at__date__range=[previous_month_start.date(), previous_month_end.date()],
            is_submitted=True
        ).count()
        
        # Calculate percentage changes
        participants_change = ((total_participants - previous_participants) / previous_participants * 100) if previous_participants > 0 else 0
        submitted_change = ((submitted_count - previous_submitted) / previous_submitted * 100) if previous_submitted > 0 else 0
        
        # For in_progress and not_started, we need to calculate based on current logic
        previous_total = previous_participants
        previous_not_started = previous_total - previous_submitted if previous_total > 0 else 0
        previous_in_progress = 0  # This is harder to calculate historically, so we'll use 0 as baseline
        
        in_progress_change = ((in_progress_count - previous_in_progress) / previous_in_progress * 100) if previous_in_progress > 0 else 0
        not_started_change = ((not_started_count - previous_not_started) / previous_not_started * 100) if previous_not_started > 0 else 0
        
        progress_data = []
        for i in range(7):
            date = datetime.now().date() - timedelta(days=6-i)
            submissions_on_date = TrainingEvaluation.objects.filter(
                training=training,
                submitted_at__date=date,
                is_submitted=True
            ).count()
            progress_data.append({
                'date': date.strftime('%m/%d'),
                'submissions': submissions_on_date
            })
        
        context = {
            'training': training,
            'participants_data': participants_data,
            'page_obj': page_obj,
            'search': search_query,
            'total_participants': total_participants,
            'submitted_count': submitted_count,
            'not_started_count': not_started_count,
            'in_progress_count': in_progress_count,
            'progress_percentage': round((submitted_count / total_participants) * 100, 1) if total_participants > 0 else 0,
            'status_counts': status_counts,
            'department_counts': department_counts,
            'progress_data': progress_data,
            # Month-over-month changes
            'participants_change': participants_change,
            'submitted_change': submitted_change,
            'in_progress_change': in_progress_change,
            'not_started_change': not_started_change,
        }
        
        return render(request, 'training/training_details.html', context)
    
    except Exception as e:
        # Handle any exceptions that might occur
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': f'An error occurred: {str(e)}'
            }, status=500)
        else:
            # For non-AJAX requests, re-raise the exception to show Django's error page
            raise

@login_required
@require_http_methods(["POST"])
def confirm_evaluation(request, training_id):
    
    try:
        # Get the evaluation
        evaluation = get_object_or_404(
            TrainingEvaluation,
            training_id=training_id,
            participant=request.user,
            status='participant_reviewed'
        )
        
        with transaction.atomic():
            evaluation.status = 'manager_reviewed'
            evaluation.save()
            
            supervisor_routing = EvaluationRouting.objects.filter(
                evaluation=evaluation,
                sequence=1
            ).first()
            
            if supervisor_routing and hasattr(supervisor_routing.approver, 'employment_info'):
                manager = supervisor_routing.approver.employment_info.approver
                if manager:
                    # Avoid duplicate routing records: create if not exists
                    routing, created = EvaluationRouting.objects.get_or_create(
                        evaluation=evaluation,
                        approver=manager,
                        defaults={
                            'sequence': 2,
                            'is_completed': False
                        }
                    )
                    if created:
                        Notification.objects.create(
                            recipient=manager,
                            sender=request.user,
                            title="Training Evaluation for Manager Review",
                            message=f"Training evaluation for {evaluation.participant.firstname} {evaluation.participant.lastname} - {evaluation.training.title} is ready for your review.",
                            notification_type="approval",
                            module="training"
                        )
                    else:
                        # routing already exists; no further action needed
                        print(f"Manager routing already exists for evaluation {evaluation.id} and manager {manager.firstname} {manager.lastname}")
        
        return JsonResponse({
            'success': True,
            'message': 'Evaluation confirmed successfully and sent to manager for approval.'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Failed to confirm evaluation: {str(e)}'
        })