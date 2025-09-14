import json
import openpyxl
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from django.template.loader import render_to_string
from django.views.decorators.http import require_http_methods
from django.db.models import Q, Avg, Count
from django.utils import timezone
import datetime as dt
from django.core.paginator import Paginator
from django.db import transaction
from .models import Evaluation, EmployeeEvaluation, TaskList, TaskRating, TrainingRequest, EvaluationInstance
from .forms import (
    TaskListImportForm, EvaluationCreateForm, TaskRatingForm, 
    SupervisorEvaluationForm, ManagerApprovalForm, TrainingRequestForm,
    EmployeeFilterForm
)
from userlogin.models import EmployeeLogin
from openpyxl.styles import Font, PatternFill, Alignment
from .models import EvaluationInstance, EmployeeEvaluation
from .utils import create_missing_instances, update_overdue_instances
from datetime import timedelta, datetime
import calendar
from openpyxl.worksheet.datavalidation import DataValidation
from notification.models import Notification
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

@login_required
def evaluation_dashboard(request):
    if request.user.hr_admin:
        return redirect('admin_evaluation')
    else:
        return redirect('user_evaluation')
    
@login_required
def evaluation_user_view(request):
    
    create_missing_instances()
    update_overdue_instances()
    
    user_instances = EvaluationInstance.objects.filter(
        employee=request.user,
        evaluation__is_active=True
    ).select_related('evaluation').order_by('-created_at')
    
    total_evaluations = user_instances.count()
    pending_evaluations = user_instances.filter(status='pending').count()
    in_progress_evaluations = user_instances.filter(status='in_progress').count()
    completed_evaluations = user_instances.filter(status='completed').count()
    overdue_evaluations = user_instances.filter(status='overdue').count()
    
    evaluations_data = []
    for instance in user_instances:
        evaluation_data = {
            'instance': instance,
            'evaluation': instance.evaluation,
            'status': instance.status,
            'due_date': instance.due_date,
            'period_start': instance.period_start,
            'period_end': instance.period_end,
            'is_overdue': instance.status == 'overdue',
            'can_start': instance.status == 'pending',
            'is_completed': instance.status == 'completed',
        }
        evaluations_data.append(evaluation_data)

    # Sort the training list by desired status order for the UI:
    # pending -> in_progress -> completed -> overdue
    status_order = {
        'pending': 0,
        'in_progress': 1,
        'completed': 2,
        'overdue': 3,
    }

    def _eval_status_key(item):
        return status_order.get(item.get('status'), 99)

    evaluations_data.sort(key=_eval_status_key)
    
    is_approver = False
    pending_approver_assessments = []
    pending_routing_count = 0
    approval_tab_label = "Employee Evaluation"
    
    if hasattr(request.user, 'employment_info') and request.user.employment_info:
        # Count supervisor pending evaluations
        supervisor_pending_count = EmployeeEvaluation.objects.filter(
            supervisor=request.user,
            status='supervisor_review'
        ).select_related('evaluation', 'employee')

        # Count manager pending evaluations
        manager_pending_count = EmployeeEvaluation.objects.filter(
            manager=request.user,
            status='manager_review'
        ).select_related('evaluation', 'employee')

        # Combine pending evaluations for display
        evals_for_supervisor = EmployeeEvaluation.objects.filter(
            supervisor=request.user
        ).select_related('evaluation', 'employee')

        evals_for_manager = EmployeeEvaluation.objects.filter(
            manager=request.user
        ).select_related('evaluation', 'employee')

        # Calculate total pending count
        pending_routing_count = supervisor_pending_count.count() + manager_pending_count.count()

        if evals_for_supervisor.exists() or evals_for_manager.exists():
            is_approver = True
            approval_tab_label = 'Employee Evaluation' if evals_for_supervisor.exists() else 'Employee Evaluation'

            if evals_for_supervisor.exists() and evals_for_manager.exists():
                pending_qs = (evals_for_supervisor | evals_for_manager).distinct()
            elif evals_for_supervisor.exists():
                pending_qs = evals_for_supervisor
            else:
                pending_qs = evals_for_manager

            # Ensure the pending assessments are ordered by our desired status priority
            # Priority: supervisor_review, manager_review, approved, disapproved
            status_priority = {
                'supervisor_review': 0,
                'manager_review': 1,
                'approved': 2,
                'disapproved': 3,
            }

            # Convert to list (if it's a queryset) so we can apply a stable sort with our mapping
            pending_list = list(pending_qs)

            # Default unseen statuses get a large index so they appear at the end
            def _status_key(obj):
                return status_priority.get(obj.status, 99)

            pending_list.sort(key=_status_key)

            pending_approver_assessments = pending_list

    assessment_data = []
    for assessment in pending_approver_assessments:
        approver = None
        if assessment.status == 'supervisor_review' and hasattr(assessment.employee, 'employment_info'):
            approver = assessment.employee.employment_info.approver if assessment.employee.employment_info else None
        elif assessment.status == 'manager_review' and hasattr(assessment.employee, 'employment_info') and assessment.employee.employment_info.approver:
            supervisor = assessment.employee.employment_info.approver
            if hasattr(supervisor, 'employment_info') and supervisor.employment_info:
                approver = supervisor.employment_info.approver
        
        assessment_info = {
            'evaluation': assessment,
            'approver': approver,
            'is_completed': assessment.status in ['approved', 'disapproved'],
        }
        assessment_data.append(assessment_info)
    
    context = {
        'evaluations_data': evaluations_data,
        'total_evaluations': total_evaluations,
        'pending_evaluations': pending_evaluations,
        'pendinsg_evaluations': pending_evaluations,
        'in_progress_evaluations': in_progress_evaluations,
        'completed_evaluations': completed_evaluations,
        'overdue_evaluations': overdue_evaluations,
        'is_approver': is_approver,
        'pending_approver_assessments': assessment_data,
        'pending_routing_count': pending_routing_count,
        'approval_tab_label': approval_tab_label,
        'user_has_approver': hasattr(request.user, 'employment_info') and request.user.employment_info and request.user.employment_info.approver is not None,
    }
    
    return render(request, 'evaluation/user_evaluation.html', context)

@login_required
def evaluation_admin_view(request):
    
    # Get search query
    search_query = request.GET.get('search', '').strip()
    
    # Get current date info
    now = timezone.now()
    first_of_this_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    first_of_last_month = (first_of_this_month - timedelta(days=1)).replace(day=1)
    
    # Get evaluations with search and pagination
    evaluations_query = Evaluation.objects.filter(is_active=True)
    
    if search_query:
        evaluations_query = evaluations_query.filter(
            Q(title__icontains=search_query) |
            Q(created_by__firstname__icontains=search_query) |
            Q(created_by__lastname__icontains=search_query) |
            Q(created_by__username__icontains=search_query)
        )
    
    evaluations_query = evaluations_query.order_by('-created_at')
    
    # Pagination
    paginator = Paginator(evaluations_query, 10)  # 10 items per page
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Get all users except admins with their employment info
    users_query = EmployeeLogin.objects.filter(
        active=True,
        hr_admin=False,
        is_superuser=False
    ).select_related(
        'employment_info__department',
        'employment_info__line', 
        'employment_info__approver'
    ).prefetch_related('task_lists')
    
    if search_query:
        users_query = users_query.filter(
            Q(firstname__icontains=search_query) |
            Q(lastname__icontains=search_query) |
            Q(username__icontains=search_query) |
            Q(idnumber__icontains=search_query)
        )
    
    users_query = users_query.order_by('firstname', 'lastname')
    
    # Pagination for users
    users_paginator = Paginator(users_query, 10)  # 10 items per page
    users_page_number = request.GET.get('users_page', 1)
    users_page_obj = users_paginator.get_page(users_page_number)
    
    # Get evaluations for current page
    evaluations = page_obj.object_list
    
    # Get users for current page
    users = users_page_obj.object_list

    # If AJAX request for table only, return rendered rows HTML
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' and request.GET.get('partial') == 'table':
        rows_html = render_to_string('evaluation/partials/evaluation_table_rows.html', {'evaluations': evaluations})
        pagination_html = render_to_string('evaluation/partials/evaluation_pagination.html', {'page_obj': page_obj, 'search': search_query})
        return JsonResponse({'rows_html': rows_html, 'pagination_html': pagination_html})
    
    # If AJAX request for tasklist table only
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' and request.GET.get('partial') == 'tasklist_table':
        rows_html = render_to_string('evaluation/partials/tasklist_table_rows.html', {'users': users})
        pagination_html = render_to_string('evaluation/partials/evaluation_pagination.html', {'page_obj': users_page_obj, 'search': search_query})
        return JsonResponse({'rows_html': rows_html, 'pagination_html': pagination_html})
    
    # Total Evaluations (all evaluation records)
    total_evaluations = EmployeeEvaluation.objects.count()
    total_this_month = EmployeeEvaluation.objects.filter(created_at__gte=first_of_this_month).count()
    total_last_month = EmployeeEvaluation.objects.filter(created_at__gte=first_of_last_month, created_at__lt=first_of_this_month).count()
    
    # Calculate percentage for total evaluations
    total_percent = 0
    total_positive = True
    if total_last_month > 0:
        total_percent = round(((total_this_month - total_last_month) / total_last_month) * 100, 1)
        total_positive = total_percent >= 0
    elif total_this_month > 0:
        total_percent = 100.0
        total_positive = True
    
    # Active Evaluations (pending or in progress)
    active_evaluations = EmployeeEvaluation.objects.filter(
        status__in=['pending', 'supervisor_review', 'manager_review']
    ).count()
    active_this_month = EmployeeEvaluation.objects.filter(
        created_at__gte=first_of_this_month,
        status__in=['pending', 'supervisor_review', 'manager_review']
    ).count()
    active_last_month = EmployeeEvaluation.objects.filter(
        created_at__gte=first_of_last_month, 
        created_at__lt=first_of_this_month,
        status__in=['pending', 'supervisor_review', 'manager_review']
    ).count()
    
    # Calculate percentage for active evaluations
    active_percent = 0
    active_positive = True
    if active_last_month > 0:
        active_percent = round(((active_this_month - active_last_month) / active_last_month) * 100, 1)
        active_positive = active_percent >= 0
    elif active_this_month > 0:
        active_percent = 100.0
        active_positive = True
    
    # Completed Evaluations (approved or disapproved)
    completed_evaluations = EmployeeEvaluation.objects.filter(
        status__in=['approved', 'disapproved']
    ).count()
    completed_this_month = EmployeeEvaluation.objects.filter(
        created_at__gte=first_of_this_month,
        status__in=['approved', 'disapproved']
    ).count()
    completed_last_month = EmployeeEvaluation.objects.filter(
        created_at__gte=first_of_last_month, 
        created_at__lt=first_of_this_month,
        status__in=['approved', 'disapproved']
    ).count()
    
    # Calculate percentage for completed evaluations
    completed_percent = 0
    completed_positive = True
    if completed_last_month > 0:
        completed_percent = round(((completed_this_month - completed_last_month) / completed_last_month) * 100, 1)
        completed_positive = completed_percent >= 0
    elif completed_this_month > 0:
        completed_percent = 100.0
        completed_positive = True
    
    # Pending Evaluations (not yet submitted or in early stages)
    pending_evaluations = EmployeeEvaluation.objects.filter(
        status='pending'
    ).count()
    pending_this_month = EmployeeEvaluation.objects.filter(
        created_at__gte=first_of_this_month,
        status='pending'
    ).count()
    pending_last_month = EmployeeEvaluation.objects.filter(
        created_at__gte=first_of_last_month, 
        created_at__lt=first_of_this_month,
        status='pending'
    ).count()
    
    # Calculate percentage for pending evaluations (note: for pending, negative is good)
    pending_percent = 0
    pending_positive = True
    if pending_last_month > 0:
        pending_percent = round(((pending_this_month - pending_last_month) / pending_last_month) * 100, 1)
        pending_positive = pending_percent <= 0  # Negative is good for pending items
    elif pending_this_month > 0:
        pending_percent = 100.0
        pending_positive = False
    
    # Chart data for response status distribution
    response_status_counts = EmployeeEvaluation.objects.values('status').annotate(count=Count('status'))
    status_data = {item['status']: item['count'] for item in response_status_counts}
    
    # Get total non-admin users for Y-axis reference
    total_non_admin_users = EmployeeLogin.objects.filter(
        active=True,
        hr_admin=False,
        is_superuser=False
    ).count()
    
    # Color palette for different evaluations
    color_palette = [
        '#3b82f6', '#10b981', '#f59e0b', '#ef4444', '#8b5cf6',
        '#06b6d4', '#84cc16', '#f97316', '#ec4899', '#6366f1'
    ]
    
    # Get all evaluations for chart data
    all_evaluations = Evaluation.objects.filter(is_active=True).order_by('created_at')
    
    # Prepare evaluation datasets for different time periods
    evaluation_datasets = []
    month_evaluation_datasets = []
    quarter_evaluation_datasets = []
    year_evaluation_datasets = []
    
    # Monthly labels (last 12 months)
    month_labels = []
    for i in range(11, -1, -1):
        month_start = (now - timedelta(days=30*i)).replace(day=1)
        month_labels.append(month_start.strftime('%b %Y'))
    
    # Quarterly labels (last 4 quarters) aligned to fiscal year May->Apr
    quarter_labels = []
    fiscal_start_month = 5  # May
    # determine current fiscal year start (e.g., if month >= May, fiscal year starts this year)
    current_fy_start = now.year if now.month >= fiscal_start_month else now.year - 1
    # months since fiscal year start (0..11)
    months_since_fy_start = (now.month - fiscal_start_month) % 12
    current_fiscal_quarter = months_since_fy_start // 3 + 1
    for i in range(3, -1, -1):
        q = current_fiscal_quarter - i
        fy = current_fy_start
        if q <= 0:
            q += 4
            fy -= 1
        quarter_labels.append(f'Q{q} FY{fy}')
    
    # Yearly labels (last 3 fiscal years, May->Apr)
    year_labels = []
    for i in range(2, -1, -1):
        fy = current_fy_start - i
        year_labels.append(f'FY{fy}-{(fy+1)%100:02d}')
    
    for idx, evaluation in enumerate(all_evaluations):
        color = color_palette[idx % len(color_palette)]
        
        # Get employee evaluations for this evaluation
        employee_evaluations = EmployeeEvaluation.objects.filter(evaluation=evaluation)
        
        # Monthly data
        monthly_counts = []
        for i in range(11, -1, -1):
            month_start = (now - timedelta(days=30*i)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(seconds=1)
            count = employee_evaluations.filter(created_at__gte=month_start, created_at__lte=month_end).count()
            monthly_counts.append(count)
        
        # Quarterly data (aligned to fiscal year May->Apr)
        quarterly_counts = []
        for i in range(3, -1, -1):
            q = current_fiscal_quarter - i
            fy = current_fy_start
            if q <= 0:
                q += 4
                fy -= 1
            # compute start month for fiscal quarter
            start_month = fiscal_start_month + (q - 1) * 3
            start_year = fy
            if start_month > 12:
                start_month -= 12
                start_year += 1
            quarter_start = dt.datetime(start_year, start_month, 1, tzinfo=dt.timezone.utc)
            # quarter_end is start + 3 months - 1 second
            next_month = start_month + 3
            next_year = start_year
            if next_month > 12:
                next_month -= 12
                next_year += 1
            quarter_end = dt.datetime(next_year, next_month, 1, tzinfo=dt.timezone.utc) - timedelta(seconds=1)
            count = employee_evaluations.filter(created_at__gte=quarter_start, created_at__lte=quarter_end).count()
            quarterly_counts.append(count)
        
        # Yearly (fiscal) data (May->Apr)
        yearly_counts = []
        for i in range(2, -1, -1):
            fy = current_fy_start - i
            year_start = dt.datetime(fy, fiscal_start_month, 1, tzinfo=dt.timezone.utc)
            year_end = dt.datetime(fy + 1, fiscal_start_month - 1, 30 if fiscal_start_month - 1 == 4 else 31, 23, 59, 59, tzinfo=dt.timezone.utc) if False else dt.datetime(fy + 1, 4, 30, 23, 59, 59, tzinfo=dt.timezone.utc)
            # explicit May->Apr end (April 30)
            year_end = dt.datetime(fy + 1, 4, 30, 23, 59, 59, tzinfo=dt.timezone.utc)
            count = employee_evaluations.filter(created_at__gte=year_start, created_at__lte=year_end).count()
            yearly_counts.append(count)
        
        # Add to datasets
        month_evaluation_datasets.append({
            'title': evaluation.title,
            'color': color,
            'monthly_counts': monthly_counts
        })
        
        quarter_evaluation_datasets.append({
            'title': evaluation.title,
            'color': color,
            'quarterly_counts': quarterly_counts
        })
        
        year_evaluation_datasets.append({
            'title': evaluation.title,
            'color': color,
            'yearly_counts': yearly_counts
        })
    
    context = {
        'total_trainings': total_evaluations,
        'active_trainings': active_evaluations,
        'completed_evaluations': completed_evaluations,
        'pending_responses': pending_evaluations,
        'trainings_percent': total_percent,
        'trainings_positive': total_positive,
        'active_percent': active_percent,
        'active_positive': active_positive,
        'completed_percent': completed_percent,
        'completed_positive': completed_positive,
        'pending_percent': pending_percent,
        'pending_positive': pending_positive,
        'response_status_counts': status_data,
        'total_non_admin_users': total_non_admin_users,
        'month_labels': month_labels,
        'quarter_labels': quarter_labels,
        'year_labels': year_labels,
        'month_evaluation_datasets': month_evaluation_datasets,
        'quarter_evaluation_datasets': quarter_evaluation_datasets,
        'year_evaluation_datasets': year_evaluation_datasets,
        'evaluations': evaluations,
        'page_obj': page_obj,
        'search': search_query,
        'users': users,
        'users_page_obj': users_page_obj,
        'fiscal_year_start': now.year if now.month >= 4 else now.year - 1,
        'fiscal_year_end': now.year + 1 if now.month >= 4 else now.year,
    'created_evaluation_message': request.session.pop('created_evaluation_message', None),
    }
    
    return render(request, 'evaluation/admin_evaluation.html', context)


@login_required
@require_http_methods(["POST"])
def create_evaluation(request):
    try:
        form = EvaluationCreateForm(request.POST)
        print("Form is valid:", form.is_valid())
        if form.is_valid():
            evaluation = form.save(commit=False)
            evaluation.created_by = request.user
            evaluation.save()

            request.session['created_evaluation_message'] = 'Evaluation created successfully!'
            return redirect('admin_evaluation')
        else:
            print("Form errors:", form.errors)
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
            return redirect('admin_evaluation')
    except Exception as e:
        messages.error(request, f'Error creating evaluation: {str(e)}')
        return redirect('admin_evaluation')


@login_required
def get_evaluation_detail(request, evaluation_id):
    """Return evaluation data as JSON for populating the edit modal"""
    evaluation = get_object_or_404(Evaluation, pk=evaluation_id)
    data = {
        'id': evaluation.id,
        'title': evaluation.title,
        'description': evaluation.description,
        'start_year': evaluation.start_year,
        'end_year': evaluation.end_year,
        'duration': evaluation.duration,
        'is_active': evaluation.is_active,
    }
    return JsonResponse({'evaluation': data})


@login_required
@require_http_methods(['POST'])
def update_evaluation(request, evaluation_id):
    evaluation = get_object_or_404(Evaluation, pk=evaluation_id)

    # Reuse EvaluationCreateForm to validate incoming data
    form = EvaluationCreateForm(request.POST, instance=evaluation)
    if form.is_valid():
        form.save()
        # On success, return JSON so frontend can refresh the table
        return JsonResponse({'success': True, 'message': 'Evaluation updated successfully'})
    else:
        # Return validation errors
        errors = {field: [str(e) for e in errs] for field, errs in form.errors.items()}
        return JsonResponse({'success': False, 'errors': errors}, status=400)


@login_required
@require_http_methods(['POST'])
def delete_evaluation(request, evaluation_id):
    """Delete an evaluation (admin action). Returns JSON."""
    evaluation = get_object_or_404(Evaluation, pk=evaluation_id)
    try:
        evaluation.delete()
        return JsonResponse({'success': True, 'message': 'Evaluation deleted successfully'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
def export_tasklist_template(request):
    """Export Excel template for tasklist import with two sheets: Tasklists and Registered Employee"""
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create Tasklists sheet
    tasklist_sheet = wb.create_sheet("Tasklists")
    
    # Create Registered Employee sheet
    employee_sheet = wb.create_sheet("Registered Employee")
    
    # Define styles
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow background
    border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin')
    )
    
    # Populate Registered Employee sheet
    employee_sheet['A1'] = "Id Number"
    employee_sheet['B1'] = "Employee Name"
    
    # Apply header styling
    for col in ['A', 'B']:
        cell = employee_sheet[f'{col}1']
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    # Get non-admin employees
    employees = EmployeeLogin.objects.filter(
        active=True,
        hr_admin=False,
        wire_admin=False,
        clinic_admin=False,
        iad_admin=False,
        accounting_admin=False,
        mis_admin=False
    ).exclude(
        status='disapproved'
    ).order_by('idnumber')
    
    # Add employee data
    for row_num, employee in enumerate(employees, start=2):
        employee_sheet[f'A{row_num}'] = employee.idnumber or ""
        employee_sheet[f'B{row_num}'] = employee.full_name
        
        # Apply border to data cells
        for col in ['A', 'B']:
            employee_sheet[f'{col}{row_num}'].border = border
    
    # Auto-adjust column widths for employee sheet
    for col in ['A', 'B']:
        max_length = 0
        column = employee_sheet.column_dimensions[col]
        for cell in employee_sheet[col]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        column.width = max_length + 2
    
    # Populate Tasklists sheet
    tasklist_sheet['A1'] = "Id Number"
    tasklist_sheet['B1'] = "Tasklist"
    
    # Apply header styling
    for col in ['A', 'B']:
        cell = tasklist_sheet[f'{col}1']
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    # Add sample data rows
    sample_data = [
        ("EMP001", "Complete monthly performance review"),
        ("EMP002", "Attend team meeting and provide updates"),
        ("EMP003", "Submit project documentation")
    ]
    
    for row_num, (id_num, task) in enumerate(sample_data, start=2):
        tasklist_sheet[f'A{row_num}'] = id_num
        tasklist_sheet[f'B{row_num}'] = task
        
        # Apply border to data cells
        for col in ['A', 'B']:
            tasklist_sheet[f'{col}{row_num}'].border = border
    
    instructions_start_row = len(sample_data) + 3
    
    tasklist_sheet[f'A{instructions_start_row}'] = "INSTRUCTIONS:"
    tasklist_sheet[f'A{instructions_start_row}'].font = Font(bold=True, size=12)
    
    instructions = [
        "1. Fill in the table above with employee evaluations:",
        "   - Id Number: Select from the dropdown (must match Registered Employee sheet)",
        "   - Tasklist: Enter the evaluation task or description",
        "",
        "2. Important Notes:",
        "   - Only use Id Numbers from the Registered Employee sheet",
        "   - Files with invalid Id Numbers will not be uploaded",
        "   - You can add multiple rows for different employees/tasks",
        "   - Save the file as .xlsx format before uploading",
        "",
        "3. Upload Process:",
        "   - Click 'Import Tasklist' button in the application",
        "   - Select your completed Excel file",
        "   - Click 'Upload Files' to process"
    ]
    
    for i, instruction in enumerate(instructions, start=instructions_start_row + 1):
        tasklist_sheet[f'A{i}'] = instruction
    
    # Add data validation for Id Number column in Tasklists sheet
    
    dv = DataValidation(
        type="list",
        formula1="'Registered Employee'!$A$2:$A$" + str(len(employees) + 1),
        allow_blank=True
    )
    dv.error = 'Please select a valid Id Number from the Registered Employee sheet'
    dv.errorTitle = 'Invalid Id Number'
    dv.prompt = 'Select an Id Number from the Registered Employee sheet'
    dv.promptTitle = 'Id Number Selection'
    
    tasklist_sheet.add_data_validation(dv)
    dv.add('A2:A1000')
    
    for col in ['A', 'B']:
        max_length = 0
        column = tasklist_sheet.column_dimensions[col]
        for cell in tasklist_sheet[col]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        column.width = max_length + 2
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=tasklist_import_template.xlsx'
    
    # Save workbook to response
    wb.save(response)
    
    return response


@login_required
@require_http_methods(["POST"])
def upload_tasklist(request):
    """Upload and process tasklist Excel file"""
    
    if 'files' not in request.FILES:
        messages.error(request, 'No file uploaded')
        return redirect(reverse('admin_evaluation') + '?tab=tasklist')
    
    uploaded_file = request.FILES['files']
    
    # Validate file extension
    if not uploaded_file.name.lower().endswith(('.xlsx', '.xls')):
        messages.error(request, 'Please upload a valid Excel file (.xlsx or .xls)')
        return redirect(reverse('admin_evaluation') + '?tab=tasklist')
    
    # Validate file size (10MB limit)
    if uploaded_file.size > 10 * 1024 * 1024:
        messages.error(request, 'File size exceeds 10MB limit')
        return redirect(reverse('admin_evaluation') + '?tab=tasklist')
    
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
        
        # Check if Tasklists sheet exists
        if 'Tasklists' not in wb.sheetnames:
            messages.error(request, 'Excel file must contain a "Tasklists" sheet')
            return redirect(reverse('admin_evaluation') + '?tab=tasklist')
        
        tasklist_sheet = wb['Tasklists']
        
        # Validate headers
        header_1 = tasklist_sheet.cell(row=1, column=1).value
        header_2 = tasklist_sheet.cell(row=1, column=2).value
        
        if str(header_1).strip().lower() != 'id number' or str(header_2).strip().lower() != 'tasklist':
            messages.error(request, 'Invalid Excel format. Headers must be "Id Number" and "Tasklist"')
            return redirect(reverse('admin_evaluation') + '?tab=tasklist')
        
        # Process the data
        success_count = 0
        error_count = 0
        errors = []
        
        for row_num in range(2, tasklist_sheet.max_row + 1):
            id_number = tasklist_sheet.cell(row=row_num, column=1).value
            task_text = tasklist_sheet.cell(row=row_num, column=2).value
            
            # Skip empty rows
            if not id_number or not task_text:
                continue
            
            # Validate data types
            id_number = str(id_number).strip()
            task_text = str(task_text).strip()
            
            if not id_number or not task_text:
                error_count += 1
                errors.append(f"Row {row_num}: Empty Id Number or Tasklist")
                continue
            
            try:
                # Find employee by idnumber
                employee = EmployeeLogin.objects.get(idnumber=id_number)
                
                # Check if tasklist already exists for this employee
                existing_task = TaskList.objects.filter(
                    employee=employee, 
                    tasklist=task_text
                ).first()
                
                if existing_task:
                    error_count += 1
                    errors.append(f"Row {row_num}: Task already exists for employee {employee.full_name}")
                    continue
                
                # Create new tasklist entry
                TaskList.objects.create(
                    employee=employee,
                    tasklist=task_text
                )
                success_count += 1
                
            except EmployeeLogin.DoesNotExist:
                error_count += 1
                errors.append(f"Row {row_num}: Employee with ID {id_number} not found")
            except Exception as e:
                error_count += 1
                errors.append(f"Row {row_num}: Error processing - {str(e)}")
        
        # Prepare success message
        if success_count > 0:
            messages.success(request, f'Successfully imported {success_count} tasklist entries')
        
        if error_count > 0:
            messages.warning(request, f'{error_count} entries failed to import')
            # Store errors in session for detailed display
            request.session['tasklist_import_errors'] = errors[:10]  # Limit to first 10 errors

        # Redirect back to admin_evaluation with tab query param
        return redirect(reverse('admin_evaluation') + '?tab=tasklist')

    except openpyxl.utils.exceptions.InvalidFileException:
        messages.error(request, 'Invalid Excel file format')
        return redirect(reverse('admin_evaluation') + '?tab=tasklist')
    except Exception as e:
        messages.error(request, f'Error processing file: {str(e)}')
        return redirect(reverse('admin_evaluation') + '?tab=tasklist')


@login_required
@require_http_methods(["GET"])
def get_user_tasklists(request, user_id):
    """Get tasklists for a specific user"""
    try:
        user = get_object_or_404(EmployeeLogin, id=user_id)
        tasklists = TaskList.objects.filter(employee=user).order_by('id')
        
        tasklist_data = [
            {
                'id': task.id,
                'tasklist': task.tasklist
            }
            for task in tasklists
        ]
        
        return JsonResponse({
            'success': True,
            'user_name': user.full_name,
            'user_id': user.id,
            'tasklists': tasklist_data
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["POST"])
def update_tasklist_item(request, tasklist_id):
    """Update a specific tasklist item"""
    try:
        tasklist = get_object_or_404(TaskList, id=tasklist_id)
        
        data = json.loads(request.body)
        new_tasklist_text = data.get('tasklist', '').strip()
        
        if not new_tasklist_text:
            return JsonResponse({
                'success': False,
                'error': 'Tasklist text cannot be empty'
            }, status=400)
        
        tasklist.tasklist = new_tasklist_text
        tasklist.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Tasklist updated successfully',
            'tasklist': {
                'id': tasklist.id,
                'tasklist': tasklist.tasklist
            }
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["POST"])
def create_tasklist_item(request, user_id):
    """Create a new tasklist item for a specific user"""
    try:
        user = get_object_or_404(EmployeeLogin, id=user_id)
        
        data = json.loads(request.body)
        tasklist_text = data.get('tasklist', '').strip()
        
        if not tasklist_text:
            return JsonResponse({
                'success': False,
                'error': 'Tasklist text cannot be empty'
            }, status=400)
        
        # Check if tasklist already exists for this user
        existing_task = TaskList.objects.filter(
            employee=user, 
            tasklist=tasklist_text
        ).first()
        
        if existing_task:
            return JsonResponse({
                'success': False,
                'error': 'This tasklist already exists for this user'
            }, status=400)
        
        # Create new tasklist entry
        new_tasklist = TaskList.objects.create(
            employee=user,
            tasklist=tasklist_text
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Tasklist created successfully',
            'tasklist': {
                'id': new_tasklist.id,
                'tasklist': new_tasklist.tasklist
            }
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["POST"])
def delete_tasklist_item(request, tasklist_id):
    """Delete a tasklist item"""
    try:
        tasklist = get_object_or_404(TaskList, id=tasklist_id)
        
        # Optional: Check if user has permission to delete this tasklist
        # For now, allowing any logged-in user to delete (as per admin context)
        
        tasklist_text = tasklist.tasklist  # Store for response
        tasklist.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Tasklist "{tasklist_text}" deleted successfully'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
def get_evaluation_instance(request, instance_id):
    """Get evaluation instance data for starting evaluation"""
    try:
        instance = get_object_or_404(EvaluationInstance, id=instance_id, employee=request.user)
        
        # Check if user has an approver set
        if not hasattr(request.user, 'employment_info') or not request.user.employment_info or not request.user.employment_info.approver:
            return JsonResponse({
                'error': 'User must set her/his approver to start evaluation.'
            }, status=400)
        
        # Get or create employee evaluation
        employee_eval = instance.get_or_create_employee_evaluation()
        
        # Get user's task lists
        task_lists = TaskList.objects.filter(employee=request.user, is_active=True).order_by('tasklist')
        
        # Create form data for task ratings
        task_ratings_data = []
        for task in task_lists:
            try:
                task_rating = TaskRating.objects.get(employee_evaluation=employee_eval, task=task)
                rating_value = task_rating.rating
                comments = task_rating.comments
            except TaskRating.DoesNotExist:
                rating_value = 0
                comments = ''
            
            task_ratings_data.append({
                'task_id': task.id,
                'task_name': task.tasklist,
                'rating': rating_value,
                'comments': comments
            })
        
        return JsonResponse({
            'success': True,
            'evaluation': {
                'id': instance.evaluation.id,
                'title': instance.evaluation.title,
                'description': instance.evaluation.description,
                'duration': instance.evaluation.get_duration_display(),
            },
            'instance': {
                'id': instance.id,
                'period_start': instance.period_start.strftime('%Y-%m-%d'),
                'period_end': instance.period_end.strftime('%Y-%m-%d'),
                'due_date': instance.due_date.strftime('%Y-%m-%d'),
                'status': instance.status,
            },
            'employee_evaluation_id': employee_eval.id,
            'task_ratings': task_ratings_data,
            'form_html': render_to_string('evaluation/evaluation_form.html', {
                'instance': instance,
                'employee_eval': employee_eval,
                'task_ratings': task_ratings_data
            }, request=request)
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["POST"])
def submit_evaluation_instance(request, employee_evaluation_id):
    try:
        employee_eval = get_object_or_404(
            EmployeeEvaluation, 
            id=employee_evaluation_id, 
            employee=request.user
        )
        
        if not employee_eval.self_completed_at:
            employee_eval.self_completed_at = timezone.now()
            employee_eval.status = 'supervisor_review'
            if hasattr(request.user, 'employment_info') and request.user.employment_info and getattr(request.user.employment_info, 'approver', None):
                employee_eval.supervisor = request.user.employment_info.approver
            employee_eval.save()
        
        task_lists = TaskList.objects.filter(employee=request.user, is_active=True)
        
        for task in task_lists:
            rating = request.POST.get(f'task_rating_{task.id}')
            comments = request.POST.get(f'task_comments_{task.id}', '')
            
            if rating:
                rating = int(rating)
                task_rating, created = TaskRating.objects.update_or_create(
                    employee_evaluation=employee_eval,
                    task=task,
                    defaults={
                        'rating': rating,
                        'comments': comments
                    }
                )

        Notification.objects.create(
            title="Self-Evaluation Submitted",
            sender=request.user,
            recipient= request.user.employment_info.approver,
            message=f'{request.user.firstname} {request.user.lastname} has completed and submitted their evaluation for your review',
            module='evaluation',
            notification_type='approval'
        )
        
        if employee_eval.evaluation_instance:
            employee_eval.evaluation_instance.status = 'completed' if employee_eval.status in ['approved', 'disapproved'] else 'in_progress'
            employee_eval.evaluation_instance.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Evaluation submitted successfully!'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
def view_completed_evaluation(request, instance_id):
    """View completed evaluation instance"""
    try:
        instance = get_object_or_404(EvaluationInstance, id=instance_id, employee=request.user)
        employee_eval = instance.get_or_create_employee_evaluation()
        
        # Get task ratings
        task_ratings = TaskRating.objects.filter(employee_evaluation=employee_eval).select_related('task')
        
        task_ratings_data = []
        for rating in task_ratings:
            task_ratings_data.append({
                'task_name': rating.task.tasklist,
                'rating': rating.rating,
                'comments': rating.comments
            })
        
        return JsonResponse({
            'success': True,
            'evaluation': {
                'id': instance.evaluation.id,
                'title': instance.evaluation.title,
                'description': instance.evaluation.description,
            },
            'instance': {
                'id': instance.id,
                'period_start': instance.period_start.strftime('%Y-%m-%d'),
                'period_end': instance.period_end.strftime('%Y-%m-%d'),
                'status': instance.status,
            },
            'employee_evaluation': {
                'status': employee_eval.status,
                'completed_at': employee_eval.self_completed_at.strftime('%Y-%m-%d %H:%M') if employee_eval.self_completed_at else None,
                'average_rating': employee_eval.average_rating,
            },
            'task_ratings': task_ratings_data,
            'form_html': render_to_string('evaluation/evaluation_view.html', {
                'instance': instance,
                'employee_eval': employee_eval,
                'task_ratings': task_ratings_data
            }, request=request)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

@login_required
def get_supervisor_evaluation(request, evaluation_id):
    """Get evaluation for supervisor review"""
    try:
        employee_eval = get_object_or_404(
            EmployeeEvaluation,
            id=evaluation_id,
            status__in=['supervisor_review', 'manager_review']
        )
        
        # Check if user is the approver
        is_supervisor_approver = False
        is_manager_approver = False
        
        if hasattr(employee_eval.employee, 'employment_info') and employee_eval.employee.employment_info:
            supervisor = employee_eval.employee.employment_info.approver
            if supervisor == request.user:
                is_supervisor_approver = True
            elif hasattr(supervisor, 'employment_info') and supervisor.employment_info and supervisor.employment_info.approver == request.user:
                is_manager_approver = True
        
        if not (is_supervisor_approver or is_manager_approver):
            return JsonResponse({
                'success': False,
                'error': 'You are not authorized to review this evaluation'
            }, status=403)
        
        is_read_only = employee_eval.status not in ['supervisor_review', 'manager_review']
        
        form = SupervisorEvaluationForm(instance=employee_eval)
        
        # Get task ratings
        task_ratings = employee_eval.task_ratings.all().select_related('task')
        task_ratings_data = []
        for rating in task_ratings:
            task_ratings_data.append({
                'task': rating.task.tasklist,
                'task_id': rating.task.id,
                'rating': rating.rating,
                'supervisor_rating': rating.supervisor_rating,
                'comments': rating.comments
            })
        
        context_data = {
            'form': form,
            'employee_eval': employee_eval,
            'task_ratings': task_ratings_data,
            'is_read_only': is_read_only,
            'is_supervisor_review': employee_eval.status == 'supervisor_review',
            'is_manager_review': employee_eval.status == 'manager_review'
        }
        
        return JsonResponse({
            'success': True,
            'html': render_to_string('evaluation/supervisor_evaluation_content.html', context_data, request=request)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

@login_required
@require_http_methods(["POST"])
def submit_supervisor_evaluation(request, evaluation_id):
    try:
        employee_eval = get_object_or_404(
            EmployeeEvaluation,
            id=evaluation_id,
            status__in=['supervisor_review', 'manager_review']
        )
        
        is_supervisor_approver = False
        is_manager_approver = False
        
        if hasattr(employee_eval.employee, 'employment_info') and employee_eval.employee.employment_info:
            supervisor = employee_eval.employee.employment_info.approver
            if supervisor == request.user:
                is_supervisor_approver = True
            elif hasattr(supervisor, 'employment_info') and supervisor.employment_info and supervisor.employment_info.approver == request.user:
                is_manager_approver = True
        
        if not (is_supervisor_approver or is_manager_approver):
            return JsonResponse({
                'success': False,
                'error': 'You are not authorized to review this evaluation'
            }, status=403)
        
        if not hasattr(request.user, 'employment_info') or not request.user.employment_info or not request.user.employment_info.approver:
            return JsonResponse({
                'success': False,
                'error': 'Please set your approver first in your profile settings before submitting evaluations.'
            }, status=400)
        
        form = SupervisorEvaluationForm(request.POST, instance=employee_eval)
        if form.is_valid():
            employee_eval = form.save(commit=False)
            
            if is_supervisor_approver:
                employee_eval.supervisor_completed_at = timezone.now()
                
                # Check if user's position level is 3 (manager level)
                if (hasattr(request.user, 'employment_info') and 
                    request.user.employment_info and 
                    hasattr(request.user.employment_info, 'position') and 
                    request.user.employment_info.position and 
                    request.user.employment_info.position.level == '3'):
                    
                    # If supervisor is level 3, they act as both supervisor and manager
                    employee_eval.manager = request.user
                    employee_eval.manager_completed_at = timezone.now()
                    employee_eval.status = 'approved'
                else:
                    # Normal flow - send to manager for review
                    employee_eval.manager = request.user.employment_info.approver
                    if employee_eval.status == 'supervisor_review':
                        employee_eval.status = 'manager_review'
            
            employee_eval.save()
            
            for key, value in request.POST.items():
                if key.startswith('supervisor_task_rating_') and value:
                    try:
                        task_id = int(key.replace('supervisor_task_rating_', ''))
                        rating_value = int(value)
                        
                        task_rating = TaskRating.objects.filter(
                            employee_evaluation=employee_eval,
                            task_id=task_id
                        ).first()
                        
                        if task_rating:
                            task_rating.supervisor_rating = rating_value
                            task_rating.save()
                            
                    except (ValueError, TaskRating.DoesNotExist):
                        continue
            
            Notification.objects.create(
                title="Evaluation Submitted",
                sender=request.user,
                recipient= request.user.employment_info.approver if (
                    not hasattr(request.user.employment_info, 'position') or 
                    not request.user.employment_info.position or 
                    request.user.employment_info.position.level != '3'
                ) else employee_eval.employee,
                message=f'{request.user.firstname} {request.user.lastname} has finished evaluating {employee_eval.employee.firstname} {employee_eval.employee.lastname}' + (
                    ' and is ready for your approval' if (
                        not hasattr(request.user.employment_info, 'position') or 
                        not request.user.employment_info.position or 
                        request.user.employment_info.position.level != '3'
                    ) else ' and the evaluation has been completed'
                ),
                module='evaluation',
                notification_type='approval'
            )
            
            # Update evaluation instance status
            if employee_eval.evaluation_instance:
                if employee_eval.status == 'approved':
                    employee_eval.evaluation_instance.status = 'completed'
                else:
                    employee_eval.evaluation_instance.status = 'in_progress'
                employee_eval.evaluation_instance.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Evaluation assessment submitted successfully!'
            })
        else:
            return JsonResponse({
                'success': False,
                'error': 'Please correct the errors in the form.',
                'form_errors': form.errors
            }, status=400)
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
def view_supervisor_evaluation(request, evaluation_id):
    try:
        employee_eval = get_object_or_404(
            EmployeeEvaluation,
            id=evaluation_id
        )
        
        is_supervisor_approver = False
        is_manager_approver = False
        is_employee = employee_eval.employee == request.user
        
        if hasattr(employee_eval.employee, 'employment_info') and employee_eval.employee.employment_info:
            supervisor = employee_eval.employee.employment_info.approver
            if supervisor == request.user:
                is_supervisor_approver = True
            elif hasattr(supervisor, 'employment_info') and supervisor.employment_info and supervisor.employment_info.approver == request.user:
                is_manager_approver = True
        
        if not (is_supervisor_approver or is_manager_approver or is_employee):
            return JsonResponse({
                'success': False,
                'error': 'You are not authorized to view this evaluation'
            }, status=403)
        
        task_ratings = employee_eval.task_ratings.all().select_related('task')
        task_ratings_data = []
        for rating in task_ratings:
            task_ratings_data.append({
                'task': rating.task.tasklist,
                'task_id': rating.task.id,
                'rating': rating.rating,
                'supervisor_rating': rating.supervisor_rating,
                'comments': rating.comments
            })

        can_edit = False
        if employee_eval.status == 'supervisor_review' and is_supervisor_approver:
            can_edit = True
        elif employee_eval.status == 'manager_review' and is_manager_approver:
            can_edit = True
        
        context_data = {
            'employee_eval': employee_eval,
            'task_ratings': task_ratings_data,
            'can_edit': can_edit,
            'is_supervisor_view': is_supervisor_approver,
            'is_manager_view': is_manager_approver,
            'is_employee_view': is_employee
        }
        
        return JsonResponse({
            'success': True,
            'html': render_to_string('evaluation/supervisor_evaluation_view.html', context_data, request=request)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
def get_manager_evaluation(request, evaluation_id):
    """Get evaluation for manager review (read-only view before approval)"""
    try:
        employee_eval = get_object_or_404(
            EmployeeEvaluation,
            id=evaluation_id,
            status='manager_review'
        )
        
        # Check if user is the manager
        is_manager_approver = False
        
        if hasattr(employee_eval.employee, 'employment_info') and employee_eval.employee.employment_info:
            supervisor = employee_eval.employee.employment_info.approver
            if hasattr(supervisor, 'employment_info') and supervisor.employment_info and supervisor.employment_info.approver == request.user:
                is_manager_approver = True
        
        if not is_manager_approver:
            return JsonResponse({
                'success': False,
                'error': 'You are not authorized to review this evaluation'
            }, status=403)
        
        # Get task ratings
        task_ratings = employee_eval.task_ratings.all().select_related('task')
        task_ratings_data = []
        for rating in task_ratings:
            task_ratings_data.append({
                'task': rating.task.tasklist,
                'task_id': rating.task.id,
                'rating': rating.rating,
                'supervisor_rating': rating.supervisor_rating,
                'comments': rating.comments
            })
        
        context_data = {
            'employee_eval': employee_eval,
            'task_ratings': task_ratings_data,
            'is_manager_view': True
        }
        
        return JsonResponse({
            'success': True,
            'html': render_to_string('evaluation/manager_evaluation_content.html', context_data, request=request)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["POST"])
def manager_approve_evaluation(request, evaluation_id):
    try:
        employee_eval = get_object_or_404(
            EmployeeEvaluation,
            id=evaluation_id,
            status='manager_review'
        )
        
        is_manager_approver = False
        
        if hasattr(employee_eval.employee, 'employment_info') and employee_eval.employee.employment_info:
            supervisor = employee_eval.employee.employment_info.approver
            if hasattr(supervisor, 'employment_info') and supervisor.employment_info and supervisor.employment_info.approver == request.user:
                is_manager_approver = True
        
        if not is_manager_approver:
            return JsonResponse({
                'success': False,
                'error': 'You are not authorized to approve this evaluation'
            }, status=403)
        
        action = request.POST.get('action')
        manager_comments = request.POST.get('manager_comments', '').strip()
        
        if action not in ['approve', 'disapprove']:
            return JsonResponse({
                'success': False,
                'error': 'Invalid action specified'
            }, status=400)
        
        # Validate comments for disapproval
        if action == 'disapprove' and not manager_comments:
            return JsonResponse({
                'success': False,
                'error': 'Comments are required when disapproving an evaluation'
            }, status=400)
        
        # Update evaluation based on action
        if action == 'approve':
            employee_eval.status = 'approved'
            employee_eval.manager_completed_at = timezone.now()
            employee_eval.manager_comments = manager_comments
            message = 'Evaluation approved successfully!'

            Notification.objects.create(
                title="Evaluation Approved",
                sender=request.user,
                recipient= employee_eval.employee,
                message=f'{request.user.firstname} {request.user.lastname} has approved your performance evaluation',
                module='evaluation',
                notification_type='approved'
            )

        else:
            employee_eval.manager = None
            employee_eval.manager_completed_at = None
            employee_eval.manager_comments = manager_comments
            # Clear supervisor_completed_at so the model's save() logic does not
            # re-promote the status to manager_review. We want it to remain
            # supervisor_review so the supervisor can re-evaluate.
            employee_eval.supervisor_completed_at = None
            employee_eval.status = "supervisor_review"
            message = 'Evaluation disapproved and returned to supervisor for re-evaluation.'
            
            Notification.objects.create(
                title="Evaluation Disapproved",
                sender=request.user,
                recipient= employee_eval.supervisor,
                message=f'{request.user.firstname} {request.user.lastname} has disapproved your evaluation for {employee_eval.employee.firstname} {employee_eval.employee.lastname}. Please review and re-evaluate.',
                module='evaluation',
                notification_type='disapproved'
            )

        employee_eval.save()
        
        if employee_eval.evaluation_instance:
            if employee_eval.status == 'approved':
                employee_eval.evaluation_instance.status = 'completed'
            else:
                employee_eval.evaluation_instance.status = 'in_progress'
            employee_eval.evaluation_instance.save()
        
        return JsonResponse({
            'success': True,
            'message': message
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

def generate_period_tabs(evaluation):
    
    distinct_periods = EvaluationInstance.objects.filter(
        evaluation=evaluation
    ).values_list('period_start', 'period_end').distinct().order_by('period_start')
    
    tabs = []
    
    for period_start, period_end in distinct_periods:
        period_start_date = period_start.date() if hasattr(period_start, 'date') else period_start
        period_end_date = period_end.date() if hasattr(period_end, 'date') else period_end
        
        if evaluation.duration == 'daily':
            label = period_start_date.strftime('%b %d, %Y')
            value = period_start_date.strftime('%Y-%m-%d')
        
        elif evaluation.duration == 'monthly':
            label = period_start_date.strftime('%B %Y')
            value = period_start_date.strftime('%Y-%m')
        
        elif evaluation.duration == 'quarterly':
            start_month_name = calendar.month_name[period_start_date.month][:3]
            end_month_name = calendar.month_name[period_end_date.month][:3]
            label = f'{start_month_name}-{end_month_name} {period_start_date.year}'
            quarter = ((period_start_date.month - 1) // 3) + 1
            value = f'{period_start_date.year}-Q{quarter}'
        
        elif evaluation.duration == 'yearly':
            label = str(period_start_date.year)
            value = str(period_start_date.year)
        
        tabs.append({
            'label': label,
            'value': value,
            'period_start': period_start_date,
            'period_end': period_end_date
        })
    
    return tabs

def parse_period_value(period_value, duration):
    
    if duration == 'daily':
        date = datetime.strptime(period_value, '%Y-%m-%d').date()
        return date, date
    
    elif duration == 'monthly':
        year, month = map(int, period_value.split('-'))
        start_date = datetime(year, month, 1).date()
        if month == 12:
            end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
        else:
            end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)
        return start_date, end_date
    
    elif duration == 'quarterly':
        year, quarter = period_value.split('-Q')
        year = int(year)
        quarter = int(quarter)
        
        start_month = (quarter - 1) * 3 + 1
        end_month = start_month + 2
        
        start_date = datetime(year, start_month, 1).date()
        end_date = datetime(year, end_month, calendar.monthrange(year, end_month)[1]).date()
        return start_date, end_date
    
    elif duration == 'yearly':
        year = int(period_value)
        start_date = datetime(year, 1, 1).date()
        end_date = datetime(year, 12, 31).date()
        return start_date, end_date
    
    return None, None

@login_required
def evaluation_details(request, evaluation_id):
    evaluation = get_object_or_404(Evaluation, id=evaluation_id)
    
    period_tabs = generate_period_tabs(evaluation)
    
    selected_period = request.POST.get('period', request.GET.get('period', period_tabs[0]['value'] if period_tabs else ''))
    
    # Get all evaluation instances for this evaluation (for total assigned count)
    all_evaluation_instances = EvaluationInstance.objects.filter(evaluation=evaluation).select_related(
        'employee', 'employee__employment_info', 'employee__employment_info__department', 'employee__employment_info__position'
    )
    
    # Initialize evaluation_instances with all instances
    evaluation_instances = all_evaluation_instances
    
    # Filter instances by selected period if provided
    if selected_period:
        # Find the specific period from tabs instead of parsing
        selected_tab = None
        for tab in period_tabs:
            if tab['value'] == selected_period:
                selected_tab = tab
                break
        
        if selected_tab:
            # Use the exact period dates from the tab
            evaluation_instances = evaluation_instances.filter(
                period_start__date=selected_tab['period_start'],
                period_end__date=selected_tab['period_end']
            )
        else:
            # Fallback to all instances if no matching tab found
            evaluation_instances = all_evaluation_instances
    
    # Total Assigned: Count of EvaluationInstance records for the selected period
    total_participants = evaluation_instances.count()
    
    # Get current period statistics
    current_completed_count = evaluation_instances.filter(status='completed').count()
    current_in_progress_count = evaluation_instances.filter(status='in_progress').count()
    current_pending_count = evaluation_instances.filter(status='pending').count()
    
    # Calculate percentage differences with previous period
    participants_change = 0
    submitted_change = 0
    in_progress_change = 0
    pending_change = 0
    
    if selected_period and period_tabs:
        # Find current period index
        current_period_index = None
        for i, tab in enumerate(period_tabs):
            if tab['value'] == selected_period:
                current_period_index = i
                break
        
        if current_period_index is not None and current_period_index > 0:
            # Get previous period
            prev_tab = period_tabs[current_period_index - 1]
            
            # Use exact period dates from the previous tab
            prev_instances = all_evaluation_instances.filter(
                period_start__date=prev_tab['period_start'],
                period_end__date=prev_tab['period_end']
            )
            
            # Calculate previous period counts
            prev_total_count = prev_instances.count()
            prev_completed_count = prev_instances.filter(status='completed').count()
            prev_in_progress_count = prev_instances.filter(status='in_progress').count()
            prev_pending_count = prev_instances.filter(status='pending').count()
            
            # Calculate percentage changes
            if prev_total_count > 0:
                participants_change = round(((total_participants - prev_total_count) / prev_total_count) * 100, 1)
            if prev_completed_count > 0:
                submitted_change = round(((current_completed_count - prev_completed_count) / prev_completed_count) * 100, 1)
            if prev_in_progress_count > 0:
                in_progress_change = round(((current_in_progress_count - prev_in_progress_count) / prev_in_progress_count) * 100, 1)
            if prev_pending_count > 0:
                pending_change = round(((current_pending_count - prev_pending_count) / prev_pending_count) * 100, 1)
    
    # For display purposes, use current period counts
    submitted_count = current_completed_count
    in_progress_count = current_in_progress_count
    pending_count = current_pending_count
    
    # Calculate progress percentage for current period
    if evaluation_instances.count() > 0:
        completed_instances = evaluation_instances.filter(status__in=['in_progress', 'completed']).count()
        progress_percentage = round((completed_instances / evaluation_instances.count()) * 100)
    else:
        progress_percentage = 0
    
    participants_data = []
    for instance in evaluation_instances:
        try:
            emp_eval = EmployeeEvaluation.objects.get(
                evaluation_instance=instance
            )
            status_display = emp_eval.get_status_display()
            is_submitted = emp_eval.self_completed_at is not None
            submitted_at = emp_eval.self_completed_at
            # Allow viewing if there's any self-assessment data (regardless of supervisor/manager completion)
            evaluation_id_for_view = emp_eval.id if emp_eval.self_completed_at is not None else None
        except EmployeeEvaluation.DoesNotExist:
            status_display = instance.get_status_display()
            is_submitted = instance.status in ['completed']
            submitted_at = None
            evaluation_id_for_view = None
        
        participants_data.append({
            'idnumber': instance.employee.idnumber,
            'name': instance.employee.full_name,
            'username': instance.employee.username,
            'department': instance.employee.employment_info.department.department_name if hasattr(instance.employee, 'employment_info') and instance.employee.employment_info and instance.employee.employment_info.department else 'N/A',
            'position': instance.employee.employment_info.position.position if hasattr(instance.employee, 'employment_info') and instance.employee.employment_info and instance.employee.employment_info.position else 'N/A',
            'status': status_display,
            'is_submitted': is_submitted,
            'submitted_at': submitted_at,
            'evaluation_id': evaluation_id_for_view
        })
    
    search = request.POST.get('search', request.GET.get('search', ''))
    if search:
        participants_data = [
            p for p in participants_data 
            if search.lower() in p['name'].lower() or 
               search.lower() in p['idnumber'].lower() or
               search.lower() in p['department'].lower() or
               search.lower() in p['position'].lower()
        ]
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'success': True,
            'total_participants': total_participants,
            'submitted_count': submitted_count,
            'in_progress_count': in_progress_count,
            'pending_count': pending_count,
            'progress_percentage': progress_percentage,
            'participants_change': participants_change,
            'submitted_change': submitted_change,
            'in_progress_change': in_progress_change,
            'pending_change': pending_change,
            'participants_data': participants_data,
            'period_tabs': period_tabs,
            'selected_period': selected_period,
        })
    
    paginator = Paginator(participants_data, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'evaluation': evaluation,
        'period_tabs': period_tabs,
        'selected_period': selected_period,
        'total_participants': total_participants,
        'submitted_count': submitted_count,
        'in_progress_count': in_progress_count,
        'pending_count': pending_count,
        'progress_percentage': progress_percentage,
        'participants_change': participants_change,
        'submitted_change': submitted_change,
        'in_progress_change': in_progress_change,
        'pending_change': pending_change,
        'participants_data': page_obj,
        'page_obj': page_obj,
        'search': search,
    }
    
    return render(request, 'evaluation/evaluation_details.html', context)

@login_required
def export_evaluation_excel(request, evaluation_id):
    try:
        evaluation = get_object_or_404(Evaluation, id=evaluation_id)
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        summary_sheet = wb.create_sheet("Summary")
        create_summary_sheet(summary_sheet, evaluation)
        
        assessment_sheet = wb.create_sheet("Supervisor Assessment")
        create_supervisor_assessment_sheet(assessment_sheet, evaluation)
        
        import re
        clean_title = re.sub(r'[^\w\s-]', '', evaluation.title).strip()
        clean_title = re.sub(r'[-\s]+', '_', clean_title)
        
        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'Evaluation_{clean_title}_{evaluation.fiscal_year_label}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error exporting evaluation {evaluation_id}: {str(e)}")
        logger.error(traceback.format_exc())
        messages.error(request, f'Error exporting evaluation: {str(e)}')
        return redirect('evaluation_details', evaluation_id=evaluation_id)

def create_summary_sheet(sheet, evaluation):
    
    header_font = Font(bold=True, size=12)
    sub_header_font = Font(italic=True, size=11)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    sheet['A1'] = "RYONAN ELECTRIC PHILIPPINES CORPORATION"
    sheet['A1'].font = header_font
    sheet['A1'].alignment = Alignment(horizontal='left', vertical='center')
    sheet['A2'] = f"Evaluation overall summary for {evaluation.title}"
    sheet['A2'].font = sub_header_font
    sheet['A2'].alignment = Alignment(horizontal='left', vertical='center')
    
    evaluation_instances = EvaluationInstance.objects.filter(
        evaluation=evaluation
    ).select_related(
        'employee', 'employee__employment_info', 
        'employee__employment_info__department'
    ).prefetch_related('employee_evaluation__task_ratings')
    
    employees_data = {}
    for instance in evaluation_instances:
        employee_id = instance.employee.id
        if employee_id not in employees_data:
            employees_data[employee_id] = {
                'employee': instance.employee,
                'instances': []
            }
        employees_data[employee_id]['instances'].append(instance)
    
    if evaluation.duration == 'yearly':
        headers_row1 = ['ID Number', 'Name', 'Department', 'Line', 'Approver', 'Evaluation', 'Behavioral']
        headers_row2 = ['', '', '', '', '', 'Total Rating', 'Total Rating']
        start_row = 5
    else:
        headers_row1 = ['ID Number', 'Name', 'Department', 'Line', 'Approver', 'Evaluation', '', '', '', '', 'Behavioral', '', '', '', '']
        headers_row2 = ['', '', '', '', '', 'Q1', 'Q2', 'Q3', 'Q4', 'Total', 'Q1', 'Q2', 'Q3', 'Q4', 'Total']
        start_row = 5
    
    for col, header in enumerate(headers_row1, 1):
        if header:
            cell = sheet.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = yellow_fill
            cell.alignment = center_alignment
            cell.border = border
    
    for col, header in enumerate(headers_row2, 1):
        if header:
            cell = sheet.cell(row=start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = yellow_fill
            cell.alignment = center_alignment
            cell.border = border
        
    sheet.merge_cells('A4:A5')  # ID Number
    sheet.merge_cells('B4:B5')  # Name
    sheet.merge_cells('C4:C5')  # Department
    sheet.merge_cells('D4:D5')  # Line
    sheet.merge_cells('E4:E5')  # Approver
    

    if evaluation.duration == 'yearly':
        pass
    else:
        sheet.merge_cells('F4:J4')
        sheet.merge_cells('K4:O4')
    
    row = start_row + 1
    for employee_id, employee_data in employees_data.items():
        employee = employee_data['employee']
        instances = employee_data['instances']
        emp_info = getattr(employee, 'employment_info', None)
        
        # Basic employee info
        sheet.cell(row=row, column=1, value=employee.idnumber).border = border
        sheet.cell(row=row, column=2, value=employee.full_name).border = border
        sheet.cell(row=row, column=3, value=emp_info.department.department_name if emp_info and emp_info.department else 'N/A').border = border
        sheet.cell(row=row, column=4, value=emp_info.line.line_name if emp_info and emp_info.line else 'N/A').border = border
        sheet.cell(row=row, column=5, value=emp_info.approver.full_name if emp_info and emp_info.approver else 'N/A').border = border
        
        if evaluation.duration == 'yearly':
            eval_total = calculate_yearly_evaluation_rating_for_employee(instances)
            behavioral_total = calculate_yearly_behavioral_rating_for_employee(instances)
            
            eval_value = 0 if eval_total == 'N/A' else eval_total
            behavioral_value = 0 if behavioral_total == 'N/A' else behavioral_total
            
            eval_cell = sheet.cell(row=row, column=6, value=eval_value)
            eval_cell.border = border
            eval_cell.alignment = center_alignment
            
            behavioral_cell = sheet.cell(row=row, column=7, value=behavioral_value)
            behavioral_cell.border = border
            behavioral_cell.alignment = center_alignment
        else:
            eval_q1, eval_q2, eval_q3, eval_q4, eval_total = calculate_quarterly_evaluation_ratings_for_employee(instances, evaluation.duration)
            behav_q1, behav_q2, behav_q3, behav_q4, behav_total = calculate_quarterly_behavioral_ratings_for_employee(instances, evaluation.duration)
            
            eval_values = [
                0 if eval_q1 == 'N/A' else eval_q1,
                0 if eval_q2 == 'N/A' else eval_q2,
                0 if eval_q3 == 'N/A' else eval_q3,
                0 if eval_q4 == 'N/A' else eval_q4,
                0 if eval_total == 'N/A' else eval_total
            ]
            
            behav_values = [
                0 if behav_q1 == 'N/A' else behav_q1,
                0 if behav_q2 == 'N/A' else behav_q2,
                0 if behav_q3 == 'N/A' else behav_q3,
                0 if behav_q4 == 'N/A' else behav_q4,
                0 if behav_total == 'N/A' else behav_total
            ]
            
            for i, value in enumerate(eval_values):
                cell = sheet.cell(row=row, column=6+i, value=value)
                cell.border = border
                cell.alignment = center_alignment
            
            for i, value in enumerate(behav_values):
                cell = sheet.cell(row=row, column=11+i, value=value)
                cell.border = border
                cell.alignment = center_alignment
        
        row += 1
    
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[column_letter].width = adjusted_width

def create_supervisor_assessment_sheet(sheet, evaluation):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # Set up styles
    header_font = Font(bold=True, size=12)
    sub_header_font = Font(italic=True, size=11)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    wrap_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    data_alignment = Alignment(horizontal='left', vertical='center')
    
    # Company header
    sheet['A1'] = "RYONAN ELECTRIC PHILIPPINES CORPORATION"
    sheet['A1'].font = header_font
    sheet['A1'].alignment = Alignment(horizontal='left', vertical='center')
    
    # Sub-header
    sheet['A2'] = f"Evaluation overall summary for {evaluation.title}"
    sheet['A2'].font = sub_header_font
    sheet['A2'].alignment = Alignment(horizontal='left', vertical='center')
    
    # Headers
    headers_row1 = ['ID Number', 'Name', 'Department', 'Line', 'Approver', 'Assessment', '', '', '', '', '', 'Behavioral', '', '', '', '']
    headers_row2 = ['', '', '', '', '', 'Strengths', 'Weaknesses', 'Training Required', 'Supervisor Comments', 'Employee Comment', 'Manager Comment', 'Cost Consciousness', 'Dependability', 'Communication', 'Work Ethics', 'Attendance']
    
    # Set headers first (before merging)
    for col, header in enumerate(headers_row1, 1):
        if header:  # Only set non-empty headers
            cell = sheet.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = yellow_fill
            cell.alignment = center_alignment
            cell.border = border
    
    for col, header in enumerate(headers_row2, 1):
        if header:  # Only set non-empty headers
            cell = sheet.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = yellow_fill
            cell.alignment = center_alignment
            cell.border = border
    
    # Merge cells for basic info headers (vertically across 2 rows)
    sheet.merge_cells('A4:A5')  # ID Number
    sheet.merge_cells('B4:B5')  # Name
    sheet.merge_cells('C4:C5')  # Department
    sheet.merge_cells('D4:D5')  # Line
    sheet.merge_cells('E4:E5')  # Approver
    
    # Merge cells for main categories after setting headers
    sheet.merge_cells('F4:K4')  # Assessment
    sheet.merge_cells('L4:P4')  # Behavioral
    
    # Get evaluation instances and group by employee to avoid duplicates
    evaluation_instances = EvaluationInstance.objects.filter(
        evaluation=evaluation
    ).select_related(
        'employee', 'employee__employment_info', 
        'employee__employment_info__department', 'employee__employment_info__line'
    ).prefetch_related('employee_evaluation')
    
    # Group instances by employee to ensure one row per employee
    employees_data = {}
    for instance in evaluation_instances:
        employee_id = instance.employee.id
        if employee_id not in employees_data:
            employees_data[employee_id] = {
                'employee': instance.employee,
                'instances': []
            }
        employees_data[employee_id]['instances'].append(instance)
    
    # Populate data - one row per employee
    row = 6
    for employee_id, employee_data in employees_data.items():
        employee = employee_data['employee']
        instances = employee_data['instances']
        emp_info = getattr(employee, 'employment_info', None)
        
        # Get the most recent employee evaluation for this employee
        employee_eval = None
        supervisor_name = 'N/A'
        for instance in instances:
            eval_obj = getattr(instance, 'employee_evaluation', None)
            if eval_obj:
                employee_eval = eval_obj
                if eval_obj.supervisor:
                    supervisor_name = eval_obj.supervisor.full_name
                break
        
        # Basic employee info
        id_cell = sheet.cell(row=row, column=1, value=employee.idnumber)
        id_cell.border = border
        id_cell.alignment = data_alignment
        
        name_cell = sheet.cell(row=row, column=2, value=employee.full_name)
        name_cell.border = border
        name_cell.alignment = data_alignment
        
        dept_cell = sheet.cell(row=row, column=3, value=emp_info.department.department_name if emp_info and emp_info.department else 'N/A')
        dept_cell.border = border
        dept_cell.alignment = data_alignment
        
        line_cell = sheet.cell(row=row, column=4, value=emp_info.line.line_name if emp_info and emp_info.line else 'N/A')
        line_cell.border = border
        line_cell.alignment = data_alignment
        
        # Approver
        approver_cell = sheet.cell(row=row, column=5, value=emp_info.approver.full_name if emp_info and emp_info.approver else 'N/A')
        approver_cell.border = border
        approver_cell.alignment = data_alignment
        
        if employee_eval:
            # Assessment data with text wrapping
            strengths_cell = sheet.cell(row=row, column=6, value=employee_eval.strengths or 'N/A')
            strengths_cell.border = border
            strengths_cell.alignment = wrap_alignment
            
            weaknesses_cell = sheet.cell(row=row, column=7, value=employee_eval.weaknesses or 'N/A')
            weaknesses_cell.border = border
            weaknesses_cell.alignment = wrap_alignment
            
            training_cell = sheet.cell(row=row, column=8, value=employee_eval.training_required or 'N/A')
            training_cell.border = border
            training_cell.alignment = wrap_alignment
            
            supervisor_comments_cell = sheet.cell(row=row, column=9, value=employee_eval.supervisor_comments or 'N/A')
            supervisor_comments_cell.border = border
            supervisor_comments_cell.alignment = wrap_alignment
            
            employee_comments_cell = sheet.cell(row=row, column=10, value=employee_eval.employee_comments or 'N/A')
            employee_comments_cell.border = border
            employee_comments_cell.alignment = wrap_alignment
            
            manager_comments_cell = sheet.cell(row=row, column=11, value=employee_eval.manager_comments or 'N/A')
            manager_comments_cell.border = border
            manager_comments_cell.alignment = wrap_alignment
            
            # Behavioral data - collect all instances for this employee with text wrapping
            all_employee_evals = [getattr(inst, 'employee_evaluation', None) for inst in instances if getattr(inst, 'employee_evaluation', None)]
            
            # Cost Consciousness
            cost_comments = format_multiple_comments([ev.cost_consciousness_comments for ev in all_employee_evals if ev.cost_consciousness_comments])
            cost_cell = sheet.cell(row=row, column=12, value=cost_comments or 'N/A')
            cost_cell.border = border
            cost_cell.alignment = wrap_alignment
            
            # Dependability
            depend_comments = format_multiple_comments([ev.dependability_comments for ev in all_employee_evals if ev.dependability_comments])
            depend_cell = sheet.cell(row=row, column=13, value=depend_comments or 'N/A')
            depend_cell.border = border
            depend_cell.alignment = wrap_alignment
            
            # Communication
            comm_comments = format_multiple_comments([ev.communication_comments for ev in all_employee_evals if ev.communication_comments])
            comm_cell = sheet.cell(row=row, column=14, value=comm_comments or 'N/A')
            comm_cell.border = border
            comm_cell.alignment = wrap_alignment
            
            # Work Ethics
            ethics_comments = format_multiple_comments([ev.work_ethics_comments for ev in all_employee_evals if ev.work_ethics_comments])
            ethics_cell = sheet.cell(row=row, column=15, value=ethics_comments or 'N/A')
            ethics_cell.border = border
            ethics_cell.alignment = wrap_alignment
            
            # Attendance
            attend_comments = format_multiple_comments([ev.attendance_comments for ev in all_employee_evals if ev.attendance_comments])
            attend_cell = sheet.cell(row=row, column=16, value=attend_comments or 'N/A')
            attend_cell.border = border
            attend_cell.alignment = wrap_alignment
        else:
            # No evaluation data - add with proper alignment for consistency
            for col in range(6, 17):
                cell = sheet.cell(row=row, column=col, value='N/A')
                cell.border = border
                if col >= 6:  # Text columns should have wrap alignment
                    cell.alignment = wrap_alignment
        
        # Set row height to accommodate wrapped text
        sheet.row_dimensions[row].height = 60
        row += 1
    
    # Set specific column widths for better text wrapping
    # Basic info columns
    sheet.column_dimensions['A'].width = 12  # ID Number
    sheet.column_dimensions['B'].width = 20  # Name
    sheet.column_dimensions['C'].width = 15  # Department
    sheet.column_dimensions['D'].width = 12  # Line
    sheet.column_dimensions['E'].width = 18  # Approver
    
    # Assessment columns with wider widths for text wrapping
    sheet.column_dimensions['F'].width = 25  # Strengths
    sheet.column_dimensions['G'].width = 25  # Weaknesses
    sheet.column_dimensions['H'].width = 25  # Training Required
    sheet.column_dimensions['I'].width = 25  # Supervisor Comments
    sheet.column_dimensions['J'].width = 25  # Employee Comment
    sheet.column_dimensions['K'].width = 25  # Manager Comment
    
    # Behavioral columns with wider widths for text wrapping
    sheet.column_dimensions['L'].width = 25  # Cost Consciousness
    sheet.column_dimensions['M'].width = 25  # Dependability
    sheet.column_dimensions['N'].width = 25  # Communication
    sheet.column_dimensions['O'].width = 25  # Work Ethics
    sheet.column_dimensions['P'].width = 25  # Attendance

def format_multiple_comments(comments_list):
    """Format multiple comments as bulleted list"""
    if not comments_list:
        return None
    
    if len(comments_list) == 1:
        return comments_list[0]
    
    return '\n'.join([f'• {comment}' for comment in comments_list if comment.strip()])

def calculate_yearly_evaluation_rating(instance):
    """Calculate yearly evaluation rating"""
    employee_eval = getattr(instance, 'employee_evaluation', None)
    if not employee_eval:
        return 'N/A'
    
    task_ratings = employee_eval.task_ratings.filter(supervisor_rating__isnull=False)
    if not task_ratings.exists():
        return 'N/A'
    
    total_rating = sum(tr.supervisor_rating for tr in task_ratings)
    return round(total_rating / task_ratings.count(), 2)

def calculate_yearly_behavioral_rating(instance):
    """Calculate yearly behavioral rating"""
    employee_eval = getattr(instance, 'employee_evaluation', None)
    if not employee_eval:
        return 'N/A'
    
    avg_behavioral = employee_eval.average_supervisor_criteria_rating
    return round(avg_behavioral, 2) if avg_behavioral is not None else 'N/A'

def get_fiscal_quarter_from_date(date):
    """
    Determine which fiscal quarter a date falls into
    Q1: May – July
    Q2: August – October  
    Q3: November – January
    Q4: February – April
    """
    month = date.month
    
    if month in [5, 6, 7]:  # May, June, July
        return 1
    elif month in [8, 9, 10]:  # August, September, October
        return 2
    elif month in [11, 12, 1]:  # November, December, January
        return 3
    elif month in [2, 3, 4]:  # February, March, April
        return 4
    else:
        return None

def calculate_quarterly_evaluation_ratings(instance, duration):
    """Calculate quarterly breakdown for evaluation ratings based on actual period dates"""
    employee_eval = getattr(instance, 'employee_evaluation', None)
    if not employee_eval:
        return 'N/A', 'N/A', 'N/A', 'N/A', 'N/A'
    
    # Get all evaluation instances for this employee and evaluation
    evaluation_instances = instance.evaluation.instances.filter(
        employee=instance.employee
    ).select_related('employee_evaluation')
    
    # Initialize quarterly data
    q1_data = []
    q2_data = []
    q3_data = []
    q4_data = []
    
    # Group instances by fiscal quarter based on their period dates
    for eval_instance in evaluation_instances:
        eval_obj = getattr(eval_instance, 'employee_evaluation', None)
        if not eval_obj:
            continue
            
        task_ratings = eval_obj.task_ratings.filter(supervisor_rating__isnull=False)
        if not task_ratings.exists():
            continue
            
        # Determine quarter based on period_start date
        quarter = get_fiscal_quarter_from_date(eval_instance.period_start)
        
        # Calculate average rating for this instance
        total_supervisor_ratings = sum(tr.supervisor_rating for tr in task_ratings)
        total_tasks = task_ratings.count()
        avg_rating = total_supervisor_ratings / total_tasks
        
        # Apply duration-specific calculation
        if duration == 'monthly':
            # For monthly: divide by months in period
            period_months = max(1, (eval_instance.period_end - eval_instance.period_start).days // 30)
            adjusted_rating = avg_rating / period_months
        elif duration == 'daily':
            # For daily: divide by days in period
            period_days = max(1, (eval_instance.period_end - eval_instance.period_start).days)
            adjusted_rating = avg_rating / period_days
        else:
            # For quarterly/yearly: use rating as is
            adjusted_rating = avg_rating
            
        # Add to appropriate quarter
        if quarter == 1:
            q1_data.append(adjusted_rating)
        elif quarter == 2:
            q2_data.append(adjusted_rating)
        elif quarter == 3:
            q3_data.append(adjusted_rating)
        elif quarter == 4:
            q4_data.append(adjusted_rating)
    
    # Calculate quarterly averages
    q1 = round(sum(q1_data) / len(q1_data), 2) if q1_data else 'N/A'
    q2 = round(sum(q2_data) / len(q2_data), 2) if q2_data else 'N/A'
    q3 = round(sum(q3_data) / len(q3_data), 2) if q3_data else 'N/A'
    q4 = round(sum(q4_data) / len(q4_data), 2) if q4_data else 'N/A'
    
    # Calculate total from quarters that have data
    valid_quarters = [q for q in [q1, q2, q3, q4] if q != 'N/A']
    total = round(sum(valid_quarters) / len(valid_quarters), 2) if valid_quarters else 'N/A'
    
    return q1, q2, q3, q4, total

def calculate_quarterly_behavioral_ratings(instance, duration):
    """Calculate quarterly breakdown for behavioral ratings based on actual period dates"""
    # Get all evaluation instances for this employee and evaluation
    evaluation_instances = instance.evaluation.instances.filter(
        employee=instance.employee
    ).select_related('employee_evaluation')
    
    # Initialize quarterly data
    q1_data = []
    q2_data = []
    q3_data = []
    q4_data = []
    
    # Group instances by fiscal quarter based on their period dates
    for eval_instance in evaluation_instances:
        eval_obj = getattr(eval_instance, 'employee_evaluation', None)
        if not eval_obj:
            continue
            
        avg_behavioral = eval_obj.average_supervisor_criteria_rating
        if avg_behavioral is None:
            continue
            
        # Determine quarter based on period_start date
        quarter = get_fiscal_quarter_from_date(eval_instance.period_start)
        
        # Apply duration-specific calculation
        if duration == 'monthly':
            # For monthly: divide by months in period
            period_months = max(1, (eval_instance.period_end - eval_instance.period_start).days // 30)
            adjusted_rating = avg_behavioral / period_months
        elif duration == 'daily':
            # For daily: divide by days in period
            period_days = max(1, (eval_instance.period_end - eval_instance.period_start).days)
            adjusted_rating = avg_behavioral / period_days
        else:
            # For quarterly/yearly: use rating as is
            adjusted_rating = avg_behavioral
            
        # Add to appropriate quarter
        if quarter == 1:
            q1_data.append(adjusted_rating)
        elif quarter == 2:
            q2_data.append(adjusted_rating)
        elif quarter == 3:
            q3_data.append(adjusted_rating)
        elif quarter == 4:
            q4_data.append(adjusted_rating)
    
    # Calculate quarterly averages
    q1 = round(sum(q1_data) / len(q1_data), 2) if q1_data else 'N/A'
    q2 = round(sum(q2_data) / len(q2_data), 2) if q2_data else 'N/A'
    q3 = round(sum(q3_data) / len(q3_data), 2) if q3_data else 'N/A'
    q4 = round(sum(q4_data) / len(q4_data), 2) if q4_data else 'N/A'
    
    # Calculate total from quarters that have data
    valid_quarters = [q for q in [q1, q2, q3, q4] if q != 'N/A']
    total = round(sum(valid_quarters) / len(valid_quarters), 2) if valid_quarters else 'N/A'
    
    return q1, q2, q3, q4, total

@login_required
def get_evaluation_assessment_details(request, evaluation_id):
    """Get detailed assessment information for offcanvas display"""
    try:
        employee_evaluation = get_object_or_404(EmployeeEvaluation, id=evaluation_id)
        
        # Get task ratings with supervisor ratings
        task_ratings = employee_evaluation.task_ratings.select_related('task').all()
        
        # Calculate averages
        self_avg_rating = employee_evaluation.average_rating
        supervisor_avg_rating = employee_evaluation.average_supervisor_rating
        supervisor_criteria_avg = employee_evaluation.average_supervisor_criteria_rating
        
        # Prepare task ratings data
        task_ratings_data = []
        for rating in task_ratings:
            task_ratings_data.append({
                'task': rating.task.tasklist,
                'self_rating': rating.rating,
                'supervisor_rating': rating.supervisor_rating,
                'comments': rating.comments or 'No comments'
            })
        
        context = {
            'employee_evaluation': employee_evaluation,
            'task_ratings': task_ratings_data,
            'self_avg_rating': round(self_avg_rating, 1) if self_avg_rating else 0,
            'supervisor_avg_rating': round(supervisor_avg_rating, 1) if supervisor_avg_rating else None,
            'supervisor_criteria_avg': round(supervisor_criteria_avg, 1) if supervisor_criteria_avg else None,
        }
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            html = render_to_string('evaluation/evaluation_assessment_details.html', context, request=request)
            return JsonResponse({
                'success': True,
                'html': html
            })
        else:
            return render(request, 'evaluation/evaluation_assessment_details.html', context)
            
    except Exception as e:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
        else:
            messages.error(request, f'Error loading evaluation details: {str(e)}')
            return redirect('admin_evaluation')

def calculate_yearly_evaluation_rating_for_employee(instances):
    """Calculate yearly evaluation rating for all instances of an employee"""
    total_rating = 0
    count = 0
    
    for instance in instances:
        employee_eval = getattr(instance, 'employee_evaluation', None)
        if not employee_eval:
            continue
            
        task_ratings = employee_eval.task_ratings.filter(supervisor_rating__isnull=False)
        if task_ratings.exists():
            total_rating += sum(tr.supervisor_rating for tr in task_ratings) / task_ratings.count()
            count += 1
    
    return round(total_rating / count, 1) if count > 0 else 'N/A'

def calculate_yearly_behavioral_rating_for_employee(instances):
    """Calculate yearly behavioral rating for all instances of an employee"""
    total_rating = 0
    count = 0
    
    for instance in instances:
        employee_eval = getattr(instance, 'employee_evaluation', None)
        if not employee_eval:
            continue
            
        avg_rating = employee_eval.average_supervisor_criteria_rating
        if avg_rating and avg_rating > 0:
            total_rating += avg_rating
            count += 1
    
    return round(total_rating / count, 1) if count > 0 else 'N/A'

def calculate_quarterly_evaluation_ratings_for_employee(instances, duration):
    """Calculate quarterly evaluation ratings for all instances of an employee"""
    # Initialize quarterly data
    q1_data = []
    q2_data = []
    q3_data = []
    q4_data = []
    
    # Group instances by fiscal quarter based on their period dates
    for instance in instances:
        employee_eval = getattr(instance, 'employee_evaluation', None)
        if not employee_eval:
            continue
            
        task_ratings = employee_eval.task_ratings.filter(supervisor_rating__isnull=False)
        if not task_ratings.exists():
            continue
            
        # Determine quarter based on period_start date
        quarter = get_fiscal_quarter_from_date(instance.period_start)
        
        # Calculate average rating for this instance
        total_supervisor_ratings = sum(tr.supervisor_rating for tr in task_ratings)
        total_tasks = task_ratings.count()
        avg_rating = total_supervisor_ratings / total_tasks
        
        # Add to appropriate quarter
        if quarter == 1:
            q1_data.append(avg_rating)
        elif quarter == 2:
            q2_data.append(avg_rating)
        elif quarter == 3:
            q3_data.append(avg_rating)
        elif quarter == 4:
            q4_data.append(avg_rating)
    
    # Calculate quarterly averages
    q1_avg = round(sum(q1_data) / len(q1_data), 1) if q1_data else 'N/A'
    q2_avg = round(sum(q2_data) / len(q2_data), 1) if q2_data else 'N/A'
    q3_avg = round(sum(q3_data) / len(q3_data), 1) if q3_data else 'N/A'
    q4_avg = round(sum(q4_data) / len(q4_data), 1) if q4_data else 'N/A'
    
    # Calculate total average - sum of all 4 quarters divided by 4
    # Replace 'N/A' with 0 for calculation, then divide by 4
    quarter_values = [
        q1_avg if q1_avg != 'N/A' else 0,
        q2_avg if q2_avg != 'N/A' else 0,
        q3_avg if q3_avg != 'N/A' else 0,
        q4_avg if q4_avg != 'N/A' else 0
    ]
    total_avg = round(sum(quarter_values) / 4, 1)
    
    return q1_avg, q2_avg, q3_avg, q4_avg, total_avg

def calculate_quarterly_behavioral_ratings_for_employee(instances, duration):
    """Calculate quarterly behavioral ratings for all instances of an employee"""
    # Initialize quarterly data
    q1_data = []
    q2_data = []
    q3_data = []
    q4_data = []
    
    # Group instances by fiscal quarter based on their period dates
    for instance in instances:
        employee_eval = getattr(instance, 'employee_evaluation', None)
        if not employee_eval:
            continue
            
        avg_rating = employee_eval.average_supervisor_criteria_rating
        if not avg_rating or avg_rating <= 0:
            continue
            
        # Determine quarter based on period_start date
        quarter = get_fiscal_quarter_from_date(instance.period_start)
        
        # Add to appropriate quarter
        if quarter == 1:
            q1_data.append(avg_rating)
        elif quarter == 2:
            q2_data.append(avg_rating)
        elif quarter == 3:
            q3_data.append(avg_rating)
        elif quarter == 4:
            q4_data.append(avg_rating)
    
    # Calculate quarterly averages
    q1_avg = round(sum(q1_data) / len(q1_data), 1) if q1_data else 'N/A'
    q2_avg = round(sum(q2_data) / len(q2_data), 1) if q2_data else 'N/A'
    q3_avg = round(sum(q3_data) / len(q3_data), 1) if q3_data else 'N/A'
    q4_avg = round(sum(q4_data) / len(q4_data), 1) if q4_data else 'N/A'
    
    # Calculate total average - sum of all 4 quarters divided by 4
    # Replace 'N/A' with 0 for calculation, then divide by 4
    quarter_values = [
        q1_avg if q1_avg != 'N/A' else 0,
        q2_avg if q2_avg != 'N/A' else 0,
        q3_avg if q3_avg != 'N/A' else 0,
        q4_avg if q4_avg != 'N/A' else 0
    ]
    total_avg = round(sum(quarter_values) / 4, 1)
    
    return q1_avg, q2_avg, q3_avg, q4_avg, total_avg