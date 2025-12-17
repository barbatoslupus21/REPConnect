from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.db.models import Q, Count
from django.views.decorators.http import require_http_methods
from datetime import datetime, date, timedelta
import calendar
import json
import pandas as pd
import openpyxl
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .models import Holiday, Timelogs, TodoItem
from .forms import HolidayForm
from django.utils import timezone
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from userlogin.models import EmployeeLogin
from .forms import TimelogForm, TimelogImportForm

# Helper function to format datetime in MM/DD/YYYY h:mmam/pm format
def format_timelog_datetime(dt):
    """Format datetime for timelog export in MM/DD/YYYY h:mmam/pm format"""
    if not dt:
        return ''
    return dt.strftime('%m/%d/%Y %I:%M%p').lower()
import pytz

# Helpers to map between UI entry values and model choices
def ui_to_model_entry(entry):
    if not entry:
        return None
    e = str(entry).strip().lower()
    if e in ['timein', 'time in', 'in', 'i', 'time-in', 'time_in', 'tin']:
        return 'IN'
    if e in ['timeout', 'time out', 'out', 'o', 'time-out', 'time_out', 'tout']:
        return 'OUT'
    return None

def model_to_ui_entry(entry):
    if not entry:
        return ''
    e = str(entry).upper()
    if e == 'IN':
        return 'timein'
    if e == 'OUT':
        return 'timeout'
    # fallback to lowercase original
    return str(entry).lower()

def is_night_shift_timein(time_obj):
    """Check if a time-in is for night shift (4:00 PM or later in local timezone)"""
    if not time_obj:
        return False
    # Convert to Manila timezone if timezone-aware
    if timezone.is_aware(time_obj):
        manila_tz = pytz.timezone('Asia/Manila')
        local_time = time_obj.astimezone(manila_tz)
        return local_time.hour >= 16  # 4:00 PM = 16:00
    return time_obj.hour >= 16  # 4:00 PM = 16:00

def is_early_morning_timeout(time_obj):
    """Check if a timeout is in early morning (before 8:00 AM in local timezone)"""
    if not time_obj:
        return False
    # Convert to Manila timezone if timezone-aware
    if timezone.is_aware(time_obj):
        manila_tz = pytz.timezone('Asia/Manila')
        local_time = time_obj.astimezone(manila_tz)
        return local_time.hour < 8  # Before 8:00 AM
    return time_obj.hour < 8  # Before 8:00 AM

def get_local_date(time_obj):
    """Get the date in local timezone (Manila)"""
    if not time_obj:
        return None
    if timezone.is_aware(time_obj):
        manila_tz = pytz.timezone('Asia/Manila')
        local_time = time_obj.astimezone(manila_tz)
        return local_time.date()
    return time_obj.date()

def check_timelog_completeness(user, target_date, all_logs_dict):
    """
    Check if timelogs are complete for a given date, accounting for both day and night shifts.
    
    Args:
        user: Employee user object
        target_date: date object to check
        all_logs_dict: Dictionary with date strings as keys and lists of log objects as values
    
    Returns:
        'complete', 'incomplete', or 'none'
    """
    date_str = target_date.strftime('%Y-%m-%d')
    next_date = target_date + timedelta(days=1)
    next_date_str = next_date.strftime('%Y-%m-%d')
    prev_date = target_date - timedelta(days=1)
    prev_date_str = prev_date.strftime('%Y-%m-%d')
    
    # Get logs for previous date, current date and next date
    prev_logs = all_logs_dict.get(prev_date_str, [])
    current_logs = all_logs_dict.get(date_str, [])
    next_logs = all_logs_dict.get(next_date_str, [])
    
    # Extract time-ins and time-outs for current date
    current_timeins = [log for log in current_logs if log.entry == 'IN']
    current_timeouts = [log for log in current_logs if log.entry == 'OUT']
    
    # Extract time-ins and time-outs for previous date
    prev_timeins = [log for log in prev_logs if log.entry == 'IN']
    prev_timeouts = [log for log in prev_logs if log.entry == 'OUT']
    
    # Extract time-outs for next date (for current day night shift)
    next_timeouts = [log for log in next_logs if log.entry == 'OUT']
    
    # Check if current day has an early morning timeout (before 8 AM) that belongs to previous night shift
    early_morning_timeout_claimed = False
    for timeout in current_timeouts:
        if is_early_morning_timeout(timeout.time):  # Before 8 AM in local timezone
            # Check if there's a night shift time-in on previous day
            for prev_timein in prev_timeins:
                if is_night_shift_timein(prev_timein.time):
                    # Check if previous day's night shift doesn't already have a timeout on same day
                    has_same_day_timeout = False
                    for prev_timeout in prev_timeouts:
                        if prev_timeout.time > prev_timein.time:
                            has_same_day_timeout = True
                            break
                    
                    # Only claim this timeout if previous night shift doesn't have same-day timeout
                    if not has_same_day_timeout:
                        early_morning_timeout_claimed = True
                        break
            if early_morning_timeout_claimed:
                break
    
    # Filter out early morning timeouts that belong to previous night shift
    filtered_current_timeouts = current_timeouts
    if early_morning_timeout_claimed:
        filtered_current_timeouts = [log for log in current_timeouts if not is_early_morning_timeout(log.time)]
    
    # If no logs at all on current date (after filtering), return 'none'
    if not current_timeins and not filtered_current_timeouts:
        return 'none'
    
    # Check for night shift pattern (time-in at or after 4:00 PM on current date)
    night_shift_timein = None
    for timein in current_timeins:
        if is_night_shift_timein(timein.time):
            night_shift_timein = timein
            break
    
    if night_shift_timein:
        # Night shift: Look for timeout on next day (before 8:00 AM in local time) or same day (after time-in)
        has_timeout = False
        
        timein_local_date = get_local_date(night_shift_timein.time)
        
        # Check for timeout on current date (same local date, after time-in)
        for timeout in filtered_current_timeouts:
            timeout_local_date = get_local_date(timeout.time)
            # Must be same local date and after time-in
            if timeout_local_date == timein_local_date and timeout.time > night_shift_timein.time:
                has_timeout = True
                break
        
        # If no same-day timeout, check next day (early morning in local timezone)
        if not has_timeout:
            # Check timeouts in current_logs that are next day in local time
            for timeout in filtered_current_timeouts:
                timeout_local_date = get_local_date(timeout.time)
                if timeout_local_date == timein_local_date + timedelta(days=1) and is_early_morning_timeout(timeout.time):
                    has_timeout = True
                    break
            
            # Also check next_logs
            if not has_timeout:
                for timeout in next_timeouts:
                    if is_early_morning_timeout(timeout.time):  # Before 8:00 AM in local timezone
                        has_timeout = True
                        break
        
        return 'complete' if has_timeout else 'incomplete'
    else:
        # Day shift: Both time-in and time-out should be on the same date
        has_timein = len(current_timeins) > 0
        has_timeout = len(filtered_current_timeouts) > 0
        
        if has_timein and has_timeout:
            return 'complete'
        elif has_timein or has_timeout:
            return 'incomplete'
        else:
            return 'none'

@login_required
def calendar_view(request):
    current_date = timezone.localdate()
    year_raw = request.GET.get('year', '')
    month_raw = request.GET.get('month', '')
    try:
        year = int(year_raw) if year_raw else current_date.year
    except ValueError:
        year = current_date.year
    try:
        month = int(month_raw) if month_raw else current_date.month
    except ValueError:
        month = current_date.month
    cal = calendar.Calendar(firstweekday=6)
    month_days = cal.monthdayscalendar(year, month)

    holidays = Holiday.objects.all()
    holidays_dict = {}
    for holiday in holidays:
        base_date = holiday.date
        rep = getattr(holiday, 'repetition', 'none')
        for week in month_days:
            for day in week:
                if day == 0:
                    continue
                this_date = date(year, month, day)
                add = False
                if rep == 'none' and this_date == base_date:
                    add = True
                elif rep == 'yearly' and this_date.month == base_date.month and this_date.day == base_date.day:
                    add = True
                elif rep == 'monthly' and this_date.day == base_date.day:
                    add = True
                elif rep == 'weekly' and this_date.weekday() == base_date.weekday():
                    add = True
                elif rep == 'daily':
                    add = True
                if add:
                    date_str = this_date.strftime('%Y-%m-%d')
                    if date_str not in holidays_dict:
                        holidays_dict[date_str] = []
                    holidays_dict[date_str].append({
                        'id': holiday.id,
                        'name': holiday.name,
                        'type': holiday.holiday_type,
                        'description': holiday.description or '',
                        'repetition': getattr(holiday, 'repetition', 'none')
                    })
    user = request.user
    timelog_status = {}
    days_in_month = [date(year, month, day) for week in month_days for day in week if day != 0]
    
    # Fetch timelogs for previous month, current month and next month (for night shift checking)
    # Previous month needed to check if early morning timeouts belong to previous day's night shift
    first_of_month = date(year, month, 1)
    prev_month_date = first_of_month - timedelta(days=1)
    next_month_date = first_of_month + timedelta(days=32)
    next_month_date = next_month_date.replace(day=1)
    
    user_timelogs = Timelogs.objects.filter(
        employee=user,
        time__gte=date(prev_month_date.year, prev_month_date.month, 1),
        time__lt=date(next_month_date.year, next_month_date.month, 1) + timedelta(days=31)
    )
    
    # Organize logs by date
    logs_by_date = {}
    for log in user_timelogs:
        # Only count entries that have valid time values
        if log.time and str(log.time).strip():
            d = log.time.date()
            d_str = d.strftime('%Y-%m-%d')
            if d_str not in logs_by_date:
                logs_by_date[d_str] = []
            logs_by_date[d_str].append(log)
    
    # Check completeness for each day in the month
    for d in days_in_month:
        status = check_timelog_completeness(user, d, logs_by_date)
        if status == 'incomplete':
            timelog_status[d.strftime('%Y-%m-%d')] = 'incomplete'
        else:
            timelog_status[d.strftime('%Y-%m-%d')] = 'none'
    
    # Build calendar_timelogs for the month only
    calendar_timelogs = {}
    for d_str, logs in logs_by_date.items():
        # Only include logs from the current month
        log_date = datetime.strptime(d_str, '%Y-%m-%d').date()
        if log_date.year == year and log_date.month == month:
            calendar_timelogs[d_str] = []
            for log in logs:
                if log.time and str(log.time).strip():
                    calendar_timelogs[d_str].append({
                        'entry': model_to_ui_entry(log.entry),
                        'time': log.time.isoformat(),
                    })

    today_date = timezone.localdate()
    context = {
        'year': year,
        'month': month,
        'month_name': calendar.month_name[month],
        'month_days': month_days,
        'holidays': json.dumps(holidays_dict),
        'timelog_status': json.dumps(timelog_status),
        'calendar_timelogs': json.dumps(calendar_timelogs),
        'today': today_date.strftime('%Y-%m-%d'),
        'today_date': today_date,
        'hr_admin': request.user.hr_admin,
        'holiday_form': HolidayForm() if request.user.hr_admin else None,
    }
    return render(request, 'usercalendar/calendar.html', context)

@login_required
@require_http_methods(["POST"])
def add_holiday(request):
    if not request.user.hr_admin:
        messages.error(request, 'You do not have permission to add holidays.')
        return redirect('calendar_view')
    
    form = HolidayForm(request.POST, request.FILES)
    
    if form.is_valid():
        holiday = form.save(commit=False)
        holiday.created_by = request.user
        holiday.save()
        messages.success(request, f'Holiday "{holiday.name}" added successfully.')
        
        # Extract year and month from the holiday date to redirect to the correct month
        if holiday.date:
            year = holiday.date.year
            month = holiday.date.month
            url = reverse('calendar_view') + f'?year={year}&month={month}'
            return redirect(url)
    
    # If form is invalid or no date, show errors and redirect to current month
    for field, errors in form.errors.items():
        for error in errors:
            messages.error(request, f'{field.title()}: {error}')
    
    # Get current year/month from request or use today's date
    current_date = timezone.localdate()
    year = request.GET.get('year', current_date.year)
    month = request.GET.get('month', current_date.month)
    
    url = reverse('calendar_view') + f'?year={year}&month={month}'
    return redirect(url)

@login_required
def get_holidays_api(request):
    year = int(request.GET.get('year'))
    month = int(request.GET.get('month'))
    
    holidays = Holiday.objects.filter(
        date__year=year,
        date__month=month
    ).values('id', 'date', 'name', 'holiday_type', 'description')
    
    holidays_dict = {}
    for holiday in holidays:
        date_str = holiday['date'].strftime('%Y-%m-%d')
        if date_str not in holidays_dict:
            holidays_dict[date_str] = []
        holidays_dict[date_str].append({
            'id': holiday['id'],
            'name': holiday['name'],
            'type': holiday['holiday_type'],
            'description': holiday['description'] or ''
        })
    
    return JsonResponse(holidays_dict)

@login_required
@require_http_methods(["POST"])
def edit_holiday(request, pk):
    if not request.user.hr_admin:
        messages.error(request, 'You do not have permission to edit holidays.')
        return redirect('calendar_view')
    holiday = get_object_or_404(Holiday, pk=pk)
    form = HolidayForm(request.POST, request.FILES, instance=holiday)
    if form.is_valid():
        form.save()
        messages.success(request, f'Holiday "{holiday.name}" updated successfully.')
        
        # Redirect to the month where the holiday is located
        if holiday.date:
            year = holiday.date.year
            month = holiday.date.month
            url = reverse('calendar_view') + f'?year={year}&month={month}'
            return redirect(url)
    
    # If form is invalid, show errors and redirect to current month
    for field, errors in form.errors.items():
        for error in errors:
            messages.error(request, f'{field.title()}: {error}')
    
    # Get current year/month from request or use today's date
    current_date = timezone.localdate()
    year = request.GET.get('year', current_date.year)
    month = request.GET.get('month', current_date.month)
    
    url = reverse('calendar_view') + f'?year={year}&month={month}'
    return redirect(url)

@login_required
@require_http_methods(["POST"])
def delete_holiday(request, pk):
    if not request.user.hr_admin:
        messages.error(request, 'You do not have permission to delete holidays.')
        return redirect('calendar_view')
    holiday = get_object_or_404(Holiday, pk=pk)
    
    # Store the date before deleting to redirect to the correct month
    holiday_date = holiday.date
    holiday.delete()
    messages.success(request, 'Event deleted successfully.')
    
    # Redirect to the month where the holiday was located
    if holiday_date:
        year = holiday_date.year
        month = holiday_date.month
        url = reverse('calendar_view') + f'?year={year}&month={month}'
        return redirect(url)
    
    # If no date, redirect to current month
    current_date = timezone.localdate()
    year = request.GET.get('year', current_date.year)
    month = request.GET.get('month', current_date.month)
    
    url = reverse('calendar_view') + f'?year={year}&month={month}'
    return redirect(url)

@login_required
def get_todos_api(request):
    user = request.user
    date_str = request.GET.get('date')
    if not date_str:
        return JsonResponse({'error': 'Missing date'}, status=400)
    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Invalid date'}, status=400)
    todos = TodoItem.objects.filter(user=user, date=date_obj).order_by('time')
    data = [
        {
            'id': todo.id,
            'description': todo.description,
            'time': todo.time.strftime('%H:%M') if todo.time else '',
            'completed': todo.completed,
        }
        for todo in todos
    ]
    return JsonResponse({'todos': data})

@login_required
@require_http_methods(["POST"])
def add_todo_api(request):
    user = request.user
    data = json.loads(request.body.decode('utf-8'))
    description = data.get('description', '').strip()
    time_str = data.get('time', '').strip()
    date_str = data.get('date', '').strip()
    if not description or not date_str:
        return JsonResponse({'error': 'Missing required fields'}, status=400)
    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Invalid date'}, status=400)
    time_obj = None
    if time_str:
        try:
            time_obj = datetime.strptime(time_str, '%H:%M').time()
        except ValueError:
            return JsonResponse({'error': 'Invalid time'}, status=400)
    todo = TodoItem.objects.create(user=user, date=date_obj, time=time_obj, description=description)
    return JsonResponse({'success': True, 'id': todo.id})

@login_required
@require_http_methods(["POST"])
def toggle_todo_api(request):
    user = request.user
    data = json.loads(request.body.decode('utf-8'))
    todo_id = data.get('id')
    try:
        todo = TodoItem.objects.get(id=todo_id, user=user)
    except TodoItem.DoesNotExist:
        return JsonResponse({'error': 'To-do not found'}, status=404)
    todo.completed = not todo.completed
    todo.save()
    return JsonResponse({'success': True, 'completed': todo.completed})

@login_required
def get_timelogs_api(request):
    user = request.user
    date_str = request.GET.get('date')
    if not date_str:
        return JsonResponse({'error': 'Missing date'}, status=400)
    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Invalid date'}, status=400)
    
    # Get logs for previous date, current date and next date (for night shift checking)
    prev_date = date_obj - timedelta(days=1)
    next_date = date_obj + timedelta(days=1)
    
    prev_logs = Timelogs.objects.filter(employee=user, time__date=prev_date).order_by('time')
    current_logs = Timelogs.objects.filter(employee=user, time__date=date_obj).order_by('time')
    next_logs = Timelogs.objects.filter(employee=user, time__date=next_date).order_by('time')
    
    # Build timelogs for current date
    data = []
    claimed_by_previous_night_shift = None
    
    # First, check if PREVIOUS date has a night shift that should claim current date's early morning timeout
    prev_night_shift_timein = None
    for log in prev_logs:
        if log.time and str(log.time).strip() and log.entry == 'IN':
            if is_night_shift_timein(log.time):
                prev_night_shift_timein = log
                break
    
    # If previous date has night shift, check if it should claim current date's early morning timeout
    if prev_night_shift_timein:
        # Check if previous date already has a timeout (on same day as the night shift time-in)
        prev_has_same_day_timeout = False
        for log in prev_logs:
            if log.entry == 'OUT' and log.time > prev_night_shift_timein.time:
                prev_has_same_day_timeout = True
                break
        
        # If previous night shift doesn't have same-day timeout, check current date's early morning timeout
        if not prev_has_same_day_timeout:
            for log in current_logs:
                if log.time and str(log.time).strip() and log.entry == 'OUT':
                    if is_early_morning_timeout(log.time):  # Before 8 AM in local timezone
                        # This timeout belongs to previous date's night shift
                        claimed_by_previous_night_shift = log.time.isoformat()
                        break
    
    # Add current date logs (excluding those claimed by previous night shift)
    for log in current_logs:
        if log.time and str(log.time).strip():
            # Skip if this timeout was claimed by previous night shift
            if claimed_by_previous_night_shift and log.time.isoformat() == claimed_by_previous_night_shift:
                continue
            data.append({
                'entry': model_to_ui_entry(log.entry),
                'time': log.time.isoformat(),
            })
    
    # Now check if CURRENT date has night shift that needs NEXT day's timeout
    current_night_shift_timein = None
    for log in current_logs:
        if log.time and str(log.time).strip() and log.entry == 'IN':
            if is_night_shift_timein(log.time):
                current_night_shift_timein = log
                break
    
    # If current date has night shift time-in, check for next day timeout (before 8 AM)
    if current_night_shift_timein:
        # Check if there's already a timeout on current date (after time-in)
        has_same_day_timeout = False
        for log in current_logs:
            if log.entry == 'OUT' and log.time > current_night_shift_timein.time:
                has_same_day_timeout = True
                break
        
        # If no same-day timeout, look for next day early morning timeout
        if not has_same_day_timeout:
            for log in next_logs:
                if log.time and str(log.time).strip() and log.entry == 'OUT':
                    if is_early_morning_timeout(log.time):  # Before 8 AM in local timezone
                        # Add this next-day timeout to current date's display
                        data.append({
                            'entry': model_to_ui_entry(log.entry),
                            'time': log.time.isoformat(),
                        })
                        break
    
    # Build previous and next day data for reference
    prev_day_data = []
    next_day_data = []
    
    # Check if previous day has night shift that needs current day's early timeout
    prev_night_shift_needs_timeout = False
    if prev_night_shift_timein:
        prev_has_timeout = False
        for log in prev_logs:
            if log.entry == 'OUT' and log.time > prev_night_shift_timein.time:
                prev_has_timeout = True
                break
        if not prev_has_timeout:
            prev_night_shift_needs_timeout = True
    
    # Build prev_day_data - include the claimed timeout from current date
    for log in prev_logs:
        if log.time and str(log.time).strip():
            prev_day_data.append({
                'entry': model_to_ui_entry(log.entry),
                'time': log.time.isoformat(),
            })
    
    # If previous night shift claimed current day's early timeout, add it to prev_day_data
    if prev_night_shift_needs_timeout and claimed_by_previous_night_shift:
        for log in current_logs:
            if log.time.isoformat() == claimed_by_previous_night_shift:
                prev_day_data.append({
                    'entry': model_to_ui_entry(log.entry),
                    'time': log.time.isoformat(),
                })
                break
    
    for log in next_logs:
        if log.time and str(log.time).strip():
            next_day_data.append({
                'entry': model_to_ui_entry(log.entry),
                'time': log.time.isoformat(),
            })
    
    return JsonResponse({
        'prev_day_timelogs': prev_day_data,
        'timelogs': data,
        'next_day_timelogs': next_day_data
    })

# TIME LOGS

def hr_admin_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return JsonResponse({'error': 'Authentication required'}, status=401)
        if not (request.user.hr_admin):
            return JsonResponse({'error': 'HR admin access required'}, status=403)
        return view_func(request, *args, **kwargs)
    return wrapper

@login_required
@hr_admin_required
def timelogs_page(request):
    return render(request, 'usercalendar/timelogs.html')

@login_required
@hr_admin_required
@require_http_methods(["GET"])
def get_employees_with_timelogs(request):
    try:
        search = request.GET.get('search', '').strip()
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        department = request.GET.get('department')
        
        # Base queryset
        employees = EmployeeLogin.objects.filter(
            active=True,
            wire_admin=False,
            clinic_admin=False,
            iad_admin=False,
            accounting_admin=False,
            hr_admin=False,
            hr_manager=False,
            mis_admin=False
        )
        
        # Apply search filter
        if search:
            employees = employees.filter(
                Q(firstname__icontains=search) |
                Q(lastname__icontains=search) |
                Q(name__icontains=search) |
                Q(email__icontains=search) |
                Q(idnumber__icontains=search)
            )
        
        # Apply department filter (if implemented in your model)
        if department and department != 'all':
            # Add department filtering logic based on your model structure
            pass
        
        # Get employees with their time logs
        employees_data = []
        for employee in employees:
            timelogs_query = employee.timelogs_set.all()
            
            # Apply date filter to time logs
            if start_date:
                try:
                    start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
                    timelogs_query = timelogs_query.filter(time__gte=start_datetime)
                except ValueError:
                    pass
            
            if end_date:
                try:
                    end_datetime = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
                    timelogs_query = timelogs_query.filter(time__lt=end_datetime)
                except ValueError:
                    pass
            
            timelogs = timelogs_query.order_by('-time')
            
            employee_data = {
                'id': employee.id,
                'name': f"{employee.firstname or ''} {employee.lastname or ''}".strip(),
                'email': employee.email,
                'idnumber': employee.idnumber,
                'avatar': employee.avatar.url if employee.avatar else None,
                'timelogs': [
                    {
                        'id': log.id,
                        'time': log.time.isoformat(),
                        'date': log.time.strftime('%Y-%m-%d'),
                        'time_only': log.time.strftime('%H:%M'),
                        'entry': model_to_ui_entry(log.entry),
                        'created_at': log.created_at.isoformat() if log.created_at else None,
                    }
                    for log in timelogs
                ]
            }
            employees_data.append(employee_data)
        
        return JsonResponse({
            'success': True,
            'employees': employees_data,
            'total': len(employees_data)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error loading employees: {str(e)}'
        }, status=500)

@login_required
@hr_admin_required
@require_http_methods(["POST"])
def import_timelogs(request):
    """Import time logs from Excel file"""
    try:
        if 'file' not in request.FILES:
            return JsonResponse({
                'success': False,
                'message': 'No file provided'
            }, status=400)
        
        excel_file = request.FILES['file']
        
        # Validate file type
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({
                'success': False,
                'message': 'Invalid file format. Please upload an Excel file (.xlsx or .xls)'
            }, status=400)
        
        # Read Excel file
        try:
            df = pd.read_excel(excel_file)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error reading Excel file: {str(e)}'
            }, status=400)
        
        # Validate required columns
        required_columns = ['ID Number', 'Name', 'Date and Time', 'Entry']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            return JsonResponse({
                'success': False,
                'message': f'Missing required columns: {", ".join(missing_columns)}'
            }, status=400)
        
        # Process each row
        success_count = 0
        error_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                # Read fields with safe checks
                id_number = ''
                name_val = ''
                datetime_raw = row.get('Date and Time') if 'Date and Time' in df.columns else ''
                entry_raw = row.get('Entry') if 'Entry' in df.columns else ''

                if not pd.isna(row.get('ID Number')):
                    id_number = str(row.get('ID Number')).strip()
                if 'Name' in df.columns and not pd.isna(row.get('Name')):
                    name_val = str(row.get('Name')).strip()

                # Validate employee
                if not id_number:
                    errors.append({
                        'id_number': id_number,
                        'name': name_val,
                        'datetime': str(datetime_raw) if not pd.isna(datetime_raw) else '',
                        'entry': str(entry_raw) if not pd.isna(entry_raw) else '',
                        'error': 'Missing ID Number'
                    })
                    error_count += 1
                    continue

                employee = EmployeeLogin.objects.filter(idnumber=id_number).first()
                if not employee:
                    errors.append({
                        'id_number': id_number,
                        'name': name_val,
                        'datetime': str(datetime_raw) if not pd.isna(datetime_raw) else '',
                        'entry': str(entry_raw) if not pd.isna(entry_raw) else '',
                        'error': f'Employee with ID {id_number} not found'
                    })
                    error_count += 1
                    continue

                # Parse datetime
                datetime_str = str(datetime_raw).strip() if not pd.isna(datetime_raw) else ''
                try:
                    if pd.isna(datetime_raw) or datetime_str == '':
                        raise ValueError('Empty datetime')

                    if isinstance(datetime_raw, pd.Timestamp):
                        log_datetime = datetime_raw.to_pydatetime()
                        if log_datetime.tzinfo is None:
                            log_datetime = timezone.make_aware(log_datetime, timezone.get_current_timezone())
                    else:
                        datetime_formats = [
                            '%m/%d/%Y %I:%M:%S%p',    # 08/15/2025 5:30:00pm
                            '%m/%d/%Y %I:%M:%S %p',   # 08/15/2025 5:30:00 pm
                            '%m/%d/%Y %I:%M%p',       # 08/15/2025 5:30pm
                            '%m/%d/%Y %I:%M %p',      # 08/15/2025 5:30 pm
                            '%Y-%m-%d %H:%M:%S',
                            '%Y-%m-%d %H:%M',
                            '%m/%d/%Y %H:%M:%S',
                            '%m/%d/%Y %H:%M',
                        ]
                        datetime_str_clean = datetime_str.strip()
                        for fmt in datetime_formats:
                            try:
                                log_datetime = datetime.strptime(datetime_str_clean, fmt)
                                log_datetime = timezone.make_aware(log_datetime, timezone.get_current_timezone())
                                break
                            except ValueError:
                                continue
                        else:
                            raise ValueError(f'Unable to parse datetime: {datetime_str}')
                except (ValueError, AttributeError) as e:
                    errors.append({
                        'id_number': id_number,
                        'name': name_val,
                        'datetime': datetime_str,
                        'entry': str(entry_raw) if not pd.isna(entry_raw) else '',
                        'error': f'Invalid datetime format: {datetime_str}'
                    })
                    error_count += 1
                    continue

                # Normalize entry type to model choices (IN/OUT)
                entry_val = str(entry_raw).strip() if not pd.isna(entry_raw) else ''
                entry_lower = entry_val.lower()
                if entry_lower in ['timein', 'time in', 'in', 'i', 'time-in', 'time_in', 'tin']:
                    entry_norm = 'IN'
                elif entry_lower in ['timeout', 'time out', 'out', 'o', 'time-out', 'time_out', 'tout']:
                    entry_norm = 'OUT'
                else:
                    errors.append({
                        'id_number': id_number,
                        'name': name_val,
                        'datetime': datetime_str,
                        'entry': entry_val,
                        'error': f'Invalid entry type: {entry_val}'
                    })
                    error_count += 1
                    continue

                # Create or update time log
                timelog, created = Timelogs.objects.get_or_create(
                    employee=employee,
                    time=log_datetime,
                    defaults={'entry': entry_norm}
                )
                if not created:
                    if timelog.entry != entry_norm:
                        timelog.entry = entry_norm
                        timelog.save()

                success_count += 1

            except Exception as e:
                errors.append({
                    'id_number': id_number if 'id_number' in locals() else '',
                    'name': name_val if 'name_val' in locals() else '',
                    'datetime': datetime_str if 'datetime_str' in locals() else '',
                    'entry': entry_val if 'entry_val' in locals() else '',
                    'error': str(e)
                })
                error_count += 1
        
        response_data = {
            'success': True,
            'message': f'Import completed. {success_count} records processed successfully.',
            'success_count': success_count,
            'error_count': error_count
        }
        
        if errors:
            # Limit the number of error rows returned to avoid huge payloads
            limit = 100
            response_data['errors'] = errors[:limit]
            if len(errors) > limit:
                response_data['more_errors'] = len(errors) - limit
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Import failed: {str(e)}'
        }, status=500)

@login_required
@hr_admin_required
@require_http_methods(["GET"])
def export_template(request):
    """Export Excel template for time logs import"""
    try:
        # Create a new workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Time Logs Template"
        
        # Define headers
        headers = ['ID Number', 'Name', 'Date and Time', 'Entry']
        
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Add sample data
        sample_data = [
            ['001', 'John Doe', '08/15/2025 9:00:00am', 'IN'],
            ['001', 'John Doe', '08/15/2025 5:00:00pm', 'OUT'],
            ['002', 'Jane Smith', '08/15/2025 8:30:00am', 'IN'],
            ['002', 'Jane Smith', '08/15/2025 5:30:00pm', 'OUT'],
        ]

        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border        # Adjust column widths
        column_widths = [12, 25, 20, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # Add instructions
        instructions_row = len(sample_data) + 4
        instructions = [
            "Instructions:",
            "1. ID Number: Employee ID number (must match existing employee)",
            "2. Name: Employee full name (for reference only)",
            "3. Date and Time: Format MM/DD/YYYY h:mmam/pm (e.g., 08/15/2025 9:00am or 08/15/2025 5:30pm)",
            "4. Entry: 'timein' or 'timeout' (case insensitive)",
            "",
            "Note: Remove sample data before importing your actual data."
        ]
        
        for i, instruction in enumerate(instructions):
            cell = ws.cell(row=instructions_row + i, column=1, value=instruction)
            if i == 0:  # Instructions header
                cell.font = Font(bold=True)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="timelog_template.xlsx"'
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error generating template: {str(e)}'
        }, status=500)

@login_required
@hr_admin_required
@require_http_methods(["POST"])
def add_timelog(request):
    """Add a new time log entry"""
    try:
        data = json.loads(request.body)
        
        # Get employee
        employee = get_object_or_404(EmployeeLogin, id=data.get('employee_id'))
        
        # Parse datetime
        date_str = data.get('date')
        time_str = data.get('time')
        
        if not date_str or not time_str:
            return JsonResponse({
                'success': False,
                'message': 'Date and time are required'
            }, status=400)
        
        try:
            datetime_str = f"{date_str} {time_str}"
            log_datetime = datetime.strptime(datetime_str, '%Y-%m-%d %H:%M')
            log_datetime = timezone.make_aware(log_datetime, timezone.get_current_timezone())
        except ValueError:
            return JsonResponse({
                'success': False,
                'message': 'Invalid date or time format'
            }, status=400)
        
        # Validate entry type and convert to model value
        entry = data.get('entry')
        model_entry = ui_to_model_entry(entry)
        if model_entry is None:
            return JsonResponse({
                'success': False,
                'message': 'Invalid entry type'
            }, status=400)
        
        # Check for duplicate entries
        existing = Timelogs.objects.filter(
            employee=employee,
            time=log_datetime
        ).exists()
        
        if existing:
            return JsonResponse({
                'success': False,
                'message': 'A time log entry already exists for this date and time'
            }, status=400)
        
        # Create time log
        timelog = Timelogs.objects.create(
            employee=employee,
            time=log_datetime,
            entry=model_entry
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Time log added successfully',
            'timelog': {
                'id': timelog.id,
                'time': timelog.time.isoformat(),
                'entry': model_to_ui_entry(timelog.entry)
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error adding time log: {str(e)}'
        }, status=500)

@login_required
@hr_admin_required
@require_http_methods(["GET"])
def get_timelog(request, timelog_id):
    """Get a specific time log entry"""
    try:
        timelog = get_object_or_404(Timelogs, id=timelog_id)
        
        return JsonResponse({
            'id': timelog.id,
            'employee_name': f"{timelog.employee.firstname or ''} {timelog.employee.lastname or ''}".strip(),
            'employee_id': timelog.employee.id,
            'date': timelog.time.strftime('%Y-%m-%d'),
            'time': timelog.time.strftime('%H:%M'),
            'entry': model_to_ui_entry(timelog.entry),
            'created_at': timelog.created_at.isoformat() if timelog.created_at else None
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error loading time log: {str(e)}'
        }, status=500)

@csrf_exempt
@login_required
@hr_admin_required
@require_http_methods(["PUT"])
def update_timelog(request, timelog_id):
    """Update a time log entry"""
    try:
        timelog = get_object_or_404(Timelogs, id=timelog_id)
        data = json.loads(request.body)

        # Parse datetime
        date_str = data.get('date')
        time_str = data.get('time')

        if not date_str or not time_str:
            return JsonResponse({
                'success': False,
                'message': 'Date and time are required'
            }, status=400)

        try:
            datetime_str = f"{date_str} {time_str}"
            log_datetime = datetime.strptime(datetime_str, '%Y-%m-%d %H:%M')
            log_datetime = timezone.make_aware(log_datetime, timezone.get_current_timezone())
        except ValueError:
            return JsonResponse({
                'success': False,
                'message': 'Invalid date or time format'
            }, status=400)

        # Validate entry type and convert to model value
        entry = data.get('entry')
        model_entry = ui_to_model_entry(entry)
        if model_entry is None:
            return JsonResponse({
                'success': False,
                'message': 'Invalid entry type'
            }, status=400)

        # Check for duplicate entries (excluding current entry)
        existing = Timelogs.objects.filter(
            employee=timelog.employee,
            time=log_datetime
        ).exclude(id=timelog.id).exists()

        if existing:
            return JsonResponse({
                'success': False,
                'message': 'A time log entry already exists for this date and time'
            }, status=400)

        # Update time log
        timelog.time = log_datetime
        timelog.entry = model_entry
        timelog.save()

        return JsonResponse({
            'success': True,
            'message': 'Time log updated successfully'
        })

    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error updating time log: {str(e)}'
        }, status=500)

@login_required
@hr_admin_required
@require_http_methods(["DELETE"])
def delete_timelog(request, timelog_id):
    """Delete a time log entry"""
    try:
        timelog = get_object_or_404(Timelogs, id=timelog_id)
        timelog.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Time log deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error deleting time log: {str(e)}'
        }, status=500)