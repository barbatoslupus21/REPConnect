from django.http import HttpResponseBadRequest
from django.contrib.auth.decorators import user_passes_test
from openpyxl.utils import get_column_letter
from django.views.decorators.http import require_POST
from django.views.decorators.http import require_POST
from django.template.loader import render_to_string
from django.http import JsonResponse
from django.views.decorators.http import require_GET
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.mail import EmailMessage
from django.conf import settings
from django.db.models import Q
from django.core.paginator import Paginator
from django.utils import timezone
from userlogin.models import EmployeeLogin
import pandas as pd
from reportlab.lib.pagesizes import letter
import io
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from io import BytesIO
from .models import Payslip, Loan, Allowance, OJTPayslipData, AllowanceType, LoanType, LoanDeduction, Savings, OJTRate, SavingsType
from .forms import PayslipUploadForm, EmployeeSearchForm, EmailSelectionForm, SavingsUploadForm
from notification.models import Notification
from django.template.loader import render_to_string
from django.views.decorators.http import require_GET
from userlogin.models import EmployeeLogin
from .models import Payslip
from django.core.files.storage import default_storage
from decimal import Decimal
from django.db import transaction, models
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.platypus.frames import Frame
from reportlab.platypus.doctemplate import PageTemplate, BaseDocTemplate
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import os

@login_required(login_url="user-login")
def finance_dashboard(request):
    user = request.user
    context = {}

    # Determine user type and role
    if hasattr(user, 'personalinformation'):
        employment_type = getattr(user.personalinformation, 'employment_type', 'regular')
    else:
        employment_type = 'regular'

    is_admin = user.accounting_admin

    if is_admin:
        # Admin view
        from django.utils import timezone
        from datetime import timedelta
        from django.db.models import Sum
        now = timezone.now()
        first_of_this_month = now.replace(day=1)
        first_of_last_month = (first_of_this_month - timedelta(days=1)).replace(day=1)
        last_of_last_month = first_of_this_month - timedelta(days=1)

        # Employees
        employees_this_month = EmployeeLogin.objects.filter(is_active=True, date_joined__gte=first_of_this_month).count()
        employees_last_month = EmployeeLogin.objects.filter(is_active=True, date_joined__gte=first_of_last_month, date_joined__lt=first_of_this_month).count()
        total_employees = EmployeeLogin.objects.filter(is_active=True).count()
        employees_percent = 0
        employees_positive = True
        if employees_last_month:
            employees_percent = round(((employees_this_month - employees_last_month) / employees_last_month) * 100, 1)
            employees_percent = min(employees_percent, 100)
            employees_positive = employees_this_month >= employees_last_month

        # Employee list for table (exclude current user and all admin flags)
        employee_list = EmployeeLogin.objects.filter(
            is_active=True
        ).exclude(
            id=request.user.id
        ).exclude(
            wire_admin=True
        ).exclude(
            clinic_admin=True
        ).exclude(
            iad_admin=True
        ).exclude(
            accounting_admin=True
        ).exclude(
            hr_admin=True
        ).exclude(
            hr_manager=True
        ).exclude(
            mis_admin=True
        )

        # Search filter
        search = request.GET.get('search', '').strip()
        if search:
            employee_list = employee_list.filter(
                Q(idnumber__icontains=search) |
                Q(firstname__icontains=search) |
                Q(lastname__icontains=search) |
                Q(username__icontains=search) |
                Q(email__icontains=search)
            )

        employee_list = employee_list.order_by('lastname', 'firstname')

        # Paginate employee list (10 per page)
        paginator = Paginator(employee_list, 10)
        page_number = request.GET.get('page')
        employee_page_obj = paginator.get_page(page_number)

        # Payslips
        payslips_this_month = Payslip.objects.filter(cutoff_date__gte=first_of_this_month).count()
        payslips_last_month = Payslip.objects.filter(cutoff_date__gte=first_of_last_month, cutoff_date__lt=first_of_this_month).count()
        total_payslips = Payslip.objects.count()
        payslips_percent = 0
        payslips_positive = True
        if payslips_last_month:
            payslips_percent = round(((payslips_this_month - payslips_last_month) / payslips_last_month) * 100, 1)
            payslips_percent = min(payslips_percent, 100)
            payslips_positive = payslips_this_month >= payslips_last_month

        # Loans
        loans_this_month = Loan.objects.filter(created_at__gte=first_of_this_month).count()
        loans_last_month = Loan.objects.filter(created_at__gte=first_of_last_month, created_at__lt=first_of_this_month).count()
        total_loans = Loan.objects.count()
        loans_percent = 0
        loans_positive = True
        if loans_last_month:
            loans_percent = round(((loans_this_month - loans_last_month) / loans_last_month) * 100, 1)
            loans_percent = min(loans_percent, 100)
            loans_positive = loans_this_month >= loans_last_month

        # Allowances
        allowances_this_month = Allowance.objects.filter(created_at__gte=first_of_this_month).count()
        allowances_last_month = Allowance.objects.filter(created_at__gte=first_of_last_month, created_at__lt=first_of_this_month).count()
        total_allowances = Allowance.objects.count()
        allowances_percent = 0
        allowances_positive = True
        if allowances_last_month:
            allowances_percent = round(((allowances_this_month - allowances_last_month) / allowances_last_month) * 100, 1)
            allowances_percent = min(allowances_percent, 100)
            allowances_positive = allowances_this_month >= allowances_last_month

        # Chart data: total loan balances for month, quarter, year
        # Month
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        # Quarter
        quarter = (now.month - 1) // 3 + 1
        quarter_start_month = 3 * (quarter - 1) + 1
        quarter_start = now.replace(month=quarter_start_month, day=1, hour=0, minute=0, second=0, microsecond=0)
        # Year
        year_start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)

        month_total_balance = Loan.objects.filter(is_active=True, created_at__gte=month_start).aggregate(total=Sum('current_balance'))['total'] or 0
        quarter_total_balance = Loan.objects.filter(is_active=True, created_at__gte=quarter_start).aggregate(total=Sum('current_balance'))['total'] or 0
        year_total_balance = Loan.objects.filter(is_active=True, created_at__gte=year_start).aggregate(total=Sum('current_balance'))['total'] or 0

        context.update({
            'is_admin': True,
            'total_employees': total_employees,
            'employees_percent': employees_percent,
            'employees_positive': employees_positive,
            'total_payslips': total_payslips,
            'payslips_percent': payslips_percent,
            'payslips_positive': payslips_positive,
            'total_loans': total_loans,
            'loans_percent': loans_percent,
            'loans_positive': loans_positive,
            'total_allowances': total_allowances,
            'allowances_percent': allowances_percent,
            'allowances_positive': allowances_positive,
            'employee_list': employee_page_obj,
            'employee_page_obj': employee_page_obj,
            'search': search,
            'month_total_balance': float(month_total_balance),
            'quarter_total_balance': float(quarter_total_balance),
            'year_total_balance': float(year_total_balance),
        })
    else:
        # Employee view
        context.update({
            'is_admin': False,
            'employment_type': employment_type,
        })

        if employment_type == 'ojt':
            context['ojt_payslips'] = OJTPayslipData.objects.filter(employee=user).order_by('-created_at')
        else:
            context.update({
                'payslips': Payslip.objects.filter(employee=user).order_by('-cutoff_date'),
                'loans': Loan.objects.filter(employee=user).order_by('-date_issued'),
                'allowances': Allowance.objects.filter(employee=user).order_by('-start_date'),
            })

    return render(request, 'finance/admin-finance.html', context)

@login_required(login_url="user-login")
def user_finance(request):
    user = request.user
    now = timezone.now()
    if hasattr(user, 'personalinformation'):
        employment_type = getattr(user.personalinformation, 'employment_type', 'regular')
    else:
        employment_type = 'regular'

    # Get payslips, allowances, loans, and savings for the current user
    payslips = Payslip.objects.filter(employee=user).order_by('-cutoff_date')
    from django.db.models import Sum
    allowances = Allowance.objects.filter(employee=user).order_by('-created_at')
    loans = Loan.objects.filter(employee=user).order_by('-created_at')
    
    # Filter savings based on withdrawal date criteria
    from datetime import timedelta
    one_week_ago = now - timedelta(weeks=1)
    
    # Get all savings and filter based on withdrawal criteria
    all_savings = Savings.objects.filter(employee=user).order_by('-created_at')
    savings = []
    
    for saving in all_savings:
        if saving.is_withdrawn:
            # If withdrawn, only show if withdrawal was within the last week
            if saving.withdrawal_date and saving.withdrawal_date >= one_week_ago:
                savings.append(saving)
        else:
            # If not withdrawn, always show
            savings.append(saving)
    
    savings_sorted = sorted(savings, key=lambda s: (s.is_withdrawn, -s.created_at.timestamp() if s.created_at else 0))

    # Dashboard stats: this month vs last month for each metric
    first_of_this_month = now.replace(day=1)
    first_of_last_month = (first_of_this_month - timedelta(days=1)).replace(day=1)
    # Payslips
    payslips_this_month = Payslip.objects.filter(employee=user, cutoff_date__gte=first_of_this_month).count()
    payslips_last_month = Payslip.objects.filter(employee=user, cutoff_date__gte=first_of_last_month, cutoff_date__lt=first_of_this_month).count()
    total_payslips = payslips.count()
    payslips_percent = 0
    payslips_positive = True
    if payslips_last_month:
        payslips_percent = round(((payslips_this_month - payslips_last_month) / payslips_last_month) * 100, 1)
        payslips_percent = min(payslips_percent, 100)
        payslips_positive = payslips_this_month >= payslips_last_month

    # Allowances - total only counts allowances without deposit_date (balance allowances)
    allowances_this_month = Allowance.objects.filter(employee=user, created_at__gte=first_of_this_month).count()
    allowances_last_month = Allowance.objects.filter(employee=user, created_at__gte=first_of_last_month, created_at__lt=first_of_this_month).count()
    total_allowances = allowances.filter(deposit_date__isnull=True).aggregate(total=Sum('amount'))['total'] or 0
    allowances_percent = 0
    allowances_positive = True
    if allowances_last_month:
        allowances_percent = round(((allowances_this_month - allowances_last_month) / allowances_last_month) * 100, 1)
        allowances_percent = min(allowances_percent, 100)
        allowances_positive = allowances_this_month >= allowances_last_month

    # Savings
    savings_this_month = Savings.objects.filter(employee=user, created_at__gte=first_of_this_month).count()
    savings_last_month = Savings.objects.filter(employee=user, created_at__gte=first_of_last_month, created_at__lt=first_of_this_month).count()
    total_savings = Savings.objects.filter(employee=user, is_withdrawn=False).aggregate(total=Sum('amount'))['total'] or 0
    savings_percent = 0
    savings_positive = True
    if savings_last_month:
        savings_percent = round(((savings_this_month - savings_last_month) / savings_last_month) * 100, 1)
        savings_percent = min(savings_percent, 100)
        savings_positive = savings_this_month >= savings_last_month

    # Loans
    loans_this_month = Loan.objects.filter(employee=user, created_at__gte=first_of_this_month).count()
    loans_last_month = Loan.objects.filter(employee=user, created_at__gte=first_of_last_month, created_at__lt=first_of_this_month).count()
    total_loans = Loan.objects.filter(employee=user, is_active=True).aggregate(total=models.Sum('current_balance'))['total'] or 0
    loans_percent = 0
    loans_positive = True
    if loans_last_month:
        loans_percent = round(((loans_this_month - loans_last_month) / loans_last_month) * 100, 1)
        loans_percent = min(loans_percent, 100)
        loans_positive = loans_this_month >= loans_last_month

    # Precompute paid and percent for each loan (similar to employee_finance_details)
    total_active_loan_balance = 0
    for loan in loans:
        try:
            principal = float(getattr(loan, 'principal_amount', 0) or 0)
            current_balance = float(getattr(loan, 'current_balance', 0) or 0)
            if principal > 0:
                paid = principal - current_balance
                percent = (current_balance / principal * 100)
            else:
                paid = 0
                percent = 100  # If no principal, consider fully paid
        except Exception:
            paid = 0
            percent = 100
        loan.paid = paid
        loan.percent = percent
        loan.percent_paid = 100 - percent
        if getattr(loan, 'is_active', False):
            total_active_loan_balance += current_balance

    content = {
        'now': now,
        'employment_type': employment_type,
        'payslips': payslips,
        'ojtpayslips': OJTPayslipData.objects.filter(employee=user).order_by('-created_at'),
        'allowances': allowances,
        'loans': loans,
        'savings': savings_sorted,
        'total_active_loan_balance': total_active_loan_balance,
        'total_payslips': total_payslips,
        'payslips_percent': payslips_percent,
        'payslips_positive': payslips_positive,
        'total_allowances': total_allowances,
        'allowances_percent': allowances_percent,
        'allowances_positive': allowances_positive,
        'total_savings': total_savings,
        'savings_percent': savings_percent,
        'savings_positive': savings_positive,
        'total_loans': total_loans,
        'loans_percent': loans_percent,
        'loans_positive': loans_positive,
    }
    return render(request, 'finance/user-finance.html', content)

# OJT Payslip Upload
@login_required(login_url="user-login")
def ojt_payslip_upload(request):
    if not request.user.accounting_admin:
        return JsonResponse({'success': False, 'message': 'Permission denied'})

    if request.method == 'POST':
        files = request.FILES.getlist('files')
        cut_off = request.POST.get('cutoff_date', '')
        errors = []
        all_error_rows = []  # Collect all error rows from all files
        created = 0
        updated = 0

        if not cut_off:
            return JsonResponse({'success': False, 'message': 'Cut-off period is required'})

        for file in files:
            if not file.name.lower().endswith('.xlsx'):
                errors.append(f'Only Excel (.xlsx) files are supported. File: {file.name}')
                continue
                
            try:
                success, file_errors, file_created, file_updated, error_rows = process_ojt_excel(file, cut_off)
                if success:
                    created += file_created
                    updated += file_updated
                else:
                    errors.extend(file_errors)
                    if error_rows:
                        all_error_rows.extend(error_rows)
            except Exception as e:
                errors.append(f'Error processing {file.name}: {str(e)}')

        if errors:
            return JsonResponse({
                'success': False, 
                'errors': errors, 
                'created': created,
                'updated': updated,
                'error_rows': all_error_rows
            })
        
        # Create notification for successful upload
        if created > 0 or updated > 0:
            Notification.objects.create(
                title="OJT Payslip Upload",
                message=f"Payslip for OJT has been uploaded for cut-off period {cut_off}.",
                sender=request.user,
                recipient=request.user,
                module="finance",
                for_all=True
            )
            
        return JsonResponse({
            'success': True, 
            'created': created,
            'updated': updated,
            'message': f'Successfully processed {created} new records and updated {updated} existing records.'
        })
        
    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)

def process_ojt_excel(file, cut_off):

    errors = []
    error_rows = []  # Store rows with errors for download
    created = 0
    updated = 0
    
    try:
        workbook = openpyxl.load_workbook(file, data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.values)
        
        if len(rows) < 2:
            return False, ['Excel file is empty or has no data'], 0, 0, []
        
        headers = rows[0]
        data_rows = rows[1:]
        
        with transaction.atomic():
            for row_idx, row in enumerate(data_rows, start=2):
                try:
                    # Skip completely empty rows
                    if not row or all(cell is None or (isinstance(cell, str) and cell.strip() == '') for cell in row):
                        continue

                    employee_id = row[0] if row and len(row) > 0 else None
                    if not employee_id:
                        remark = 'Missing employee ID'
                        errors.append(f'Row {row_idx}: {remark}')
                        error_row = list(row[:6]) if row else ['', '', '', '', '', '']
                        # Pad to 6 columns if needed
                        while len(error_row) < 6:
                            error_row.append('')
                        error_row.append(remark)  # Add remarks column
                        error_rows.append(error_row)
                        continue
                    
                    try:
                        employee = EmployeeLogin.objects.get(idnumber=str(employee_id))
                    except EmployeeLogin.DoesNotExist:
                        remark = f'Employee {employee_id} not found'
                        errors.append(f'Row {row_idx}: {remark}')
                        error_row = list(row[:6]) if row else ['', '', '', '', '', '']
                        while len(error_row) < 6:
                            error_row.append('')
                        error_row.append(remark)
                        error_rows.append(error_row)
                        continue
                    
                    def safe_decimal(value):
                        if value is None or value == '':
                            return Decimal('0')
                        try:
                            return Decimal(str(value))
                        except:
                            return Decimal('0')
                    
                    payslip_data = {
                        'employee': employee,
                        'cut_off': cut_off,
                        'regular_day': safe_decimal(row[1] if len(row) > 1 else None),
                        'allowance_day': safe_decimal(row[2] if len(row) > 2 else None),
                        'total_allowance': safe_decimal(row[3] if len(row) > 3 else None),
                        'nd_allowance': safe_decimal(row[4] if len(row) > 4 else None),
                        'grand_total': safe_decimal(row[5] if len(row) > 5 else None),
                        'basic_school_share': safe_decimal(row[6] if len(row) > 6 else None),
                        'basic_ojt_share': safe_decimal(row[7] if len(row) > 7 else None),
                        'deduction': safe_decimal(row[8] if len(row) > 8 else None),
                        'net_ojt_share': safe_decimal(row[9] if len(row) > 9 else None),
                        'rice_allowance': safe_decimal(row[10] if len(row) > 10 else None),
                        'ot_allowance': safe_decimal(row[11] if len(row) > 11 else None),
                        'nd_ot_allowance': safe_decimal(row[12] if len(row) > 12 else None),
                        'special_holiday': safe_decimal(row[13] if len(row) > 13 else None),
                        'legal_holiday': safe_decimal(row[14] if len(row) > 14 else None),
                        'satoff_allowance': safe_decimal(row[15] if len(row) > 15 else None),
                        'rd_ot': safe_decimal(row[16] if len(row) > 16 else None),
                        'adjustment': safe_decimal(row[17] if len(row) > 17 else None),
                        'deduction_2': safe_decimal(row[18] if len(row) > 18 else None),
                        'ot_pay_allowance': safe_decimal(row[19] if len(row) > 19 else None),
                        'total_allow': safe_decimal(row[20] if len(row) > 20 else None),
                        'holiday_hours': safe_decimal(row[21] if len(row) > 21 else None),
                        'rd_ot_days': safe_decimal(row[22] if len(row) > 22 else None),
                        'perfect_attendance': safe_decimal(row[23] if len(row) > 23 else None),
                    }
                    
                    # Check if record exists
                    existing_record = OJTPayslipData.objects.filter(
                        employee=employee,
                        cut_off=cut_off
                    ).first()
                    
                    if existing_record:
                        # Update existing record
                        for field, value in payslip_data.items():
                            if field not in ['employee', 'cut_off']:
                                setattr(existing_record, field, value)
                        existing_record.save()
                        updated += 1
                    else:
                        # Create new record
                        OJTPayslipData.objects.create(**payslip_data)
                        created += 1
                        
                except Exception as e:
                    remark = str(e)
                    errors.append(f'Row {row_idx}: {remark}')
                    error_row = list(row[:6]) if row else ['', '', '', '', '', '']
                    while len(error_row) < 6:
                        error_row.append('')
                    error_row.append(remark)
                    error_rows.append(error_row)
                    continue
        
        workbook.close()
        
        if errors:
            return False, errors, created, updated, error_rows
        return True, [], created, updated, []
        
    except Exception as e:
        return False, [f'Error reading Excel file: {str(e)}'], 0, 0, []

def ojt_payslip_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OJTPayslipTemplate"
    
    headers = [
        'ID_NO',
        'Regular_Day',
        'ALLOWANCE_DAY',
        'TOTAL_ALLOWANCE',
        'ND_ALLOWANCE',
        'GRAND_TOTAL',
        'BASIC_SCHOOL_SHARE',
        'BASIC_OJT_SHARE',
        'DEDUCTION',
        'NET_OJT_SHARE',
        'RICE_ALLOWANCE',
        'OT_ALLOWANCE',
        'ND_OT_ALLOWANCE',
        'SPECIAL_HOLIDAY',
        'LEGAL_HOLIDAY',
        'SATOFF_ALLOWANCE',
        'RD_OT',
        'ADJUSTMENT',
        'DEDUCTION_2',
        'OT_PAY_ALLOWANCE',
        'TOTAL_ALLOW',
        'HOLIDAY_DATE',
        'RD_OT_DATE',
        'PERFECT_ATTENDANCE',
    ]
    
    # Define styles for header (yellow background, bold text, borders)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add headers with yellow background, bold text, and borders
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = bold_font
        cell.fill = yellow_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add sample data row
    sample_data = [
        '960921',           # ID_NO
        '22',               # Regular_Day
        '5000',             # ALLOWANCE_DAY
        '110000',           # TOTAL_ALLOWANCE
        '1500',             # ND_ALLOWANCE
        '111500',           # GRAND_TOTAL
        '55750',            # BASIC_SCHOOL_SHARE
        '55750',            # BASIC_OJT_SHARE
        '0',                # DEDUCTION
        '55750',            # NET_OJT_SHARE
        '0',                # RICE_ALLOWANCE
        '0',                # OT_ALLOWANCE
        '0',                # ND_OT_ALLOWANCE
        '0',                # SPECIAL_HOLIDAY
        '0',                # LEGAL_HOLIDAY
        '0',                # SATOFF_ALLOWANCE
        '0',                # RD_OT
        '0',                # ADJUSTMENT
        '0',                # DEDUCTION_2
        '0',                # OT_PAY_ALLOWANCE
        '0',                # TOTAL_ALLOW
        'N/A',              # HOLIDAY_DATE
        'N/A',              # RD_OT_DATE
        '0',                # PERFECT_ATTENDANCE
    ]
    
    for col_num, value in enumerate(sample_data, 1):
        cell = ws.cell(row=2, column=col_num, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Adjust column widths for better readability
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 18
    
    # Add instructions below the table
    instruction_row = 4
    instructions = [
        "INSTRUCTIONS FOR FILLING OUT THE OJT PAYSLIP TEMPLATE:",
        "",
        "1. REQUIRED COLUMNS - These columns MUST have values:",
        "   • ID_NO: Employee ID number (e.g., 960921)",
        "   • Regular_Day: Number of regular working days (numeric value)",
        "   • ALLOWANCE_DAY: Daily allowance amount (numeric value)",
        "",
        "2. DATE FORMAT:",
        "   • HOLIDAY_DATE and RD_OT_DATE: Use format YYYY-MM-DD (e.g., 2025-10-21)",
        "   • If no holiday or rest day overtime, enter 'N/A'",
        "",
        "3. NUMERIC COLUMNS:",
        "   • All monetary values and counts should be numeric (e.g., 5000, 1500)",
        "   • If a column has no value or is not applicable, enter '0' (zero)",
        "   • Do NOT leave cells empty - use '0' or 'N/A' as appropriate",
        "",
        "4. IMPORTANT NOTES:",
        "   • Do NOT modify or delete the header row",
        "   • Start entering data from row 2 onwards (row 2 contains sample data)",
        "   • Ensure all ID_NO values correspond to valid employee IDs in the system",
        "   • Save the file in Excel format (.xlsx) before uploading",
        "",
        "5. SAMPLE DATA:",
        "   • Row 2 contains sample data for reference",
        "   • You may delete the sample row before uploading actual data",
        "",
    ]
    
    for idx, instruction in enumerate(instructions):
        cell = ws.cell(row=instruction_row + idx, column=1, value=instruction)
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        if idx == 0:  # Title row
            cell.font = Font(bold=True, size=12)
    
    # Merge cells for instructions to span across columns
    for idx in range(len(instructions)):
        ws.merge_cells(start_row=instruction_row + idx, start_column=1, 
                       end_row=instruction_row + idx, end_column=len(headers))
    
    # Set row height for instructions
    for idx in range(len(instructions)):
        ws.row_dimensions[instruction_row + idx].height = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="ojt_payslip_template.xlsx"'
    return response

@login_required(login_url="user-login")
def ojt_payslip_details(request, payslip_id):
    try:
        payslip = OJTPayslipData.objects.get(id=payslip_id)
        data = {
            'cut_off': str(payslip.cut_off),
            'deduction': str(payslip.deduction),
            'deduction_2': str(payslip.deduction_2),
            'net_ojt_share': str(payslip.net_ojt_share),
            'ot_pay_allowance': str(payslip.ot_pay_allowance),
            'total_allow': str(payslip.total_allow),
            'regular_day': str(payslip.regular_day),
            'allowance_day': str(payslip.allowance_day),
            'total_allowance': str(payslip.total_allowance),
            'nd_allowance': str(payslip.nd_allowance),
            'grand_total': str(payslip.grand_total),
            'basic_school_share': str(payslip.basic_school_share),
            'basic_ojt_share': str(payslip.basic_ojt_share),
            'rice_allowance': str(payslip.rice_allowance),
            'ot_allowance': str(payslip.ot_allowance),
            'nd_ot_allowance': str(payslip.nd_ot_allowance),
            'special_holiday': str(payslip.special_holiday),
            'legal_holiday': str(payslip.legal_holiday),
            'satoff_allowance': str(payslip.satoff_allowance),
            'rd_ot': str(payslip.rd_ot),
            'perfect_attendance': str(payslip.perfect_attendance),
            'adjustment': str(payslip.adjustment),
            'holiday_hours': str(payslip.holiday_hours),
            'rd_ot_days': str(payslip.rd_ot_days),
        }
        return JsonResponse({'success': True, 'data': data})
    except OJTPayslipData.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Payslip not found'}, status=404)

@login_required(login_url="user-login")
def ajax_ojt_payslip_details(request, payslip_id):
    try:
        payslip = OJTPayslipData.objects.get(id=payslip_id)
        current_user = request.user
        
        # Get OJT rates for the "main" site
        try:
            main_rates = OJTRate.objects.get(site='main')
            rates_data = {
                'reg_nd_rate': str(main_rates.reg_nd_rate),
                'reg_nd_ot_rate': str(main_rates.reg_nd_ot_rate),
                'reg_ot_rate': str(main_rates.reg_ot_rate),
                'rest_ot_rate': str(main_rates.rest_ot_rate),
                'legal_rate': str(main_rates.legal_rate),
                'sat_off_rate': str(main_rates.sat_off_rate),
            }
        except OJTRate.DoesNotExist:
            # If no "main" site rates found, use defaults
            rates_data = {
                'reg_nd_rate': '0.00',
                'reg_nd_ot_rate': '0.00',
                'reg_ot_rate': '0.00',
                'rest_ot_rate': '0.00',
                'legal_rate': '0.00',
                'sat_off_rate': '0.00',
            }
        
        # Get line name from current user's employment info
        try:
            if hasattr(current_user, 'employment_info') and current_user.employment_info:
                if hasattr(current_user.employment_info, 'line') and current_user.employment_info.line:
                    line_name = current_user.employment_info.line.line_name
                else:
                    line_name = '-'
            else:
                line_name = '-'
        except Exception as e:
            print(f"Error getting line info: {e}")
            line_name = '-'
        
        data = {
            'cut_off': str(payslip.cut_off),
            'regular_day': str(payslip.regular_day),
            'allowance_day': str(payslip.allowance_day),
            'holiday_hours': str(payslip.holiday_hours),
            'rd_ot_days': str(payslip.rd_ot_days),
            'basic_ojt_share': str(payslip.basic_ojt_share),
            'nd_allowance': str(payslip.nd_allowance),
            'rice_allowance': str(payslip.rice_allowance),
            'perfect_attendance': str(payslip.perfect_attendance),
            'satoff_allowance': str(payslip.satoff_allowance),
            'nd_ot_allowance': str(payslip.nd_ot_allowance),
            'ot_allowance': str(payslip.ot_allowance),
            'rd_ot': str(payslip.rd_ot),
            'special_holiday': str(payslip.special_holiday),
            'legal_holiday': str(payslip.legal_holiday),
            'adjustment': str(payslip.adjustment),
            'basic_school_share': str(payslip.basic_school_share),
            'deduction': str(payslip.deduction),
            'deduction_2': str(payslip.deduction_2),
            'net_ojt_share': str(payslip.net_ojt_share),
            'ot_pay_allowance': str(payslip.ot_pay_allowance),
            'total_allow': str(payslip.total_allow),
            'total_allowance': str(payslip.total_allowance),
            'grand_total': str(payslip.grand_total),
            'employee': {
                'idnumber': current_user.idnumber or '-',
                'full_name': f"{current_user.firstname or ''} {current_user.lastname or ''}".strip() or '-',
                'line': line_name
            }
        }
        
        # Add rate data to the response
        data.update(rates_data)
        
        return JsonResponse({'success': True, 'payslip': data})
    except OJTPayslipData.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Payslip not found'}, status=404)
    except Exception as e:

        # Add rate data to the response
        data.update(rates_data)
        
        return JsonResponse({'success': True, 'payslip': data})
    except OJTPayslipData.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Payslip not found'}, status=404)
    except Exception as e:
        # Log the error for debugging
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error in ajax_ojt_payslip_details: {str(e)}")
        return JsonResponse({'success': False, 'error': f'Server error: {str(e)}'}, status=500)
    
# REGULAR Payslip Upload
@login_required(login_url="user-login")
def regular_payslip_upload(request):
    """Handle payslip file upload with validation and progress tracking"""
    if not request.user.accounting_admin:
        return JsonResponse({'success': False, 'message': 'Permission denied'})

    if request.method == 'POST':
        files = request.FILES.getlist('files')
        cutoff_date = request.POST.get('cutoff_date')
        errors = []
        success_count = 0
        processed = 0
        successful_uploads = []

        if not files:
            return JsonResponse({'success': False, 'message': 'No files selected'})
        if not cutoff_date:
            return JsonResponse({'success': False, 'message': 'Cutoff date is required'})

        for file in files:
            try:
                # Validate file type
                if not file.name.lower().endswith('.pdf'):
                    errors.append({'filename': file.name, 'error': 'Only PDF files are allowed'})
                    processed += 1
                    continue

                # Parse filename: "idnumber_lastname.pdf" (e.g., "960921_mier.pdf" or "960921_dela cruz.pdf")
                # Lastname can contain spaces, the important part is the idnumber
                import re
                filename_pattern = r'^([A-Za-z0-9]+)_(.+)\.pdf$'
                match = re.match(filename_pattern, file.name, re.IGNORECASE)
                if not match:
                    errors.append({'filename': file.name, 'error': 'Invalid filename format. Expected: idnumber_lastname.pdf (e.g., 960921_mier.pdf)'})
                    processed += 1
                    continue
                employee_id, lastname = match.groups()

                # Find employee by ID number (the important validation)
                from userlogin.models import EmployeeLogin
                try:
                    employee = EmployeeLogin.objects.get(idnumber=employee_id)
                except EmployeeLogin.DoesNotExist:
                    errors.append({'filename': file.name, 'error': f'Employee with ID {employee_id} not found in the system'})
                    processed += 1
                    continue

                # Save file with new name: {cutoff_date}_{original_filename}
                final_filename = f"{cutoff_date}_{file.name}"
                from django.core.files.base import ContentFile
                from django.core.files.storage import default_storage
                file_path = default_storage.save(f"payslips/{final_filename}", ContentFile(file.read()))

                # Create Payslip record
                from .models import Payslip
                payslip = Payslip.objects.create(
                    employee=employee,
                    cutoff_date=cutoff_date,
                    file_path=file_path,
                    uploaded_by=request.user
                )
                successful_uploads.append({'filename': file.name, 'employee_id': employee_id, 'employee_name': f"{employee.firstname} {employee.lastname}", 'final_filename': final_filename, 'action': 'Created'})
                success_count += 1
                processed += 1
            except Exception as e:
                errors.append({'filename': file.name, 'error': str(e)})
                processed += 1

        # Create notification for successful upload
        if success_count > 0:
            Notification.objects.create(
                title="Regular Payslip Upload",
                message=f"Regular payslip has been uploaded for cutoff Period: {cutoff_date}.",
                sender=request.user,
                recipient=request.user,
                module="finance",
                for_all=True
            )

        # Determine overall success status
        # If all uploads failed, return success: False
        # If some succeeded, return success: True with error details
        if success_count == 0 and len(errors) > 0:
            return JsonResponse({
                'success': False,
                'message': f'Upload failed. {len(errors)} file(s) have errors.',
                'success_count': success_count,
                'error_count': len(errors),
                'errors': errors,
                'successful_uploads': successful_uploads
            })

        return JsonResponse({
            'success': True if success_count > 0 else False,
            'message': f'Payslips uploaded: {success_count}, Errors: {len(errors)}',
            'success_count': success_count,
            'error_count': len(errors),
            'errors': errors,
            'successful_uploads': successful_uploads
        })

    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required(login_url="user-login")
def download_failed_payslips(request):
    """Generate Excel file with failed payslip uploads"""
    if not request.user.accounting_admin:
        return JsonResponse({'success': False, 'message': 'Permission denied'})

    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            errors = data.get('errors', [])

            # Create a new workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Failed Uploads"

            # Add headers
            headers = ['Filename', 'Error']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            # Add data
            for row_num, error in enumerate(errors, 2):
                ws.cell(row=row_num, column=1, value=error.get('filename', 'Unknown'))
                ws.cell(row=row_num, column=2, value=error.get('error', 'Unknown error'))

            # Adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            # Create response
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="payslip_upload_errors.xlsx"'

            return response

        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error generating file: {str(e)}'}, status=500)

    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=400)

def ajax_employee_payslips(request, employee_id):
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'error': 'Unauthorized'}, status=403)
    try:
        employee = EmployeeLogin.objects.get(id=employee_id)
        search = request.GET.get('search', '').strip()
        payslip_qs = Payslip.objects.filter(employee=employee)
        if search:
            from django.db.models import Q
            payslip_qs = payslip_qs.filter(
                Q(cutoff_date__icontains=search) |
                Q(date_uploaded__icontains=search) |
                Q(uploaded_by__firstname__icontains=search) |
                Q(uploaded_by__lastname__icontains=search) |
                Q(file_path__icontains=search)
            )
        payslip_list = payslip_qs.order_by('-cutoff_date')
        page_number = request.GET.get('page')
        from django.core.paginator import Paginator
        paginator = Paginator(payslip_list, 10)
        payslips = paginator.get_page(page_number)
        html = render_to_string('finance/partials/payslips_table.html', {
            'payslips': payslips,
            'search': search,
        }, request=request)
        return JsonResponse({'success': True, 'html': html})
    except EmployeeLogin.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Employee not found.'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
# Loan Management
@login_required(login_url="user-login")
def loan_principal_upload(request):
    """Upload loan principal amounts (creates or stacks loans)"""
    try:
        if not request.user.accounting_admin:
            return JsonResponse({'success': False, 'message': 'Permission denied'})

        if request.method == 'POST':
            try:
                file = request.FILES.get('file')
                if not file:
                    files = request.FILES.getlist('files')
                    if files:
                        file = files[0]
                if not file or not file.name.lower().endswith('.xlsx'):
                    return JsonResponse({'success': False, 'message': 'Please upload an Excel (.xlsx) file'})
                success, errors, created, updated, stacked, not_uploaded_rows = process_loan_principal_excel(file)
                if errors:
                    return JsonResponse({
                        'success': False,
                        'errors': errors,
                        'created': created,
                        'updated': updated,
                        'stacked': stacked,
                        'not_uploaded_rows': not_uploaded_rows
                    })
                
                # Create notification for successful upload
                if created > 0 or updated > 0 or stacked > 0:
                    Notification.objects.create(
                        title="Principal Balance Upload",
                        message=f"New principal loan balance has been successfully uploaded.",
                        sender=request.user,
                        recipient=request.user,
                        module="finance",
                        for_all=True
                    )
                
                return JsonResponse({
                    'success': True,
                    'created': created,
                    'updated': updated,
                    'stacked': stacked,
                    'not_uploaded_rows': not_uploaded_rows,
                    'message': f'Successfully processed {created} new loans, updated {updated} loans, and stacked {stacked} loans.'
                })
            except Exception as e:
                import traceback
                import sys
                print('Loan principal upload error:', e)
                traceback.print_exc()
                return JsonResponse({'success': False, 'error': f'Unexpected error: {str(e)}'})
        return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)
    except Exception as e:
        import traceback
        import sys
        print('Loan principal upload error (outer):', e)
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': f'Unexpected error (outer): {str(e)}'})

@login_required(login_url="user-login")
def loan_deduction_upload(request):
    try:
        if not request.user.accounting_admin:
            return JsonResponse({'success': False, 'message': 'Permission denied'})

        if request.method == 'POST':
            try:
                file = request.FILES.get('file')
                if not file:
                    files = request.FILES.getlist('files')
                    if files:
                        file = files[0]
                        
                cutoff_date = request.POST.get('cutoff_date', '')
                
                if not file or not (file.name.lower().endswith('.xlsx') or file.name.lower().endswith('.xls') or file.name.lower().endswith('.csv')):
                    return JsonResponse({'success': False, 'message': 'Please upload an Excel (.xlsx, .xls) or CSV file'})
                    
                if not cutoff_date:
                    return JsonResponse({'success': False, 'message': 'Cutoff date is required'})
                
                success, errors, processed, added_deductions = process_loan_deduction_excel(file, cutoff_date)
                
                if errors:
                    return JsonResponse({
                        'success': False, 
                        'errors': errors, 
                        'processed': processed,
                        'added_deductions': added_deductions
                    })
                
                # Create notification for successful upload
                if processed > 0:
                    Notification.objects.create(
                        title="Deduction Upload",
                        message=f"Loan deductions has been successfully uploaded for cutoff period {cutoff_date}.",
                        sender=request.user,
                        recipient=request.user,
                        module="finance",
                        for_all=True
                    )
                    
                return JsonResponse({
                    'success': True, 
                    'processed': processed,
                    'added_deductions': added_deductions,
                    'message': f'Successfully processed {processed} loan deductions for {cutoff_date}.'
                })
                
            except Exception as e:
                import traceback
                import sys
                print('Loan deduction upload error:', e)
                traceback.print_exc()
                return JsonResponse({'success': False, 'error': f'Unexpected error: {str(e)}'})
                
        return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)
        
    except Exception as e:
        import traceback
        import sys
        print('Loan deduction upload error (outer):', e)
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': f'Unexpected error (outer): {str(e)}'})

def process_loan_principal_excel(file):

    errors = []
    created = 0
    updated = 0
    stacked = 0
    not_uploaded_rows = []
    try:
        workbook = openpyxl.load_workbook(file, data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.values)
        if len(rows) < 2:
            return False, ['Excel file is empty or has no data'], 0, 0, 0, []
        headers = rows[0]
        data_rows = rows[1:]
        print(f"DEBUG: Headers found: {headers}")
        def parse_decimal(val):
            if val is None or val == '':
                return Decimal('0')
            try:
                return Decimal(str(val).replace(',', '').replace(' ', '').strip())
            except Exception:
                return None
        with transaction.atomic():
            for row_idx, row in enumerate(data_rows, start=2):
                try:
                    # Skip completely empty rows
                    if not row or all(cell is None or (isinstance(cell, str) and cell.strip() == '') for cell in row):
                        continue
                    if len(row) < 4:
                        continue
                    employee_id = row[0]
                    employee_name = row[1]
                    loan_type_name = row[2]
                    amount = row[3]
                    monthly_deduction = row[4] if len(row) > 4 else 0
                    if not all([employee_id, loan_type_name, amount]):
                        remark = 'Missing required data (Id Number, Loan Type, or Amount)'
                        errors.append(f'Row {row_idx}: {remark}')
                        error_row = list(row[:5]) if row else ['', '', '', '', '']
                        while len(error_row) < 5:
                            error_row.append('')
                        error_row.append(remark)
                        not_uploaded_rows.append(error_row)
                        continue
                    try:
                        employee = EmployeeLogin.objects.get(idnumber=str(employee_id))
                    except EmployeeLogin.DoesNotExist:
                        remark = f'Employee {employee_id} not found'
                        errors.append(f'Row {row_idx}: {remark}')
                        error_row = [employee_id, employee_name, loan_type_name, amount, monthly_deduction, remark]
                        not_uploaded_rows.append(error_row)
                        continue
                    loan_type, _ = LoanType.objects.get_or_create(
                        loan_type=str(loan_type_name).strip(),
                        defaults={'is_stackable': True}
                    )
                    amount_val = parse_decimal(amount)
                    monthly_deduction_val = parse_decimal(monthly_deduction)
                    if amount_val is None:
                        remark = f'Invalid principal amount value ({amount})'
                        errors.append(f'Row {row_idx}: {remark}')
                        error_row = [employee_id, employee_name, loan_type_name, amount, monthly_deduction, remark]
                        not_uploaded_rows.append(error_row)
                        continue
                    if monthly_deduction and monthly_deduction_val is None:
                        remark = f'Invalid monthly deduction value ({monthly_deduction})'
                        errors.append(f'Row {row_idx}: {remark}')
                        error_row = [employee_id, employee_name, loan_type_name, amount, monthly_deduction, remark]
                        not_uploaded_rows.append(error_row)
                        continue
                    amount = amount_val
                    monthly_deduction = monthly_deduction_val if monthly_deduction else Decimal('0')
                    existing_loan = Loan.objects.filter(
                        employee=employee,
                        loan_type=loan_type,
                        is_active=True
                    ).first()
                    # If loan type is not stackable and there is an active loan with nonzero balance, skip and collect row
                    if existing_loan and not loan_type.is_stackable and existing_loan.current_balance > 0:
                        remark = f'Active non-stackable loan exists for {employee_id} ({loan_type_name})'
                        errors.append(f'Row {row_idx}: {remark}')
                        error_row = [employee_id, employee_name, loan_type_name, amount, monthly_deduction, remark]
                        not_uploaded_rows.append(error_row)
                        continue
                    if existing_loan and loan_type.is_stackable:
                        existing_loan.add_principal(amount)
                        if monthly_deduction > 0:
                            existing_loan.monthly_deduction = monthly_deduction
                            existing_loan.save()
                        stacked += 1
                    else:
                        Loan.objects.create(
                            employee=employee,
                            loan_type=loan_type,
                            principal_amount=amount,
                            current_balance=amount,
                            monthly_deduction=monthly_deduction
                        )
                        created += 1
                except Exception as e:
                    remark = str(e)
                    errors.append(f'Row {row_idx}: {remark}')
                    error_row = list(row[:5]) if row else ['', '', '', '', '']
                    while len(error_row) < 5:
                        error_row.append('')
                    error_row.append(remark)
                    not_uploaded_rows.append(error_row)
                    continue
        workbook.close()
        if errors:
            return False, errors, created, updated, stacked, not_uploaded_rows
        return True, [], created, updated, stacked, not_uploaded_rows
    except Exception as e:
        return False, [f'Error reading Excel file: {str(e)}'], 0, 0, 0, []

def process_loan_deduction_excel(file, cutoff_date):

    errors = []
    processed = 0
    added_deductions = []
    
    try:
        if file.name.lower().endswith('.csv'):
            # Handle CSV files
            import csv
            import io
            content = file.read().decode('utf-8')
            csv_data = csv.reader(io.StringIO(content))
            rows = list(csv_data)
        else:
            # Handle Excel files
            workbook = openpyxl.load_workbook(file, data_only=True)
            worksheet = workbook.active
            rows = list(worksheet.values)
            workbook.close()
        
        if len(rows) < 2:
            return False, ['File is empty or has no data'], 0, []
        
        headers = rows[0]
        data_rows = rows[1:]
        
        with transaction.atomic():
            for row_idx, row in enumerate(data_rows, start=2):
                try:
                    # Skip completely empty rows
                    if not row or all(cell is None or (isinstance(cell, str) and cell.strip() == '') for cell in row):
                        continue
                    # Expecting: Id Number, Name, Loan Type, Deduction
                    if len(row) < 4:
                        continue
                    employee_id = row[0]
                    name = row[1]  # not used for lookup
                    loan_type_name = row[2]
                    deduction_amount = row[3]
                    if not all([employee_id, loan_type_name, deduction_amount]):
                        remark = 'Missing required data (Id Number, Loan Type, or Deduction)'
                        errors.append(f'Row {row_idx}: {remark}')
                        # Always output 4 columns + remark for error file
                        added_deductions.append([
                            row[0] if len(row) > 0 else '',
                            row[1] if len(row) > 1 else '',
                            row[2] if len(row) > 2 else '',
                            row[3] if len(row) > 3 else '',
                            remark
                        ])
                        continue
                    # Find employee
                    try:
                        employee = EmployeeLogin.objects.get(idnumber=str(employee_id))
                    except EmployeeLogin.DoesNotExist:
                        remark = f'Employee {employee_id} not found'
                        errors.append(f'Row {row_idx}: {remark}')
                        added_deductions.append([
                            employee_id, name, loan_type_name, deduction_amount, remark
                        ])
                        continue
                    # Find loan type
                    try:
                        loan_type = LoanType.objects.get(loan_type=str(loan_type_name).strip())
                    except LoanType.DoesNotExist:
                        remark = f'Loan type {loan_type_name} not found'
                        errors.append(f'Row {row_idx}: {remark}')
                        added_deductions.append([
                            employee_id, name, loan_type_name, deduction_amount, remark
                        ])
                        continue
                    # Convert deduction amount
                    try:
                        deduction_amount = Decimal(str(deduction_amount))
                    except:
                        remark = 'Invalid deduction amount'
                        errors.append(f'Row {row_idx}: {remark}')
                        added_deductions.append([
                            employee_id, name, loan_type_name, deduction_amount, remark
                        ])
                        continue
                    # Find oldest active loan (order by created_at ascending)
                    loan = Loan.objects.filter(
                        employee=employee,
                        loan_type=loan_type,
                        is_active=True,
                        current_balance__gt=0
                    ).order_by('created_at').first()
                    if not loan:
                        remark = f'No active loan found for {employee_id} - {loan_type_name}'
                        errors.append(f'Row {row_idx}: {remark}')
                        added_deductions.append([
                            employee_id, name, loan_type_name, deduction_amount, remark
                        ])
                        continue
                    # Check if deduction already exists for this cutoff
                    if LoanDeduction.objects.filter(loan=loan, cut_off=cutoff_date).exists():
                        remark = f'Deduction already exists for {cutoff_date}'
                        errors.append(f'Row {row_idx}: {remark}')
                        added_deductions.append([
                            employee_id, name, loan_type_name, deduction_amount, remark
                        ])
                        continue
                    # Apply deduction
                    actual_deduction = loan.apply_deduction(deduction_amount, cutoff_date)
                    if actual_deduction > 0:
                        processed += 1
                    else:
                        remark = 'No deduction applied (loan may be fully paid)'
                        errors.append(f'Row {row_idx}: {remark}')
                        added_deductions.append([
                            employee_id, name, loan_type_name, deduction_amount, remark
                        ])
                except Exception as e:
                    remark = str(e)
                    errors.append(f'Row {row_idx}: {remark}')
                    # Defensive: always output 4 columns + remark
                    added_deductions.append([
                        row[0] if len(row) > 0 else '',
                        row[1] if len(row) > 1 else '',
                        row[2] if len(row) > 2 else '',
                        row[3] if len(row) > 3 else '',
                        remark
                    ])
                    continue
        
        if errors:
            return False, errors, processed, added_deductions
        return True, [], processed, added_deductions
        
    except Exception as e:
        return False, [f'Error reading file: {str(e)}'], 0, []
    
@login_required(login_url="user-login")
def export_loan_principal_template(request):
    """Export Excel template for loan principal upload"""
    if not request.user.accounting_admin:
        return JsonResponse({'success': False, 'message': 'Permission denied'})
    
    # Create workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Principal Balance"
    
    # Define headers
    headers = ['Id Number', 'Name', 'Loan Type', 'Principal Balance', 'Monthly Deduction']
    
    # Style definitions
    header_font = Font(bold=True, color='000000')
    header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow background
    border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Add headers with styling
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
    
    # Get all loan types for validation
    loan_types = LoanType.objects.all().values_list('loan_type', flat=True)
    loan_type_list = ','.join([f'"{lt}"' for lt in loan_types])
    
    # Add data validation for Loan Type column (column C)
    if loan_types:
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(loan_types)}"',
            allow_blank=False
        )
        dv.error = 'Please select a loan type from the dropdown list only!'
        dv.errorTitle = 'Invalid Loan Type'
        dv.prompt = 'Choose from the dropdown list'
        dv.promptTitle = 'Loan Type Selection'
        
        # Apply validation to column C (rows 2 to 1000)
        worksheet.add_data_validation(dv)
        dv.add('C2:C1000')
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 15  # Id Number
    worksheet.column_dimensions['B'].width = 30  # Name
    worksheet.column_dimensions['C'].width = 20  # Loan Type
    worksheet.column_dimensions['D'].width = 18  # Principal Balance
    worksheet.column_dimensions['E'].width = 18  # Monthly Deduction

    # Add sample data row with borders
    sample_row = 2
    sample_data = ['960001', 'John Doe', '', '10000.00', '500.00']
    
    for col_num, value in enumerate(sample_data, 1):
        cell = worksheet.cell(row=sample_row, column=col_num, value=value)
        cell.border = border
        if col_num == 4:  # Principal Balance column
            cell.alignment = Alignment(horizontal='right')
    
    # Add instructions in a separate area
    instructions_row = 4
    instruction_text = [
        "INSTRUCTIONS:",
        "1. Fill in the Id Number (Employee ID)",
        "2. Fill in the Name (Employee Name)",  
        "3. Select Loan Type from dropdown ONLY - do not type manually!",
        "4. Enter Principal Balance amount",
        "5. Save file and upload through the system",
        "",
        "IMPORTANT: You must select loan type from dropdown menu only!",
        "Typing loan type manually will cause upload errors."
    ]
    
    for i, instruction in enumerate(instruction_text):
        cell = worksheet.cell(row=instructions_row + i, column=1, value=instruction)
        if i == 0:  # Title
            cell.font = Font(bold=True, color='FF0000')
        elif "IMPORTANT:" in instruction:
            cell.font = Font(bold=True, color='FF0000')
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Loan_Principal_Template.xlsx"'
    
    # Save workbook to response
    workbook.save(response)
    return response

@login_required(login_url="user-login")  
def export_loan_deduction_template(request):
    """Export Excel template for loan deduction upload"""
    if not request.user.accounting_admin:
        return JsonResponse({'success': False, 'message': 'Permission denied'})
    
    # Create workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Loan Deduction"
    
    # Define headers
    headers = ['Id Number', 'Name', 'Loan Type', 'Deduction']
    
    # Style definitions
    header_font = Font(bold=True, color='000000')
    header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow background
    border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Add headers with styling
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
    
    # Get all loan types for validation
    loan_types = LoanType.objects.all().values_list('loan_type', flat=True)
    
    # Add data validation for Loan Type column (column D)
    if loan_types:
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(loan_types)}"',
            allow_blank=False
        )
        dv.error = 'Please select a loan type from the dropdown list only!'
        dv.errorTitle = 'Invalid Loan Type'
        dv.prompt = 'Choose from the dropdown list'
        dv.promptTitle = 'Loan Type Selection'
        
        # Apply validation to column D (rows 2 to 1000)
        worksheet.add_data_validation(dv)
        dv.add('C2:C1000')
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 15
    worksheet.column_dimensions['B'].width = 30
    worksheet.column_dimensions['C'].width = 20
    worksheet.column_dimensions['D'].width = 15

    sample_row = 2
    sample_data = ['960001', 'John Doe', '', '2500.00']
    
    for col_num, value in enumerate(sample_data, 1):
        cell = worksheet.cell(row=sample_row, column=col_num, value=value)
        cell.border = border
        if col_num == 5:  # Deduction column
            cell.alignment = Alignment(horizontal='right')
    
    # Add instructions in a separate area
    instructions_row = 4
    instruction_text = [
        "INSTRUCTIONS:",
        "1. Fill in the Id Number (Employee ID)",
        "2. Fill in the Name (Employee Name)",
        "3. Fill in the Cut Off period (e.g., January 2025)",
        "4. Select Loan Type from dropdown ONLY - do not type manually!",
        "5. Enter Deduction amount",
        "6. Save file and upload through the system",
        "",
        "IMPORTANT: You must select loan type from dropdown menu only!",
        "Typing loan type manually will cause upload errors.",
        "",
        "NOTE: Deductions will be applied to active loans of the selected type.",
        "If deduction exceeds remaining balance, only the remaining balance will be deducted."
    ]
    
    for i, instruction in enumerate(instruction_text):
        cell = worksheet.cell(row=instructions_row + i, column=1, value=instruction)
        if i == 0:  # Title
            cell.font = Font(bold=True, color='FF0000')
        elif "IMPORTANT:" in instruction or "NOTE:" in instruction:
            cell.font = Font(bold=True, color='FF0000')
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Loan_Deduction_Template.xlsx"'
    
    # Save workbook to response
    workbook.save(response)
    return response

@login_required(login_url="user-login")
@require_POST
def delete_loan(request, loan_id):
    from django.urls import reverse
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    if not request.user.accounting_admin:
        if is_ajax:
            return JsonResponse({'success': False, 'message': 'Permission denied.'}, status=403)
        else:
            messages.error(request, 'Permission denied.')
            return redirect('admin_finance')
    try:
        loan = Loan.objects.get(id=loan_id)
        employee = loan.employee
        loan.delete()
        loans = Loan.objects.filter(employee=employee).order_by('-created_at')
        # Recalculate total active loan balance
        total_active_loan_balance = sum(l.current_balance for l in loans if l.is_active)
        if is_ajax:
            html = render_to_string('finance/partials/loans_table.html', {
                'loans': loans,
                'total_active_loan_balance': total_active_loan_balance,
            }, request=request)
            return JsonResponse({'success': True, 'message': 'Loan deleted successfully.', 'html': html})
        else:
            messages.success(request, 'Loan deleted successfully.')
            url = reverse('employee_finance_details', args=[employee.id]) + '?tab=loans'
            return redirect(url)
    except Loan.DoesNotExist:
        if is_ajax:
            return JsonResponse({'success': False, 'message': 'Loan not found.'}, status=404)
        else:
            messages.error(request, 'Loan not found.')
            return redirect('admin_finance')
    except Exception as e:
        if is_ajax:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
        else:
            messages.error(request, f'Error: {str(e)}')
            return redirect('admin_finance')
   
# Allowance Upload
@login_required(login_url="user-login")
def allowances_upload(request):
    def is_ajax(req):
        return req.headers.get('x-requested-with') == 'XMLHttpRequest' or req.headers.get('accept', '').find('application/json') != -1

    if not request.user.accounting_admin:
        if is_ajax(request):
            return JsonResponse({'success': False, 'message': "You don't have permission to access this page."}, status=403, content_type='application/json')
        else:
            messages.error(request, "You don't have permission to access this page.")
            return redirect('dashboard')


    if request.method == 'POST':
        try:
            from .models import Allowance
            # Delete all Allowance records before import
            Allowance.objects.all().delete()

            files = request.FILES.getlist('files')
            success_count = 0
            error_count = 0
            error_details = []
            all_error_rows = []  # Collect all error rows from all files

            for file in files:
                try:
                    if file.name.endswith(('.xlsx', '.xls', '.csv')):
                        # Process allowance file (now only inserts into Allowance model)
                        success, errors, error_rows = process_allowance_file(file)
                        if success:
                            success_count += 1
                        else:
                            error_count += 1
                            if errors:
                                for error in errors:
                                    error_details.append({'filename': file.name, 'error': error})
                            if error_rows:
                                all_error_rows.extend(error_rows)
                    else:
                        error_count += 1
                        error_details.append({'filename': file.name, 'error': f"Invalid file type: {file.name}"})
                except Exception as e:
                    error_count += 1
                    error_details.append({'filename': file.name, 'error': f"Error processing {file.name}: {str(e)}"})

            # Always return JSON for POST requests (fixes frontend invalid response format)
            if error_count > 0 or len(all_error_rows) > 0:
                return JsonResponse({
                    'success': False,
                    'message': f'Failed to upload {error_count} file(s)',
                    'error_count': error_count,
                    'success_count': success_count,
                    'errors': error_details,
                    'error_rows': all_error_rows
                }, content_type='application/json')
            else:
                # Create notification for successful upload
                if success_count > 0:
                    Notification.objects.create(
                        title="Allowances Upload",
                        message=f"New allowances have been successfully uploaded.",
                        sender=request.user,
                        recipient=request.user,
                        module="finance",
                        for_all=True
                    )
                
                return JsonResponse({
                    'success': True,
                    'message': f'Successfully uploaded {success_count} allowance file(s)',
                    'success_count': success_count,
                    'error_count': error_count,
                    'errors': error_details
                }, content_type='application/json')

        except Exception as e:
            # Always return JSON for POST errors
            return JsonResponse({
                'success': False,
                'message': f'Unexpected error: {str(e)}',
                'error_count': 1,
                'success_count': 0,
                'errors': [{'filename': 'system', 'error': str(e)}]
            }, content_type='application/json')

    # For non-POST, always return JSON for AJAX, else legacy
    if is_ajax(request):
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405, content_type='application/json')
    return JsonResponse({'success': False, 'message': 'Invalid request method'})

def process_allowance_file(file):
    from .models import Allowance, AllowanceType
    from userlogin.models import EmployeeLogin
    import pandas as pd
    success = True
    errors = []
    error_rows = []  # Store rows with errors for download
    try:
        df = pd.read_excel(file) if file.name.endswith(('.xlsx', '.xls')) else pd.read_csv(file)
        for index, row in df.iterrows():
            try:
                idnumber = str(row.get('Id Number', '')).strip()
                name = str(row.get('Name', '')).strip()
                allowance_type_name = str(row.get('Allowance Type', '')).strip()
                amount_val = row.get('Amount', 0)
                is_percentage_value = str(row.get('Is Percentage', 'No')).strip().lower()
                deposit_date_value = row.get('Deposit Date', '')
                period_covered_value = row.get('Period Covered', '')
                
                if not idnumber:
                    continue
                try:
                    employee = EmployeeLogin.objects.get(idnumber=idnumber)
                except EmployeeLogin.DoesNotExist:
                    remark = f"Employee not found for ID: {idnumber}"
                    errors.append(remark)
                    error_rows.append([idnumber, name, allowance_type_name, amount_val, is_percentage_value, deposit_date_value, period_covered_value, remark])
                    continue
                if not allowance_type_name:
                    remark = f"Allowance type is required for employee {idnumber}"
                    errors.append(remark)
                    error_rows.append([idnumber, name, allowance_type_name, amount_val, is_percentage_value, deposit_date_value, period_covered_value, remark])
                    continue
                try:
                    allowance_type = AllowanceType.objects.get(allowance_type=allowance_type_name)
                except AllowanceType.DoesNotExist:
                    remark = f"Allowance type '{allowance_type_name}' not found"
                    errors.append(f"{remark} for employee {idnumber}")
                    error_rows.append([idnumber, name, allowance_type_name, amount_val, is_percentage_value, deposit_date_value, period_covered_value, remark])
                    continue
                try:
                    amount = float(amount_val or 0)
                except Exception:
                    remark = f"Invalid amount value"
                    errors.append(f"{remark} for employee {idnumber}")
                    error_rows.append([idnumber, name, allowance_type_name, amount_val, is_percentage_value, deposit_date_value, period_covered_value, remark])
                    continue
                
                # Handle is_percentage - convert Yes/No to boolean
                is_percentage = is_percentage_value in ['yes', 'y', 'true', '1']
                
                # Handle deposit date - keep as null if empty, otherwise parse
                deposit_date = None
                
                if deposit_date_value and str(deposit_date_value).strip() and str(deposit_date_value).strip().lower() not in ['', 'nan', 'none', 'null']:
                    try:
                        deposit_date = pd.to_datetime(deposit_date_value).date()
                    except Exception:
                        remark = f"Invalid deposit date '{deposit_date_value}'"
                        errors.append(f"{remark} for employee {idnumber}")
                        error_rows.append([idnumber, name, allowance_type_name, amount_val, is_percentage_value, deposit_date_value, period_covered_value, remark])
                        continue
                
                # Handle period covered - optional field
                period_covered = None
                if period_covered_value and str(period_covered_value).strip() and str(period_covered_value).strip().lower() not in ['', 'nan', 'none', 'null']:
                    period_covered = str(period_covered_value).strip()
                
                Allowance.objects.create(
                    employee=employee,
                    allowance_type=allowance_type,
                    amount=amount,
                    is_percentage=is_percentage,
                    deposit_date=deposit_date,
                    period_covered=period_covered
                )
            except Exception as e:
                remark = str(e)
                errors.append(f"Error processing row {index + 2}: {remark}")
                error_rows.append([idnumber, name, allowance_type_name, amount_val, is_percentage_value, deposit_date_value, period_covered_value, remark])
                success = False
        return success, errors, error_rows
    except Exception as e:
        return False, [f"Error reading file: {str(e)}"], []

@login_required(login_url="user-login")  
def export_allowance_template(request):
    """Export Excel template for allowance import"""
    if not request.user.accounting_admin:
        return JsonResponse({'success': False, 'message': 'Permission denied'})

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Allowance Import"

    headers = ['Id Number', 'Name', 'Allowance Type', 'Amount', 'Is Percentage', 'Deposit Date', 'Period Covered']
    header_font = Font(bold=True, color='000000')
    header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')

    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    # Data validation for Allowance Type
    allowance_types = AllowanceType.objects.all().values_list('allowance_type', flat=True)
    if allowance_types:
        dv = DataValidation(
            type="list",
            formula1=f'"{','.join([str(a) for a in allowance_types])}"',
            allow_blank=False
        )
        dv.error = 'Please select an allowance type from the dropdown list only!'
        dv.errorTitle = 'Invalid Allowance Type'
        dv.prompt = 'Choose from the dropdown list'
        dv.promptTitle = 'Allowance Type Selection'
        worksheet.add_data_validation(dv)
        dv.add('C2:C1000')

    # Data validation for Is Percentage (Yes/No dropdown)
    is_percentage_dv = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    is_percentage_dv.error = 'Please select Yes or No from the dropdown list!'
    is_percentage_dv.errorTitle = 'Invalid Selection'
    is_percentage_dv.prompt = 'Select Yes if amount is a percentage, No if fixed amount'
    is_percentage_dv.promptTitle = 'Is Percentage'
    worksheet.add_data_validation(is_percentage_dv)
    is_percentage_dv.add('E2:E1000')

    worksheet.column_dimensions['A'].width = 15
    worksheet.column_dimensions['B'].width = 30
    worksheet.column_dimensions['C'].width = 25
    worksheet.column_dimensions['D'].width = 18
    worksheet.column_dimensions['E'].width = 15
    worksheet.column_dimensions['F'].width = 18
    worksheet.column_dimensions['G'].width = 20

    # Add sample data rows
    sample_row = 2
    # Sample 1: Fixed amount allowance
    sample_data_1 = ['960001', 'John Doe', '', '2000', 'No', '2025-07-31', '']
    for col_num, value in enumerate(sample_data_1, 1):
        cell = worksheet.cell(row=sample_row, column=col_num, value=value)
        cell.border = border
        if col_num == 4:
            cell.alignment = Alignment(horizontal='right')
        if col_num == 5:
            cell.alignment = center_alignment

    # Sample 2: Percentage allowance
    sample_row_2 = 3
    sample_data_2 = ['960002', 'Jane Smith', '', '10', 'Yes', '', 'January 2025']
    for col_num, value in enumerate(sample_data_2, 1):
        cell = worksheet.cell(row=sample_row_2, column=col_num, value=value)
        cell.border = border
        if col_num == 4:
            cell.alignment = Alignment(horizontal='right')
        if col_num == 5:
            cell.alignment = center_alignment

    # Add instructions
    instructions_row = 5
    instruction_text = [
        "INSTRUCTIONS:",
        "1. Fill in the Id Number (Employee ID)",
        "2. Fill in the Name (Employee Name)",
        "3. Select Allowance Type from dropdown ONLY - do not type manually!",
        "4. Enter Amount value (numeric only, no symbols)",
        "5. Select 'Is Percentage' from dropdown:",
        "   - 'Yes': The Amount is a percentage (e.g., 10 means 10% of base salary)",
        "   - 'No': The Amount is a fixed value (e.g., 2000 means ₱2,000)",
        "6. Enter Deposit Date in YYYY-MM-DD format (e.g., 2025-07-31) OR leave empty",
        "7. If no Deposit Date, enter Period Covered (e.g., 'January 2025', 'Q1 2025')",
        "8. Save file and upload through the system",
        "",
        "IMPORTANT: You must select allowance type from dropdown menu only!",
        "Typing allowance type manually will cause upload errors.",
        "",
        "HOW 'IS PERCENTAGE' WORKS:",
        "- When 'Is Percentage' = Yes: The amount represents a percentage.",
        "  Example: Amount = 10 with Is Percentage = Yes means 10% of base salary.",
        "- When 'Is Percentage' = No: The amount is a fixed peso value.",
        "  Example: Amount = 2000 with Is Percentage = No means ₱2,000 flat allowance.",
        "",
        "NOTE: Deposit Date must be in YYYY-MM-DD format (e.g., 2025-07-31).",
        "NOTE: Use Period Covered for balance allowances without a specific deposit date."
    ]
    for i, instruction in enumerate(instruction_text):
        cell = worksheet.cell(row=instructions_row + i, column=1, value=instruction)
        if i == 0:
            cell.font = Font(bold=True, color='FF0000')
        elif "IMPORTANT:" in instruction or "NOTE:" in instruction:
            cell.font = Font(bold=True, color='FF0000')
        else:
            cell.font = Font(bold=False, color='000000')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Allowance_Import_Template.xlsx"'
    workbook.save(response)
    return response

@login_required(login_url="user-login")
@require_POST
def delete_allowance(request, allowance_id):
    from django.urls import reverse
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    if not request.user.accounting_admin:
        if is_ajax:
            return JsonResponse({'success': False, 'message': 'Permission denied.'}, status=403)
        else:
            messages.error(request, 'Permission denied.')
            return redirect('admin_finance')
    try:
        allowance = Allowance.objects.get(id=allowance_id)
        employee = allowance.employee
        allowance.delete()
        allowances = Allowance.objects.filter(employee=employee).order_by('-created_at')
        total_allowances = sum(a.amount for a in allowances)
        if is_ajax:
            html = render_to_string('finance/partials/allowances_table.html', {
                'allowances': allowances,
                'total_allowances': total_allowances,
            }, request=request)
            return JsonResponse({'success': True, 'message': 'Allowance deleted successfully.', 'html': html})
        else:
            messages.success(request, 'Allowance deleted successfully.')
            url = reverse('employee_finance_details', args=[employee.id]) + '?tab=allowances'
            return redirect(url)
    except Allowance.DoesNotExist:
        if is_ajax:
            return JsonResponse({'success': False, 'message': 'Allowance not found.'}, status=404)
        else:
            messages.error(request, 'Allowance not found.')
            return redirect('admin_finance')
    except Exception as e:
        if is_ajax:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
        else:
            messages.error(request, f'Error: {str(e)}')
            return redirect('admin_finance')
    
# Send to Mail
@login_required(login_url="user-login")
@require_POST
def send_payslip(request, payslip_id):
    payslip = get_object_or_404(Payslip, id=payslip_id)

    # Check permissions
    if not (request.user == payslip.employee or request.user.accounting_admin):
        return JsonResponse({'success': False, 'message': 'Permission denied'})

    try:
        # Get selected email address from request
        selected_email = request.POST.get('selected_email', '').strip()
        
        if not selected_email:
            return JsonResponse({'success': False, 'message': 'Email address is required'})
    
        # Prepare email context
        email_context = {
            'payslip': payslip,
            'employee': payslip.employee,
        }
        
        # Render email template
        html_message = render_to_string('finance/email_template.html', email_context)
        
        # Create email subject  
        subject = f"Payslip - {payslip.cutoff_date.strftime('%b %d, %Y')} - {payslip.employee.full_name}"
        
        # Create email message with configured sender
        from django.core.mail import EmailMessage
        from django.conf import settings
        
        # Use specific email configuration from settings
        email_message = EmailMessage(
            subject=subject,
            body=html_message,
            from_email=settings.EMAIL_HOST_USER,  # Use configured sender email
            to=[selected_email],
        )
        
        # Set email as HTML
        email_message.content_subtype = 'html'
        
        # Attach payslip file if it exists
        if payslip.file_path and hasattr(payslip.file_path, 'path'):
            try:
                email_message.attach_file(payslip.file_path.path)
            except Exception as e:
                return JsonResponse({'success': False, 'message': f'Error attaching file: {str(e)}'})
        
        # Send email
        email_message.send()
        
        # Update payslip status
        payslip.is_send_to_mail = True
        payslip.save()
        
        return JsonResponse({
            'success': True, 
            'message': f'Payslip successfully sent to {selected_email}'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error sending email: {str(e)}'})


@login_required(login_url="user-login")
def employee_finance_details(request, employee_id):
    user = request.user
    if not (user.accounting_admin):
        messages.error(request, "You don't have permission to access this page.")
        return redirect('dashboard')


    try:
        employee = EmployeeLogin.objects.get(id=employee_id)
        search = request.GET.get('search', '').strip()
        payslip_qs = Payslip.objects.filter(employee=employee)
        if search:
            from django.db.models import Q
            payslip_qs = payslip_qs.filter(
                Q(cutoff_date__icontains=search) |
                Q(date_uploaded__icontains=search) |
                Q(uploaded_by__firstname__icontains=search) |
                Q(uploaded_by__lastname__icontains=search)
            )
        payslip_list = payslip_qs.order_by('-cutoff_date')
        page_number = request.GET.get('page')
        from django.core.paginator import Paginator
        paginator = Paginator(payslip_list, 10)
        payslips = paginator.get_page(page_number)
        loans = Loan.objects.filter(employee=employee).order_by('-created_at')
        # Precompute paid and percent for each loan (percent = (current_balance / principal_amount) * 100, paid = principal_amount - current_balance)
        total_active_loan_balance = 0
        for loan in loans:
            try:
                principal = float(getattr(loan, 'principal_amount', 0) or 0)
                current_balance = float(getattr(loan, 'current_balance', 0) or 0)
                if principal > 0:
                    paid = principal - current_balance
                    percent = (current_balance / principal * 100)
                else:
                    paid = 0
                    percent = 100  # If no principal, consider fully paid
            except Exception:
                paid = 0
                percent = 100
            loan.paid = paid
            loan.percent = percent
            loan.percent_paid = 100 - percent
            if getattr(loan, 'is_active', False):
                total_active_loan_balance += current_balance


        allowances = Allowance.objects.filter(employee=employee).order_by('-created_at')
        total_allowances = sum(a.amount for a in allowances)

        # Add savings queryset for this employee
        savings = Savings.objects.filter(employee=employee).order_by('-created_at')

        from datetime import datetime

        # OJT Payslip Data
        ojt_payslips = None
        if hasattr(employee, 'employment_info') and getattr(employee.employment_info, 'employment_type', None) == 'OJT':
            ojt_payslips = OJTPayslipData.objects.filter(employee=employee).order_by('-created_at')

        context = {
            'employee': employee,
            'payslips': payslips,
            'loans': loans,
            'allowances': allowances,
            'total_allowances': total_allowances,
            'total_active_loan_balance': total_active_loan_balance,
            'now': datetime.now(),
            'search': search,
            'ojt_payslips': ojt_payslips,
            'savings': savings,
        }

        return render(request, 'finance/employee_finance_details.html', context)

    except EmployeeLogin.DoesNotExist:
        messages.error(request, "Employee not found.")
        return redirect('admin_finance')
    except Exception as e:
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect('admin_finance')
    
@login_required(login_url="user-login")
def employee_allowances(request, employee_id):
    if not request.user.accounting_admin:
        return JsonResponse({'success': False, 'message': 'Permission denied'})

    employee = get_object_or_404(EmployeeLogin, id=employee_id)
    allowances = Allowance.objects.filter(employee=employee).order_by('-start_date')

    # Group allowances by type
    allowance_groups = {}
    for allowance in allowances:
        allowance_type = allowance.get_allowance_type_display()
        if allowance_type not in allowance_groups:
            allowance_groups[allowance_type] = []
        allowance_groups[allowance_type].append({
            'id': allowance.id,
            'amount': str(allowance.amount),
            'frequency': allowance.get_frequency_display(),
            'start_date': allowance.start_date.strftime('%Y-%m-%d'),
            'end_date': allowance.end_date.strftime('%Y-%m-%d') if allowance.end_date else None,
            'is_active': allowance.is_active,
            'description': allowance.description,
        })

    return JsonResponse({
        'success': True,
        'employee': {
            'name': employee.full_name,
            'idnumber': employee.idnumber,
        },
        'allowance_groups': allowance_groups
    })

@login_required(login_url="user-login")
def employees_list(request):
    if not request.user.accounting_admin:
        return JsonResponse({'success': False, 'message': 'Permission denied'})

    search_form = EmployeeSearchForm(request.GET)
    tab = request.GET.get('tab')
    if tab == 'loans':
        employees = EmployeeLogin.objects.filter(
            is_active=True,
            wire_admin=False,
            clinic_admin=False,
            iad_admin=False,
            accounting_admin=False,
            hr_admin=False,
            hr_manager=False,
            mis_admin=False,
            loans__isnull=False
        ).distinct().order_by('lastname', 'firstname')
    else:
        employees = EmployeeLogin.objects.filter(
            is_active=True
        ).order_by('lastname', 'firstname')

    if search_form.is_valid():
        search = search_form.cleaned_data.get('search')
        department = search_form.cleaned_data.get('department')
        employment_type = search_form.cleaned_data.get('employment_type')

        if search:
            employees = employees.filter(
                Q(firstname__icontains=search) |
                Q(lastname__icontains=search) |
                Q(idnumber__icontains=search) |
                Q(username__icontains=search)
            )

        if department:
            employees = employees.filter(
                employment_info__department__name__icontains=department
            )

        if employment_type:
            employees = employees.filter(
                employment_info__employment_type=employment_type
            )

    paginator = Paginator(employees, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    employees_data = []
    for employee in page_obj:
        emp_info = getattr(employee, 'employment_info', None)
        department = getattr(emp_info.department, 'name', '-') if emp_info and emp_info.department else '-'
        emp_type = getattr(emp_info, 'employment_type', 'regular') if emp_info else 'regular'

        employees_data.append({
            'id': employee.id,
            'name': f"{employee.firstname or ''} {employee.lastname or ''}".strip() or employee.username or 'Unknown',
            'idnumber': employee.idnumber,
            'department': department,
            'employment_type': emp_type.title(),
            'email': employee.email,
        })

    return JsonResponse({
        'success': True,
        'employees': employees_data,
        'has_next': page_obj.has_next(),
        'has_previous': page_obj.has_previous(),
        'page_number': page_obj.number,
        'total_pages': paginator.num_pages,
        'total_count': paginator.count,
    })

@login_required(login_url="user-login")
@require_GET
def employee_table_partial(request):
    user = request.user
    employee_list = EmployeeLogin.objects.filter(
        is_active=True
    ).exclude(
        id=user.id
    ).exclude(
        accounting_admin=True
    ).exclude(
        hr_admin=True
    )

    search = request.GET.get('search', '').strip()
    if search:
        employee_list = employee_list.filter(
            Q(idnumber__icontains=search) |
            Q(firstname__icontains=search) |
            Q(lastname__icontains=search) |
            Q(username__icontains=search) |
            Q(email__icontains=search)
        )

    employee_list = employee_list.order_by('lastname', 'firstname')
    paginator = Paginator(employee_list, 10)
    page_number = request.GET.get('page')
    employee_page_obj = paginator.get_page(page_number)

    html = render_to_string('finance/partials/employee_table.html', {
        'employee_page_obj': employee_page_obj,
        'search': search,
    }, request=request)
    return JsonResponse({'html': html})

@login_required(login_url="user-login")
def chart_data(request):
    try:
        if not request.user.accounting_admin:
            return JsonResponse({'success': False, 'message': 'Permission denied'})

        category = request.GET.get('category', '')
        filter_type = request.GET.get('type', '')
        period = request.GET.get('period', 'month')

        from datetime import datetime, timedelta
        from django.db.models import Sum, Count
        from django.db.models.functions import TruncDate, TruncMonth, TruncYear

        now = timezone.now()

        import calendar
        if period == 'month':
            # This Month: 1st to last day of current month
            start_date = now.replace(day=1)
            last_day = calendar.monthrange(now.year, now.month)[1]
            end_date = now.replace(day=last_day)
            # Prepare all days in month
            all_days = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]
        elif period == 'quarter':
            # This Quarter: last 3 months (labels: months only)
            # Get the last 3 months including current
            months = [(now.year, now.month - i) for i in reversed(range(3))]
            # Adjust for year wrap
            months = [(y if m > 0 else y - 1, m if m > 0 else m + 12) for (y, m) in months]
            start_date = datetime(months[0][0], months[0][1], 1, tzinfo=now.tzinfo)
            last_day = calendar.monthrange(months[-1][0], months[-1][1])[1]
            end_date = datetime(months[-1][0], months[-1][1], last_day, tzinfo=now.tzinfo)
            all_months = [(y, m) for (y, m) in months]
        else:  # year
            # This Year: all months of current year
            start_date = now.replace(month=1, day=1)
            end_date = now.replace(month=12, day=calendar.monthrange(now.year, 12)[1])
            all_months = [(now.year, m) for m in range(1, 13)]

        chart_data = {
            'labels': [],
            'datasets': []
        }

        if category == 'loans':
            if filter_type:
                loans = Loan.objects.filter(
                    loan_type__loan_type=filter_type,
                    created_at__gte=start_date,
                    created_at__lte=end_date
                )
            else:
                loans = Loan.objects.filter(
                    created_at__gte=start_date,
                    created_at__lte=end_date
                )

            # Group by month/week/day based on period using Django's date functions
            if period == 'month':
                loans_data = loans.annotate(
                    period=TruncDate('created_at')
                ).values('period').annotate(
                    total_amount=Sum('current_balance'),
                    count=Count('id')
                ).order_by('period')
                # Map date to value
                data_map = {item['period'].date() if hasattr(item['period'], 'date') else item['period']: float(item['total_amount'] or 0) for item in loans_data}
                labels = [d.strftime('%b %d') for d in all_days]
                data = [data_map.get(d.date() if hasattr(d, 'date') else d, 0) for d in all_days]
            elif period == 'quarter':
                loans_data = loans.annotate(
                    period=TruncMonth('created_at')
                ).values('period').annotate(
                    total_amount=Sum('current_balance'),
                    count=Count('id')
                ).order_by('period')
                # Map (year, month) to value
                data_map = {(item['period'].year, item['period'].month): float(item['total_amount'] or 0) for item in loans_data}
                labels = [datetime(y, m, 1).strftime('%b %Y') for (y, m) in all_months]
                data = [data_map.get((y, m), 0) for (y, m) in all_months]
            else:  # year
                loans_data = loans.annotate(
                    period=TruncMonth('created_at')
                ).values('period').annotate(
                    total_amount=Sum('current_balance'),
                    count=Count('id')
                ).order_by('period')
                # Map (year, month) to value
                data_map = {(item['period'].year, item['period'].month): float(item['total_amount'] or 0) for item in loans_data}
                labels = [datetime(now.year, m, 1).strftime('%b') for (y, m) in all_months]
                data = [data_map.get((now.year, m), 0) for (y, m) in all_months]

            chart_data['labels'] = labels
            chart_data['datasets'] = [{
                'label': 'Total Loan Balance',
                'data': data,
                'borderColor': '#2563eb',
                'backgroundColor': 'rgba(37, 99, 235, 0.1)',
                'barBackgroundColor': '#2563eb',
                'barBorderColor': '#2563eb',
                'tension': 0.4,
                'fill': True
            }]

        elif category == 'allowances':
            if filter_type:
                allowances = Allowance.objects.filter(
                    allowance_type__allowance_type=filter_type,
                    created_at__gte=start_date,
                    created_at__lte=end_date
                )
            else:
                allowances = Allowance.objects.filter(
                    created_at__gte=start_date,
                    created_at__lte=end_date
                )

            # Group by month/week/day based on period using Django's date functions
            if period == 'month':
                allowances_data = allowances.annotate(
                    period=TruncDate('created_at')
                ).values('period').annotate(
                    total_amount=Sum('amount'),
                    count=Count('id')
                ).order_by('period')
                data_map = {item['period'].date() if hasattr(item['period'], 'date') else item['period']: float(item['total_amount'] or 0) for item in allowances_data}
                labels = [d.strftime('%b %d') for d in all_days]
                data = [data_map.get(d.date() if hasattr(d, 'date') else d, 0) for d in all_days]
            elif period == 'quarter':
                allowances_data = allowances.annotate(
                    period=TruncMonth('created_at')
                ).values('period').annotate(
                    total_amount=Sum('amount'),
                    count=Count('id')
                ).order_by('period')
                data_map = {(item['period'].year, item['period'].month): float(item['total_amount'] or 0) for item in allowances_data}
                labels = [datetime(y, m, 1).strftime('%b %Y') for (y, m) in all_months]
                data = [data_map.get((y, m), 0) for (y, m) in all_months]
            else:  # year
                allowances_data = allowances.annotate(
                    period=TruncMonth('created_at')
                ).values('period').annotate(
                    total_amount=Sum('amount'),
                    count=Count('id')
                ).order_by('period')
                data_map = {(item['period'].year, item['period'].month): float(item['total_amount'] or 0) for item in allowances_data}
                labels = [datetime(now.year, m, 1).strftime('%b') for (y, m) in all_months]
                data = [data_map.get((now.year, m), 0) for (y, m) in all_months]

            chart_data['labels'] = labels
            chart_data['datasets'] = [{
                'label': 'Total Allowance Amount',
                'data': data,
                'borderColor': '#22c55e',
                'backgroundColor': 'rgba(34, 197, 94, 0.1)',
                'barBackgroundColor': '#22c55e',
                'barBorderColor': '#22c55e',
                'tension': 0.4,
                'fill': True
            }]

        return JsonResponse({
            'success': True,
            'chart_data': chart_data
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url="user-login")
def filter_options(request):
    if not request.user.accounting_admin:
        return JsonResponse({'success': False, 'message': 'Permission denied'})
    
    category = request.GET.get('category', '')
    
    if category == 'loans':
        loan_types = LoanType.objects.all().values_list('loan_type', flat=True)
        return JsonResponse({
            'success': True,
            'options': list(loan_types)
        })
    elif category == 'allowances':
        allowance_types = AllowanceType.objects.all().values_list('allowance_type', flat=True)
        return JsonResponse({
            'success': True,
            'options': list(allowance_types)
        })
    
    return JsonResponse({
        'success': True,
        'options': []
    })

@login_required(login_url="user-login")
def employee_details(request, employee_id):
    if not request.user.accounting_admin:
        return JsonResponse({'success': False, 'message': 'Permission denied'})

    try:
        employee = EmployeeLogin.objects.get(id=employee_id)
        payslips = Payslip.objects.filter(employee=employee).order_by('-cutoff_date')[:5]
        # Inline HTML for payslip list (no employee name or 'Payslips' title)
        payslip_html = ''
        if payslips:
            payslip_html += '<div class="payslips-grid">'
            for payslip in payslips:
                payslip_html += f'''
                <div class="payslip-card">
                    <div class="card-header">
                        <h4>{payslip.cutoff_date.strftime('%b %d')} - {payslip.cutoff_date.strftime('%b %d, %Y')}</h4>
                        <span class="payslip-type-badge {payslip.employee_type}">{payslip.get_employee_type_display()}</span>
                    </div>
                    <div class="card-body">
                        <div class="payslip-details">
                            <div class="detail-row">
                                <span class="label">Period:</span>
                                <span class="value">{payslip.cutoff_date.strftime('%Y-%m-%d')} to {payslip.cutoff_date.strftime('%Y-%m-%d')}</span>
                            </div>
                            <div class="detail-row">
                                <span class="label">Amount:</span>
                                <span class="value highlight">₱{payslip.amount:.2f}</span>
                            </div>
                            <div class="detail-row">
                                <span class="label">Uploaded:</span>
                                <span class="value">{payslip.date_uploaded.strftime('%b %d, %Y')}</span>
                            </div>
                        </div>
                    </div>
                </div>
                '''
            payslip_html += '</div>'
        else:
            payslip_html += ''
        return JsonResponse({
            'success': True,
            'employee': {
                'id': employee.id,
                'name': f"{employee.firstname} {employee.lastname}",
                'idnumber': employee.idnumber,
                'email': employee.email,
            },
            'html': payslip_html
        })
    except EmployeeLogin.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Employee not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required(login_url="user-login")
def delete_payslip(request, payslip_id):
    if not request.user.accounting_admin:
        return JsonResponse({'success': False, 'message': "You don't have permission to perform this action."}, status=403)

    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            payslip = Payslip.objects.get(id=payslip_id)
            employee = payslip.employee
            employee_id = employee.id
            employee_name = f"{employee.firstname} {employee.lastname}"

            # Delete the payslip file if it exists
            if payslip.file_path and default_storage.exists(payslip.file_path.name):
                default_storage.delete(payslip.file_path.name)

            # Delete the payslip record
            payslip.delete()

            # Get updated payslips for the employee
            payslip_list = Payslip.objects.filter(employee=employee).order_by('-cutoff_date')
            page_number = request.GET.get('page')
            paginator = Paginator(payslip_list, 10)
            payslips = paginator.get_page(page_number)
            context = {
                'employee': employee,
                'payslips': payslips,
            }
            # Render only the payslip table body
            html = render_to_string('finance/partials/payslips_table.html', context, request=request)
            return JsonResponse({'success': True, 'message': f"Payslip for {employee_name} has been deleted.", 'html': html})

        except Payslip.DoesNotExist:
            return JsonResponse({'success': False, 'message': "Payslip not found."}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': f"An error occurred while deleting the payslip: {str(e)}"}, status=500)
    return JsonResponse({'success': False, 'message': "Invalid request."}, status=400)


@login_required(login_url="user-login")
@require_GET
def loan_deductions_list(request, loan_id):
    from .models import LoanDeduction
    try:
        loan = Loan.objects.get(id=loan_id)
        # Only allow if user can view this employee's finance
        user = request.user
        deductions = LoanDeduction.objects.filter(loan=loan).order_by('-created_at')
        total_deduction = deductions.aggregate(total=models.Sum('amount'))['total'] or 0
        html = render_to_string('finance/partials/loan_deductions_list.html', {
            'deductions': deductions,
            'total_deduction': total_deduction,
        })
        return JsonResponse({'success': True, 'html': html})
    except Loan.DoesNotExist:
        return JsonResponse({'success': False, 'error': "Loan not found."}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
@login_required(login_url="user-login")
def export_finance_report(request):
    export_type = request.GET.get('type')
    if export_type == 'employee_total_loans':
        return export_employee_total_loans(request)
    elif export_type == 'employee_savings':
        return export_employee_savings_report(request)
    elif export_type == 'total_ojt_allowances':
        return export_total_ojt_allowances_report(request)
    else:
        return HttpResponseBadRequest('Invalid export type')
    
# Export Employee Total Loans Report (Excel)
@login_required(login_url="user-login")
@user_passes_test(lambda u: u.is_superuser or u.accounting_admin or u.hr_admin)
def export_employee_total_loans(request):
    # Only export loans with nonzero balance
    loans = Loan.objects.filter(current_balance__gt=0).select_related('employee', 'loan_type')
    data = []
    for loan in loans:
        deductions = loan.deductions.aggregate(total=models.Sum('amount'))['total'] or 0
        remaining_balance = loan.principal_amount - deductions
        data.append({
            'idnumber': loan.employee.idnumber or '',
            'name': f"{loan.employee.firstname or ''} {loan.employee.lastname or ''}".strip(),
            'loan_type': loan.loan_type.loan_type,
            'monthly_deduction': loan.monthly_deduction,
            'principal_amount': loan.principal_amount,
            'total_deduction': deductions,
            'remaining_balance': remaining_balance,
        })

    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Total Loans"

    # Title and subtitle
    ws.merge_cells('A1:G1')
    ws['A1'] = "RYONAN ELECTRIC PHILIPPINES CORPORATION"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:G2')
    ws['A2'] = "Employee Loan Balances"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    # Empty row (row 3)

    # Header row (row 4)
    headers = [
        'Id Number', 'Name', 'Loan Type', 'Monthly Deduction', 'Principal Balance', 'Total Deduction', 'Remaining Balance'
    ]
    ws.append(['']*7)  # row 3 blank
    ws.append(headers)  # row 4

    # Style header row
    header_font = Font(bold=True, color='000000')
    header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow
    border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )
    for col_num in range(1, 8):
        cell = ws.cell(row=4, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data rows (start at row 5)
    for i, row in enumerate(data, start=5):
        ws.append([
            row['idnumber'],
            row['name'],
            row['loan_type'],
            float(row['monthly_deduction']),
            float(row['principal_amount']),
            float(row['total_deduction']),
            float(row['remaining_balance']),
        ])
        for col_num in range(1, 8):
            cell = ws.cell(row=i, column=col_num)
            cell.border = border

    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=employee_total_loans.xlsx'
    wb.save(response)
    return response

# Export Employee Savings Report (Excel)
@login_required(login_url="user-login")
@user_passes_test(lambda u: u.is_superuser or u.accounting_admin or u.hr_admin)
def export_employee_savings_report(request):
    """Export Employee Savings Total Report"""
    # Get only active savings (not withdrawn)
    from django.db import models
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    
    savings = Savings.objects.filter(is_withdrawn=False).select_related('employee')
    data = []
    for saving in savings:
        data.append({
            'idnumber': saving.employee.idnumber or '',
            'name': f"{saving.employee.firstname or ''} {saving.employee.lastname or ''}".strip(),
            'total_savings': saving.amount,
            'date_started': saving.created_at.strftime('%Y-%m-%d') if saving.created_at else '',
        })

    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Savings Report"

    # Title and subtitle
    ws.merge_cells('A1:D1')
    ws['A1'] = "RYONAN ELECTRIC PHILIPPINES CORPORATION"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:D2')
    ws['A2'] = "Employee Savings Total Report"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    # Empty row (row 3)
    ws.append(['']*4)  # row 3 blank

    # Header row (row 4)
    headers = ['Id Number', 'Name', 'Total Savings', 'Date Started']
    ws.append(headers)  # row 4

    # Style header row
    header_font = Font(bold=True, color='000000')
    header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow
    border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )
    for col_num in range(1, 5):
        cell = ws.cell(row=4, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data rows (start at row 5)
    for i, row in enumerate(data, start=5):
        ws.append([
            row['idnumber'],
            row['name'],
            float(row['total_savings']),
            row['date_started'],
        ])
        for col_num in range(1, 5):
            cell = ws.cell(row=i, column=col_num)
            cell.border = border

    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=employee_savings_report.xlsx'
    wb.save(response)
    return response

# Export Total OJT Allowances Report (Excel)
@login_required(login_url="user-login")
@user_passes_test(lambda u: u.is_superuser or u.accounting_admin or u.hr_admin)
def export_total_ojt_allowances_report(request):
    """Export Total OJT Allowances Report"""
    from django.db import models
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    
    # Group OJT payslips by cut_off and aggregate data
    ojt_data = (OJTPayslipData.objects
                .values('cut_off')
                .annotate(
                    number_of_ojt=models.Count('id'),
                    total_allowances=models.Sum('total_allow'),
                    date_uploaded=models.Max('created_at')
                )
                .order_by('cut_off'))
    
    data = []
    for item in ojt_data:
        data.append({
            'cut_off': item['cut_off'],
            'number_of_ojt': item['number_of_ojt'],
            'total_allowances': item['total_allowances'] or 0,
            'date_uploaded': item['date_uploaded'].strftime('%Y-%m-%d') if item['date_uploaded'] else '',
        })

    wb = Workbook()
    ws = wb.active
    ws.title = "Total OJT Allowances Report"

    # Title and subtitle
    ws.merge_cells('A1:D1')
    ws['A1'] = "RYONAN ELECTRIC PHILIPPINES CORPORATION"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:D2')
    ws['A2'] = "Total OJT Allowances Report"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    # Empty row (row 3)
    ws.append(['']*4)  # row 3 blank

    # Header row (row 4)
    headers = ['Cut Off', 'Number of OJT', 'Total Allowances', 'Date Uploaded']
    ws.append(headers)  # row 4

    # Style header row
    header_font = Font(bold=True, color='000000')
    header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow
    border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )
    for col_num in range(1, 5):
        cell = ws.cell(row=4, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data rows (start at row 5)
    for i, row in enumerate(data, start=5):
        ws.append([
            row['cut_off'],
            row['number_of_ojt'],
            float(row['total_allowances']),
            row['date_uploaded'],
        ])
        for col_num in range(1, 5):
            cell = ws.cell(row=i, column=col_num)
            cell.border = border

    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=total_ojt_allowances_report.xlsx'
    wb.save(response)
    return response

@login_required(login_url="user-login")
def export_savings_template(request):
    """Export savings template Excel file"""
    user = request.user
    if not (user.accounting_admin):
        messages.error(request, "You don't have permission to access this page.")
        return redirect('dashboard')

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Savings Template"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers - added Savings Type column
    headers = ['Id Number', 'Name', 'Savings', 'Savings Type']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Get all savings types for data validation
    savings_types = list(SavingsType.objects.all().order_by('savings_type').values_list('savings_type', flat=True))
    
    # Sample data with savings type
    first_savings_type = savings_types[0] if savings_types else 'Personal Savings'
    sample_data = [
        ['EMP001', 'John Doe', '5000.00', first_savings_type],
        ['EMP002', 'Jane Smith', '3500.00', first_savings_type],
        ['EMP003', 'Mike Johnson', '7500.00', first_savings_type]
    ]

    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border

    # Add data validation for Savings Type column (column D)
    if savings_types:
        savings_type_validation = DataValidation(
            type="list",
            formula1='"' + ','.join(savings_types) + '"',
            allow_blank=True
        )
        savings_type_validation.error = "Please select a valid savings type from the list"
        savings_type_validation.errorTitle = "Invalid Savings Type"
        savings_type_validation.prompt = "Select a savings type"
        savings_type_validation.promptTitle = "Savings Type"
        ws.add_data_validation(savings_type_validation)
        savings_type_validation.add('D2:D1000')  # Apply to column D from row 2 to 1000

    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Add instructions sheet
    ws_instructions = wb.create_sheet("Instructions")
    instructions = [
        "SAVINGS IMPORT TEMPLATE INSTRUCTIONS",
        "",
        "1. REQUIRED COLUMNS:",
        "   - Id Number: Employee ID number (must exist in system)",
        "   - Name: Employee full name (for reference only)",
        "   - Savings: Amount to be added to employee's savings",
        "   - Savings Type: Type of savings (select from dropdown list)",
        "",
        "2. IMPORTANT NOTES:",
        "   - Do not modify the column headers",
        "   - Id Number must match exactly with existing employee records",
        "   - Savings amount should be numeric (no currency symbols)",
        "   - Savings Type must be selected from the dropdown list",
        "   - If employee has existing savings of the same type, the amount will be added",
        "   - If employee has no savings of that type, new savings record will be created",
        "",
        "3. AVAILABLE SAVINGS TYPES:",
    ]
    
    # Add each savings type to the instructions
    for st in savings_types:
        instructions.append(f"   - {st}")
    
    instructions.extend([
        "",
        "4. SUPPORTED FILE FORMATS:",
        "   - Excel (.xlsx, .xls)",
        "   - CSV (.csv)",
        "",
        "5. EXAMPLE:",
        f"   EMP001, John Doe, 5000.00, {first_savings_type}",
        f"   This will add ₱5,000.00 to John Doe's {first_savings_type} savings account"
    ])

    for row, instruction in enumerate(instructions, 1):
        cell = ws_instructions.cell(row=row, column=1, value=instruction)
        if row == 1:
            cell.font = Font(bold=True, size=14)
        elif instruction.startswith(("1.", "2.", "3.", "4.", "5.")):
            cell.font = Font(bold=True)

    # Auto-width instructions
    ws_instructions.column_dimensions['A'].width = 60

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="savings_template.xlsx"'
    wb.save(response)
    return response

@login_required(login_url="user-login")
@require_POST
def savings_upload(request):
    if request.method == 'POST':
        form = SavingsUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['file']
            
            try:
                # Read Excel file
                wb = openpyxl.load_workbook(excel_file)
                
                # Look for "Savings Template" sheet specifically
                sheet_name = None
                if 'Savings Template' in wb.sheetnames:
                    sheet_name = 'Savings Template'
                elif 'Sheet1' in wb.sheetnames:  # Fallback to Sheet1
                    sheet_name = 'Sheet1'
                else:
                    # Use the first sheet if neither exists
                    sheet_name = wb.sheetnames[0]
                
                ws = wb[sheet_name]
                
                success_count = 0
                error_count = 0
                errors = []
                error_rows = []  # Store rows with errors for download
                
                # Skip header row, start from row 2
                for row_num in range(2, ws.max_row + 1):
                    id_number = ws.cell(row=row_num, column=1).value  # Column A: Id Number
                    name = ws.cell(row=row_num, column=2).value       # Column B: Name (for reference only)
                    savings_amount = ws.cell(row=row_num, column=3).value  # Column C: Savings
                    savings_type_name = ws.cell(row=row_num, column=4).value  # Column D: Savings Type
                    
                    # Skip empty rows
                    if not id_number or not savings_amount:
                        continue
                    
                    try:
                        # Find employee by ID number
                        employee = EmployeeLogin.objects.get(idnumber=str(id_number).strip())
                        
                        # Convert amount to Decimal
                        original_savings_amount = savings_amount
                        savings_amount = Decimal(str(savings_amount))
                        
                        if savings_amount <= 0:
                            error_count += 1
                            remark = "Savings amount must be greater than 0"
                            errors.append(f"Row {row_num}: {remark}.")
                            error_rows.append([id_number, name, original_savings_amount, savings_type_name, remark])
                            continue
                        
                        # Find savings type if provided
                        savings_type = None
                        if savings_type_name:
                            original_savings_type_name = savings_type_name
                            savings_type_name = str(savings_type_name).strip()
                            try:
                                savings_type = SavingsType.objects.get(savings_type=savings_type_name)
                            except SavingsType.DoesNotExist:
                                error_count += 1
                                remark = f"Savings type '{savings_type_name}' not found"
                                errors.append(f"Row {row_num}: {remark}.")
                                error_rows.append([id_number, name, original_savings_amount, original_savings_type_name, remark])
                                continue
                        
                        # Try to find existing non-withdrawn savings for this employee with the same savings type
                        existing_savings = Savings.objects.filter(
                            employee=employee,
                            savings_type=savings_type,
                            is_withdrawn=False
                        ).first()
                        
                        if existing_savings:
                            # Add to existing non-withdrawn savings of the same type
                            existing_savings.amount += savings_amount
                            existing_savings.save()
                        else:
                            # Create new savings record since no non-withdrawn savings of this type exist
                            Savings.objects.create(
                                employee=employee,
                                savings_type=savings_type,
                                amount=savings_amount,
                                deposit_date=timezone.now().date(),
                                is_withdrawn=False
                            )
                        
                        success_count += 1
                        
                    except EmployeeLogin.DoesNotExist:
                        error_count += 1
                        remark = f"Employee with ID '{id_number}' not found"
                        errors.append(f"Row {row_num}: {remark}.")
                        error_rows.append([id_number, name, savings_amount, savings_type_name, remark])
                    except (ValueError, TypeError) as e:
                        error_count += 1
                        remark = f"Invalid savings amount '{savings_amount}': {str(e)}"
                        errors.append(f"Row {row_num}: {remark}")
                        error_rows.append([id_number, name, savings_amount, savings_type_name, remark])
                    except Exception as e:
                        error_count += 1
                        remark = str(e)
                        errors.append(f"Row {row_num}: {remark}")
                        error_rows.append([id_number, name, savings_amount, savings_type_name, remark])
                
                # Prepare response message
                if success_count > 0 and error_count == 0:
                    messages.success(request, f'Successfully imported {success_count} savings records.')
                elif success_count > 0 and error_count > 0:
                    messages.warning(request, f'Imported {success_count} records successfully. {error_count} records failed.')
                    for error in errors[:5]:  # Show first 5 errors
                        messages.error(request, error)
                else:
                    messages.error(request, 'No records were imported.')
                    for error in errors[:5]:  # Show first 5 errors
                        messages.error(request, error)
                
                # Create notification for successful upload
                if success_count > 0:
                    Notification.objects.create(
                        title="Savings Upload",
                        message=f"New savings has been successfully uploaded.",
                        sender=request.user,
                        recipient=request.user,
                        module="finance",
                        for_all=True
                    )
                
                # Return response with error_rows if there are errors
                if error_count > 0:
                    return JsonResponse({
                        'success': False,
                        'message': f'Processed {success_count + error_count} records. {success_count} successful, {error_count} failed.',
                        'success_count': success_count,
                        'error_count': error_count,
                        'errors': errors,
                        'error_rows': error_rows,
                        'redirect_url': '/finance/admin/'
                    })
                
                return JsonResponse({
                    'success': True,
                    'message': f'Processed {success_count + error_count} records. {success_count} successful, {error_count} failed.',
                    'redirect_url': '/finance/admin/'
                })
                
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Error processing file: {str(e)}'
                })
        else:
            return JsonResponse({
                'success': False,
                'message': 'Please select a valid Excel file.'
            })
    
    return JsonResponse({
        'success': False,
        'message': 'Invalid request method.'
    })

@login_required(login_url="user-login")
@require_POST
def withdraw_savings(request, savings_id):
    """Withdraw savings - set is_withdrawn to True but keep the amount"""
    if not (request.user.accounting_admin or request.user.hr_admin or request.user.hr_manager):
        return JsonResponse({'success': False, 'message': 'Permission denied'}, status=403)
    
    try:
        savings = Savings.objects.get(id=savings_id)
        employee = savings.employee
        
        if savings.is_withdrawn:
            return JsonResponse({'success': False, 'message': 'Savings already withdrawn'})
        
        # Get the amount before withdrawal
        withdrawal_amount = savings.amount
        
        # Call the withdraw method from the model
        if savings.withdraw():
            messages.success(request, f'Savings of ₱{float(withdrawal_amount):,.2f} successfully withdrawn for {employee.firstname} {employee.lastname}')
            
            # Use reverse to get the correct URL
            from django.urls import reverse
            redirect_url = reverse('employee_finance_details', args=[employee.id]) + '?tab=allowances'
            
            return JsonResponse({
                'success': True, 
                'message': f'Savings of ₱{float(withdrawal_amount):,.2f} successfully withdrawn',
                'redirect_url': redirect_url
            })
        else:
            return JsonResponse({'success': False, 'message': 'Failed to withdraw savings'})
            
    except Savings.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Savings record not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)

@login_required(login_url="user-login")
def send_ojt_payslip_email(request, payslip_id):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
    
    try:
        import json
        data = json.loads(request.body)
        email = data.get('email')
        
        if not email:
            return JsonResponse({'success': False, 'message': 'Email address is required'})
        
        payslip = OJTPayslipData.objects.get(id=payslip_id, employee=request.user)
        
        pdf_content = generate_ojt_payslip_pdf(payslip)
        
        subject = f"OJT Payslip - {payslip.cut_off}"
        message = f"""
Dear {request.user.firstname} {request.user.lastname},

Please find attached your OJT payslip for the period: {payslip.cut_off}

Thank you for your hard work and dedication.

Best regards,
RYONAN ELECTRIC PHILIPPINES CORPORATION
Finance Department
        """
        
        email_msg = EmailMessage(
            subject=subject,
            body=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[email],
        )
        
        # Attach PDF
        filename = f"OJT_Payslip_{request.user.idnumber}_{payslip.cut_off.replace('/', '_')}.pdf"
        email_msg.attach(filename, pdf_content, 'application/pdf')
        
        email_msg.send()
        
        return JsonResponse({
            'success': True,
            'message': f'OJT payslip sent successfully to {email}'
        })
        
    except OJTPayslipData.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'OJT payslip not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error sending email: {str(e)}'})

def generate_ojt_payslip_pdf(payslip):
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Company header exactly like modal
    logo_path = os.path.join(settings.STATIC_ROOT or settings.STATICFILES_DIRS[0], 'images', 'icon', 'ryonanlogo.png')
    
    try:
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=0.8*inch, height=0.8*inch)
            
            # Create header with logo on left and centered text
            company_info = Paragraph("""
                <para align=center>
                <b>RYONAN ELECTRIC PHILIPPINES CORPORATION</b><br/>
                105 East Main Avenue, Special Export Processing Zone<br/>
                Laguna, Technopark, Biñan, Laguna
                </para>
            """, ParagraphStyle('CompanyInfo', fontSize=12, fontName='Helvetica', alignment=TA_CENTER))
            
            header_data = [[logo, company_info]]
            header_table = Table(header_data, colWidths=[1*inch, 5*inch])
            header_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                ('ALIGN', (1, 0), (1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ]))
        else:
            # Fallback without logo - center everything
            company_info = Paragraph("""
                <para align=center>
                <b>RYONAN ELECTRIC PHILIPPINES CORPORATION</b><br/>
                105 East Main Avenue, Special Export Processing Zone<br/>
                Laguna, Technopark, Biñan, Laguna
                </para>
            """, ParagraphStyle('CompanyInfo', fontSize=12, fontName='Helvetica', alignment=TA_CENTER))
            
            header_table = Table([[company_info]], colWidths=[6*inch])
            header_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
    except:
        # Fallback without logo - center everything
        company_info = Paragraph("""
            <para align=center>
            <b>RYONAN ELECTRIC PHILIPPINES CORPORATION</b><br/>
            105 East Main Avenue, Special Export Processing Zone<br/>
            Laguna, Technopark, Biñan, Laguna
            </para>
        """, ParagraphStyle('CompanyInfo', fontSize=12, fontName='Helvetica', alignment=TA_CENTER))
        
        header_table = Table([[company_info]], colWidths=[6*inch])
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
    
    elements.append(header_table)
    elements.append(Spacer(1, 20))
    
    # Employee info grid (2x2) exactly like modal
    try:
        if hasattr(payslip.employee, 'employment_info') and payslip.employee.employment_info:
            if hasattr(payslip.employee.employment_info, 'line') and payslip.employee.employment_info.line:
                line_name = payslip.employee.employment_info.line.line_name
            else:
                line_name = '-'
        else:
            line_name = '-'
    except Exception:
        line_name = '-'
    
    employee_data = [
        ['ID Number:', payslip.employee.idnumber or '-', 'Cut-Off:', payslip.cut_off],
        ['Name:', f"{payslip.employee.firstname} {payslip.employee.lastname}", 'Line:', line_name]
    ]
    
    employee_table = Table(employee_data, colWidths=[1.2*inch, 2.3*inch, 1.2*inch, 2.3*inch])
    employee_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    elements.append(employee_table)
    elements.append(Spacer(1, 20))
    
    # Payslip sections exactly like modal - no peso signs
    regular_day_data = [
        ['Regular Day', ''],
        ['REGULAR # of Days Work', f"{payslip.regular_day}"],
        ['ALLOWANCE/DAY', f"{float(payslip.allowance_day):,.2f}"],
        ['Total:', f"{float(payslip.regular_day) * float(payslip.allowance_day):,.2f}"],
        ['REG ND ALLOWANCE', f"{float(payslip.nd_allowance):,.2f}"],
        ['GRAND TOTAL', f"{float(payslip.grand_total):,.2f}"],
        ['BASIC ALLOW.SCHOOL SHARE', f"{float(payslip.basic_school_share):,.2f}"],
        ['BASIC ALLOW. OJT SHARE', f"{float(payslip.basic_ojt_share):,.2f}"],
        ['DEDUCTION', f"{float(payslip.deduction):,.2f}"],
        ['NET BASIC ALLOW. OJT SHARE', f"{float(payslip.net_ojt_share):,.2f}"],
    ]
    
    allowances_data = [
        ['Allowances', ''],
        ['RICE ALLOWANCE', f"{float(payslip.rice_allowance):,.2f}"],
        ['Reg OT ALLOWANCE', f"{float(payslip.ot_allowance):,.2f}"],
        ['REG ND OT ALLOWANCE', f"{float(payslip.nd_ot_allowance):,.2f}"],
        ['SPECIAL HOLIDAY', f"{float(payslip.special_holiday):,.2f}"],
        ['LEGAL HOLIDAY', f"{float(payslip.legal_holiday):,.2f}"],
        ['SAT-OFF ALLOWANCE', f"{float(payslip.satoff_allowance):,.2f}"],
        ['RD OT', f"{float(payslip.rd_ot):,.2f}"],
        ['PERFECT ATTENDANCE', f"{float(payslip.perfect_attendance):,.2f}"],
        ['ADJUSTMENT', f"{float(payslip.adjustment):,.2f}"],
        ['DEDUCTION 2', f"{float(payslip.deduction_2):,.2f}"],
        ['NET OJT OT PAY ALLOWANCE', f"{float(payslip.ot_pay_allowance):,.2f}"],
    ]
    
    # Create side-by-side tables exactly like modal
    sections_data = []
    max_rows = max(len(regular_day_data), len(allowances_data))
    
    for i in range(max_rows):
        row = []
        if i < len(regular_day_data):
            row.extend(regular_day_data[i])
        else:
            row.extend(['', ''])
        
        row.append('')  # Spacer column
        
        if i < len(allowances_data):
            row.extend(allowances_data[i])
        else:
            row.extend(['', ''])
        
        sections_data.append(row)
    
    sections_table = Table(sections_data, colWidths=[2.3*inch, 1.2*inch, 0.2*inch, 2.3*inch, 1.2*inch])
    sections_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (1, -1), 0.5, colors.black),
        ('GRID', (3, 0), (4, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),   # Labels left-aligned
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),  # Values right-aligned
        ('ALIGN', (3, 0), (3, -1), 'LEFT'),   # Labels left-aligned
        ('ALIGN', (4, 0), (4, -1), 'RIGHT'),  # Values right-aligned
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        
        # Header rows styling exactly like modal
        ('FONTNAME', (0, 0), (1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (3, 0), (4, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 0), (1, 0), colors.lightgrey),
        ('BACKGROUND', (3, 0), (4, 0), colors.lightgrey),
        
        # Highlight NET BASIC ALLOW. OJT SHARE and NET OJT OT PAY ALLOWANCE like modal
        ('FONTNAME', (0, 9), (1, 9), 'Helvetica-Bold'),  # NET BASIC ALLOW. OJT SHARE
        ('BACKGROUND', (0, 9), (1, 9), colors.lightyellow),
        ('FONTNAME', (3, 11), (4, 11), 'Helvetica-Bold'),  # NET OJT OT PAY ALLOWANCE
        ('BACKGROUND', (3, 11), (4, 11), colors.lightyellow),
    ]))
    
    elements.append(sections_table)
    elements.append(Spacer(1, 20))
    
    # Total allowance - Right aligned without table, like image 1
    net_basic_ojt_share = float(payslip.net_ojt_share) if payslip.net_ojt_share else 0.0
    net_ojt_ot_pay_allowance = float(payslip.ot_pay_allowance) if payslip.ot_pay_allowance else 0.0
    total_allowance_value = net_basic_ojt_share + net_ojt_ot_pay_allowance
    
    # Create right-aligned total allowance paragraph
    total_allowance_style = ParagraphStyle(
        'TotalAllowance',
        fontSize=16,
        fontName='Helvetica-Bold',
        alignment=TA_RIGHT,
        spaceBefore=10,
        spaceAfter=10
    )
    
    total_allowance_text = f"TOTAL ALLOWANCE: {total_allowance_value:,.2f}"
    total_allowance_para = Paragraph(total_allowance_text, total_allowance_style)
    
    elements.append(total_allowance_para)
    
    # Build PDF
    doc.build(elements)
    
    pdf_content = buffer.getvalue()
    buffer.close()
    
    return pdf_content
