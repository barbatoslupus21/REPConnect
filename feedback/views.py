from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.utils import timezone
from django.views.decorators.http import require_http_methods
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime
import io

from .models import Feedback
from .forms import FeedbackForm, ExportDateRangeForm, AdminNotesForm

def is_hr_manager(user):
    return hasattr(user, 'hr_manager') and user.hr_manager

@login_required
def feedback_list(request):
    if not is_hr_manager(request.user):
        messages.error(request, "You don't have permission to access this page.")
        return redirect('dashboard')
    
    feedback_list = Feedback.objects.all()
    search_query = request.GET.get('search', '')
    
    if search_query:
        feedback_list = feedback_list.filter(
            Q(subject__icontains=search_query) |
            Q(message__icontains=search_query) |
            Q(submitter__firstname__icontains=search_query) |
            Q(submitter__lastname__icontains=search_query)
        )
    
    paginator = Paginator(feedback_list, 10)
    page_number = request.GET.get('page')
    feedback = paginator.get_page(page_number)
    
    context = {
        'feedback': feedback,
        'search_query': search_query,
        'total_feedback': feedback_list.count(),
    }
    
    return render(request, 'feedback/feedback-admin.html', context)

@login_required
def feedback_detail(request, feedback_id):
    if not is_hr_manager(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    feedback = get_object_or_404(Feedback, id=feedback_id)
    
    if not feedback.is_read:
        feedback.is_read = True
        feedback.save()
    
    if request.method == 'POST':
        notes_form = AdminNotesForm(request.POST, instance=feedback)
        if notes_form.is_valid():
            notes_form.save()
            return JsonResponse({
                'success': True,
                'message': 'Admin notes updated successfully.',
                'admin_notes': feedback.admin_notes
            })
        else:
            return JsonResponse({
                'success': False,
                'errors': notes_form.errors
            })
    
    # Return JSON data for modal
    return JsonResponse({
        'success': True,
        'data': {
            'id': feedback.id,
            'subject': feedback.subject,
            'message': feedback.message,
            'display_name': feedback.display_name,
            'display_email': feedback.display_email,
            'is_anonymous': feedback.is_anonymous,
            'is_read': feedback.is_read,
            'admin_notes': feedback.admin_notes or '',
            'created_at': feedback.created_at.strftime('%B %d, %Y at %H:%M'),
            'updated_at': feedback.updated_at.strftime('%B %d, %Y at %H:%M'),
        }
    })

@login_required
def submit_feedback(request):
    if request.method == 'POST':
        form = FeedbackForm(request.POST, user=request.user)
        if form.is_valid():
            feedback = form.save(commit=False)

            # If anonymous checkbox checked, do not set submitter
            if form.cleaned_data.get('is_anonymous'):
                feedback.submitter = None
                feedback.is_anonymous = True
            else:
                feedback.submitter = request.user
                feedback.is_anonymous = False

            feedback.save()

            # If AJAX request, return JSON
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Feedback submitted'})

            messages.success(request, "Your feedback has been submitted successfully.")
            return redirect('submit_feedback')
    else:
        form = FeedbackForm(user=request.user)

    context = {
        'form': form,
    }

    return render(request, 'feedback/submit_feedback.html', context)

@require_http_methods(["POST"])
@login_required
def export_feedback(request):
    if not is_hr_manager(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    form = ExportDateRangeForm(request.POST)
    if form.is_valid():
        start_date = form.cleaned_data['start_date']
        end_date = form.cleaned_data['end_date']
        
        feedback_data = Feedback.objects.filter(
            created_at__date__range=[start_date, end_date]
        ).order_by('-created_at')
        
        response = generate_excel_report(feedback_data, start_date, end_date)
        return response
    else:
        return JsonResponse({'error': 'Invalid date range'}, status=400)

def generate_excel_report(feedback_data, start_date, end_date):
    wb = openpyxl.Workbook()
    
    summary_sheet = wb.active
    summary_sheet.title = "Feedback Summary"
    
    feedback_sheet = wb.create_sheet(title="Feedback Details")
    
    create_summary_sheet(summary_sheet, feedback_data)
    create_feedback_sheet(feedback_sheet, feedback_data, start_date, end_date)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="feedback_report_{start_date}_to_{end_date}.xlsx"'
    
    return response

def create_summary_sheet(sheet, feedback_data):
    sheet['A1'] = "RYONAN ELECTRIC PHILIPPINES"
    sheet['A1'].font = Font(bold=True, size=16)
    sheet['A1'].alignment = Alignment(horizontal='center')
    sheet.merge_cells('A1:D1')
    
    sheet['A3'] = "FEEDBACK SUMMARY"
    sheet['A3'].font = Font(bold=True, size=14)
    sheet['A3'].alignment = Alignment(horizontal='center')
    sheet.merge_cells('A3:D3')
    
    headers = ['Metric', 'Count', 'Percentage', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cell.border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )
    
    total_feedback = feedback_data.count()
    anonymous_count = feedback_data.filter(is_anonymous=True).count()
    named_count = total_feedback - anonymous_count
    read_count = feedback_data.filter(is_read=True).count()
    unread_count = total_feedback - read_count
    
    summary_data = [
        ('Total Feedback', total_feedback, '100%'),
        ('Anonymous Submissions', anonymous_count, f"{(anonymous_count/total_feedback*100):.1f}%" if total_feedback > 0 else '0%'),
        ('Named Submissions', named_count, f"{(named_count/total_feedback*100):.1f}%" if total_feedback > 0 else '0%'),
        ('Read', read_count, f"{(read_count/total_feedback*100):.1f}%" if total_feedback > 0 else '0%'),
        ('Unread', unread_count, f"{(unread_count/total_feedback*100):.1f}%" if total_feedback > 0 else '0%'),
    ]
    
    row = 6
    for metric, count, percentage in summary_data:
        sheet.cell(row=row, column=1, value=metric)
        sheet.cell(row=row, column=2, value=count)
        sheet.cell(row=row, column=3, value=percentage)
        sheet.cell(row=row, column=4, value="")
        
        for col in range(1, 5):
            cell = sheet.cell(row=row, column=col)
            if metric == 'Total Feedback':
                cell.font = Font(bold=True)
            cell.border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )
        
        row += 1
    
    for col in range(1, 5):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

def create_feedback_sheet(sheet, feedback_data, start_date, end_date):
    sheet['A1'] = "RYONAN ELECTRIC PHILIPPINES"
    sheet['A1'].font = Font(bold=True, size=16)
    sheet['A1'].alignment = Alignment(horizontal='center')
    sheet.merge_cells('A1:F1')
    
    sheet['A2'] = f"Employee Feedback Summary as of {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"
    sheet['A2'].font = Font(bold=True, size=12)
    sheet['A2'].alignment = Alignment(horizontal='center')
    sheet.merge_cells('A2:F2')
    
    headers = ['Date Submitted', 'Submitter Name', 'Email', 'Subject', 'Message', 'Status']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cell.border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )
    
    row = 5
    for feedback in feedback_data:
        status = "Read" if feedback.is_read else "Unread"
        
        sheet.cell(row=row, column=1, value=feedback.created_at.strftime('%Y-%m-%d %H:%M'))
        sheet.cell(row=row, column=2, value=feedback.display_name)
        sheet.cell(row=row, column=3, value=feedback.display_email)
        sheet.cell(row=row, column=4, value=feedback.subject)
        sheet.cell(row=row, column=5, value=feedback.message)
        sheet.cell(row=row, column=6, value=status)
        
        for col in range(1, 7):
            cell = sheet.cell(row=row, column=col)
            cell.border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        row += 1
    
    column_widths = [20, 25, 30, 30, 50, 12]
    for col, width in enumerate(column_widths, 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width