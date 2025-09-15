from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname, get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import json
from .models import PersonalInformation, ContactPerson, EducationalBackground, FamilyBackground, EmploymentInformation
from .forms import PersonalInformationForm, ContactPersonForm, EducationalBackgroundForm, FamilyBackgroundForm, EmploymentInformationForm, CreateEmployeeForm
from userlogin.models import EmployeeLogin
from certificate.models import Certificate
from notification.models import Notification
from generalsettings.models import Department, Position, Line
from django.core.paginator import Paginator
from django.db.models import Q
from django.contrib.auth.hashers import make_password
from django.db import transaction
from django.views.decorators.http import require_GET
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import authenticate
from django.contrib.auth.hashers import check_password
from django.views.decorators.http import require_POST

@login_required(login_url="user-login")
def profile_view(request):
    user = request.user
    personal_info = PersonalInformation.objects.filter(user=user).first()
    contact_person = ContactPerson.objects.filter(user=user).first()
    family_background = FamilyBackground.objects.filter(user=user).first()
    employment_info = EmploymentInformation.objects.filter(user=user).first()
    # Order education records: ongoing education first, then by graduation year (most recent first)
    education_records = EducationalBackground.objects.filter(user=user).order_by(
        '-year_graduated',  # Then by year in descending order (most recent first)
        '-created_at'  # Finally by creation date for records with same year
    )
    personal_form = PersonalInformationForm(instance=personal_info, user=user)
    contact_form = ContactPersonForm(instance=contact_person)
    family_form = FamilyBackgroundForm(instance=family_background)
    employment_form = EmploymentInformationForm(instance=employment_info, user_role='employee')
    certificates = Certificate.objects.filter(employee=user).order_by("-created_at")[:5]
    notifications = Notification.objects.filter(recipient=user)
    approvers = EmployeeLogin.objects.filter(hr_admin=True)
    departments = Department.objects.prefetch_related('lines').all()
    context = {
        'user': user,
        'personal_info': personal_info,
        'contact_person': contact_person,
        'family_background': family_background,
        'employment_info': employment_info,
        'education_records': education_records,
        'personal_form': personal_form,
        'contact_form': contact_form,
        'family_form': family_form,
        'employment_form': employment_form,
        'certificates': certificates,
        'notifications': notifications,
        'approvers': approvers,
        'departments': departments,
    }
    return render(request, 'userprofile/user-profile.html', context)

@login_required(login_url="user-login")
def admin_employees(request):
    if not request.user.hr_admin:
        messages.error(request, 'Access denied. HR permissions required.')
        return redirect('user_profile')

    search_query = request.GET.get('search', '')
    employees = EmployeeLogin.objects.filter(
        wire_admin=False,
        clinic_admin=False,
        iad_admin=False,
        accounting_admin=False,
        hr_admin=False,
        hr_manager=False,
        mis_admin=False,
        is_active=True
    )

    if search_query:
        employees = employees.filter(
            Q(firstname__icontains=search_query) |
            Q(lastname__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(idnumber__icontains=search_query)
        )

    paginator = Paginator(employees, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    departments = Department.objects.all()
    positions = Position.objects.all()
    lines = Line.objects.all()

    context = {
        'employees': page_obj,
        'search_query': search_query,
        'departments': departments,
        'positions': positions,
        'lines': lines,
    }

    return render(request, 'userprofile/admin-profile.html', context)

@login_required(login_url="user-login")
def update_personal_info(request):
    if request.method == 'POST':
        user = request.user
        personal_info, created = PersonalInformation.objects.get_or_create(user=user)
        contact_person, created = ContactPerson.objects.get_or_create(user=user)
        family_background, created = FamilyBackground.objects.get_or_create(user=user)

        personal_form = PersonalInformationForm(request.POST, instance=personal_info, user=user)
        contact_form = ContactPersonForm(request.POST, instance=contact_person)
        family_form = FamilyBackgroundForm(request.POST, instance=family_background)

        if personal_form.is_valid() and contact_form.is_valid() and family_form.is_valid():
            with transaction.atomic():
                personal_info = personal_form.save()
                contact_person = contact_form.save()
                family_background = family_form.save()

                user.firstname = personal_form.cleaned_data.get('first_name', user.firstname)
                user.lastname = personal_form.cleaned_data.get('last_name', user.lastname)
                user.email = personal_form.cleaned_data.get('email', user.email)
                user.save()

            return JsonResponse({'success': True, 'message': 'Personal information updated successfully'})
        else:
            errors = {}
            errors.update(personal_form.errors)
            errors.update(contact_form.errors)
            errors.update(family_form.errors)
            return JsonResponse({'success': False, 'errors': errors})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required(login_url="user-login")
def update_employment_info(request):
    if request.method == 'POST':
        user = request.user
        employment_info, created = EmploymentInformation.objects.get_or_create(user=user)

        user_role = 'admin' if request.user.hr_admin else 'employee'
        employment_form = EmploymentInformationForm(request.POST, instance=employment_info, user_role=user_role)

        if employment_form.is_valid():
            employment_form.save()
            return JsonResponse({'success': True, 'message': 'Employment information updated successfully'})
        else:
            return JsonResponse({'success': False, 'errors': employment_form.errors})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required(login_url="user-login")
def add_education(request):
    if request.method == 'POST':
        form = EducationalBackgroundForm(request.POST)
        if form.is_valid():
            education = form.save(commit=False)
            education.user = request.user
            education.save()
            return JsonResponse({
                'success': True,
                'message': 'Education record added successfully',
                'education': {
                    'id': education.id,
                    'level': education.get_level_display(),
                    'school_name': education.school_name,
                    'degree_course': education.degree_course or '',
                    'year_graduated': education.year_graduated or '',
                    'honors_awards': education.honors_awards or ''
                }
            })
        else:
            return JsonResponse({'success': False, 'errors': form.errors})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required(login_url="user-login")
def update_education(request, education_id):
    education = get_object_or_404(EducationalBackground, id=education_id, user=request.user)

    if request.method == 'POST':
        form = EducationalBackgroundForm(request.POST, instance=education)
        if form.is_valid():
            education = form.save()
            return JsonResponse({
                'success': True,
                'message': 'Education record updated successfully',
                'education': {
                    'id': education.id,
                    'level': education.get_level_display(),
                    'school_name': education.school_name,
                    'degree_course': education.degree_course or '',
                    'year_graduated': education.year_graduated or '',
                    'honors_awards': education.honors_awards or ''
                }
            })
        else:
            return JsonResponse({'success': False, 'errors': form.errors})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required(login_url="user-login")
def delete_education(request, education_id):
    if request.method == 'DELETE':
        education = get_object_or_404(EducationalBackground, id=education_id, user=request.user)
        education.delete()
        return JsonResponse({'success': True, 'message': 'Education record deleted successfully'})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required(login_url="user-login")
def admin_view_employee(request, employee_id):
    if not request.user.hr_admin:
        return JsonResponse({'success': False, 'message': 'Access denied'})

    employee = get_object_or_404(EmployeeLogin, id=employee_id)
    personal_info, created = PersonalInformation.objects.get_or_create(user=employee)
    contact_person, created = ContactPerson.objects.get_or_create(user=employee)
    family_background, created = FamilyBackground.objects.get_or_create(user=employee)
    employment_info, created = EmploymentInformation.objects.get_or_create(user=employee)
    positions = Position.objects.all()
    departments = Department.objects.all()
    lines = Line.objects.all()
    education_records = EducationalBackground.objects.filter(user=employee).order_by(
        'year_graduated',  # Null values (ongoing) appear first
        '-year_graduated',  # Then by year in descending order (most recent first)
        '-created_at'  # Finally by creation date for records with same year
    )

    employment_form = EmploymentInformationForm(instance=employment_info, user_role='admin')
    certificates = Certificate.objects.filter(employee=employee).order_by('-created_at')[:5]

    context = {
        'employee': employee,
        'personal_info': personal_info,
        'contact_person': contact_person,
        'family_background': family_background,
        'employment_info': employment_info,
        'education_records': education_records,
        'employment_form': employment_form,
        'is_admin_view': True,
        'positions':positions,
        'departments':departments,
        'lines':lines,
        'certificates': certificates
    }

    return render(request, 'userprofile/admin_employee_view.html', context)

@login_required(login_url="user-login")
def admin_update_employee(request, employee_id):
    if not request.user.hr_admin:
        return JsonResponse({'success': False, 'message': 'Access denied'})

    if request.method == 'POST':
        employee = get_object_or_404(EmployeeLogin, id=employee_id)
        employment_info, created = EmploymentInformation.objects.get_or_create(user=employee)

        employment_form = EmploymentInformationForm(request.POST, instance=employment_info, user_role='admin')

        if employment_form.is_valid():
            employment_form.save()
            return JsonResponse({'success': True, 'message': 'Employee information updated successfully'})
        else:
            return JsonResponse({'success': False, 'errors': employment_form.errors})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required(login_url="user-login")
def admin_create_employee(request):
    if not request.user.hr_admin:
        return JsonResponse({'success': False, 'message': 'Access denied'})

    if request.method == 'POST':
        form = CreateEmployeeForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                employee = form.save(commit=False)
                employee.username = form.cleaned_data['lastname'].lower()
                employee.password = make_password(f"Repco{form.cleaned_data['idnumber']}")
                employee.save()

                employment_info = EmploymentInformation.objects.create(
                    user=employee,
                    position=form.cleaned_data['position'],
                    department=form.cleaned_data['department'],
                    line=form.cleaned_data['line'],
                    employment_type=form.cleaned_data['employment_type'],
                    date_hired=form.cleaned_data['date_hired'],
                    tin_number=form.cleaned_data.get('tin_number', ''),
                    sss_number=form.cleaned_data.get('sss_number', ''),
                    hdmf_number=form.cleaned_data.get('hdmf_number', ''),
                    philhealth_number=form.cleaned_data.get('philhealth_number', ''),
                    bank_account=form.cleaned_data.get('bank_account', ''),
                )

            return JsonResponse({
                'success': True,
                'message': f'Employee created successfully. Username: {employee.username}, Password: Repco{employee.idnumber}'
            })
        else:
            return JsonResponse({'success': False, 'errors': form.errors})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required(login_url="user-login")
def admin_deactivate_employee(request, employee_id):
    if not request.user.hr_admin:
        return JsonResponse({'success': False, 'message': 'Access denied'})

    if request.method == 'POST':
        employee = get_object_or_404(EmployeeLogin, id=employee_id)
        employee.active = False
        employee.save()
        return JsonResponse({'success': True, 'message': 'Employee deactivated successfully'})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required(login_url="user-login")
def admin_delete_employee(request, employee_id):
    if not request.user.hr_admin:
        return JsonResponse({'success': False, 'message': 'Access denied'})

    if request.method == 'POST':
        employee = get_object_or_404(EmployeeLogin, id=employee_id)
        employee.delete()
        return JsonResponse({'success': True, 'message': 'Employee deleted successfully'})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required(login_url="user-login")
def admin_reset_password(request, employee_id):
    if not request.user.hr_admin:
        return JsonResponse({'success': False, 'message': 'Access denied'})

    if request.method == 'POST':
        employee = get_object_or_404(EmployeeLogin, id=employee_id)
        new_password = f"Repco_{employee.idnumber}"
        employee.password = make_password(new_password)
        employee.save()
        return JsonResponse({'success': True, 'message': f'Password reset successfully. New password: {new_password}'})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required(login_url="user-login")
def admin_activate_employee(request, employee_id):
    if not request.user.hr_admin:
        return JsonResponse({'success': False, 'message': 'Access denied'})
    if request.method == 'POST':
        employee = get_object_or_404(EmployeeLogin, id=employee_id)
        employee.active = True
        employee.save()
        return JsonResponse({'success': True, 'message': 'Employee activated successfully'})
    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required(login_url="user-login")
def admin_lock_employee(request, employee_id):
    if not request.user.hr_admin:
        return JsonResponse({'success': False, 'message': 'Access denied'})

    if request.method == 'POST':
        employee = get_object_or_404(EmployeeLogin, id=employee_id)
        employee.locked = True
        employee.save()
        return JsonResponse({'success': True, 'message': 'Employee account locked successfully'})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required(login_url="user-login")
def admin_unlock_employee(request, employee_id):
    if not request.user.hr_admin:
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    if request.method == 'POST':
        employee = get_object_or_404(EmployeeLogin, id=employee_id)
        employee.locked = False
        employee.save()
        return JsonResponse({'success': True, 'message': 'Employee account unlocked successfully'})
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required(login_url="user-login")
def update_profile_section(request):
    if request.method == 'POST':
        user = request.user
        group = request.POST.get('group')
        
        try:
            with transaction.atomic():
                if group == 'personal-info':
                    # Handle Personal Information and Contact Person
                    personal_info, created = PersonalInformation.objects.get_or_create(user=user)
                    contact_person, contact_created = ContactPerson.objects.get_or_create(user=user)
                    
                    # Update Personal Information
                    personal_info.middle_name = request.POST.get('middle_name', '').strip() or None
                    personal_info.nickname = request.POST.get('nickname', '').strip() or None
                    personal_info.work_email = request.POST.get('work_email', '').strip() or None
                    personal_info.gender = request.POST.get('gender', '').strip() or None
                    birth_date_str = request.POST.get('birth_date')
                    if birth_date_str:
                        try:
                            from datetime import datetime
                            personal_info.birth_date = datetime.strptime(birth_date_str, '%Y-%m-%d').date()
                        except (ValueError, TypeError):
                            personal_info.birth_date = None
                    else:
                        personal_info.birth_date = None
                    personal_info.birth_place = request.POST.get('birth_place', '').strip() or None
                    
                    # Handle contact number with country code
                    contact_country_code = request.POST.get('contact_country_code', '+63')
                    contact_number = request.POST.get('contact_number', '').strip()
                    if contact_number:
                        personal_info.contact_country_code = contact_country_code
                        personal_info.contact_number = contact_number
                    else:
                        personal_info.contact_country_code = None
                        personal_info.contact_number = None
                    
                    # Present Address
                    personal_info.present_country = request.POST.get('present_country', 'Philippines')
                    personal_info.present_province = request.POST.get('present_province', '').strip() or None
                    personal_info.present_city = request.POST.get('present_city', '').strip() or None
                    personal_info.present_barangay = request.POST.get('present_barangay', '').strip() or None
                    personal_info.present_street = request.POST.get('present_street', '').strip() or None
                    personal_info.present_block_lot = request.POST.get('present_block_lot', '').strip() or None
                    
                    # Provincial Address
                    personal_info.provincial_country = request.POST.get('provincial_country', 'Philippines')
                    personal_info.provincial_province = request.POST.get('provincial_province', '').strip() or None
                    personal_info.provincial_city = request.POST.get('provincial_city', '').strip() or None
                    personal_info.provincial_barangay = request.POST.get('provincial_barangay', '').strip() or None
                    personal_info.provincial_street = request.POST.get('provincial_street', '').strip() or None
                    personal_info.provincial_block_lot = request.POST.get('provincial_block_lot', '').strip() or None
                    
                    personal_info.save()
                    
                    # Update Contact Person
                    contact_name = request.POST.get('contact_name', '').strip()
                    contact_relationship = request.POST.get('contact_relationship', '').strip()
                    
                    # Set default values for required fields if empty
                    contact_person.name = contact_name if contact_name else 'N/A'
                    contact_person.relationship = contact_relationship if contact_relationship else 'Other'
                    
                    # Handle contact person contact number with country code
                    contact_contact_country_code = request.POST.get('contact_contact_country_code', '+63')
                    contact_contact_number = request.POST.get('contact_contact_number', '').strip()
                    if contact_contact_number:
                        contact_person.contact_country_code = contact_contact_country_code
                        contact_person.contact_number = contact_contact_number
                    else:
                        # Set default values for required fields if empty
                        contact_person.contact_country_code = '+63'
                        contact_person.contact_number = 'N/A'
                    
                    contact_person.address = request.POST.get('contact_address', '').strip() or None
                    contact_person.save()
                    
                    # Update user basic information
                    first_name = request.POST.get('first_name', '').strip()
                    last_name = request.POST.get('last_name', '').strip()
                    email = request.POST.get('email', '').strip()
                    
                    if first_name:
                        user.firstname = first_name
                    if last_name:
                        user.lastname = last_name
                    if email:
                        user.email = email
                    
                    user.save()
                    
                    # Prepare updated data for response
                    updated_data = {
                        'first_name': user.firstname or '-',
                        'last_name': user.lastname or '-',
                        'middle_name': personal_info.middle_name or '-',
                        'nickname': personal_info.nickname or '-',
                        'work_email': personal_info.work_email or '-',
                        'gender': personal_info.gender or '-',
                        'birth_date': personal_info.birth_date.strftime('%B %d, %Y') if personal_info.birth_date else '-',
                        'birth_place': personal_info.birth_place or '-',
                        'contact_number': f"{personal_info.contact_country_code} {personal_info.contact_number}" if personal_info.contact_number else '-',
                        'present_country': personal_info.present_country or '-',
                        'present_province': personal_info.present_province or '-',
                        'present_city': personal_info.present_city or '-',
                        'present_barangay': personal_info.present_barangay or '-',
                        'present_street': personal_info.present_street or '-',
                        'present_block_lot': personal_info.present_block_lot or '-',
                        'provincial_country': personal_info.provincial_country or '-',
                        'provincial_province': personal_info.provincial_province or '-',
                        'provincial_city': personal_info.provincial_city or '-',
                        'provincial_barangay': personal_info.provincial_barangay or '-',
                        'provincial_street': personal_info.provincial_street or '-',
                        'provincial_block_lot': personal_info.provincial_block_lot or '-',
                        'contact_name': contact_person.name or '-',
                        'contact_relationship': contact_person.relationship or '-',
                        'contact_contact_number': f"{contact_person.contact_country_code} {contact_person.contact_number}" if contact_person.contact_number else '-',
                        'contact_address': contact_person.address or '-',
                    }
                    
                    action = 'created' if created else 'updated'
                    return JsonResponse({
                        'success': True, 
                        'message': f'Personal information {action} successfully',
                        'updated_data': updated_data
                    })
                
                elif group == 'employment-info':
                    # Handle Employment Information
                    employment_info, created = EmploymentInformation.objects.get_or_create(user=user)
                    
                    # Check if user is HR admin to determine which fields can be edited
                    is_hr_admin = user.hr_admin
                    
                    # Fields that can be edited by regular users
                    department_name = request.POST.get('department', '').strip()
                    if department_name:
                        try:
                            employment_info.department = Department.objects.get(department_name=department_name)
                        except Department.DoesNotExist:
                            employment_info.department = None
                    else:
                        employment_info.department = None
                    
                    line_name = request.POST.get('line', '').strip()
                    if line_name:
                        try:
                            employment_info.line = Line.objects.get(line_name=line_name)
                        except Line.DoesNotExist:
                            employment_info.line = None
                    else:
                        employment_info.line = None
                    
                    approver_id = request.POST.get('approver')
                    if approver_id:
                        try:
                            employment_info.approver = EmployeeLogin.objects.get(id=approver_id)
                        except EmployeeLogin.DoesNotExist:
                            employment_info.approver = None
                    else:
                        employment_info.approver = None
                    
                    # Fields that can only be edited by HR admins
                    if is_hr_admin:
                        position_name = request.POST.get('position', '').strip()
                        if position_name:
                            try:
                                employment_info.position = Position.objects.get(position=position_name)
                            except Position.DoesNotExist:
                                employment_info.position = None
                        else:
                            employment_info.position = None
                        
                        # Handle required employment_type field
                        employment_type = request.POST.get('employment_type', '').strip()
                        employment_info.employment_type = employment_type if employment_type else 'Regular'
                        
                        # Handle required date_hired field
                        date_hired_str = request.POST.get('date_hired')
                        if date_hired_str:
                            try:
                                from datetime import datetime
                                employment_info.date_hired = datetime.strptime(date_hired_str, '%Y-%m-%d').date()
                            except (ValueError, TypeError):
                                # Set a default date if parsing fails
                                from datetime import date
                                employment_info.date_hired = date.today()
                        else:
                            # Set a default date if not provided
                            from datetime import date
                            employment_info.date_hired = date.today()
                        
                        employment_info.tin_number = request.POST.get('tin_number', '').strip() or None
                        employment_info.sss_number = request.POST.get('sss_number', '').strip() or None
                        employment_info.hdmf_number = request.POST.get('hdmf_number', '').strip() or None
                        employment_info.philhealth_number = request.POST.get('philhealth_number', '').strip() or None
                        employment_info.bank_account = request.POST.get('bank_account', '').strip() or None
                    else:
                        # For non-HR users, set default values for required fields if they don't exist
                        if not employment_info.employment_type:
                            employment_info.employment_type = 'Regular'
                        if not employment_info.date_hired:
                            from datetime import date
                            employment_info.date_hired = date.today()
                    
                    employment_info.save()
                    
                    # Prepare updated data for response
                    updated_data = {
                        'department': employment_info.department.department_name if employment_info.department else '-',
                        'line': employment_info.line.line_name if employment_info.line else '-',
                        'approver': f"{employment_info.approver.firstname} {employment_info.approver.lastname}" if employment_info.approver else '-',
                    }
                    
                    # Add HR admin only fields to response if user is HR admin
                    if is_hr_admin:
                        updated_data.update({
                            'position': employment_info.position.position if employment_info.position else '-',
                            'employment_type': employment_info.employment_type or '-',
                            'date_hired': employment_info.date_hired.strftime('%B %d, %Y') if employment_info.date_hired else '-',
                            'tin_number': employment_info.tin_number or '-',
                            'sss_number': employment_info.sss_number or '-',
                            'hdmf_number': employment_info.hdmf_number or '-',
                            'philhealth_number': employment_info.philhealth_number or '-',
                            'bank_account': employment_info.bank_account or '-',
                        })
                    
                    action = 'created' if created else 'updated'
                    return JsonResponse({
                        'success': True, 
                        'message': f'Employment information {action} successfully',
                        'updated_data': updated_data
                    })
                
                elif group == 'family-info':
                    # Handle Family Background
                    family_background, created = FamilyBackground.objects.get_or_create(user=user)
                    
                    family_background.mother_name = request.POST.get('mother_name', '').strip() or None
                    family_background.father_name = request.POST.get('father_name', '').strip() or None
                    family_background.spouse_name = request.POST.get('spouse_name', '').strip() or None
                    family_background.children_names = request.POST.get('children_names', '').strip() or None
                    family_background.save()
                    
                    # Prepare updated data for response
                    updated_data = {
                        'mother_name': family_background.mother_name or '-',
                        'father_name': family_background.father_name or '-',
                        'spouse_name': family_background.spouse_name or '-',
                        'children_names': family_background.children_names or '-',
                    }
                    
                    action = 'created' if created else 'updated'
                    return JsonResponse({
                        'success': True, 
                        'message': f'Family background {action} successfully',
                        'updated_data': updated_data
                    })
                
                else:
                    return JsonResponse({
                        'success': False, 
                        'message': 'Invalid section specified'
                    })
                    
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'message': f'Error saving data: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=400)

@login_required(login_url="user-login")
def download_employee_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Employee Import Template'

    # Header row
    headers = [
        'Id Number', 'First Name', 'Last Name', 'Email Address', 'Department', 'Line', 'Position',
        'Employment Type', 'Date Hired', 'TIN Number', 'SSS Number', 'HDMF Number',
        'PhilHealth Number', 'Bank Account'
    ]
    ws.append(headers)

    # Sample data row
    sample = [
        'EMP001', 'Juan', 'Dela Cruz', 'juan.delacruz@email.com', '', '', '',
        'Regular', '2024-01-01', '123-456-789', '12-3456789-0', '1234-5678-9012',
        '12-345678901-2', '1234567890'
    ]
    ws.append(sample)

    # Instruction row (below sample data, not merged, only in first column)
    instruction = [
        'INSTRUCTIONS: Fill out the template below. For Department, Line, and Position, use the dropdowns provided. Do not type values manually. Employment Type also uses a dropdown. Date Hired must be in YYYY-MM-DD format.'
    ] + [''] * (len(headers) - 1)
    ws.append(instruction)

    # Style header row (row 1)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=False)
    for cell in ws[2]:
        cell.alignment = Alignment(wrap_text=False)
    for cell in ws[3]:
        cell.alignment = Alignment(wrap_text=False)

    # Validation data sheet
    validation_ws = wb.create_sheet(title='ValidationData')
    departments = list(Department.objects.values_list('department_name', flat=True))
    lines = list(Line.objects.values_list('line_name', flat=True))
    positions = list(Position.objects.values_list('position', flat=True))
    employment_types = ['Regular', 'Probationary', 'OJT']

    for idx, value in enumerate(departments, 1):
        validation_ws.cell(row=1, column=idx, value=value)
    for idx, value in enumerate(lines, 1):
        validation_ws.cell(row=2, column=idx, value=value)
    for idx, value in enumerate(positions, 1):
        validation_ws.cell(row=3, column=idx, value=value)
    for idx, value in enumerate(employment_types, 1):
        validation_ws.cell(row=4, column=idx, value=value)
    validation_ws.sheet_state = 'hidden'

    # Data validation for Department (E), Line (F), Position (G), Employment Type (H)
    max_row = 1000
    start_row = 2
    # Department
    dept_col = get_column_letter(5)
    if departments:
        dept_last_col = get_column_letter(len(departments))
        dept_range = f"=ValidationData!$A$1:${dept_last_col}$1"
        dv_dept = DataValidation(type="list", formula1=dept_range, allow_blank=True)
        ws.add_data_validation(dv_dept)
        dv_dept.add(f'{dept_col}{start_row}:{dept_col}{max_row}')
    # Line
    line_col = get_column_letter(6)
    if lines:
        line_last_col = get_column_letter(len(lines))
        line_range = f"=ValidationData!$A$2:${line_last_col}$2"
        dv_line = DataValidation(type="list", formula1=line_range, allow_blank=True)
        ws.add_data_validation(dv_line)
        dv_line.add(f'{line_col}{start_row}:{line_col}{max_row}')
    # Position
    pos_col = get_column_letter(7)
    if positions:
        pos_last_col = get_column_letter(len(positions))
        pos_range = f"=ValidationData!$A$3:${pos_last_col}$3"
        dv_pos = DataValidation(type="list", formula1=pos_range, allow_blank=True)
        ws.add_data_validation(dv_pos)
        dv_pos.add(f'{pos_col}{start_row}:{pos_col}{max_row}')
    # Employment Type
    emp_col = get_column_letter(8)
    if employment_types:
        emp_last_col = get_column_letter(len(employment_types))
        emp_range = f"=ValidationData!$A$4:${emp_last_col}$4"
        dv_emp = DataValidation(type="list", formula1=emp_range, allow_blank=True)
        ws.add_data_validation(dv_emp)
        dv_emp.add(f'{emp_col}{start_row}:{emp_col}{max_row}')

    # Set column widths for readability
    for col in ws.columns:
        max_length = 0
        col_letter = None
        for cell in col:
            if hasattr(cell, 'column_letter'):
                col_letter = cell.column_letter
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        if col_letter:
            ws.column_dimensions[col_letter].width = max(15, max_length + 2)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=employee_import_template.xlsx'
    wb.save(response)
    return response

@login_required(login_url="user-login")
def api_provinces(request):
    """API endpoint to get list of Philippine provinces"""
    provinces = [
        "Abra", "Agusan del Norte", "Agusan del Sur", "Aklan", "Albay", "Antique",
        "Apayao", "Aurora", "Basilan", "Bataan", "Batanes", "Batangas", "Benguet",
        "Biliran", "Bohol", "Bukidnon", "Bulacan", "Cagayan", "Camarines Norte",
        "Camarines Sur", "Camiguin", "Capiz", "Catanduanes", "Cavite", "Cebu",
        "Compostela Valley", "Cotabato", "Davao del Norte", "Davao del Sur",
        "Davao Occidental", "Davao Oriental", "Dinagat Islands", "Eastern Samar",
        "Guimaras", "Ifugao", "Ilocos Norte", "Ilocos Sur", "Iloilo", "Isabela",
        "Kalinga", "La Union", "Laguna", "Lanao del Norte", "Lanao del Sur",
        "Leyte", "Maguindanao", "Marinduque", "Masbate", "Metro Manila",
        "Misamis Occidental", "Misamis Oriental", "Mountain Province",
        "Negros Occidental", "Negros Oriental", "Northern Samar", "Nueva Ecija",
        "Nueva Vizcaya", "Occidental Mindoro", "Oriental Mindoro", "Palawan",
        "Pampanga", "Pangasinan", "Quezon", "Quirino", "Rizal", "Romblon",
        "Samar", "Sarangani", "Siquijor", "Sorsogon", "South Cotabato",
        "Southern Leyte", "Sultan Kudarat", "Sulu", "Surigao del Norte",
        "Surigao del Sur", "Tarlac", "Tawi-Tawi", "Zambales", "Zamboanga del Norte",
        "Zamboanga del Sur", "Zamboanga Sibugay"
    ]

    return JsonResponse({
        'success': True,
        'provinces': sorted(provinces)
    })

@login_required(login_url="user-login")
def api_cities(request):
    """API endpoint to get list of cities/municipalities by province"""
    province = request.GET.get('province', '')

    # Philippine cities/municipalities by province
    cities_by_province = {
        "Metro Manila": [
            "Caloocan", "Las Piñas", "Makati", "Malabon", "Mandaluyong", "Manila",
            "Marikina", "Muntinlupa", "Navotas", "Parañaque", "Pasay", "Pasig",
            "Quezon City", "San Juan", "Taguig", "Valenzuela"
        ],
        "Cebu": [
            "Alcantara", "Alcoy", "Alegria", "Aloguinsan", "Argao", "Asturias",
            "Badian", "Balamban", "Bantayan", "Barili", "Bogo", "Boljoon",
            "Borbon", "Carcar", "Carmen", "Catmon", "Cebu City", "Compostela",
            "Consolacion", "Cordova", "Daanbantayan", "Dalaguete", "Danao",
            "Dumanjug", "Ginatilan", "Lapu-Lapu", "Liloan", "Madridejos",
            "Malabuyoc", "Mandaue", "Medellin", "Minglanilla", "Moalboal",
            "Naga", "Oslob", "Pilar", "Pinamungajan", "Poro", "Ronda",
            "Samboan", "San Fernando", "San Francisco", "San Remigio",
            "Santa Fe", "Santander", "Sibonga", "Sogod", "Tabogon", "Tabuelan",
            "Talisay", "Toledo", "Tuburan", "Tudela"
        ],
        "Laguna": [
            "Alaminos", "Bay", "Biñan", "Cabuyao", "Calamba", "Calauan",
            "Cavinti", "Famy", "Kalayaan", "Liliw", "Los Baños", "Luisiana",
            "Lumban", "Mabitac", "Magdalena", "Majayjay", "Nagcarlan",
            "Paete", "Pagsanjan", "Pakil", "Pangil", "Pila", "Rizal",
            "San Pablo", "San Pedro", "Santa Cruz", "Santa Maria",
            "Santa Rosa", "Siniloan", "Victoria"
        ],
        "Bulacan": [
            "Angat", "Balagtas", "Baliuag", "Bocaue", "Bulakan", "Bustos",
            "Calumpit", "Doña Remedios Trinidad", "Guiguinto", "Hagonoy",
            "Marilao", "Meycauayan", "Norzagaray", "Obando", "Pandi",
            "Paombong", "Plaridel", "Pulilan", "San Ildefonso", "San Jose del Monte",
            "San Miguel", "San Rafael", "Santa Maria"
        ],
        "Cavite": [
            "Alfonso", "Amadeo", "Bacoor", "Carmona", "Cavite City", "Dasmariñas",
            "General Emilio Aguinaldo", "General Mariano Alvarez", "General Trias",
            "Imus", "Indang", "Kawit", "Magallanes", "Maragondon", "Mendez",
            "Naic", "Noveleta", "Rosario", "Silang", "Tagaytay", "Tanza", "Ternate", "Trece Martires"
        ]
    }

    # Add more provinces as needed
    if province not in cities_by_province:
        # For provinces not in our data, return a generic list
        cities = [
            "City/Municipality 1", "City/Municipality 2", "City/Municipality 3"
        ]
    else:
        cities = cities_by_province[province]

    return JsonResponse({
        'success': True,
        'cities': sorted(cities)
    })

@login_required(login_url="user-login")
def api_approvers(request):
    """API endpoint to get approvers based on department and position level"""
    department_name = request.GET.get('department')
    
    if not department_name:
        return JsonResponse({'approvers': []})
    
    try:
        department = Department.objects.get(department_name=department_name)
        approvers = EmployeeLogin.objects.filter(
            employment_info__department=department,
            employment_info__position__level__in=['2', '3']
        ).exclude(id=request.user.id)
        
        approver_list = []
        for approver in approvers:
            approver_list.append({
                'id': approver.id,
                'name': f"{approver.firstname} {approver.lastname}",
                'position': approver.employment_info.position.position if approver.employment_info.position else '',
                'level': approver.employment_info.position.level if approver.employment_info.position else ''
            })
        
        return JsonResponse({'approvers': approver_list})
    except Department.DoesNotExist:
        return JsonResponse({'approvers': []})

import uuid
import pandas as pd
from django.core.cache import cache
from django.utils import timezone
from datetime import datetime
import threading
import time
import logging

logger = logging.getLogger(__name__)

@login_required(login_url="user-login")
def import_employees(request):
    if not request.user.hr_admin:
        return JsonResponse({'success': False, 'message': 'Access denied'})

    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})

    if 'files' not in request.FILES:
        return JsonResponse({'success': False, 'message': 'No files uploaded'})

    try:
        uploaded_file = request.FILES['files']

        if not uploaded_file.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({'success': False, 'message': 'Invalid file format. Please upload Excel files only.'})

        try:
            df = pd.read_excel(uploaded_file)
            logger.info(f"Excel file read successfully. Shape: {df.shape}")
            logger.info(f"Columns: {list(df.columns)}")
        except Exception as e:
            logger.error(f"Error reading Excel file: {str(e)}")
            return JsonResponse({'success': False, 'message': f'Error reading Excel file: {str(e)}'})

        df.columns = df.columns.str.strip().str.lower()
        
        # Map column names to expected format
        column_mapping = {
            'id number': 'idnumber',
            'first name': 'firstname', 
            'last name': 'lastname',
            'email address': 'email',
            'department': 'department',
            'line': 'line',
            'position': 'position',
            'employment type': 'employment_type',
            'date hired': 'date_hired',
            'tin number': 'tin_number',
            'sss number': 'sss_number',
            'hdmf number': 'hdmf_number',
            'philhealth number': 'philhealth_number',
            'bank account': 'bank_account'
        }
        
        df.columns = df.columns.map(lambda x: column_mapping.get(x, x))
        
        logger.info(f"Column mapping completed. Final columns: {list(df.columns)}")

        required_columns = ['idnumber', 'firstname', 'lastname', 'email']
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            logger.error(f"Missing required columns: {missing_columns}. Available columns: {list(df.columns)}")
            return JsonResponse({
                'success': False,
                'message': f'Missing required columns: {", ".join(missing_columns)}. Found columns: {", ".join(df.columns)}'
            })

        df = df.dropna(subset=required_columns, how='all')
        logger.info(f"After removing empty rows: {len(df)} rows remaining")

        if len(df) == 0:
            logger.error("No valid data rows found after removing empty rows")
            return JsonResponse({'success': False, 'message': 'No valid data rows found in the Excel file'})

        import_id = str(uuid.uuid4())

        import_data = {
            'data': df.to_dict('records'),
            'total_rows': len(df),
            'processed': 0,
            'success_count': 0,
            'error_count': 0,
            'errors': [],
            'status': 'pending',
            'created_at': timezone.now().isoformat(),
            'user_id': request.user.id
        }

        logger.info(f"Storing import data in cache with ID: {import_id}")
        cache.set(f'import_{import_id}', import_data, timeout=3600) 

        logger.info(f"Starting background thread for import ID: {import_id}")
        thread = threading.Thread(target=process_employee_import, args=(import_id,))
        thread.daemon = True
        thread.start()
        logger.info(f"Background thread started successfully for import ID: {import_id}")

        return JsonResponse({
            'success': True,
            'import_id': import_id,
            'total_rows': len(df),
            'message': 'File uploaded successfully. Processing started.'
        })

    except Exception as e:
        logger.error(f"Error processing file: {str(e)}")
        import traceback
        logger.error(f"Full traceback: {traceback.format_exc()}")
        return JsonResponse({'success': False, 'message': f'Error processing file: {str(e)}'})

@login_required(login_url="user-login")
def import_progress(request, import_id):
    """Get import progress status"""
    if not request.user.hr_admin:
        return JsonResponse({'success': False, 'message': 'Access denied'})

    import_data = cache.get(f'import_{import_id}')
    logger.info(f"Import progress request for ID: {import_id}, data found: {import_data is not None}")

    if not import_data:
        logger.error(f"Import session not found for ID: {import_id}")
        return JsonResponse({'success': False, 'message': 'Import session not found'})

    response_data = {
        'success': True,
        'status': import_data['status'],
        'total': import_data['total_rows'],
        'processed': import_data['processed'],
        'success_count': import_data['success_count'],
        'error_count': import_data['error_count'],
        'errors': import_data['errors']
    }
    
    if import_data.get('error_message'):
        response_data['error_message'] = import_data['error_message']
        logger.error(f"Import error message: {import_data['error_message']}")
    
    logger.info(f"Import progress response: {response_data}")
    return JsonResponse(response_data)

def process_employee_import(import_id):
    try:
        logger.info(f"Starting background import process for ID: {import_id}")
        import_data = cache.get(f'import_{import_id}')

        if not import_data:
            logger.error(f"Import data not found in cache for ID: {import_id}")
            return

        import_data['status'] = 'processing'
        cache.set(f'import_{import_id}', import_data, timeout=3600)
        logger.info(f"Import process started. Total rows to process: {import_data['total_rows']}")

        data_rows = import_data['data']
        total_rows = len(data_rows)

        departments = {dept.department_name.lower(): dept for dept in Department.objects.all()}
        positions = {pos.position.lower(): pos for pos in Position.objects.all()}
        lines = {line.line_name.lower(): line for line in Line.objects.all()}

        employment_type_choices = EmploymentInformation.EMPLOYMENT_TYPE_CHOICES
        valid_employment_types = {choice[0].lower(): choice[0] for choice in employment_type_choices}

        logger.info(f"Available departments: {list(departments.keys())}")
        logger.info(f"Available positions: {list(positions.keys())}")
        logger.info(f"Available lines: {list(lines.keys())}")
        logger.info(f"Valid employment types: {list(valid_employment_types.keys())}")
        
        logger.info(f"Validation data loaded - Departments: {len(departments)}, Positions: {len(positions)}, Lines: {len(lines)}")
        logger.info(f"Valid employment types: {valid_employment_types}")

        for index, row in enumerate(data_rows):
            try:
                time.sleep(0.1)

                logger.info(f"Processing row {index + 1}/{total_rows}: {row.get('firstname', 'N/A')} {row.get('lastname', 'N/A')}")

                required_fields = ['idnumber', 'firstname', 'lastname', 'email']
                missing_fields = [field for field in required_fields if not row.get(field)]
                if missing_fields:
                    logger.error(f"Row {index + 1}: Missing required fields: {missing_fields}")
                    raise ValueError(f"Missing required fields: {', '.join(missing_fields)}")

                if EmployeeLogin.objects.filter(email=row['email']).exists():
                    logger.error(f"Row {index + 1}: Employee with email {row['email']} already exists")
                    raise ValueError(f"Employee with email {row['email']} already exists")

                if EmployeeLogin.objects.filter(idnumber=row['idnumber']).exists():
                    logger.error(f"Row {index + 1}: Employee with ID number {row['idnumber']} already exists")
                    raise ValueError(f"Employee with ID number {row['idnumber']} already exists")

                errors = []

                department_obj = None
                if row.get('department'):
                    dept_key = str(row['department']).lower().strip()
                    if dept_key not in departments:
                        errors.append(f"Invalid department: {row['department']}")
                    else:
                        department_obj = departments[dept_key]

                position_obj = None
                if row.get('position'):
                    pos_key = str(row['position']).lower().strip()
                    if pos_key not in positions:
                        errors.append(f"Invalid position: {row['position']}")
                    else:
                        position_obj = positions[pos_key]

                line_obj = None
                if row.get('line'):
                    line_key = str(row['line']).lower().strip()
                    if line_key not in lines:
                        errors.append(f"Invalid line: {row['line']}")
                    else:
                        line_obj = lines[line_key]

                employment_type = None
                if row.get('employment_type'):
                    emp_type_key = str(row['employment_type']).lower().strip()
                    if emp_type_key in valid_employment_types:
                        employment_type = valid_employment_types[emp_type_key]
                    else:
                        errors.append(f"Invalid employment type: {row['employment_type']}. Valid options: {', '.join([choice[1] for choice in EmploymentInformation.EMPLOYMENT_TYPE_CHOICES])}")

                if not employment_type:
                    employment_type = 'Regular'  # Default to Regular

                if errors:
                    logger.error(f"Row {index + 1}: Validation errors: {errors}")
                    raise ValueError("; ".join(errors))

                with transaction.atomic():
                    try:
                        # Create EmployeeLogin
                        idnumber_val = str(row['idnumber']).strip()
                        default_password = f"Repco{idnumber_val}"
                        employee = EmployeeLogin.objects.create(
                            idnumber=idnumber_val,
                            username=str(row['lastname']).strip().lower(), 
                            firstname=str(row['firstname']).strip(),
                            lastname=str(row['lastname']).strip(),
                            email=str(row['email']).strip().lower(),
                            status='approved',
                            password=make_password(default_password)
                        )

                        logger.info(f"Created EmployeeLogin for {employee.firstname} {employee.lastname}")

                        date_hired = timezone.now().date()
                        if row.get('date_hired'):
                            try:
                                if isinstance(row['date_hired'], str):
                                    date_hired = datetime.strptime(row['date_hired'], '%Y-%m-%d').date()
                                else:
                                    date_hired = row['date_hired']
                            except (ValueError, TypeError) as e:
                                logger.warning(f"Invalid date format for date_hired: {row.get('date_hired')}, using current date")

                        tin_number = str(row.get('tin_number', '')).strip() if row.get('tin_number') else None
                        sss_number = str(row.get('sss_number', '')).strip() if row.get('sss_number') else None
                        hdmf_number = str(row.get('hdmf_number', '')).strip() if row.get('hdmf_number') else None
                        philhealth_number = str(row.get('philhealth_number', '')).strip() if row.get('philhealth_number') else None
                        bank_account = str(row.get('bank_account', '')).strip() if row.get('bank_account') else None

                        logger.info(f"Additional fields for {employee.firstname} {employee.lastname}: TIN={tin_number}, SSS={sss_number}, HDMF={hdmf_number}, PhilHealth={philhealth_number}, Bank={bank_account}")

                        employment_info = EmploymentInformation.objects.create(
                            user=employee,
                            department=department_obj,
                            position=position_obj,
                            line=line_obj,
                            employment_type=employment_type,
                            date_hired=date_hired,
                            tin_number=tin_number,
                            sss_number=sss_number,
                            hdmf_number=hdmf_number,
                            philhealth_number=philhealth_number,
                            bank_account=bank_account
                        )
                        logger.info(f"Created EmploymentInformation for {employee.firstname} {employee.lastname}")

                    except Exception as db_error:
                        logger.error(f"Database error for row {index + 1}: {str(db_error)}")
                        import traceback
                        logger.error(f"Database error traceback: {traceback.format_exc()}")
                        logger.error(f"Row data: {row}")
                        raise ValueError(f"Database error: {str(db_error)}")

                import_data['success_count'] += 1
                logger.info(f"Successfully imported employee {index + 1}/{total_rows}")

            except Exception as e:
                logger.error(f"Error processing row {index + 1}: {str(e)}")
                import traceback
                logger.error(f"Row {index + 1} error traceback: {traceback.format_exc()}")
                import_data['error_count'] += 1
                import_data['errors'].append({
                    'row': index + 2,  # Excel row number (accounting for header)
                    'name': f"{row.get('firstname', '')} {row.get('lastname', '')}".strip(),
                    'email': row.get('email', ''),
                    'errors': str(e)
                })

            # Update progress
            import_data['processed'] = index + 1
            cache.set(f'import_{import_id}', import_data, timeout=3600)

        # Mark as completed
        import_data['status'] = 'completed'
        cache.set(f'import_{import_id}', import_data, timeout=3600)

        logger.info(f"Import completed. Success: {import_data['success_count']}, Errors: {import_data['error_count']}")

    except Exception as e:
        logger.error(f"Import process failed: {str(e)}")
        import traceback
        logger.error(f"Import process failed traceback: {traceback.format_exc()}")
        
        # Try to update the import data with error status
        try:
            import_data = cache.get(f'import_{import_id}')
            if import_data:
                import_data['status'] = 'failed'
                import_data['error_message'] = str(e)
                cache.set(f'import_{import_id}', import_data, timeout=3600)
                logger.info(f"Updated import status to failed for ID: {import_id}")
        except Exception as cache_error:
            logger.error(f"Failed to update cache with error status: {str(cache_error)}")

@login_required(login_url="user-login")
def api_export_filters(request):
    departments = list(Department.objects.values_list('department_name', flat=True))
    lines = list(Line.objects.values_list('line_name', flat=True))
    positions = list(Position.objects.values_list('position', flat=True))
    return JsonResponse({
        'departments': departments,
        'lines': lines,
        'positions': positions,
    })

@login_required(login_url="user-login")
def export_employees(request):
    department = request.GET.get('department')
    line = request.GET.get('line')
    position = request.GET.get('position')
    employment_type = request.GET.get('employment_type')
    status = request.GET.get('status')

    filters = Q()
    if department:
        filters &= Q(employment_info__department__department_name=department)
    if line:
        filters &= Q(employment_info__line__line_name=line)
    if position:
        filters &= Q(employment_info__position__position=position)
    if employment_type:
        filters &= Q(employment_info__employment_type=employment_type)
    if status:
        if status == 'deactivated':
            filters &= Q(active=False)
        else:
            filters &= Q(status=status)

    employees = EmployeeLogin.objects.filter(filters).order_by('idnumber')

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = 'Personal Information'
    ws2 = wb.create_sheet('Employment Information')
    ws3 = wb.create_sheet('Contact Person')
    ws4 = wb.create_sheet('Educational Background')

    # Sheet 1: Personal Information (with Family Background)
    personal_headers = [
        'ID Number', 'First Name', 'Last Name', 'Middle Name', 'Nickname', 'Work Email', 'Gender', 'Birth Date', 'Birth Place',
        'Contact Country Code', 'Contact Number',
        'Present Block/Lot', 'Present Street', 'Present Barangay', 'Present City', 'Present Province', 'Present Country',
        'Provincial Block/Lot', 'Provincial Street', 'Provincial Barangay', 'Provincial City', 'Provincial Province', 'Provincial Country',
        'Mother Name', 'Father Name', 'Spouse Name', 'Children Names'
    ]
    ws1.append(personal_headers)
    for emp in employees:
        pi = getattr(emp, 'personal_info', None)
        fb = getattr(emp, 'family_background', None)
        ws1.append([
            emp.idnumber,
            emp.firstname,
            emp.lastname,
            pi.middle_name if pi else '',
            pi.nickname if pi else '',
            pi.work_email if pi else '',
            pi.gender if pi else '',
            pi.birth_date.strftime('%Y-%m-%d') if pi and pi.birth_date else '',
            pi.birth_place if pi else '',
            pi.contact_country_code if pi else '',
            pi.contact_number if pi else '',
            pi.present_block_lot if pi else '',
            pi.present_street if pi else '',
            pi.present_barangay if pi else '',
            pi.present_city if pi else '',
            pi.present_province if pi else '',
            pi.present_country if pi else '',
            pi.provincial_block_lot if pi else '',
            pi.provincial_street if pi else '',
            pi.provincial_barangay if pi else '',
            pi.provincial_city if pi else '',
            pi.provincial_province if pi else '',
            pi.provincial_country if pi else '',
            fb.mother_name if fb else '',
            fb.father_name if fb else '',
            fb.spouse_name if fb else '',
            fb.children_names if fb else '',
        ])
    for col in range(1, len(personal_headers)+1):
        ws1.column_dimensions[get_column_letter(col)].width = 18
    for cell in ws1[1]:
        cell.font = Font(bold=True)

    # Sheet 2: Employment Information
    employment_headers = [
        'ID Number', 'Name', 'Department', 'Line', 'Position', 'Employment Type', 'Status', 'Date Hired',
        'TIN Number', 'SSS Number', 'HDMF Number', 'PhilHealth Number', 'Bank Account'
    ]
    ws2.append(employment_headers)
    for emp in employees:
        info = getattr(emp, 'employment_info', None)
        name = f"{emp.lastname}, {emp.firstname}"
        if getattr(emp, 'personal_info', None) and emp.personal_info.middle_name:
            name += f" {emp.personal_info.middle_name}"
        ws2.append([
            emp.idnumber,
            name,
            info.department.department_name if info and info.department else '',
            info.line.line_name if info and info.line else '',
            info.position.position if info and info.position else '',
            info.employment_type if info else '',
            'Active' if emp.active and emp.status == 'approved' else ('Pending' if emp.status == 'pending' else 'Deactivated'),
            info.date_hired.strftime('%Y-%m-%d') if info and info.date_hired else '',
            info.tin_number if info else '',
            info.sss_number if info else '',
            info.hdmf_number if info else '',
            info.philhealth_number if info else '',
            info.bank_account if info else '',
        ])
    for col in range(1, len(employment_headers)+1):
        ws2.column_dimensions[get_column_letter(col)].width = 18
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    # Sheet 3: Contact Person
    contact_headers = [
        'ID Number', 'Name', 'Relationship', 'Contact Country Code', 'Contact Number', 'Address'
    ]
    ws3.append(contact_headers)
    for emp in employees:
        cp = getattr(emp, 'contact_person', None)
        ws3.append([
            emp.idnumber,
            cp.name if cp else '',
            cp.relationship if cp else '',
            cp.contact_country_code if cp else '',
            cp.contact_number if cp else '',
            cp.address if cp else '',
        ])
    for col in range(1, len(contact_headers)+1):
        ws3.column_dimensions[get_column_letter(col)].width = 18
    for cell in ws3[1]:
        cell.font = Font(bold=True)

    # Sheet 4: Educational Background (one row per education record)
    education_headers = [
        'ID Number', 'Level', 'School Name', 'Degree/Course', 'Year Graduated', 'Honors/Awards'
    ]
    ws4.append(education_headers)
    for emp in employees:
        educations = emp.education.all()
        for edu in educations:
            ws4.append([
                emp.idnumber,
                edu.level,
                edu.school_name,
                edu.degree_course if edu.degree_course else '',
                edu.year_graduated if edu.year_graduated else '',
                edu.honors_awards if edu.honors_awards else '',
            ])
    for col in range(1, len(education_headers)+1):
        ws4.column_dimensions[get_column_letter(col)].width = 18
    for cell in ws4[1]:
        cell.font = Font(bold=True)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=employees_export.xlsx'
    wb.save(response)
    return response

@login_required(login_url="user-login")
@require_GET
def api_lines(request):
    department = request.GET.get('department', '').strip()
    lines = []
    if department:
        try:
            dept = Department.objects.get(department_name__iexact=department)
            lines = list(dept.lines.values_list('line_name', flat=True))
        except Department.DoesNotExist:
            lines = []
    return JsonResponse({'lines': lines})

@login_required(login_url="user-login")
def admin_approve_employee(request, employee_id):
    if not request.user.hr_admin:
        return JsonResponse({'success': False, 'message': 'Access denied'})
    if request.method == 'POST':
        employee = get_object_or_404(EmployeeLogin, id=employee_id)
        employee.status = 'approved'
        employee.save()
        return JsonResponse({'success': True, 'message': 'Employee approved successfully'})
    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required(login_url="user-login")
def admin_disapprove_employee(request, employee_id):
    if not request.user.hr_admin:
        return JsonResponse({'success': False, 'message': 'Access denied'})
    if request.method == 'POST':
        employee = get_object_or_404(EmployeeLogin, id=employee_id)
        employee.status = 'disapproved'
        employee.save()
        return JsonResponse({'success': True, 'message': 'Employee disapproved successfully'})
    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required(login_url="user-login")
def change_password(request):
    return render(request, 'userprofile/change-password.html')

@csrf_exempt
@login_required(login_url="user-login")
@require_POST
def api_change_password(request):
    user = request.user
    data = request.POST
    current_password = data.get('currentPassword')
    new_password = data.get('newPassword')
    confirm_password = data.get('confirmPassword')

    if not current_password or not new_password or not confirm_password:
        return JsonResponse({'success': False, 'message': 'All fields are required.'}, status=400)

    if not check_password(current_password, user.password):
        return JsonResponse({'success': False, 'message': 'Current password is incorrect.'}, status=400)

    if new_password != confirm_password:
        return JsonResponse({'success': False, 'message': 'New password and confirm password do not match.'}, status=400)

    user.set_password(new_password)
    user.save()
    return JsonResponse({'success': True, 'message': 'Password updated successfully.'})