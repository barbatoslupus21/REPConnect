from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.urls import reverse
from django.core.paginator import Paginator
from django.db.models import Q, Count, F, Case, When, Max
from django.db import models
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.core.exceptions import ValidationError
import json
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
from userlogin.models import EmployeeLogin
import logging
import traceback
from .models import (
    Survey, Question, SurveyResponse, Answer, 
    SurveyTemplate, SurveyCategory, SurveyDraft, SurveyAnalytics
)
from .forms import SurveyForm, QuestionForm, SurveyTemplateForm, SurveyCategoryForm
from django.template.loader import render_to_string
from collections import Counter
from notification.models import Notification

@login_required
def survey_dashboard(request):
    if request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin:
        return redirect('admin_dashboard')
    else:
        return redirect('user_dashboard')

@login_required
def admin_dashboard(request):
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin):
        return redirect('user_dashboard')
    
    # Statistics
    from django.utils import timezone
    from datetime import timedelta

    now = timezone.now()
    first_of_this_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    first_of_last_month = (first_of_this_month - timedelta(days=1)).replace(day=1)

    # Totals - only count surveys created by the current user
    total_surveys = Survey.objects.filter(created_by=request.user).count()
    active_surveys = Survey.objects.filter(created_by=request.user, status='active').count()
    # Only count responses for surveys created by the current user
    total_responses = SurveyResponse.objects.filter(survey__created_by=request.user, is_complete=True).count()

    # Month vs last month comparisons
    # Surveys created this month vs last month (only current user's surveys)
    surveys_this_month = Survey.objects.filter(created_by=request.user, created_at__gte=first_of_this_month).count()
    surveys_last_month = Survey.objects.filter(created_by=request.user, created_at__gte=first_of_last_month, created_at__lt=first_of_this_month).count()
    employees_percent = 0
    employees_positive = True
    if surveys_last_month:
        employees_percent = round(((surveys_this_month - surveys_last_month) / surveys_last_month) * 100, 1)
        employees_positive = surveys_this_month >= surveys_last_month

    # Active surveys this month vs last month (based on created_at of active surveys, only current user's)
    active_this_month = Survey.objects.filter(created_by=request.user, status='active', created_at__gte=first_of_this_month).count()
    active_last_month = Survey.objects.filter(created_by=request.user, status='active', created_at__gte=first_of_last_month, created_at__lt=first_of_this_month).count()
    payslips_percent = 0
    payslips_positive = True
    if active_last_month:
        payslips_percent = round(((active_this_month - active_last_month) / active_last_month) * 100, 1)
        payslips_positive = active_this_month >= active_last_month

    # Responses this month vs last month (only for current user's surveys)
    responses_this_month = SurveyResponse.objects.filter(survey__created_by=request.user, is_complete=True, submitted_at__gte=first_of_this_month).count()
    responses_last_month = SurveyResponse.objects.filter(survey__created_by=request.user, is_complete=True, submitted_at__gte=first_of_last_month, submitted_at__lt=first_of_this_month).count()
    loans_percent = 0
    loans_positive = True
    if responses_last_month:
        loans_percent = round(((responses_this_month - responses_last_month) / responses_last_month) * 100, 1)
        loans_positive = responses_this_month >= responses_last_month
    
    # Recent surveys
    recent_surveys = Survey.objects.select_related('created_by', 'category').order_by('-created_at').filter(created_by=request.user)
    templates_qs = SurveyTemplate.objects.filter(created_by=request.user).order_by("-created_at")
    
    # Apply search filter for templates if provided
    template_search = request.GET.get('search', '').strip()
    if template_search:
        templates_qs = templates_qs.filter(
            Q(name__icontains=template_search) |
            Q(description__icontains=template_search)
        )

    # Pagination for templates tab
    templates_page_number = request.GET.get('templates_page', 1)
    templates_paginator = Paginator(templates_qs, 10)
    templates = templates_paginator.get_page(templates_page_number)

    # Survey categories
    categories = SurveyCategory.objects.annotate(
        survey_count=Count('survey')
    ).order_by('name')
    
    from .forms import SurveyForm

    form = SurveyForm()
    users = EmployeeLogin.objects.filter(is_active=True).order_by('firstname', 'lastname')

    context = {
        'total_surveys': total_surveys,
        'active_surveys': active_surveys,
        'total_responses': total_responses,
        'recent_surveys': recent_surveys,
        'categories': categories,
        'employees_percent': employees_percent,
        'employees_positive': employees_positive,
        'payslips_percent': payslips_percent,
        'payslips_positive': payslips_positive,
        'loans_percent': loans_percent,
        'loans_positive': loans_positive,
        'templates':templates,
        'form': form,
        'users': users,
    }
    
    accept = request.META.get('HTTP_ACCEPT', '')
    if 'application/json' in accept or request.GET.get('ajax') == '1':
        templates_html = render_to_string('survey/partials/admin_template_table.html', {'templates': templates}, request=request)
        templates_pagination_html = render_to_string('survey/partials/admin_templates_pagination.html', {'templates': templates}, request=request)
        return JsonResponse({
            'success': True,
            'templates_html': templates_html,
            'templates_pagination_html': templates_pagination_html,
        })

    return render(request, 'survey/admin_dashboard.html', context)


@login_required
def admin_table(request):
    """AJAX endpoint that returns the surveys table HTML and pagination as JSON.
    This mirrors the tables used in admin_dashboard but returns JSON for the JS loader.
    """
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    # Build the surveys queryset with optional search
    surveys = Survey.objects.select_related('created_by', 'category').order_by('-created_at')
    search = request.GET.get('search', '')
    if search:
        surveys = surveys.filter(
            Q(title__icontains=search) |
            Q(description__icontains=search) |
            Q(category__name__icontains=search)
        )

    paginator = Paginator(surveys, 10)
    page_number = request.GET.get('page') or 1
    page_obj = paginator.get_page(page_number)

    table_html = render_to_string('survey/partials/admin_survey_table.html', {'recent_surveys': page_obj}, request=request)
    pagination_html = render_to_string('survey/partials/admin_survey_pagination.html', {'tickets': page_obj}, request=request)

    return JsonResponse({
        'success': True,
        'table_html': table_html,
        'pagination_html': pagination_html,
    })

@login_required
def user_dashboard(request):
    # If the current user is an admin, do not show the regular user dashboard surveys
    if request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin:
        assigned_surveys = Survey.objects.none()
    else:
        # Show active surveys that are either visible to all users, or explicitly
        # assigned to the current user via the selected_users M2M.
        assigned_surveys = Survey.objects.filter(
            status='active'
        ).filter(
            Q(visibility='all') | Q(visibility='selected', selected_users=request.user)
        ).distinct().order_by('-created_at')
    
    user_responses = SurveyResponse.objects.filter(
        user=request.user,
        is_complete=True
    ).values_list('survey_id', flat=True)
    
    # Categorize surveys with additional status logic
    surveys_data = []
    
    for survey in assigned_surveys:
        # Check if survey is expired
        is_expired = survey.is_expired() if hasattr(survey, 'is_expired') else (
            survey.deadline and timezone.now() > survey.deadline
        )
        
        # Determine status
        if survey.id in user_responses:
            status = 'completed'
        elif is_expired:
            status = 'expired'
        else:
            status = 'pending'
        
        # Get user's progress for this survey
        try:
            response = SurveyResponse.objects.get(survey=survey, user=request.user)
            progress = response.get_completion_percentage()
            completed_questions = response.answers.count()
        except SurveyResponse.DoesNotExist:
            progress = 0
            completed_questions = 0
        
        total_questions = survey.questions.count()
        
        surveys_data.append({
            'survey': survey,
            'status': status,
            'progress': progress,
            'completed_questions': completed_questions,
            'total_questions': total_questions,
            'is_expired': is_expired
        })
    
    # Calculate counts for dashboard stats
    total_surveys = len(surveys_data)
    completed_surveys = sum(1 for s in surveys_data if s['status'] == 'completed')
    pending_surveys = sum(1 for s in surveys_data if s['status'] == 'pending')
    expired_surveys = sum(1 for s in surveys_data if s['status'] == 'expired')
    
    context = {
        'surveys_data': surveys_data,
        'total_surveys': total_surveys,
        'completed_surveys': completed_surveys,
        'pending_surveys': pending_surveys,
        'expired_surveys': expired_surveys,
    }
    
    return render(request, 'survey/user_dashboard.html', context)

@login_required
def survey_list(request):
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin):
        return redirect('user_dashboard')
    
    surveys = Survey.objects.select_related('created_by', 'category').order_by('-created_at')
    
    # Search functionality
    search = request.GET.get('search', '')
    if search:
        surveys = surveys.filter(
            Q(title__icontains=search) |
            Q(description__icontains=search) |
            Q(category__name__icontains=search)
        )
    
    # Filter by status
    status = request.GET.get('status', '')
    if status:
        surveys = surveys.filter(status=status)
    
    # Filter by category
    category = request.GET.get('category', '')
    if category:
        surveys = surveys.filter(category_id=category)
    
    # Pagination
    paginator = Paginator(surveys, 10)
    page_number = request.GET.get('page')
    surveys = paginator.get_page(page_number)
    
    categories = SurveyCategory.objects.all()
    
    context = {
        'surveys': surveys,
        'categories': categories,
        'search': search,
        'status_filter': status,
        'category_filter': category,
    }
    
    return render(request, 'survey/survey_list.html', context)

@login_required
def create_survey(request):
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin):
        messages.error(request, 'You do not have permission to create surveys.')
        return redirect('user_dashboard')
    
    if request.method == 'POST':
        form = SurveyForm(request.POST)
        if form.is_valid():
            survey = form.save(commit=False)
            survey.created_by = request.user
            # If visibility is 'all' make the survey active so it's visible to users
            if survey.visibility == 'all':
                survey.status = 'active'
            survey.save()
            
            # Create questions from template if template is selected
            if survey.template and survey.template.template_data.get('questions'):
                for i, q_data in enumerate(survey.template.template_data['questions']):
                    Question.objects.create(
                        survey=survey,
                        question_text=q_data.get('question_text', ''),
                        question_type=q_data.get('question_type', 'short_answer'),
                        required=q_data.get('required', True),
                        order=i,
                        description=q_data.get('description', ''),
                        options=q_data.get('options', []),
                        min_value=q_data.get('min_value'),
                        max_value=q_data.get('max_value'),
                        validation_rules=q_data.get('validation_rules', {})
                    )
            
            if survey.visibility == 'selected':
                selected_users = request.POST.getlist('selected_users')
                survey.selected_users.set(selected_users)

            # If visibility is 'all', ensure selected_users is empty
            if survey.visibility == 'all':
                survey.selected_users.clear()

            # Create notifications based on survey visibility
            if survey.visibility == 'all':
                # Create a general notification for all users
                Notification.objects.create(
                    title=f"New Survey: {survey.title}",
                    message=f"A new survey '{survey.title}' has been created and is available for you to complete.",
                    notification_type='general',
                    sender=request.user,
                    recipient=request.user,  # For general notifications, recipient can be the creator
                    module='survey_dashboard',
                    for_all=True
                )
            elif survey.visibility == 'selected':
                # Create individual notifications for each selected user
                for user_id in request.POST.getlist('selected_users'):
                    try:
                        recipient = EmployeeLogin.objects.get(id=user_id)
                        Notification.objects.create(
                            title=f"New Survey Assigned: {survey.title}",
                            message=f"You have been assigned a new survey '{survey.title}' that requires your completion.",
                            notification_type='general',
                            sender=request.user,
                            recipient=recipient,
                            module='survey_dashboard',
                            for_all=False
                        )
                    except EmployeeLogin.DoesNotExist:
                        # Skip if user doesn't exist
                        continue

            messages.success(request, 'Survey created successfully!')
            return redirect('edit_survey', survey_id=survey.id)
    else:
        form = SurveyForm()
    
    users = EmployeeLogin.objects.filter(is_active=True).order_by('firstname', 'lastname')
    categories = SurveyCategory.objects.all()
    templates = SurveyTemplate.objects.filter(is_active=True).order_by('name')
    
    context = {
        'form': form,
        'users': users,
        'categories': categories,
        'templates': templates,
    }
    
    return render(request, 'survey/create_survey.html', context)

@login_required
def edit_survey(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or survey.created_by == request.user):
        messages.error(request, 'You do not have permission to edit this survey.')
        return redirect('survey_list')
    
    if request.method == 'POST':
        form = SurveyForm(request.POST, instance=survey)
        if form.is_valid():
            survey = form.save()
            # If visibility changed to 'all', set status active so users can see it
            if survey.visibility == 'all':
                survey.status = 'active'
                survey.selected_users.clear()
                survey.save()
            elif survey.visibility == 'selected':
                selected_users = request.POST.getlist('selected_users')
                survey.selected_users.set(selected_users)
            
            messages.success(request, 'Survey updated successfully!')
            return redirect('survey_detail', survey_id=survey.id)
    else:
        form = SurveyForm(instance=survey)
    
    questions = survey.questions.order_by('order')
    users = EmployeeLogin.objects.filter(is_active=True).order_by('firstname', 'lastname')
    categories = SurveyCategory.objects.all()
    
    context = {
        'form': form,
        'survey': survey,
        'questions': questions,
        'users': users,
        'categories': categories,
    }
    
    return redirect('survey_dashboard')


@login_required
def survey_json(request, survey_id):
    """Return JSON representation of a survey for admin modal edit."""
    survey = get_object_or_404(Survey, id=survey_id)

    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or survey.created_by == request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    # Prepare simple payload: form initial values and selected user ids
    data = {
        'id': survey.id,
        'title': survey.title,
        'description': survey.description,
        'category': survey.category.id if survey.category else None,
        'template': survey.template.id if survey.template else None,
        'deadline': survey.deadline.isoformat() if survey.deadline else None,
        'visibility': survey.visibility,
        'status': survey.status,
        'settings': {
            'allow_multiple_responses': survey.allow_multiple_responses,
            'anonymous_responses': survey.anonymous_responses,
            'randomize_questions': survey.randomize_questions,
            'show_progress': survey.show_progress,
            'auto_save': survey.auto_save,
        },
        'selected_users': list(survey.selected_users.values_list('id', flat=True)),
    }

    return JsonResponse({'success': True, 'survey': data})

@login_required
def survey_details_json(request, survey_id):
    """Return JSON with survey metadata, KPIs and simple trend data for charts."""
    survey = get_object_or_404(Survey, id=survey_id)

    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or survey.created_by == request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    # Assigned users (active)
    assigned_qs = survey.get_assigned_users().filter(is_active=True)
    assigned_count = assigned_qs.count()

    # Answered users (distinct completed responses)
    answered_count = SurveyResponse.objects.filter(survey=survey, is_complete=True).values('user').distinct().count()

    # Trend data: last 14 days responses per day
    from django.db.models.functions import TruncDate
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=13)
    trend_qs = SurveyResponse.objects.filter(
        survey=survey,
        is_complete=True,
        submitted_at__date__gte=start_date,
        submitted_at__date__lte=end_date
    ).annotate(day=TruncDate('submitted_at')).values('day').annotate(cnt=Count('id')).order_by('day')

    trend_map = {item['day'].isoformat(): item['cnt'] for item in trend_qs}
    labels = []
    values = []
    for i in range(14):
        d = start_date + timedelta(days=i)
        labels.append(d.strftime('%Y-%m-%d'))
        values.append(trend_map.get(d.isoformat(), 0))

    payload = {
        'title': survey.title,
        'description': survey.description,
        'category': survey.category.name if survey.category else None,
        'status': survey.status,
        'visibility': survey.visibility,
        'deadline': survey.deadline.isoformat() if survey.deadline else None,
        'created_by': f"{survey.created_by.firstname} {survey.created_by.lastname}" if getattr(survey, 'created_by', None) else None,
        'created_at': survey.created_at.isoformat() if survey.created_at else None,
        'assigned_count': assigned_count,
        'answered_count': answered_count,
        'trend': { 'labels': labels, 'values': values }
    }

    return JsonResponse({'success': True, **payload})


@login_required
def survey_respondents_json(request, survey_id):
    """Return list of respondents (completed responses) for a survey."""
    survey = get_object_or_404(Survey, id=survey_id)

    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or survey.created_by == request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    responses = SurveyResponse.objects.filter(survey=survey, is_complete=True).select_related('user').order_by('-submitted_at')
    rows = []
    for r in responses:
        user = r.user
        name = ''
        try:
            name = f"{user.firstname} {user.lastname}".strip()
        except Exception:
            name = getattr(user, 'username', '')
            
        # Get department from employment_info.department.department_name
        department = ''
        try:
            if hasattr(user, 'employment_info') and user.employment_info and user.employment_info.department:
                department = user.employment_info.department.department_name
        except AttributeError:
            department = ''
            
        submitted_at = r.submitted_at.isoformat() if r.submitted_at else ''
        completion = round(r.get_completion_percentage() or 0)
        rows.append({
            'name': name,
            'department': department,
            'submitted_at': submitted_at,
            'completion_percent': completion
        })

    return JsonResponse({'success': True, 'rows': rows})


@login_required
def survey_stats(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or survey.created_by == request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    assigned = survey.get_assigned_users().filter(is_active=True).count()
    completed = SurveyResponse.objects.filter(survey=survey, is_complete=True).values('user').distinct().count()
    started_not_complete = SurveyResponse.objects.filter(survey=survey, is_complete=False).values('user').distinct().count()
    pending = started_not_complete
    not_started = max(0, assigned - completed - pending)

    return JsonResponse({
        'success': True,
        'total_assigned': assigned,
        'completed': completed,
        'pending': pending,
        'not_started': not_started,
    })


@login_required
def survey_chart_status(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or survey.created_by == request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    assigned = survey.get_assigned_users().filter(is_active=True).count()
    completed = SurveyResponse.objects.filter(survey=survey, is_complete=True).values('user').distinct().count()
    started_not_complete = SurveyResponse.objects.filter(survey=survey, is_complete=False).values('user').distinct().count()
    not_started = max(0, assigned - completed - started_not_complete)

    labels = ['Completed', 'Pending', 'Not Started']
    values = [completed, started_not_complete, not_started]

    return JsonResponse({'success': True, 'labels': labels, 'values': values})


@login_required
def survey_chart_timeline(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or survey.created_by == request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    period = int(request.GET.get('period', 7))
    period = max(1, min(period, 90))
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=period - 1)

    from django.db.models.functions import TruncDate
    qs = SurveyResponse.objects.filter(survey=survey, is_complete=True, submitted_at__date__gte=start_date, submitted_at__date__lte=end_date)
    qs = qs.annotate(day=TruncDate('submitted_at')).values('day').annotate(cnt=Count('id')).order_by('day')
    counts = {item['day'].isoformat(): item['cnt'] for item in qs}

    labels = []
    values = []
    for i in range(period):
        d = start_date + timedelta(days=i)
        labels.append(d.strftime('%Y-%m-%d'))
        values.append(counts.get(d.isoformat(), 0))

    return JsonResponse({'success': True, 'labels': labels, 'values': values})


@login_required
def survey_question_analysis(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or survey.created_by == request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    # Basic placeholder: build minimal analysis structure per question
    questions = []
    for q in survey.questions.order_by('order'):
        questions.append({
            'id': q.id,
            'order': q.order,
            'question_text': q.question_text,
            'question_type': q.question_type,
            'analysis': {}  # detailed analysis can be added later
        })

    return JsonResponse({'success': True, 'questions': questions})


@login_required
def survey_responses_data(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or survey.created_by == request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    qs = SurveyResponse.objects.filter(survey=survey).select_related('user').order_by('-submitted_at')
    search = request.GET.get('search', '').strip()
    if search:
        qs = qs.filter(Q(user__firstname__icontains=search) | Q(user__lastname__icontains=search) | Q(user__email__icontains=search))

    paginator = Paginator(qs, 10)
    page = int(request.GET.get('page', 1))
    page_obj = paginator.get_page(page)

    rows = []
    for r in page_obj.object_list:
        user = r.user
        name = f"{getattr(user, 'firstname', '')} {getattr(user, 'lastname', '')}".strip()
        
        # Get department from employment_info.department.department_name
        department = ''
        try:
            if hasattr(user, 'employment_info') and user.employment_info and user.employment_info.department:
                department = user.employment_info.department.department_name
        except AttributeError:
            department = ''
            
        rows.append({
            'user_id': user.id,
            'employee_name': name,
            'employee_id': getattr(user, 'employee_id', '') or '',
            'department': department,
            'status': 'completed' if r.is_complete else 'incomplete',
            'submitted_at': r.submitted_at.isoformat() if r.submitted_at else None,
            'progress': round(r.get_completion_percentage() or 0)
        })

    pagination = {
        'start_index': page_obj.start_index() if page_obj.paginator.count else 0,
        'end_index': page_obj.end_index() if page_obj.paginator.count else 0,
        'total_count': page_obj.paginator.count,
        'page_range': list(page_obj.paginator.page_range),
        'current_page': page_obj.number,
        'has_previous': page_obj.has_previous(),
        'has_next': page_obj.has_next(),
        'previous_page': page_obj.previous_page_number() if page_obj.has_previous() else None,
        'next_page': page_obj.next_page_number() if page_obj.has_next() else None,
    }

    return JsonResponse({'success': True, 'responses': rows, 'pagination': pagination})

@login_required
def take_survey(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id, status='active')
    
    if survey.visibility == 'selected' and not survey.selected_users.filter(id=request.user.id).exists():
        if not request.user.is_admin:
            messages.error(request, 'You are not assigned to this survey.')
            return redirect('user_dashboard')
    
    if survey.is_expired():
        messages.error(request, 'This survey has expired.')
        return redirect('user_dashboard')
    
    existing_response = SurveyResponse.objects.filter(
        survey=survey, 
        user=request.user, 
        is_complete=True
    ).first()
    
    # Check if this is an edit request (POST with existing completed response)
    is_edit_request = request.method == 'POST' and existing_response
    
    if existing_response and not survey.allow_multiple_responses and not is_edit_request:
        messages.info(request, 'You have already completed this survey.')
        return redirect('user_dashboard')
    
    # For editing, use the existing completed response
    if is_edit_request:
        response = existing_response
    else:
        # For new submissions, get or create an incomplete response
        response, created = SurveyResponse.objects.get_or_create(
            survey=survey,
            user=request.user,
            is_complete=False,
            defaults={
                'ip_address': request.META.get('REMOTE_ADDR'),
                'user_agent': request.META.get('HTTP_USER_AGENT', '')
            }
        )
    
    if request.method == 'POST':
        # Simple debug output
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"DEBUG: POST received for survey {survey_id}")
        
        print(f"DEBUG: POST submission for survey {survey_id}, user {request.user.id}")
        print(f"DEBUG: POST data keys: {list(request.POST.keys())}")
        print(f"DEBUG: Response ID: {response.id}, is_complete: {response.is_complete}")
        
        try:
            questions = survey.questions.order_by('order')
            all_answered = True
            
            print(f"DEBUG: Processing {questions.count()} questions")
            
            for question in questions:
                answer_key = f'question_{question.id}'
                print(f"DEBUG: Processing question {question.id} ({question.question_type}), looking for key: {answer_key}")
                
                answer, created = Answer.objects.get_or_create(
                    response=response,
                    question=question
                )
                print(f"DEBUG: Answer object {'created' if created else 'retrieved'} - ID: {answer.id}")
                
                if question.question_type == 'single_choice':
                    selected = request.POST.get(answer_key)
                    print(f"DEBUG: Single choice value: {selected}")
                    if selected:
                        answer.text_answer = selected  # Use text_answer for single choice
                        print(f"DEBUG: Single choice saved: {answer.text_answer}")
                    elif question.required:
                        all_answered = False
                        
                elif question.question_type == 'multiple_choice':
                    selected = request.POST.getlist(answer_key)
                    if not selected:
                        # Check if it's JSON-stringified (from JavaScript form submission)
                        json_selected = request.POST.get(answer_key)
                        if json_selected:
                            try:
                                import json
                                selected = json.loads(json_selected)
                            except (json.JSONDecodeError, TypeError):
                                selected = []
                    
                    if selected:
                        answer.selected_options = selected
                    elif question.required:
                        all_answered = False
                        
                elif question.question_type == 'rating_scale':
                    # Support two forms for rating answers:
                    # 1) legacy single-value rating under 'question_<id>' (integer)
                    # 2) per-item ratings (e.g. question_<id>_0, question_<id>_1, ...)
                    # If per-item ratings are present, collect them and store as JSON
                    # in Answer.selected_options keyed by item label (or index).
                    item_ratings = {}
                    prefix = f"{answer_key}_"
                    for k, v in request.POST.items():
                        if k.startswith(prefix):
                            idx = k[len(prefix):]
                            if v:
                                try:
                                    val = int(v)
                                except (TypeError, ValueError):
                                    continue
                                # Try to map index to option label when possible
                                key = idx
                                if question.options and idx.isdigit():
                                    i = int(idx)
                                    if 0 <= i < len(question.options):
                                        key = question.options[i]
                                item_ratings[key] = val

                    if item_ratings:
                        # store per-item ratings in selected_options as a mapping
                        answer.selected_options = item_ratings
                    else:
                        # fallback to legacy single-value rating
                        rating = request.POST.get(answer_key)
                        if rating:
                            try:
                                answer.rating_value = int(rating)
                            except (TypeError, ValueError):
                                pass
                        elif question.required:
                            all_answered = False
                        
                elif question.question_type == 'dropdown':
                    selected = request.POST.get(answer_key)
                    if selected:
                        answer.text_answer = selected
                    elif question.required:
                        all_answered = False
                        
                elif question.question_type in ['short_answer', 'paragraph']:
                    text = request.POST.get(answer_key, '').strip()
                    if text:
                        answer.text_answer = text
                    elif question.required:
                        all_answered = False
                        
                elif question.question_type == 'yes_no':
                    value = request.POST.get(answer_key)
                    if value is not None:
                        answer.boolean_answer = value == 'true'
                    elif question.required:
                        all_answered = False
                        
                elif question.question_type == 'date':
                    date_value = request.POST.get(answer_key)
                    if date_value:
                        answer.date_answer = datetime.strptime(date_value, '%Y-%m-%d').date()
                    elif question.required:
                        all_answered = False
                        
                elif question.question_type == 'file_upload':
                    uploaded_file = request.FILES.get(answer_key)
                    print(f"DEBUG: File upload for {answer_key}: {uploaded_file}")
                    if uploaded_file:
                        answer.file_upload = uploaded_file
                        print(f"DEBUG: File saved: {answer.file_upload.name}")
                    elif question.required:
                        all_answered = False
                
                answer.save()
                print(f"DEBUG: Saved answer for question {question.id}")
            
            print(f"DEBUG: All answered: {all_answered}")
            print(f"DEBUG: Submit survey in POST: {'submit_survey' in request.POST}")
            
            if 'submit_survey' in request.POST:
                if all_answered:
                    response.is_complete = True
                    if not response.submitted_at:  # Only set submitted_at if it's a new submission
                        response.submitted_at = timezone.now()
                    response.save()
                    
                    print(f"DEBUG: Response marked as complete and saved")
                    
                    # Return JSON success response for AJAX submission
                    if is_edit_request:
                        message = 'Survey response updated successfully!'
                    else:
                        message = 'Survey submitted successfully!'
                    
                    return JsonResponse({
                        'success': True,
                        'message': message,
                        'survey_title': survey.title
                    })
                else:
                    return JsonResponse({
                        'success': False,
                        'error': 'Please answer all required questions.'
                    }, status=400)
            else:
                # Save progress
                return JsonResponse({
                    'success': True,
                    'message': 'Progress saved successfully!'
                })
        except Exception as e:
            # Log and return traceback as JSON to help debug the 500 error from the client
            logging.exception('Error processing survey submission')
            trace = traceback.format_exc()
            return JsonResponse({'success': False, 'error': str(e), 'trace': trace}, status=500)
    
    # For GET requests, return survey data as JSON for rendering in the dashboard
    questions = survey.questions.order_by('order')
    answers = {answer.question_id: answer for answer in response.answers.all()}
    
    total_questions = questions.count()
    answered_questions = response.answers.count()
    progress = (answered_questions / total_questions * 100) if total_questions > 0 else 0
    
    # Return JSON data for rendering the survey form in the content area
    questions_data = []
    for question in questions:
        answer = answers.get(question.id)
        question_data = {
            'id': question.id,
            'question_text': question.question_text,
            'question_type': question.question_type,
            'required': question.required,
            'description': question.description,
            'options': question.options,
            'answer': None
        }
        
        if answer:
            if question.question_type == 'yes_no':
                question_data['answer'] = answer.boolean_answer
            elif question.question_type in ['short_answer', 'paragraph', 'dropdown']:
                question_data['answer'] = answer.text_answer
            elif question.question_type in ['single_choice', 'multiple_choice']:
                question_data['answer'] = answer.selected_options
            elif question.question_type == 'rating_scale':
                question_data['answer'] = answer.rating_value
            elif question.question_type == 'date':
                question_data['answer'] = answer.date_answer.isoformat() if answer.date_answer else None
        
        questions_data.append(question_data)
    
    return JsonResponse({
        'success': True,
        'survey': {
            'id': survey.id,
            'title': survey.title,
            'description': survey.description,
            'deadline': survey.deadline.isoformat() if survey.deadline else None,
            'show_progress': survey.show_progress
        },
        'questions': questions_data,
        'progress': progress,
        'response_id': response.id
    })

@login_required
def view_response(request, response_id):
    response = get_object_or_404(SurveyResponse, id=response_id)
    
    # Check permissions
    if not (request.user == response.user or request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin):
        messages.error(request, 'You do not have permission to view this response.')
        return redirect('user_dashboard')
    
    questions = response.survey.questions.order_by('order')
    answers = {answer.question_id: answer for answer in response.answers.all()}
    
    context = {
        'response': response,
        'questions': questions,
        'answers': answers,
    }
    
    return render(request, 'survey/view_response.html', context)

@login_required
@require_POST
def add_question(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or survey.created_by == request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        data = json.loads(request.body)
        
        # Get the next order number
        last_order = survey.questions.aggregate(models.Max('order'))['order__max'] or 0
        
        question = Question.objects.create(
            survey=survey,
            question_text=data.get('question_text', ''),
            question_type=data.get('question_type', 'short_answer'),
            required=data.get('required', True),
            order=last_order + 1,
            description=data.get('description', ''),
            options=data.get('options', []),
            min_value=data.get('min_value'),
            max_value=data.get('max_value'),
            validation_rules=data.get('validation_rules', {})
        )
        
        return JsonResponse({
            'success': True,
            'question': {
                'id': question.id,
                'question_text': question.question_text,
                'question_type': question.question_type,
                'required': question.required,
                'order': question.order,
                'options': question.options,
            }
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@require_POST
def update_question(request, question_id):
    question = get_object_or_404(Question, id=question_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or question.survey.created_by == request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        data = json.loads(request.body)
        
        question.question_text = data.get('question_text', question.question_text)
        question.question_type = data.get('question_type', question.question_type)
        question.required = data.get('required', question.required)
        question.description = data.get('description', question.description)
        question.options = data.get('options', question.options)
        question.min_value = data.get('min_value', question.min_value)
        question.max_value = data.get('max_value', question.max_value)
        question.validation_rules = data.get('validation_rules', question.validation_rules)
        
        question.save()
        
        return JsonResponse({'success': True})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@require_POST
def delete_question(request, question_id):
    question = get_object_or_404(Question, id=question_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or question.survey.created_by == request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    question.delete()
    return JsonResponse({'success': True})

@login_required
@require_POST
def reorder_questions(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or survey.created_by == request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        data = json.loads(request.body)
        question_ids = data.get('question_ids', [])
        
        for index, question_id in enumerate(question_ids):
            Question.objects.filter(id=question_id, survey=survey).update(order=index + 1)
        
        return JsonResponse({'success': True})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
def export_survey_responses(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or survey.created_by == request.user):
        messages.error(request, 'You do not have permission to export this survey.')
        return redirect('survey_list')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Survey Responses - {survey.title[:20]}"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Create headers
    headers = ['Response ID', 'User', 'Email', 'Submitted At', 'Completion Status']
    questions = survey.questions.order_by('order')
    
    for question in questions:
        headers.append(f"Q{question.order}: {question.question_text[:50]}")
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Write data
    responses = SurveyResponse.objects.filter(survey=survey).select_related('user').order_by('-submitted_at')
    
    for row, response in enumerate(responses, 2):
        ws.cell(row=row, column=1, value=response.id)
        ws.cell(row=row, column=2, value=f"{response.user.firstname} {response.user.lastname}")
        ws.cell(row=row, column=3, value=response.user.email)
        ws.cell(row=row, column=4, value=response.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if response.submitted_at else 'Not submitted')
        ws.cell(row=row, column=5, value='Complete' if response.is_complete else 'Incomplete')
        
        # Get answers for this response
        answers = {answer.question_id: answer for answer in response.answers.all()}
        
        for col, question in enumerate(questions, 6):
            answer = answers.get(question.id)
            if answer:
                if question.question_type in ['single_choice', 'multiple_choice']:
                    value = ', '.join(answer.selected_options) if answer.selected_options else ''
                elif question.question_type == 'rating_scale':
                    value = answer.rating_value or ''
                elif question.question_type == 'yes_no':
                    value = 'Yes' if answer.boolean_answer else 'No' if answer.boolean_answer is not None else ''
                elif question.question_type == 'date':
                    value = answer.date_answer.strftime('%Y-%m-%d') if answer.date_answer else ''
                else:
                    value = answer.text_answer or ''
                
                ws.cell(row=row, column=col, value=value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Survey_{survey.id}_{survey.title[:20]}_responses.xlsx"'
    
    wb.save(response)
    return response

@login_required
def export_survey_excel_detailed(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or survey.created_by == request.user):
        messages.error(request, 'You do not have permission to export this survey.')
        return redirect('survey_list')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Response Data"
    
    # Styling
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold_font = Font(bold=True)
    
    # Row counter
    current_row = 1
    
    # 1. Header: Ryonan Electric Philippines Corporation
    ws.cell(row=current_row, column=1, value="Ryonan Electric Philippines Corporation")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=16)
    current_row += 1
    
    # 2. Subheader: Response data for (survey name)
    ws.cell(row=current_row, column=1, value=f"Response data for {survey.title}")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
    current_row += 2
    
    # 3. Statistics Summary Table
    ws.cell(row=current_row, column=1, value="Question Analysis Summary")
    ws.cell(row=current_row, column=1).font = bold_font
    current_row += 1
    
    # Summary table headers
    summary_headers = ["Question", "Least Chosen", "Highest Chosen"]
    for col, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    current_row += 1
    
    # Get questions that support statistics (single_choice, multiple_choice, dropdown, rating_scale, yes_no)
    statistical_questions = survey.questions.filter(
        question_type__in=['single_choice', 'multiple_choice', 'dropdown', 'rating_scale', 'yes_no']
    ).order_by('order')
    
    from collections import Counter
    
    for question in statistical_questions:
        answers = Answer.objects.filter(question=question, response__is_complete=True)
        
        if question.question_type in ['single_choice', 'dropdown']:
            # Single choice and dropdown store their values in text_answer
            options_count = Counter()
            for answer in answers:
                if answer.text_answer:
                    options_count[answer.text_answer] += 1
        
        elif question.question_type == 'multiple_choice':
            # Multiple choice stores in selected_options array
            options_count = Counter()
            for answer in answers:
                if answer.selected_options:
                    options_count.update(answer.selected_options)
        
        elif question.question_type == 'rating_scale':
            # Count rating values
            options_count = Counter()
            for answer in answers:
                if answer.selected_options and isinstance(answer.selected_options, dict):
                    # Per-item ratings: count each item's ratings
                    for item, rating in answer.selected_options.items():
                        options_count[f"{item}: {rating}"] += 1
                elif answer.rating_value is not None:
                    # Single rating value
                    options_count[str(answer.rating_value)] += 1
        
        elif question.question_type == 'yes_no':
            # Count yes/no responses
            options_count = Counter()
            for answer in answers:
                if answer.boolean_answer is not None:
                    options_count['Yes' if answer.boolean_answer else 'No'] += 1
        
        # Find least and highest chosen options
        if options_count:
            most_common = options_count.most_common()
            highest = most_common[0]  # (option, count)
            least = most_common[-1]   # (option, count)
            
            total_responses = sum(options_count.values())
            highest_percentage = (highest[1] / total_responses) * 100 if total_responses > 0 else 0
            least_percentage = (least[1] / total_responses) * 100 if total_responses > 0 else 0
            
            highest_text = f"{highest[0]} ({highest_percentage:.1f}%)"
            least_text = f"{least[0]} ({least_percentage:.1f}%)"
        else:
            highest_text = "No responses"
            least_text = "No responses"
        
        # Add row to summary table
        row_data = [question.question_text, least_text, highest_text]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        current_row += 1
    
    current_row += 2  # Add spacing
    
    # 4. Detailed Response Data Table
    ws.cell(row=current_row, column=1, value="Detailed Response Data")
    ws.cell(row=current_row, column=1).font = bold_font
    current_row += 1
    
    # Get all questions for the detailed table
    questions = survey.questions.order_by('order')
    
    # Create detailed table headers
    detailed_headers = ['ID Number', 'Name', 'Department', 'Line']
    for question in questions:
        detailed_headers.append(question.question_text)
    
    # Write detailed headers
    for col, header in enumerate(detailed_headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    current_row += 1
    
    # Get all users who have responded
    responses = SurveyResponse.objects.filter(
        survey=survey, is_complete=True
    ).select_related('user').order_by('user__idnumber')
    
    for response in responses:
        user = response.user
        
        # Get employment info
        department = ''
        line = ''
        try:
            if hasattr(user, 'employment_info') and user.employment_info:
                if user.employment_info.department:
                    department = user.employment_info.department.department_name
                if user.employment_info.line:
                    line = user.employment_info.line.line_name
        except AttributeError:
            pass
        
        # Basic user info
        row_data = [
            user.idnumber or user.username,
            f"{user.firstname} {user.lastname}".strip(),
            department,
            line
        ]
        
        # Get answers for this response
        answers = {answer.question_id: answer for answer in response.answers.all()}
        
        # Add answer data for each question
        for question in questions:
            answer = answers.get(question.id)
            if answer:
                if question.question_type == 'single_choice':
                    # Single choice stores answer in text_answer field
                    value = answer.text_answer or ''
                elif question.question_type == 'multiple_choice':
                    # Multiple choice stores in selected_options array
                    value = ', '.join(answer.selected_options) if answer.selected_options else ''
                elif question.question_type == 'dropdown':
                    # Dropdown stores answer in text_answer field
                    value = answer.text_answer or ''
                elif question.question_type == 'rating_scale':
                    # Check if it's per-item ratings (stored in selected_options) or single rating
                    if answer.selected_options and isinstance(answer.selected_options, dict):
                        # Per-item ratings: format as "Item1: 4, Item2: 5"
                        rating_pairs = [f"{k}: {v}" for k, v in answer.selected_options.items()]
                        value = ', '.join(rating_pairs)
                    else:
                        # Single rating value
                        value = str(answer.rating_value) if answer.rating_value is not None else ''
                elif question.question_type == 'yes_no':
                    value = 'Yes' if answer.boolean_answer else 'No' if answer.boolean_answer is not None else ''
                elif question.question_type == 'date':
                    value = answer.date_answer.strftime('%Y-%m-%d') if answer.date_answer else ''
                elif question.question_type == 'file_upload':
                    value = answer.file_upload.name if answer.file_upload else ''
                else:
                    # For short_answer, paragraph, and other text-based questions
                    value = answer.text_answer or ''
            else:
                value = ''
            
            row_data.append(value)
        
        # Write row data
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        current_row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="survey-response-data.xlsx"'
    
    wb.save(response)
    return response

@login_required
@require_POST
def update_template(request, template_id):
    template = get_object_or_404(SurveyTemplate, id=template_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or template.created_by == request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        data = json.loads(request.body)
        
        template.name = data.get('name', template.name)
        template.description = data.get('description', template.description)
        template.template_data = data.get('template_data', template.template_data)
        template.save()
        
        return JsonResponse({'success': True, 'message': 'Template updated successfully'})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@require_POST
def delete_template(request, template_id):
    template = get_object_or_404(SurveyTemplate, id=template_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or template.created_by == request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        template.is_active = False
        template.save()
        return JsonResponse({'success': True, 'message': 'Template deleted successfully'})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@require_POST
def update_category(request, category_id):
    category = get_object_or_404(SurveyCategory, id=category_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        data = json.loads(request.body)
        
        category.name = data.get('name', category.name)
        category.color = data.get('color', category.color)
        category.icon = data.get('icon', category.icon)
        category.save()
        
        return JsonResponse({'success': True, 'message': 'Category updated successfully'})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@require_POST
def delete_category(request, category_id):
    category = get_object_or_404(SurveyCategory, id=category_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        # Check if category is in use
        surveys_using_category = Survey.objects.filter(category=category).count()
        if surveys_using_category > 0:
            return JsonResponse({
                'error': f'Cannot delete category. {surveys_using_category} surveys are using this category.'
            }, status=400)
        
        category.delete()
        return JsonResponse({'success': True, 'message': 'Category deleted successfully'})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
def create_survey_from_template(request, template_id):
    template = get_object_or_404(SurveyTemplate, id=template_id, is_active=True)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin):
        messages.error(request, 'You do not have permission to create surveys from templates.')
        return redirect('create_template')
    
    try:
        # Create new survey from template
        survey = Survey.objects.create(
            title=f"{template.name} - Survey",
            description=template.description,
            created_by=request.user,
            template=template,
            status='draft'
        )
        
        # Create questions from template data
        if template.template_data.get('questions'):
            for q_data in template.template_data['questions']:
                Question.objects.create(
                    survey=survey,
                    question_text=q_data.get('question_text', ''),
                    question_type=q_data.get('question_type', 'short_answer'),
                    required=q_data.get('required', True),
                    order=q_data.get('order', 0),
                    description=q_data.get('description', ''),
                    options=q_data.get('options', []),
                    min_value=q_data.get('min_value'),
                    max_value=q_data.get('max_value'),
                    validation_rules=q_data.get('validation_rules', {})
                )
        
        messages.success(request, 'Survey created from template successfully!')
        return redirect('edit_survey', survey_id=survey.id)
        
    except Exception as e:
        messages.error(request, f'Error creating survey from template: {str(e)}')
        return redirect('create_template')

@login_required
def save_as_template(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or survey.created_by == request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            template_name = data.get('name', f"{survey.title} Template")
            template_description = data.get('description', survey.description)
            
            # Build template data from survey questions
            questions_data = []
            for question in survey.questions.order_by('order'):
                questions_data.append({
                    'question_text': question.question_text,
                    'question_type': question.question_type,
                    'required': question.required,
                    'order': question.order,
                    'description': question.description,
                    'options': question.options,
                    'min_value': question.min_value,
                    'max_value': question.max_value,
                    'validation_rules': question.validation_rules,
                })
            
            template = SurveyTemplate.objects.create(
                name=template_name,
                description=template_description,
                created_by=request.user,
                template_data={'questions': questions_data}
            )
            
            return JsonResponse({
                'success': True, 
                'message': 'Survey saved as template successfully!',
                'template_id': template.id
            })
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)

@login_required
def survey_analytics_data(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or survey.created_by == request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        analytics_data = {}
        questions = survey.questions.order_by('order')
        
        for question in questions:
            answers = Answer.objects.filter(
                question=question, 
                response__is_complete=True
            )
            
            question_analytics = {
                'question_text': question.question_text,
                'question_type': question.question_type,
                'total_responses': answers.count(),
            }
            
            if question.question_type in ['single_choice', 'multiple_choice']:
                # Count options
                option_counts = {}
                total_selections = 0
                
                for answer in answers:
                    for option in answer.selected_options:
                        option_counts[option] = option_counts.get(option, 0) + 1
                        total_selections += 1
                
                # Calculate percentages
                option_percentages = {}
                if total_selections > 0:
                    for option, count in option_counts.items():
                        option_percentages[option] = {
                            'count': count,
                            'percentage': round((count / answers.count()) * 100, 1)
                        }
                
                question_analytics.update({
                    'chart_type': 'pie' if question.question_type == 'single_choice' else 'bar',
                    'options_data': option_percentages
                })
                
            elif question.question_type == 'rating_scale':
                # Rating distribution and average
                ratings = [answer.rating_value for answer in answers if answer.rating_value]
                
                if ratings:
                    rating_counts = {}
                    for rating in ratings:
                        rating_counts[rating] = rating_counts.get(rating, 0) + 1
                    
                    avg_rating = sum(ratings) / len(ratings)
                    
                    question_analytics.update({
                        'chart_type': 'bar',
                        'average_rating': round(avg_rating, 2),
                        'rating_distribution': rating_counts,
                        'total_ratings': len(ratings)
                    })
                
            elif question.question_type == 'yes_no':
                # Yes/No counts
                yes_count = answers.filter(boolean_answer=True).count()
                no_count = answers.filter(boolean_answer=False).count()
                
                question_analytics.update({
                    'chart_type': 'pie',
                    'yes_no_data': {
                        'Yes': {
                            'count': yes_count,
                            'percentage': round((yes_count / answers.count()) * 100, 1) if answers.count() > 0 else 0
                        },
                        'No': {
                            'count': no_count,
                            'percentage': round((no_count / answers.count()) * 100, 1) if answers.count() > 0 else 0
                        }
                    }
                })
                
            elif question.question_type in ['short_answer', 'paragraph']:
                # Text response analysis
                text_responses = [answer.text_answer for answer in answers if answer.text_answer]
                
                question_analytics.update({
                    'chart_type': 'text',
                    'text_responses': text_responses[:10],  # Show first 10 responses
                    'total_text_responses': len(text_responses)
                })
            
            analytics_data[f'question_{question.id}'] = question_analytics
        
        return JsonResponse({'success': True, 'analytics': analytics_data})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
def export_csv_responses(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or survey.created_by == request.user):
        messages.error(request, 'You do not have permission to export this survey.')
        return redirect('survey_list')
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="Survey_{survey.id}_{survey.title[:20]}_responses.csv"'
    
    writer = csv.writer(response)
    
    # Create headers
    headers = ['Response ID', 'User', 'Email', 'Submitted At', 'Completion Status']
    questions = survey.questions.order_by('order')
    
    for question in questions:
        headers.append(f"Q{question.order}: {question.question_text[:50]}")
    
    writer.writerow(headers)
    
    # Write data
    responses = SurveyResponse.objects.filter(survey=survey).select_related('user').order_by('-submitted_at')
    
    for survey_response in responses:
        row = [
            survey_response.id,
            f"{survey_response.user.firstname} {survey_response.user.lastname}",
            survey_response.user.email,
            survey_response.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if survey_response.submitted_at else 'Not submitted',
            'Complete' if survey_response.is_complete else 'Incomplete'
        ]
        
        # Get answers for this response
        answers = {answer.question_id: answer for answer in survey_response.answers.all()}
        
        for question in questions:
            answer = answers.get(question.id)
            if answer:
                if question.question_type in ['single_choice', 'multiple_choice']:
                    value = ', '.join(answer.selected_options) if answer.selected_options else ''
                elif question.question_type == 'rating_scale':
                    value = answer.rating_value or ''
                elif question.question_type == 'yes_no':
                    value = 'Yes' if answer.boolean_answer else 'No' if answer.boolean_answer is not None else ''
                elif question.question_type == 'date':
                    value = answer.date_answer.strftime('%Y-%m-%d') if answer.date_answer else ''
                else:
                    value = answer.text_answer or ''
            else:
                value = ''
            
            row.append(value)
        
        writer.writerow(row)
    
    return response

@login_required
@require_POST
def delete_survey(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or survey.created_by == request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        survey_title = survey.title
        survey.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Survey "{survey_title}" deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
def survey_participants(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or survey.created_by == request.user):
        messages.error(request, 'You do not have permission to view participants for this survey.')
        return redirect('survey_list')
    
    # Get assigned users
    assigned_users = survey.get_assigned_users()
    
    # Get completion status for each user
    completed_user_ids = set(
        survey.responses.filter(is_complete=True).values_list('user_id', flat=True)
    )
    
    participants_data = []
    for user in assigned_users:
        participants_data.append({
            'user': user,
            'completed': user.id in completed_user_ids,
            'response': survey.responses.filter(user=user).first()
        })
    
    context = {
        'survey': survey,
        'participants': participants_data,
        'total_assigned': len(participants_data),
        'total_completed': len(completed_user_ids),
        'completion_rate': (len(completed_user_ids) / len(participants_data) * 100) if participants_data else 0
    }
    
    return render(request, 'survey/survey_participants.html', context)

@login_required
def get_survey_progress(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    
    # Check if user has access to this survey
    if survey.visibility == 'selected' and not survey.selected_users.filter(id=request.user.id).exists():
        if not request.user.is_admin:
            return JsonResponse({'error': 'Access denied'}, status=403)
    
    try:
        # Get user's response
        response = SurveyResponse.objects.filter(
            survey=survey, 
            user=request.user
        ).first()
        
        if not response:
            return JsonResponse({
                'success': True,
                'progress': 0,
                'completed_questions': 0,
                'total_questions': survey.questions.count(),
                'has_response': False
            })
        
        total_questions = survey.questions.count()
        completed_questions = response.answers.count()
        progress = (completed_questions / total_questions * 100) if total_questions > 0 else 0
        
        return JsonResponse({
            'success': True,
            'progress': round(progress, 1),
            'completed_questions': completed_questions,
            'total_questions': total_questions,
            'has_response': True,
            'is_complete': response.is_complete,
            'submitted_at': response.submitted_at.isoformat() if response.submitted_at else None
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
def survey_analytics(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or survey.created_by == request.user):
        messages.error(request, 'You do not have permission to view analytics for this survey.')
        return redirect('survey_list')
    
    responses = SurveyResponse.objects.filter(survey=survey, is_complete=True)
    questions = survey.questions.order_by('order')
    
    analytics_data = {}
    
    for question in questions:
        answers = Answer.objects.filter(question=question, response__is_complete=True)
        
        if question.question_type in ['single_choice', 'multiple_choice']:
            # Count options
            option_counts = {}
            for answer in answers:
                for option in answer.selected_options:
                    option_counts[option] = option_counts.get(option, 0) + 1
            analytics_data[question.id] = {
                'type': 'chart',
                'chart_type': 'pie' if question.question_type == 'single_choice' else 'bar',
                'data': option_counts
            }
            
        elif question.question_type == 'rating_scale':
            # Average rating and distribution
            ratings = [answer.rating_value for answer in answers if answer.rating_value]
            if ratings:
                avg_rating = sum(ratings) / len(ratings)
                rating_counts = {}
                for rating in ratings:
                    rating_counts[rating] = rating_counts.get(rating, 0) + 1
                analytics_data[question.id] = {
                    'type': 'rating',
                    'average': avg_rating,
                    'distribution': rating_counts
                }
                
        elif question.question_type == 'yes_no':
            # Yes/No counts
            yes_count = answers.filter(boolean_answer=True).count()
            no_count = answers.filter(boolean_answer=False).count()
            analytics_data[question.id] = {
                'type': 'chart',
                'chart_type': 'pie',
                'data': {'Yes': yes_count, 'No': no_count}
            }
    
    context = {
        'survey': survey,
        'questions': questions,
        'total_responses': responses.count(),
        'analytics_data': json.dumps(analytics_data),
    }
    
    return render(request, 'survey/survey_analytics.html', context)

@login_required
@csrf_exempt
def save_draft(request, survey_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    survey = get_object_or_404(Survey, id=survey_id, status='active')
    
    try:
        data = json.loads(request.body)
        
        draft, created = SurveyDraft.objects.update_or_create(
            survey=survey,
            user=request.user,
            defaults={'draft_data': data}
        )
        
        return JsonResponse({'success': True, 'message': 'Draft saved successfully'})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
def load_draft(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    
    try:
        draft = SurveyDraft.objects.get(survey=survey, user=request.user)
        return JsonResponse({'success': True, 'data': draft.draft_data})
    except SurveyDraft.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'No draft found'})

# Template Management Views
@login_required
def template_list(request):
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin):
        return redirect('user_dashboard')
    
    templates = SurveyTemplate.objects.filter(is_active=True).order_by('-created_at')
    
    context = {
        'templates': templates,
    }
    
    return redirect('survey_dashboard')

@login_required
def create_template(request, template_id=None):
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin):
        messages.error(request, 'You do not have permission to create templates.')
        return redirect('user_dashboard')
    
    if request.method == 'POST':
        # Support JSON POSTs from the SPA/template-builder (application/json)
        content_type = request.META.get('CONTENT_TYPE', '')
        if content_type.startswith('application/json'):
            try:
                data = json.loads(request.body)
                name = data.get('name', '').strip() or 'Untitled Template'
                description = data.get('description', '').strip() or ''
                template_data = {
                    'questions': data.get('questions', []),
                    'settings': data.get('settings', {}),
                }

                template = SurveyTemplate.objects.create(
                    name=name,
                    description=description,
                    created_by=request.user,
                    template_data=template_data
                )

                return JsonResponse({'success': True, 'template_id': template.id})
            except Exception as e:
                return JsonResponse({'error': str(e)}, status=400)

        # fallback: traditional form submit
        form = SurveyTemplateForm(request.POST)
        if form.is_valid():
            template = form.save(commit=False)
            template.created_by = request.user
            template.save()
            
            messages.success(request, 'Template created successfully!')
            return redirect('create_template')
    else:
        form = SurveyTemplateForm()

    # If editing an existing template, load its data into the context so the JS can initialize
    edit_template = None
    template_json = None
    if template_id:
        try:
            edit_template = SurveyTemplate.objects.get(id=template_id)
            template_json = edit_template.template_data
            # prefill form fields
            form = SurveyTemplateForm(instance=edit_template)
        except SurveyTemplate.DoesNotExist:
            messages.error(request, 'Template not found.')
            return redirect('admin_dashboard')

    context = {
        'form': form,
        'edit_template': edit_template,
        'template_json': json.dumps(template_json) if template_json is not None else None,
    }

    return render(request, 'survey/create_template.html', context)

# Category Management Views
@login_required
def category_list(request):
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin):
        return redirect('user_dashboard')
    
    categories = SurveyCategory.objects.annotate(
        survey_count=Count('survey')
    ).order_by('name')
    
    context = {
        'categories': categories,
    }
    
    return render(request, 'survey/category_list.html', context)

@login_required
@require_POST
def create_category(request):
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        data = json.loads(request.body)
        
        category = SurveyCategory.objects.create(
            name=data.get('name'),
            color=data.get('color', '#6366f1'),
            icon=data.get('icon', 'fas fa-folder')
        )
        
        return JsonResponse({
            'success': True,
            'category': {
                'id': category.id,
                'name': category.name,
                'color': category.color,
                'icon': category.icon,
            }
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@require_POST
def duplicate_survey(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or survey.created_by == request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        # Create duplicate survey
        new_survey = Survey.objects.create(
            title=f"{survey.title} (Copy)",
            description=survey.description,
            category=survey.category,
            created_by=request.user,
            visibility=survey.visibility,
            deadline=survey.deadline,
            status='draft',
            allow_multiple_responses=survey.allow_multiple_responses,
            anonymous_responses=survey.anonymous_responses,
            randomize_questions=survey.randomize_questions,
            show_progress=survey.show_progress,
            auto_save=survey.auto_save,
        )
        
        # Copy questions
        for question in survey.questions.order_by('order'):
            Question.objects.create(
                survey=new_survey,
                question_text=question.question_text,
                question_type=question.question_type,
                required=question.required,
                order=question.order,
                description=question.description,
                options=question.options,
                min_value=question.min_value,
                max_value=question.max_value,
                max_file_size=question.max_file_size,
                allowed_file_types=question.allowed_file_types,
                validation_rules=question.validation_rules,
            )
        
        return JsonResponse({
            'success': True,
            'survey_id': new_survey.id,
            'message': 'Survey duplicated successfully!'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@require_POST
def update_survey_status(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or survey.created_by == request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        data = json.loads(request.body)
        new_status = data.get('status')
        
        if new_status in dict(Survey.STATUS_CHOICES):
            survey.status = new_status
            survey.save()
            return JsonResponse({'success': True, 'message': f'Survey status updated to {survey.get_status_display()}'})
        else:
            return JsonResponse({'error': 'Invalid status'}, status=400)
            
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
def survey_preview(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or survey.created_by == request.user):
        messages.error(request, 'You do not have permission to preview this survey.')
        return redirect('survey_list')
    
    questions = survey.questions.order_by('order')
    
    context = {
        'survey': survey,
        'questions': questions,
        'preview_mode': True,
    }
    
    return render(request, 'survey/survey_preview.html', context)


@login_required
def survey_preview_json(request, survey_id):
    """Return survey metadata and questions as JSON for client-side rendering."""
    survey = get_object_or_404(Survey, id=survey_id)

    # Ensure the user is allowed to view/take this survey
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or
            survey.visibility == 'all' or request.user in survey.selected_users.all()):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    questions = []
    qs = list(survey.questions.order_by('order'))
    for q in qs:
        questions.append({
            'id': q.id,
            'text': q.question_text,
            'type': q.question_type,
            'required': q.required,
            'options': q.options,
            'description': q.description,
            'min_value': q.min_value,
            'max_value': q.max_value,
        })

    # Fallback to template data if survey has no DB-backed questions
    if not questions and getattr(survey, 'template', None):
        try:
            tmpl = survey.template
            tmpl_questions = tmpl.template_data.get('questions', []) if tmpl and isinstance(tmpl.template_data, dict) else []
            for idx, tq in enumerate(tmpl_questions):
                # Expect template question keys: question_text, question_type, required, options, description
                # For some templates (rating matrices) items are stored under 'rows'
                q_type = tq.get('question_type') or tq.get('type') or 'short_answer'
                opts = []
                if q_type == 'rating_scale' and tq.get('rows'):
                    opts = tq.get('rows') or []
                else:
                    opts = tq.get('options', []) or []

                questions.append({
                    'id': f'tpl-{idx+1}',
                    'text': tq.get('question_text') or tq.get('text') or '',
                    'type': q_type,
                    'required': tq.get('required', True),
                    'options': opts,
                    'description': tq.get('description', '') or '',
                    'min_value': tq.get('min_value'),
                    'max_value': tq.get('max_value'),
                })
        except Exception:
            # If template data is malformed, ignore and return empty questions list
            pass

    data = {
        'id': survey.id,
        'title': survey.title,
        'description': survey.description,
        'questions': questions,
        'can_take': survey.status == 'active' and not survey.is_expired(),
    }

    return JsonResponse({'success': True, 'survey': data})

@login_required
def get_question_data(request, question_id):
    question = get_object_or_404(Question, id=question_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or question.survey.created_by == request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    data = {
        'id': question.id,
        'question_text': question.question_text,
        'question_type': question.question_type,
        'required': question.required,
        'description': question.description,
        'options': question.options,
        'min_value': question.min_value,
        'max_value': question.max_value,
        'max_file_size': question.max_file_size,
        'allowed_file_types': question.allowed_file_types,
        'validation_rules': question.validation_rules,
    }
    
    return JsonResponse({'success': True, 'question': data})

@login_required
def survey_responses_list(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or survey.created_by == request.user):
        messages.error(request, 'You do not have permission to view responses for this survey.')
        return redirect('survey_list')
    
    responses = SurveyResponse.objects.filter(
        survey=survey,
        is_complete=True
    ).select_related('user').order_by('-submitted_at')
    
    # Search functionality
    search = request.GET.get('search', '')
    if search:
        responses = responses.filter(
            Q(user__firstname__icontains=search) |
            Q(user__lastname__icontains=search) |
            Q(user__email__icontains=search)
        )
    
    # Date filtering
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if date_from:
        try:
            date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
            responses = responses.filter(submitted_at__date__gte=date_from)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
            responses = responses.filter(submitted_at__date__lte=date_to)
        except ValueError:
            pass
    
    # Pagination
    paginator = Paginator(responses, 20)
    page_number = request.GET.get('page')
    responses = paginator.get_page(page_number)
    
    context = {
        'survey': survey,
        'responses': responses,
        'search': search,
        'date_from': request.GET.get('date_from', ''),
        'date_to': request.GET.get('date_to', ''),
    }
    
    return render(request, 'survey/survey_responses.html', context)

@login_required
def bulk_export_surveys(request):
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            survey_ids = data.get('survey_ids', [])
            export_format = data.get('format', 'excel')
            
            if not survey_ids:
                return JsonResponse({'error': 'No surveys selected'}, status=400)
            
            surveys = Survey.objects.filter(id__in=survey_ids)
            
            if export_format == 'excel':
                return create_bulk_excel_export(surveys)
            else:
                return JsonResponse({'error': 'Unsupported format'}, status=400)
                
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)

def create_bulk_excel_export(surveys):
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    for survey in surveys:
        # Create sheet for each survey
        sheet_name = survey.title[:31]  # Excel sheet name limit
        ws = wb.create_sheet(title=sheet_name)
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Create headers
        headers = ['Response ID', 'User', 'Email', 'Submitted At']
        questions = survey.questions.order_by('order')
        
        for question in questions:
            headers.append(f"Q{question.order}: {question.question_text[:50]}")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Write data
        responses = SurveyResponse.objects.filter(
            survey=survey, 
            is_complete=True
        ).select_related('user').order_by('-submitted_at')
        
        for row, response in enumerate(responses, 2):
            ws.cell(row=row, column=1, value=response.id)
            ws.cell(row=row, column=2, value=f"{response.user.firstname} {response.user.lastname}")
            ws.cell(row=row, column=3, value=response.user.email)
            ws.cell(row=row, column=4, value=response.submitted_at.strftime('%Y-%m-%d %H:%M:%S'))
            
            # Get answers for this response
            answers = {answer.question_id: answer for answer in response.answers.all()}
            
            for col, question in enumerate(questions, 5):
                answer = answers.get(question.id)
                if answer:
                    if question.question_type in ['single_choice', 'multiple_choice']:
                        value = ', '.join(answer.selected_options) if answer.selected_options else ''
                    elif question.question_type == 'rating_scale':
                        value = answer.rating_value or ''
                    elif question.question_type == 'yes_no':
                        value = 'Yes' if answer.boolean_answer else 'No' if answer.boolean_answer is not None else ''
                    elif question.question_type == 'date':
                        value = answer.date_answer.strftime('%Y-%m-%d') if answer.date_answer else ''
                    else:
                        value = answer.text_answer or ''
                    
                    ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Bulk_Survey_Export.xlsx"'
    
    wb.save(response)
    return response

@login_required
def clear_survey_responses(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or survey.created_by == request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method == 'POST':
        try:
            response_count = survey.responses.count()
            survey.responses.all().delete()
            
            return JsonResponse({
                'success': True,
                'message': f'Cleared {response_count} responses'
            })
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)

@login_required
def get_survey_stats(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or survey.created_by == request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    total_assigned = survey.get_assigned_users().count()
    total_responses = survey.responses.filter(is_complete=True).count()
    completion_rate = (total_responses / total_assigned * 100) if total_assigned > 0 else 0
    
    # Response rate over time (last 30 days)
    thirty_days_ago = timezone.now() - timedelta(days=30)
    
    daily_responses = survey.responses.filter(
        submitted_at__gte=thirty_days_ago,
        is_complete=True
    ).extra(
        {'day': 'date(submitted_at)'}
    ).values('day').annotate(count=Count('id')).order_by('day')
    
    stats = {
        'total_assigned': total_assigned,
        'total_responses': total_responses,
        'completion_rate': round(completion_rate, 1),
        'daily_responses': list(daily_responses),
        'avg_completion_time': None,  # Could calculate average time to complete
    }
    
    return JsonResponse({'success': True, 'stats': stats})

@login_required
@require_POST
def send_survey_reminders(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or survey.created_by == request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        # Get users who haven't completed the survey
        assigned_users = survey.get_assigned_users()
        completed_user_ids = survey.responses.filter(is_complete=True).values_list('user_id', flat=True)
        pending_users = assigned_users.exclude(id__in=completed_user_ids)
        
        # Here you would integrate with your email system
        # For now, just return success with count
        reminder_count = pending_users.count()
        
        return JsonResponse({
            'success': True,
            'message': f'Reminders sent to {reminder_count} users'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
def export_survey_summary(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or survey.created_by == request.user):
        messages.error(request, 'You do not have permission to export this survey summary.')
        return redirect('survey_list')
    
    # Create workbook for summary
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Survey Summary"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Survey Info
    ws['A1'] = 'Survey Summary Report'
    ws['A1'].font = Font(bold=True, size=16)
    
    ws['A3'] = 'Survey Title:'
    ws['B3'] = survey.title
    ws['A4'] = 'Description:'
    ws['B4'] = survey.description
    ws['A5'] = 'Created By:'
    ws['B5'] = f"{survey.created_by.firstname} {survey.created_by.lastname}"
    ws['A6'] = 'Created Date:'
    ws['B6'] = survey.created_at.strftime('%Y-%m-%d')
    ws['A7'] = 'Status:'
    ws['B7'] = survey.get_status_display()
    
    # Statistics
    ws['A9'] = 'Response Statistics'
    ws['A9'].font = Font(bold=True, size=14)
    
    total_assigned = survey.get_assigned_users().count()
    total_responses = survey.get_response_count()
    completion_rate = survey.get_completion_rate()
    
    ws['A11'] = 'Total Assigned:'
    ws['B11'] = total_assigned
    ws['A12'] = 'Total Responses:'
    ws['B12'] = total_responses
    ws['A13'] = 'Completion Rate:'
    ws['B13'] = f"{completion_rate:.1f}%"
    
    # Question Summary
    current_row = 16
    ws[f'A{current_row}'] = 'Question Analysis'
    ws[f'A{current_row}'].font = Font(bold=True, size=14)
    current_row += 2
    
    questions = survey.questions.order_by('order')
    
    for question in questions:
        ws[f'A{current_row}'] = f"Q{question.order}: {question.question_text}"
        ws[f'A{current_row}'].font = Font(bold=True)
        current_row += 1
        
        ws[f'A{current_row}'] = 'Type:'
        ws[f'B{current_row}'] = question.get_question_type_display()
        current_row += 1
        
        # Get answer statistics
        answers = Answer.objects.filter(question=question, response__is_complete=True)
        answer_count = answers.count()
        
        ws[f'A{current_row}'] = 'Responses:'
        ws[f'B{current_row}'] = answer_count
        current_row += 1
        
        if question.question_type in ['single_choice', 'multiple_choice']:
            # Count each option
            option_counts = {}
            for answer in answers:
                for option in answer.selected_options:
                    option_counts[option] = option_counts.get(option, 0) + 1
            
            for option, count in option_counts.items():
                ws[f'B{current_row}'] = f"{option}: {count} ({count/answer_count*100:.1f}%)" if answer_count > 0 else f"{option}: 0"
                current_row += 1
                
        elif question.question_type == 'rating_scale':
            # Calculate average rating
            ratings = [answer.rating_value for answer in answers if answer.rating_value]
            if ratings:
                avg_rating = sum(ratings) / len(ratings)
                ws[f'B{current_row}'] = f"Average Rating: {avg_rating:.1f}/5"
                current_row += 1
        
        current_row += 1  # Add spacing between questions
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Survey_{survey.id}_Summary.xlsx"'
    
    wb.save(response)
    return response

@login_required
@require_POST
def update_template(request, template_id):
    template = get_object_or_404(SurveyTemplate, id=template_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or template.created_by == request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        data = json.loads(request.body)
        
        template.name = data.get('name', template.name)
        template.description = data.get('description', template.description)
        template.template_data = data.get('template_data', template.template_data)
        template.save()
        
        return JsonResponse({'success': True, 'message': 'Template updated successfully'})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@require_POST
def delete_template(request, template_id):
    template = get_object_or_404(SurveyTemplate, id=template_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or template.created_by == request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        template.delete()
        return JsonResponse({'success': True, 'message': 'Template deleted successfully'})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@require_POST
def update_category(request, category_id):
    category = get_object_or_404(SurveyCategory, id=category_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        data = json.loads(request.body)
        
        category.name = data.get('name', category.name)
        category.color = data.get('color', category.color)
        category.icon = data.get('icon', category.icon)
        category.save()
        
        return JsonResponse({'success': True, 'message': 'Category updated successfully'})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@require_POST
def delete_category(request, category_id):
    category = get_object_or_404(SurveyCategory, id=category_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        # Check if category is in use
        surveys_using_category = Survey.objects.filter(category=category).count()
        if surveys_using_category > 0:
            return JsonResponse({
                'error': f'Cannot delete category. {surveys_using_category} surveys are using this category.'
            }, status=400)
        
        category.delete()
        return JsonResponse({'success': True, 'message': 'Category deleted successfully'})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
def create_survey_from_template(request, template_id):
    template = get_object_or_404(SurveyTemplate, id=template_id, is_active=True)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin):
        messages.error(request, 'You do not have permission to create surveys from templates.')
        return redirect('create_template')
    
    try:
        # Create new survey from template
        survey = Survey.objects.create(
            title=f"{template.name} - Survey",
            description=template.description,
            created_by=request.user,
            template=template,
            status='draft'
        )
        
        # Create questions from template data
        if template.template_data.get('questions'):
            for q_data in template.template_data['questions']:
                Question.objects.create(
                    survey=survey,
                    question_text=q_data.get('question_text', ''),
                    question_type=q_data.get('question_type', 'short_answer'),
                    required=q_data.get('required', True),
                    order=q_data.get('order', 0),
                    description=q_data.get('description', ''),
                    options=q_data.get('options', []),
                    min_value=q_data.get('min_value'),
                    max_value=q_data.get('max_value'),
                    validation_rules=q_data.get('validation_rules', {})
                )
        
        messages.success(request, 'Survey created from template successfully!')
        return redirect('edit_survey', survey_id=survey.id)
        
    except Exception as e:
        messages.error(request, f'Error creating survey from template: {str(e)}')
        return redirect('create_template')

@login_required
def save_as_template(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or survey.created_by == request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            template_name = data.get('name', f"{survey.title} Template")
            template_description = data.get('description', survey.description)
            
            # Build template data from survey questions
            questions_data = []
            for question in survey.questions.order_by('order'):
                questions_data.append({
                    'question_text': question.question_text,
                    'question_type': question.question_type,
                    'required': question.required,
                    'order': question.order,
                    'description': question.description,
                    'options': question.options,
                    'min_value': question.min_value,
                    'max_value': question.max_value,
                    'validation_rules': question.validation_rules,
                })
            
            template = SurveyTemplate.objects.create(
                name=template_name,
                description=template_description,
                created_by=request.user,
                template_data={'questions': questions_data}
            )
            
            return JsonResponse({
                'success': True, 
                'message': 'Survey saved as template successfully!',
                'template_id': template.id
            })
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)

@login_required
def survey_analytics_data(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or survey.created_by == request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        analytics_data = {}
        questions = survey.questions.order_by('order')
        
        for question in questions:
            answers = Answer.objects.filter(
                question=question, 
                response__is_complete=True
            )
            
            question_analytics = {
                'question_text': question.question_text,
                'question_type': question.question_type,
                'total_responses': answers.count(),
            }
            
            if question.question_type in ['single_choice', 'multiple_choice']:
                # Count options
                option_counts = {}
                total_selections = 0
                
                for answer in answers:
                    for option in answer.selected_options:
                        option_counts[option] = option_counts.get(option, 0) + 1
                        total_selections += 1
                
                # Calculate percentages
                option_percentages = {}
                if total_selections > 0:
                    for option, count in option_counts.items():
                        option_percentages[option] = {
                            'count': count,
                            'percentage': round((count / answers.count()) * 100, 1)
                        }
                
                question_analytics.update({
                    'chart_type': 'pie' if question.question_type == 'single_choice' else 'bar',
                    'options_data': option_percentages
                })
                
            elif question.question_type == 'rating_scale':
                # Rating distribution and average
                ratings = [answer.rating_value for answer in answers if answer.rating_value]
                
                if ratings:
                    rating_counts = {}
                    for rating in ratings:
                        rating_counts[rating] = rating_counts.get(rating, 0) + 1
                    
                    avg_rating = sum(ratings) / len(ratings)
                    
                    question_analytics.update({
                        'chart_type': 'bar',
                        'average_rating': round(avg_rating, 2),
                        'rating_distribution': rating_counts,
                        'total_ratings': len(ratings)
                    })
                
            elif question.question_type == 'yes_no':
                # Yes/No counts
                yes_count = answers.filter(boolean_answer=True).count()
                no_count = answers.filter(boolean_answer=False).count()
                
                question_analytics.update({
                    'chart_type': 'pie',
                    'yes_no_data': {
                        'Yes': {
                            'count': yes_count,
                            'percentage': round((yes_count / answers.count()) * 100, 1) if answers.count() > 0 else 0
                        },
                        'No': {
                            'count': no_count,
                            'percentage': round((no_count / answers.count()) * 100, 1) if answers.count() > 0 else 0
                        }
                    }
                })
                
            elif question.question_type in ['short_answer', 'paragraph']:
                # Text response analysis
                text_responses = [answer.text_answer for answer in answers if answer.text_answer]
                
                question_analytics.update({
                    'chart_type': 'text',
                    'text_responses': text_responses[:10],  # Show first 10 responses
                    'total_text_responses': len(text_responses)
                })
            
            analytics_data[f'question_{question.id}'] = question_analytics
        
        return JsonResponse({'success': True, 'analytics': analytics_data})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
def export_csv_responses(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or survey.created_by == request.user):
        messages.error(request, 'You do not have permission to export this survey.')
        return redirect('survey_list')
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="Survey_{survey.id}_{survey.title[:20]}_responses.csv"'
    
    writer = csv.writer(response)
    
    # Create headers
    headers = ['Response ID', 'User', 'Email', 'Submitted At', 'Completion Status']
    questions = survey.questions.order_by('order')
    
    for question in questions:
        headers.append(f"Q{question.order}: {question.question_text[:50]}")
    
    writer.writerow(headers)
    
    # Write data
    responses = SurveyResponse.objects.filter(survey=survey).select_related('user').order_by('-submitted_at')

@login_required
def survey_detail(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    
    if not (request.user.hr_admin or request.user.iad_admin or request.user.accounting_admin or survey.created_by == request.user):
        messages.error(request, 'You do not have permission to view this survey.')
        return redirect('survey_list')
    
    context = {
        'survey': survey,
    }

    return render(request, 'survey/survey_details.html', context)

@login_required
def survey_stats(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    
    # Get all assigned users
    assigned_users = survey.get_assigned_users()
    total_assigned = assigned_users.count()
    
    # Get completed responses
    completed_responses = SurveyResponse.objects.filter(
        survey=survey, 
        is_complete=True
    ).count()
    
    # Get started but not completed responses
    started_responses = SurveyResponse.objects.filter(
        survey=survey, 
        is_complete=False
    ).count()
    
    pending = total_assigned - completed_responses - started_responses
    
    data = {
        'total_assigned': total_assigned,
        'completed': completed_responses,
        'pending': pending,
        'started': started_responses
    }
    
    return JsonResponse(data)

@login_required
def survey_chart_data(request, survey_id, chart_type):
    survey = get_object_or_404(Survey, id=survey_id)
    
    if chart_type == 'status':
        # Response status distribution
        assigned_users = survey.get_assigned_users()
        total_assigned = assigned_users.count()
        
        completed_responses = SurveyResponse.objects.filter(
            survey=survey, 
            is_complete=True
        ).count()
        
        started_responses = SurveyResponse.objects.filter(
            survey=survey, 
            is_complete=False
        ).count()
        
        not_started = total_assigned - completed_responses - started_responses
        
        data = {
            'labels': ['Completed', 'Not Started', 'Started'],
            'values': [completed_responses, not_started, started_responses]
        }
        
    elif chart_type == 'timeline':
        # Response timeline
        period_days = int(request.GET.get('period', 7))
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=period_days - 1)
        
        # Get responses by date
        responses_by_date = SurveyResponse.objects.filter(
            survey=survey,
            is_complete=True,
            submitted_at__date__gte=start_date,
            submitted_at__date__lte=end_date
        ).extra(
            select={'date': 'DATE(submitted_at)'}
        ).values('date').annotate(
            count=Count('id')
        ).order_by('date')
        
        # Create date range
        date_range = []
        current_date = start_date
        while current_date <= end_date:
            date_range.append(current_date.strftime('%Y-%m-%d'))
            current_date += timedelta(days=1)
        
        # Fill in data
        response_dict = {item['date'].strftime('%Y-%m-%d'): item['count'] for item in responses_by_date}
        values = [response_dict.get(date, 0) for date in date_range]
        labels = [datetime.strptime(date, '%Y-%m-%d').strftime('%m/%d') for date in date_range]
        
        data = {
            'labels': labels,
            'values': values
        }
    
    return JsonResponse(data)

@login_required
def survey_question_analysis(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    questions = survey.questions.order_by('order')
    
    analysis_data = []
    
    for question in questions:
        question_data = {
            'id': question.id,
            'question_text': question.question_text,
            'question_type': question.question_type,
            'order': question.order,
            'analysis': get_question_analysis(question)
        }
        analysis_data.append(question_data)
    
    return JsonResponse({'questions': analysis_data})

def get_question_analysis(question):
    """Generate analysis data for a specific question"""
    answers = Answer.objects.filter(question=question, response__is_complete=True)
    total_responses = answers.count()
    
    if total_responses == 0:
        return None
    
    analysis = {'total_responses': total_responses}
    
    if question.question_type in ['single_choice', 'dropdown']:
        # Single choice analysis
        options_data = []
        for option in question.options:
            count = answers.filter(text_answer=option).count()
            options_data.append({
                'text': option,
                'count': count
            })
        analysis['options'] = options_data
        
    elif question.question_type == 'multiple_choice':
        # Multiple choice analysis
        all_selections = []
        for answer in answers:
            if answer.selected_options:
                all_selections.extend(answer.selected_options)
        
        option_counts = Counter(all_selections)
        options_data = []
        for option in question.options:
            count = option_counts.get(option, 0)
            options_data.append({
                'text': option,
                'count': count
            })
        analysis['options'] = options_data
        
    elif question.question_type == 'rating_scale':
        # Rating scale analysis
        ratings_data = []
        min_val = question.min_value or 1
        max_val = question.max_value or 5
        
        for rating in range(min_val, max_val + 1):
            count = answers.filter(rating_value=rating).count()
            ratings_data.append({
                'value': rating,
                'count': count,
                'label': f'{rating} Star{"s" if rating != 1 else ""}'
            })
        analysis['ratings'] = ratings_data
        
    elif question.question_type == 'yes_no':
        # Yes/No analysis
        yes_count = answers.filter(boolean_answer=True).count()
        no_count = answers.filter(boolean_answer=False).count()
        analysis.update({
            'yes_count': yes_count,
            'no_count': no_count
        })
        
    elif question.question_type in ['short_answer', 'paragraph']:
        # Text analysis - sample answers
        sample_answers = list(answers.exclude(
            text_answer__isnull=True
        ).exclude(
            text_answer=''
        ).values_list('text_answer', flat=True)[:10])
        
        analysis['sample_answers'] = sample_answers
        
    elif question.question_type == 'file_upload':
        # File upload analysis
        uploaded_count = answers.exclude(file_upload__isnull=True).exclude(file_upload='').count()
        no_upload_count = total_responses - uploaded_count
        
        analysis.update({
            'uploaded_count': uploaded_count,
            'no_upload_count': no_upload_count
        })
        
    elif question.question_type == 'date':
        # Date analysis
        date_distribution = []
        date_answers = answers.exclude(date_answer__isnull=True).values('date_answer').annotate(
            count=Count('id')
        ).order_by('-count')
        
        for item in date_answers:
            date_distribution.append({
                'date': item['date_answer'].strftime('%Y-%m-%d'),
                'count': item['count']
            })
        
        analysis['date_distribution'] = date_distribution
    
    return analysis

@login_required
def survey_responses_data(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    
    # Get search query
    search_query = request.GET.get('search', '').strip()
    
    # Get all assigned users
    assigned_users = survey.get_assigned_users()
    
    # Create base queryset with user info and response status
    users_data = []
    
    for user in assigned_users:
        # Get user's response if exists
        try:
            response = SurveyResponse.objects.get(survey=survey, user=user)
            has_response = response.is_complete
            submitted_at = response.submitted_at if response.is_complete else None
            progress = response.get_completion_percentage()
            status = 'completed' if response.is_complete else 'started'
        except SurveyResponse.DoesNotExist:
            has_response = False
            submitted_at = None
            progress = 0
            status = 'not_started'
        
        # Get department from employment_info.department.department_name
        department = 'N/A'
        try:
            if hasattr(user, 'employment_info') and user.employment_info and user.employment_info.department:
                department = user.employment_info.department.department_name
        except AttributeError:
            department = 'N/A'
        
        user_data = {
            'user_id': user.id,
            'employee_name': f"{user.firstname} {user.lastname}" if user.firstname and user.lastname else user.username,
            'employee_id': user.idnumber or user.username,
            'department': department,
            'status': status,
            'submitted_at': submitted_at,
            'progress': progress,
            'has_response': has_response
        }
        
        # Apply search filter
        if search_query:
            search_fields = [
                user_data['employee_name'].lower(),
                user_data['employee_id'].lower(),
                user_data.get('department', '').lower()
            ]
            if not any(search_query.lower() in field for field in search_fields):
                continue
        
        users_data.append(user_data)
    
    # Sort: responded users first, then by name
    users_data.sort(key=lambda x: (not x['has_response'], x['employee_name']))
    
    # Paginate results
    page_number = request.GET.get('page', 1)
    paginator = Paginator(users_data, 25)  # 25 users per page
    page_obj = paginator.get_page(page_number)
    
    # Prepare pagination data
    pagination_data = {
        'current_page': page_obj.number,
        'total_pages': paginator.num_pages,
        'total_count': paginator.count,
        'has_previous': page_obj.has_previous(),
        'has_next': page_obj.has_next(),
        'previous_page': page_obj.previous_page_number() if page_obj.has_previous() else None,
        'next_page': page_obj.next_page_number() if page_obj.has_next() else None,
        'start_index': page_obj.start_index(),
        'end_index': page_obj.end_index(),
        'page_range': list(paginator.get_elided_page_range(page_obj.number, on_each_side=2))
    }
    
    return JsonResponse({
        'responses': list(page_obj),
        'pagination': pagination_data
    })

@login_required
def survey_response_detail(request, survey_id, user_id):
    survey = get_object_or_404(Survey, id=survey_id)
    user = get_object_or_404(EmployeeLogin, id=user_id)
    
    try:
        response = SurveyResponse.objects.get(survey=survey, user=user, is_complete=True)
        answers = Answer.objects.filter(response=response).select_related('question').order_by('question__order')
        
        answer_data = []
        for answer in answers:
            answer_info = {
                'question_id': answer.question.id,
                'question_text': answer.question.question_text,
                'question_type': answer.question.question_type,
                'answer': get_formatted_answer(answer),
                'max_value': answer.question.max_value,
                'file_url': answer.file_upload.url if answer.file_upload else None
            }
            answer_data.append(answer_info)
        
        return JsonResponse({
            'user_name': f"{user.firstname} {user.lastname}" if user.firstname and user.lastname else user.username,
            'submitted_at': response.submitted_at,
            'answers': answer_data
        })
        
    except SurveyResponse.DoesNotExist:
        return JsonResponse({
            'error': 'No response found for this user'
        }, status=404)

def get_formatted_answer(answer):
    """Format answer based on question type"""
    if answer.question.question_type in ['single_choice', 'dropdown', 'short_answer', 'paragraph']:
        return answer.text_answer
    elif answer.question.question_type == 'multiple_choice':
        return answer.selected_options
    elif answer.question.question_type == 'rating_scale':
        return answer.rating_value
    elif answer.question.question_type == 'yes_no':
        return answer.boolean_answer
    elif answer.question.question_type == 'date':
        return answer.date_answer.strftime('%Y-%m-%d') if answer.date_answer else None
    elif answer.question.question_type == 'file_upload':
        return answer.file_upload.name if answer.file_upload else None
    else:
        return answer.text_answer

@login_required
def export_survey_report(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            format_type = data.get('format', 'excel')
            include_responses = data.get('include_responses', True)
            include_analysis = data.get('include_analysis', True)
            include_charts = data.get('include_charts', False)
            
            if format_type == 'excel':
                return export_excel_report(survey, include_responses, include_analysis)
            elif format_type == 'csv':
                return export_csv_report(survey, include_responses)
            elif format_type == 'pdf':
                return export_pdf_report(survey, include_responses, include_analysis)
            else:
                return JsonResponse({'error': 'Invalid format'}, status=400)
                
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)

def export_excel_report(survey, include_responses, include_analysis):
    """Export survey data to Excel format"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Survey Report"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Survey info
    ws.append([f"Survey: {survey.title}"])
    ws.append([f"Description: {survey.description}"])
    ws.append([f"Created: {survey.created_at.strftime('%Y-%m-%d')}"])
    ws.append([f"Status: {survey.get_status_display()}"])
    ws.append([])  # Empty row
    
    if include_responses:
        # Response data
        ws.append(["Employee ID", "Employee Name", "Status", "Submitted At", "Progress"])
        
        # Style header row
        for cell in ws[ws.max_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Add response data
        assigned_users = survey.get_assigned_users()
        for user in assigned_users:
            try:
                response = SurveyResponse.objects.get(survey=survey, user=user)
                status = "Completed" if response.is_complete else "Started"
                submitted = response.submitted_at.strftime('%Y-%m-%d %H:%M') if response.submitted_at else "-"
                progress = f"{response.get_completion_percentage()}%"
            except SurveyResponse.DoesNotExist:
                status = "Not Started"
                submitted = "-"
                progress = "0%"
            
            ws.append([
                user.idnumber or user.username,
                f"{user.firstname} {user.lastname}" if user.firstname and user.lastname else user.username,
                status,
                submitted,
                progress
            ])
    
    # Save to response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="survey_{survey.id}_report.xlsx"'
    
    return response

def export_csv_report(survey, include_responses):
    """Export survey data to CSV format"""
    import csv
    import io
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Survey info
    writer.writerow([f"Survey: {survey.title}"])
    writer.writerow([f"Description: {survey.description}"])
    writer.writerow([f"Created: {survey.created_at.strftime('%Y-%m-%d')}"])
    writer.writerow([])  # Empty row
    
    if include_responses:
        # Headers
        writer.writerow(["Employee ID", "Employee Name", "Status", "Submitted At", "Progress"])
        
        # Response data
        assigned_users = survey.get_assigned_users()
        for user in assigned_users:
            try:
                response = SurveyResponse.objects.get(survey=survey, user=user)
                status = "Completed" if response.is_complete else "Started"
                submitted = response.submitted_at.strftime('%Y-%m-%d %H:%M') if response.submitted_at else "-"
                progress = f"{response.get_completion_percentage()}%"
            except SurveyResponse.DoesNotExist:
                status = "Not Started"
                submitted = "-"
                progress = "0%"
            
            writer.writerow([
                user.idnumber or user.username,
                f"{user.firstname} {user.lastname}" if user.firstname and user.lastname else user.username,
                status,
                submitted,
                progress
            ])
    
    output.seek(0)
    response = HttpResponse(output.getvalue(), content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="survey_{survey.id}_report.csv"'
    
    return response

def export_pdf_report(survey, include_responses, include_analysis):
    """Export survey data to PDF format"""
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    import io
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=1  # Center
    )
    story.append(Paragraph(f"Survey Report: {survey.title}", title_style))
    story.append(Spacer(1, 12))
    
    # Survey details
    story.append(Paragraph(f"<b>Description:</b> {survey.description}", styles['Normal']))
    story.append(Paragraph(f"<b>Created:</b> {survey.created_at.strftime('%Y-%m-%d')}", styles['Normal']))
    story.append(Paragraph(f"<b>Status:</b> {survey.get_status_display()}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    if include_responses:
        # Response summary
        assigned_users = survey.get_assigned_users()
        total_assigned = assigned_users.count()
        completed = SurveyResponse.objects.filter(survey=survey, is_complete=True).count()
        
        story.append(Paragraph("<b>Response Summary</b>", styles['Heading2']))
        story.append(Paragraph(f"Total Assigned: {total_assigned}", styles['Normal']))
        story.append(Paragraph(f"Completed: {completed}", styles['Normal']))
        story.append(Paragraph(f"Completion Rate: {(completed/total_assigned*100):.1f}%" if total_assigned > 0 else "0%", styles['Normal']))
    
    doc.build(story)
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="survey_{survey.id}_report.pdf"'
    
    return response


@login_required
def survey_preview(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    
    # Check if user has access to this survey
    if not (survey.visibility == 'all' or request.user in survey.selected_users.all()):
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    questions = survey.questions.order_by('order')
    
    # Prepare questions data
    questions_data = []
    for question in questions:
        question_data = {
            'id': question.id,
            'text': question.question_text,
            'type': question.question_type,
            'required': question.required,
            'options': question.options if question.options else None
        }
        questions_data.append(question_data)
    
    # Calculate estimated time (rough estimate: 1 minute per question)
    estimated_time = max(len(questions_data), 1)
    
    data = {
        'id': survey.id,
        'title': survey.title,
        'description': survey.description,
        'questions': questions_data,
        'questions_count': len(questions_data),
        'estimated_time': estimated_time,
        'anonymous': survey.anonymous_responses,
        'deadline': survey.deadline.isoformat() if survey.deadline else None
    }
    
    return JsonResponse(data)

@login_required
def survey_details(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    
    # Check if user has access to this survey
    if not (survey.visibility == 'all' or request.user in survey.selected_users.all()):
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    # Get user's progress
    try:
        response = SurveyResponse.objects.get(survey=survey, user=request.user)
        progress = response.get_completion_percentage()
        can_take = not response.is_complete
    except SurveyResponse.DoesNotExist:
        progress = 0
        can_take = True
    
    # Check if survey is still available
    is_expired = survey.deadline and timezone.now() > survey.deadline
    if is_expired:
        can_take = False
    
    # Get question types breakdown
    questions = survey.questions.all()
    question_types = {}
    for question in questions:
        q_type = question.question_type
        question_types[q_type] = question_types.get(q_type, 0) + 1
    
    data = {
        'id': survey.id,
        'title': survey.title,
        'description': survey.description,
        'created_by': f"{survey.created_by.firstname} {survey.created_by.lastname}" if survey.created_by.firstname else survey.created_by.username,
        'created_at': survey.created_at.isoformat(),
        'status': survey.status,
        'status_display': survey.get_status_display(),
        'deadline': survey.deadline.isoformat() if survey.deadline else None,
        'anonymous': survey.anonymous_responses,
        'progress': progress,
        'can_take': can_take,
        'questions_count': questions.count(),
        'question_types': question_types
    }
    
    return JsonResponse(data)

@login_required
def user_survey_response(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    
    try:
        response = SurveyResponse.objects.get(
            survey=survey, 
            user=request.user, 
            is_complete=True
        )
        
        print(f"DEBUG: Found response ID {response.id} for user {request.user.id} and survey {survey_id}")
        
        answers = Answer.objects.filter(response=response).select_related('question').order_by('question__order')
        print(f"DEBUG: Found {answers.count()} answers for response {response.id}")
        
        # Let's also check all answers for this user and survey
        all_answers = Answer.objects.filter(response__survey=survey, response__user=request.user)
        print(f"DEBUG: Total answers for this user/survey: {all_answers.count()}")
        
        # Check if there are any incomplete responses with answers
        incomplete_responses = SurveyResponse.objects.filter(survey=survey, user=request.user, is_complete=False)
        for inc_resp in incomplete_responses:
            inc_answers = Answer.objects.filter(response=inc_resp)
            print(f"DEBUG: Incomplete response {inc_resp.id} has {inc_answers.count()} answers")
        
        # Calculate completion time
        if response.submitted_at and response.started_at:
            completion_time_delta = response.submitted_at - response.started_at
            completion_minutes = int(completion_time_delta.total_seconds() / 60)
            completion_time = f"{completion_minutes} minutes"
        else:
            completion_time = "Unknown"
        
        # Get all questions for the survey (for editing form)
        questions = survey.questions.order_by('order')
        questions_data = []
        for question in questions:
            questions_data.append({
                'id': question.id,
                'text': question.question_text,
                'type': question.question_type,
                'required': question.required,
                'description': question.description,
                'options': question.options,
                'min_value': question.min_value,
                'max_value': question.max_value
            })
        
        # Format answers
        answers_data = []
        for answer in answers:
            answer_data = {
                'question_id': answer.question.id,
                'question_text': answer.question.question_text,
                'question_type': answer.question.question_type,
                'answer': get_formatted_answer_value(answer),
                'min_value': answer.question.min_value,
                'max_value': answer.question.max_value,
                'file_url': answer.file_upload.url if answer.file_upload else None
            }
            answers_data.append(answer_data)
        
        data = {
            'survey_id': survey.id,
            'survey_title': survey.title,
            'submitted_at': response.submitted_at.isoformat(),
            'completion_time': completion_time,
            'questions': questions_data,
            'answers': answers_data
        }
        
        return JsonResponse(data)
        
    except SurveyResponse.DoesNotExist:
        return JsonResponse({'error': 'No completed response found'}, status=404)

def get_formatted_answer_value(answer):
    """Format answer value based on question type"""
    if answer.question.question_type in ['single_choice', 'dropdown', 'short_answer', 'paragraph']:
        return answer.text_answer
    elif answer.question.question_type == 'multiple_choice':
        return answer.selected_options if answer.selected_options else []
    elif answer.question.question_type == 'rating_scale':
        return answer.rating_value
    elif answer.question.question_type == 'yes_no':
        return answer.boolean_answer
    elif answer.question.question_type == 'date':
        return answer.date_answer.strftime('%Y-%m-%d') if answer.date_answer else None
    elif answer.question.question_type == 'file_upload':
        return answer.file_upload.name if answer.file_upload else None
    else:
        return answer.text_answer
