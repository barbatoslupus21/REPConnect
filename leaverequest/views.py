import json
import io
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q, F, Exists, OuterRef, Subquery, Count, Q
from django.utils import timezone
from django.views.decorators.http import require_POST
from userlogin.models import EmployeeLogin
from .models import LeaveRequest, LeaveType, LeaveBalance, LeaveApprovalAction, LeaveReason, SundayException
from .forms import LeaveRequestForm, LeaveApprovalForm, LeaveSearchForm
from django.db.models import Case, When, IntegerField, Max
from datetime import date, datetime, timedelta
from collections import defaultdict
from generalsettings.models import Department, Line, Position
import calendar
from decimal import Decimal
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference
from notification.models import Notification

@login_required(login_url="user-login")
def leave_dashboard(request):
    user = request.user

    recent_leaves = LeaveRequest.objects.filter(employee=user).order_by('-date_prepared')[:4]
    leave_balances = LeaveBalance.objects.filter(employee=user).select_related('leave_type').order_by('valid_from', 'leave_type__name')
    leave_types = LeaveType.objects.filter(is_active=True).order_by('name')
    leave_reasons = LeaveReason.objects.filter(is_active=True).select_related('leave_type').order_by('leave_type__name', 'reason_text')

    from collections import defaultdict
    balance_sets = defaultdict(list)
    for balance in leave_balances:
        key = (balance.valid_from, balance.valid_to)
        balance_sets[key].append(balance)

    grouped_balances = []
    for k, v in balance_sets.items():
        total_entitled = sum(balance.entitled for balance in v)
        total_used = sum(balance.used for balance in v)
        total_remaining = sum(balance.remaining for balance in v)
        
        grouped_balances.append({
            'valid_from': k[0],
            'valid_to': k[1],
            'balances': v,
            'total_entitled': total_entitled,
            'total_used': total_used,
            'total_remaining': total_remaining
        })
    
    grouped_balances.sort(key=lambda x: x['valid_from'], reverse=True)
    
    # Handle pending approvals with pagination and search
    for_leave_approval = (
        LeaveApprovalAction.objects
        .filter(approver=user)
        .annotate(routing_priority=Case(When(status='routing', then=0), default=1, output_field=IntegerField()))
        .exclude(Q(status__iexact='cancelled') | Q(comments='Updated my leave request'))
        .select_related('leave_request__employee', 'leave_request__leave_type', 'leave_request__leave_reason')
        .order_by('routing_priority', '-created_at')
    )

    # Search functionality for approvals
    search_query = request.GET.get('search', '').strip()
    if search_query:
        for_leave_approval = for_leave_approval.filter(
            Q(leave_request__control_number__icontains=search_query) |
            Q(leave_request__employee__firstname__icontains=search_query) |
            Q(leave_request__employee__lastname__icontains=search_query) |
            Q(leave_request__employee__idnumber__icontains=search_query) |
            Q(leave_request__leave_type__name__icontains=search_query) |
            Q(leave_request__leave_reason__reason_text__icontains=search_query)
        )

    # Pagination for pending approvals
    approvals_page = request.GET.get('approvals_page', 1)
    approvals_paginator = Paginator(for_leave_approval, 10)
    pending_approvals_page_obj = approvals_paginator.get_page(approvals_page)

    pending_routing_count = LeaveApprovalAction.objects.filter(approver=user, status='routing').count()

    # Pagination for my requests
    all_requests = LeaveRequest.objects.filter(employee=user).order_by('-date_prepared')
    requests_page = request.GET.get('page', 1)
    requests_paginator = Paginator(all_requests, 10)
    my_requests_page_obj = requests_paginator.get_page(requests_page)

    # compute can_cancel flag for recent_leaves and for items in my_requests_page_obj
    now = timezone.now()
    def _can_cancel(leave):
        try:
            if leave.status == 'routing':
                return True
            if leave.status == 'approved' and leave.updated_at:
                # allow cancel if updated within the last 1 minute
                return (now - leave.updated_at).total_seconds() <= 172800
        except Exception:
            return False
        return False

    # annotate recent_leaves list
    annotated_recent = []
    for l in recent_leaves:
        l.can_cancel = _can_cancel(l)
        annotated_recent.append(l)

    # annotate page object leaves
    for l in my_requests_page_obj:
        try:
            l.can_cancel = _can_cancel(l)
        except Exception:
            l.can_cancel = False

    context={
        'leave_types': leave_types,
        'leave_reasons': leave_reasons,
        'leave_balance_sets': grouped_balances,
        'pending_approvals': pending_approvals_page_obj,
        'pending_routing_count': pending_routing_count,
        'is_approver': for_leave_approval.exists(),
        'recent_leaves': annotated_recent,
        'my_requests_page_obj': my_requests_page_obj,
        'search_query': search_query,
        'today': timezone.now().date(),
    }

    return render(request, 'leaverequest/user-leave.html', context)

@login_required(login_url="user-login")
def leave_requests_list(request):
    """List all leave requests for the current user"""
    search_form = LeaveSearchForm(request.GET)
    leave_requests = LeaveRequest.objects.filter(employee=request.user)
    
    # Apply filters
    if search_form.is_valid():
        if search_form.cleaned_data['search']:
            search_term = search_form.cleaned_data['search']
            leave_requests = leave_requests.filter(
                Q(control_number__icontains=search_term) |
                Q(reason__icontains=search_term)
            )
        
        if search_form.cleaned_data['status']:
            leave_requests = leave_requests.filter(status=search_form.cleaned_data['status'])
        
        if search_form.cleaned_data['leave_type']:
            leave_requests = leave_requests.filter(leave_type=search_form.cleaned_data['leave_type'])
        
        if search_form.cleaned_data['date_from']:
            leave_requests = leave_requests.filter(date_from__gte=search_form.cleaned_data['date_from'])
        
        if search_form.cleaned_data['date_to']:
            leave_requests = leave_requests.filter(date_to__lte=search_form.cleaned_data['date_to'])
    
    # Pagination
    paginator = Paginator(leave_requests.order_by('-date_prepared'), 10)
    page = request.GET.get('page')
    leave_requests = paginator.get_page(page)
    
    context = {
        'leave_requests': leave_requests,
        'search_form': search_form,
        'page_title': 'My Leave Requests'
    }
    
    return render(request, 'leaverequest/requests_list.html', context)

@login_required(login_url="user-login")
def apply_leave(request):
    if request.method == 'POST':
        form = LeaveRequestForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            leave_request = form.save(commit=False)
            leave_request.employee = request.user
            
            approver = None
            approval_comments = ""
            
            notice_days = (leave_request.date_from - timezone.now().date()).days
            
            # Rule 1: Urgent leave (< 1 day notice, non-clinic) → IAD Admin
            if not leave_request.leave_type.go_to_clinic and notice_days < 1:
                iad_admin = EmployeeLogin.objects.filter(iad_admin=True, is_active=True).first()
                if iad_admin:
                    approver = iad_admin
                else:
                    hr_manager = EmployeeLogin.objects.filter(hr_manager=True, is_active=True).first()
                    if hr_manager:
                        approver = hr_manager
                    else:
                        hr_admin = EmployeeLogin.objects.filter(hr_admin=True, is_active=True).first()
                        approver = hr_admin
            
            # Rule 2: Regular leave (≥ 1 day notice, non-clinic) → Employee's approver
            elif not leave_request.leave_type.go_to_clinic and notice_days >= 1:
                try:
                    employment_info = request.user.employment_info
                    if employment_info.approver and employment_info.approver.is_active:
                        approver = employment_info.approver
                    else:
                        error_msg = 'You do not have an assigned approver. Please assign an approver before submitting leave requests.'
                        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                            return JsonResponse({'success': False, 'message': error_msg})
                        messages.error(request, error_msg)
                        return redirect('user_leave')
                except:
                    error_msg = 'Your employment information is incomplete. Please contact HR to complete your profile before submitting leave requests.'
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({'success': False, 'message': error_msg})
                    messages.error(request, error_msg)
                    return redirect('user_leave')
            
            # Rule 3: Clinic leave (go_to_clinic = True) → Clinic Admin
            elif leave_request.leave_type.go_to_clinic:
                clinic_admin = EmployeeLogin.objects.filter(clinic_admin=True, is_active=True).first()
                if clinic_admin:
                    approver = clinic_admin
                else:
                    hr_manager = EmployeeLogin.objects.filter(hr_manager=True, is_active=True).first()
                    if hr_manager:
                        approver = hr_manager
                    else:
                        hr_admin = EmployeeLogin.objects.filter(hr_admin=True, is_active=True).first()
                        approver = hr_admin
            
            # Final fallback to HR manager or HR admin if no approver found
            if not approver:
                hr_manager = EmployeeLogin.objects.filter(hr_manager=True, is_active=True).first()
                if hr_manager:
                    approver = hr_manager
                else:
                    hr_admin = EmployeeLogin.objects.filter(hr_admin=True, is_active=True).first()
                    approver = hr_admin
            
            # If still no approver found, show error
            if not approver:
                messages.error(request, 'No available approver found. Please assign your supervisor as approver.')
                return redirect('user_leave')
            
            leave_request.current_approver = approver
            leave_request.save()
            
            LeaveApprovalAction.objects.create(
                leave_request=leave_request,
                approver=approver,
                sequence=1,
                status='routing',
                action='submitted',
                comments=approval_comments,
                action_at=timezone.now()
            )

            Notification.objects.create(
                title="Leave Request Submitted",
                message=f"{request.user.firstname} {request.user.lastname} has submitted a leave request for your approval.",
                notification_type="approval",
                sender=request.user,
                recipient=approver,
                for_all=False,
                module="leave"
            )
            
            messages.success(request, f'Leave request {leave_request.control_number} submitted successfully!')
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': f'Leave request {leave_request.control_number} submitted successfully!',
                    'control_number': leave_request.control_number
                })
            
            return redirect('user_leave')
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                error_messages = []
                for field, errors in form.errors.items():
                    if field == '__all__':
                        error_messages.extend(errors)
                    else:
                        field_name = form.fields[field].label or field.replace('_', ' ').title()
                        for error in errors:
                            error_messages.append(f"{field_name}: {error}")
                
                return JsonResponse({
                    'success': False,
                    'message': 'Please correct the following errors:',
                    'errors': form.errors,
                    'error_list': error_messages
                })
    else:
        form = LeaveRequestForm(user=request.user)
    
    context = {
        'form': form,
        'page_title': 'Apply for Leave'
    }
    
    return redirect('user_leave')

@login_required(login_url="user-login")
def leave_detail(request, control_number):
    leave_request = get_object_or_404(LeaveRequest, control_number=control_number)
    
    # Check permissions
    # if (leave_request.employee != request.user and 
    #     leave_request.current_approver != request.user and
    #     not (request.user.hr_admin or request.user.hr_manager or 
    #          request.user.is_superuser or request.user.is_staff)):
    #     return JsonResponse({'success': False, 'message': 'You do not have permission to view this leave request.'})

    approval_actions = leave_request.approval_actions.all().order_by('action_at')
    
    from .models import LeaveReason
    leave_reasons = LeaveReason.objects.filter(leave_type=leave_request.leave_type, is_active=True)
    
    context = {
        'leave_request': leave_request,
        'approval_actions': approval_actions,
        'leave_reasons': leave_reasons,
        'can_edit': (leave_request.employee == request.user and 
                    leave_request.status == 'routing'),
        'can_cancel': (leave_request.employee == request.user and 
                      leave_request.status in ['routing']),
        'can_approve': (leave_request.current_approver == request.user and 
                       leave_request.status == 'routing')
    }
    
    # Return detail content and approval permission flag
    # Indicate if the leave requestor has an assigned approver in their profile
    has_approver = bool(getattr(leave_request.employee, 'employment_info', None) and 
                        leave_request.employee.employment_info.approver)
    return JsonResponse({
        'success': True,
        'html': render(request, 'leaverequest/detail_content.html', context).content.decode('utf-8'),
        'can_approve': context.get('can_approve', False),
        'has_approver': has_approver
    })

@login_required(login_url="user-login")
@login_required(login_url="user-login")
def edit_leave(request, control_number):
    leave_request = get_object_or_404(LeaveRequest, control_number=control_number)

    if leave_request.employee != request.user or leave_request.status != 'routing':
        return JsonResponse({'success': False, 'message': 'You cannot edit this leave request.'})
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method.'})

    from datetime import datetime
    date_from_str = request.POST.get('date_from')
    date_to_str = request.POST.get('date_to')
    try:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
    except Exception:
        return JsonResponse({'success': False, 'message': 'Invalid date format.'})
    reason_text = request.POST.get('reason', '').strip()
    leave_reason_id = request.POST.get('leave_reason')

    if not (date_from and date_to and leave_reason_id and reason_text):
        return JsonResponse({'success': False, 'message': 'All fields are required.'})

    from .models import LeaveReason
    try:
        leave_request.date_from = date_from
        leave_request.date_to = date_to
        leave_request.reason = reason_text
        leave_request.leave_reason = LeaveReason.objects.get(pk=leave_reason_id)
        leave_request.save()

        LeaveApprovalAction.objects.create(
            leave_request=leave_request,
            approver=request.user,
            sequence=leave_request.approval_actions.count() + 1,
            status='approved',
            action='approved',
            comments='Updated my leave request',
            action_at=timezone.now()
        )
        return JsonResponse({'success': True, 'message': 'Leave request updated successfully!'})
    except LeaveReason.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Invalid leave reason selected.'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

def restore_leave_balances(leave_request):
    """
    Restore leave balances when canceling an approved leave request.
    Split the leave request across matching leave balance periods and restore balances.
    """
    from datetime import timedelta
    
    # Get all leave balances for this employee and leave type
    leave_balances = LeaveBalance.objects.filter(
        employee=leave_request.employee,
        leave_type=leave_request.leave_type
    ).order_by('valid_from')
    
    current_date = leave_request.date_from
    
    while current_date <= leave_request.date_to:
        # Find the leave balance that covers this date
        matching_balance = None
        for balance in leave_balances:
            if balance.valid_from <= current_date <= balance.valid_to:
                matching_balance = balance
                break
        
        if not matching_balance:
            current_date += timedelta(days=1)
            continue
        
        # Find the end date for this balance period
        period_end = min(leave_request.date_to, matching_balance.valid_to)
        
        # Calculate hours for this period
        period_hours = 0
        temp_date = current_date
        
        # Get holidays and sunday exceptions for the period
        from usercalendar.models import Holiday
        holidays = set(Holiday.objects.filter(
            date__range=[current_date, period_end]
        ).values_list('date', flat=True))
        
        sunday_exceptions = set(SundayException.objects.filter(
            date__range=[current_date, period_end]
        ).values_list('date', flat=True))
        
        while temp_date <= period_end:
            is_holiday = temp_date in holidays
            is_sunday = temp_date.weekday() == 6
            is_sunday_exempted = temp_date in sunday_exceptions
            
            # Count as working day if not holiday and (not sunday or sunday exempted)
            if not is_holiday and (not is_sunday or is_sunday_exempted):
                period_hours += 8  # 8 hours per working day
            
            temp_date += timedelta(days=1)
        
        # Convert hours to days (8 hours = 1 day)
        days_to_restore = period_hours / 8
        
        if days_to_restore > 0:
            # Restore the balance
            from decimal import Decimal
            days_decimal = Decimal(str(days_to_restore))
            
            # First, calculate what the new used value would be
            new_used = matching_balance.used - days_decimal
            
            # If new_used would be negative, we can only restore up to the current used amount
            if new_used < 0:
                # Restore only what was actually used
                actual_days_to_restore = matching_balance.used
                matching_balance.remaining += actual_days_to_restore
                matching_balance.used = Decimal('0.00')
            else:
                # Normal case: restore the full amount
                matching_balance.remaining += days_decimal
                matching_balance.used = new_used
            
            matching_balance.save()
        
        # Move to the next period
        current_date = period_end + timedelta(days=1)

@login_required(login_url="user-login")
@require_POST
def cancel_leave(request, control_number):
    leave_request = get_object_or_404(LeaveRequest, control_number=control_number)
    
    if (leave_request.employee != request.user or 
        leave_request.status not in ['routing', 'approved']):
        return JsonResponse({'success': False, 'message': 'Cannot cancel this leave request.'})
    
    # Store original status to determine if balance restoration is needed
    original_status = leave_request.status
    
    # If the leave was approved, restore leave balances
    if original_status == 'approved':
        try:
            restore_leave_balances(leave_request)
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Failed to restore leave balances: {str(e)}'})
    
    leave_request.status = 'cancelled'
    leave_request.cancelled_at = timezone.now()
    leave_request.current_approver = None
    leave_request.save()

    # Only delete approval actions if it was routing
    if original_status == 'routing':
        LeaveApprovalAction.objects.filter(leave_request=leave_request).delete()

    LeaveApprovalAction.objects.create(
        leave_request=leave_request,
        approver=request.user,
        sequence=leave_request.approval_actions.count() + 1,
        status=leave_request.status,
        action='cancelled',
        comments="Leave request cancelled by employee",
        action_at=timezone.now()
    )
    
    return JsonResponse({'success': True, 'message': 'Leave request cancelled successfully.'})

@login_required(login_url="user-login")
def approver_dashboard(request):
    pending_approvals = LeaveRequest.objects.filter(
        current_approver=request.user,
        status='routing'
    ).order_by('-date_prepared')

    search_form = LeaveSearchForm(request.GET)
    if search_form.is_valid():
        if search_form.cleaned_data['search']:
            search_term = search_form.cleaned_data['search']
            pending_approvals = pending_approvals.filter(
                Q(control_number__icontains=search_term) |
                Q(employee__firstname__icontains=search_term) |
                Q(employee__lastname__icontains=search_term)
            )
    
    paginator = Paginator(pending_approvals, 10)
    page = request.GET.get('page')
    pending_approvals = paginator.get_page(page)
    
    context = {
        'pending_approvals': pending_approvals,
        'search_form': search_form,
        'page_title': 'Pending Approvals'
    }
    
    return render(request, 'leaverequest/approver_dashboard.html', context)

def get_next_approver(leave_request, current_approver):
    try:
        # Rule 1: If current approver is Clinic Admin → next is IAD Admin
        if current_approver.clinic_admin:
            iad_admin = EmployeeLogin.objects.filter(iad_admin=True, active=True).first()
            return iad_admin
        
        # Rule 2: If current approver is IAD Admin → next is requestor's direct supervisor
        if current_approver.iad_admin:
            if (hasattr(leave_request.employee, 'employment_info') and 
                leave_request.employee.employment_info.approver):
                return leave_request.employee.employment_info.approver
        
        # Get position levels for rules 4 & 5
        requestor_level = None
        current_approver_level = None
        
        if (hasattr(leave_request.employee, 'employment_info') and 
            leave_request.employee.employment_info.position):
            requestor_level = int(leave_request.employee.employment_info.position.level)
        
        if (hasattr(current_approver, 'employment_info') and 
            current_approver.employment_info.position):
            current_approver_level = int(current_approver.employment_info.position.level)
        
        # Rule 4: If requestor's level ≠ 2 or 3, and current approver level = 2 → HR Admin
        if (requestor_level and requestor_level not in [2, 3] and 
            current_approver_level == 3):
            hr_admin = EmployeeLogin.objects.filter(hr_admin=True, active=True).first()
            return hr_admin
        
        # Rule 5: If requestor's level = 2, and current approver level = 3 → HR Admin  
        if (requestor_level in [2, 3] and current_approver_level == 3):
            hr_admin = EmployeeLogin.objects.filter(hr_admin=True, active=True).first()
            return hr_admin
        
        # Rule 3: For other cases → current user's direct supervisor
        if (hasattr(current_approver, 'employment_info') and 
            current_approver.employment_info.approver):
            return current_approver.employment_info.approver
        
        # No next approver found
        return None
        
    except Exception as e:
        print(f"Error determining next approver: {e}")
        return None

@login_required(login_url="user-login")
@require_POST
def process_approval(request, control_number):
    approval_action = get_object_or_404(LeaveApprovalAction, leave_request__control_number=control_number, approver=request.user)
    leave_request_obj = approval_action.leave_request

    last_sequence = LeaveApprovalAction.objects.filter(
        leave_request__control_number=control_number
    ).aggregate(Max('sequence'))['sequence__max'] or 0
    next_sequence = last_sequence + 1

    try:
        data = json.loads(request.body)
        action = data.get('action')
        comments = data.get('comments', '')
        next_approver = get_next_approver(leave_request_obj, request.user)

        if action == 'approve' and leave_request_obj.status == 'disapproved':
            leave_status = 'disapproved'
        elif action == 'approve' and leave_request_obj.status == 'routing':
            leave_status = 'routing'
        else:
            leave_status = 'approved'

        if action == 'approve':
            approval_action.status = 'approved'
            approval_action.action = 'approved'
            approval_action.action_at = timezone.now()
            approval_action.comments = comments
            approval_action.save()

            if next_approver:
                leave_request_obj.current_approver = next_approver
                leave_request_obj.status = leave_status
                message = 'Leave request approved!'

                LeaveApprovalAction.objects.create(
                    leave_request=leave_request_obj,
                    approver=next_approver,
                    sequence=next_sequence,
                    status='routing',
                    action='submitted',
                    comments='',
                    action_at=timezone.now()
                )

                Notification.objects.create(
                    title="Leave Approval",
                    message=f"{leave_request_obj.employee.firstname} {leave_request_obj.employee.lastname} leave request is waiting for your approval.",
                    notification_type="approval",
                    sender=request.user,
                    recipient=next_approver,
                    for_all=False,
                    module="leave"
                )
                
            else:
                leave_request_obj.current_approver = None
                leave_request_obj.status = leave_status
                message = 'Leave request fully approved!'

        elif action == 'disapprove':
            approval_action.status = 'disapproved'
            approval_action.action = 'disapproved'
            approval_action.action_at = timezone.now()
            approval_action.comments = comments
            approval_action.save()

            if next_approver:
                leave_request_obj.current_approver = next_approver
                leave_request_obj.status = 'disapproved'
                message = 'Leave request disapproved.'

                LeaveApprovalAction.objects.create(
                    leave_request=leave_request_obj,
                    approver=next_approver,
                    sequence=next_sequence,
                    status='routing',
                    action='submitted',
                    comments='',
                    action_at=timezone.now()
                )

                start_date_fmt = leave_request_obj.date_from.strftime("%b %d, %Y")
                end_date_fmt = leave_request_obj.date_to.strftime("%b %d, %Y")

                Notification.objects.create(
                    title="Leave Disapproved",
                    message=f"Your leave request for {start_date_fmt} to {end_date_fmt} has been disapproved.",
                    notification_type="approval",
                    sender=request.user,
                    recipient=leave_request_obj.employee,
                    for_all=False,
                    module="leave"
                )

                Notification.objects.create(
                    title="Leave Disapproved",
                    message=f"{request.user.firstname} {request.user.lastname} has disapproved the leave request submitted by {leave_request_obj.employee.firstname} {leave_request_obj.employee.lastname}. Your approval is required next.",
                    notification_type="approval",
                    sender=request.user,
                    recipient=next_approver,
                    for_all=False,
                    module="leave"
                )
                
        else:
            return JsonResponse({'success': False, 'message': 'Invalid action.'})

        leave_request_obj.updated_at = timezone.now()
        leave_request_obj.save()

        return JsonResponse({'success': True, 'message': message})

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required(login_url="user-login")
def admin_dashboard(request):
    if not (request.user.hr_admin or request.user.hr_manager or 
            request.user.is_superuser or request.user.is_staff):
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('admin_leave')
    
    # Get all leave requests with related data
    leave_requests = LeaveRequest.objects.all().select_related(
        'employee', 'leave_type', 'leave_reason', 'employee__employment_info__department'
    ).prefetch_related('approval_actions')
    
    # Handle search (applied across all records, not just current page)
    search_query = request.GET.get('search', '').strip()
    if search_query:
        leave_requests = leave_requests.filter(
            Q(control_number__icontains=search_query) |
            Q(employee__firstname__icontains=search_query) |
            Q(employee__lastname__icontains=search_query) |
            Q(employee__username__icontains=search_query) |
            Q(employee__idnumber__icontains=search_query)
        )
    
    # Handle filters (only applied when explicitly submitted via modal)
    filter_status = request.GET.get('status', '')
    filter_leave_type = request.GET.get('leave_type', '')
    filter_department = request.GET.get('department', '')
    filter_date_from = request.GET.get('date_from', '')
    filter_date_to = request.GET.get('date_to', '')
    
    # Apply filters if they exist
    if filter_status:
        leave_requests = leave_requests.filter(status=filter_status)
    
    if filter_leave_type:
        leave_requests = leave_requests.filter(leave_type_id=filter_leave_type)
    
    if filter_department:
        leave_requests = leave_requests.filter(
            employee__employment_info__department__department_name=filter_department
        )
    
    if filter_date_from:
        leave_requests = leave_requests.filter(date_from__gte=filter_date_from)
    
    if filter_date_to:
        leave_requests = leave_requests.filter(date_to__lte=filter_date_to)

    approver_exists = LeaveApprovalAction.objects.filter(
        leave_request=OuterRef('pk'),
        approver=request.user, status='routing'
    )
    leave_requests = leave_requests.annotate(
        user_is_approver=Exists(approver_exists)
    ).order_by('-user_is_approver', '-date_prepared')

    # --- Dashboard stats: current month vs previous month ---
    today = timezone.now().date()
    # Current month range
    current_start = today.replace(day=1)
    if today.month == 12:
        next_month_first = today.replace(year=today.year + 1, month=1, day=1)
    else:
        next_month_first = today.replace(month=today.month + 1, day=1)
    current_end = next_month_first - timedelta(days=1)
    # Previous month range
    if current_start.month == 1:
        prev_start = current_start.replace(year=current_start.year - 1, month=12, day=1)
    else:
        prev_start = current_start.replace(month=current_start.month - 1, day=1)
    prev_end = current_start - timedelta(days=1)

    def count_in_range(qs, start, end, status=None):
        base = LeaveRequest.objects.filter(date_prepared__date__gte=start, date_prepared__date__lte=end)
        if status:
            base = base.filter(status=status)
        return base.count()

    def pct_change(curr, prev):
        if prev == 0:
            if curr == 0:
                return 0.0
            return 100.0
        return round(((curr - prev) / prev) * 100.0, 1)

    total_curr = count_in_range(LeaveRequest.objects.all(), current_start, current_end)
    total_prev = count_in_range(LeaveRequest.objects.all(), prev_start, prev_end)
    total_pct = pct_change(total_curr, total_prev)

    routing_curr = count_in_range(LeaveRequest.objects.all(), current_start, current_end, status='routing')
    routing_prev = count_in_range(LeaveRequest.objects.all(), prev_start, prev_end, status='routing')
    routing_pct = pct_change(routing_curr, routing_prev)

    approved_curr = count_in_range(LeaveRequest.objects.all(), current_start, current_end, status='approved')
    approved_prev = count_in_range(LeaveRequest.objects.all(), prev_start, prev_end, status='approved')
    approved_pct = pct_change(approved_curr, approved_prev)

    disapproved_curr = count_in_range(LeaveRequest.objects.all(), current_start, current_end, status='disapproved')
    disapproved_prev = count_in_range(LeaveRequest.objects.all(), prev_start, prev_end, status='disapproved')
    disapproved_pct = pct_change(disapproved_curr, disapproved_prev)

    # Pagination for table
    paginator = Paginator(leave_requests, 15)
    page = request.GET.get('page')
    leave_requests_page = paginator.get_page(page)
    
    # Data for filters
    leave_types = LeaveType.objects.filter(is_active=True).order_by('name')
    departments = Department.objects.all().order_by('department_name')
    
    stats = {
        'total_requests': LeaveRequest.objects.count(),
        'pending_requests': LeaveRequest.objects.filter(status='routing').count(),
        'approved_requests': LeaveRequest.objects.filter(status='approved').count(),
        'disapproved_requests': LeaveRequest.objects.filter(status='disapproved').count(),
    }
    
    # Leave Balances with search functionality
    balance_search_query = request.GET.get('balance_search', '').strip()
    leave_balances = LeaveBalance.objects.select_related('employee', 'leave_type', 'employee__employment_info__department').all()
    
    if balance_search_query:
        leave_balances = leave_balances.filter(
            Q(employee__idnumber__icontains=balance_search_query) |
            Q(employee__firstname__icontains=balance_search_query) |
            Q(employee__lastname__icontains=balance_search_query) |
            Q(leave_type__name__icontains=balance_search_query) |
            Q(employee__employment_info__department__department_name__icontains=balance_search_query)
        )
    
    leave_balances = leave_balances.order_by('-valid_from', '-created_at')
    
    # Pagination for leave balances
    balance_page = request.GET.get('balance_page', 1)
    balance_paginator = Paginator(leave_balances, 10)  # 10 per page as requested
    leave_balances_page = balance_paginator.get_page(balance_page)
    
    # Handle AJAX request for balance search FIRST (before general AJAX handler)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' and 'balance_search' in request.GET:
        from django.template.loader import render_to_string
        balance_table_html = render_to_string('leaverequest/partials/balance_table.html', {
            'leave_balances': leave_balances_page,
            'user': request.user
        }, request=request)
        return JsonResponse({
            'success': True,
            'html': balance_table_html,
            'count': leave_balances.count(),
            'has_next': leave_balances_page.has_next(),
            'has_previous': leave_balances_page.has_previous(),
            'page_number': leave_balances_page.number,
            'total_pages': balance_paginator.num_pages
        })
    
    # Handle general AJAX requests (for leave requests search)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        from django.template.loader import render_to_string
        table_html = render_to_string('leaverequest/partials/leave_table.html', {
            'leave_requests': leave_requests_page,
            'user': request.user
        }, request=request)
        return JsonResponse({
            'success': True,
            'html': table_html,
            'has_next': leave_requests_page.has_next(),
            'has_previous': leave_requests_page.has_previous(),
            'page_number': leave_requests_page.number,
            'total_pages': paginator.num_pages
        })
    
    context = {
        'leave_requests': leave_requests_page,
        'leave_balances': leave_balances_page,
        'balance_search': balance_search_query,
        'stats': stats,
        'page_title': 'Leave Management - Admin',
        # Cards: current month numbers and percent change vs prev. month
        'total_requests': total_curr,
        'total_percent': total_pct,
        'total_positive': total_pct >= 0,
        'routing_requests': routing_curr,
        'routing_percent': routing_pct,
        'routing_positive': routing_pct >= 0,
        'approved_requests': approved_curr,
        'approved_percent': approved_pct,
        'approved_positive': approved_pct >= 0,
        'disapproved_requests': disapproved_curr,
        'disapproved_percent': disapproved_pct,
        'disapproved_positive': disapproved_pct >= 0,
        # Filters
        'leave_types': leave_types,
        'departments': departments,
        # Search and filter values to maintain state
        'search': search_query,
        'filter_status': filter_status,
        'filter_leave_type': filter_leave_type,
        'filter_department': filter_department,
        'filter_date_from': filter_date_from,
        'filter_date_to': filter_date_to,
        'departments': departments,
        # Search and filter values to maintain state
        'search': search_query,
        'filter_status': filter_status,
        'filter_leave_type': filter_leave_type,
        'filter_department': filter_department,
        'filter_date_from': filter_date_from,
        'filter_date_to': filter_date_to,
    }
    
    return render(request, 'leaverequest/admin-leave.html', context)

@login_required(login_url="user-login")
def admin_chart_data(request):
    """AJAX endpoint to get chart data for admin dashboard"""
    if request.method != 'GET':
        return JsonResponse({'error': 'Invalid request method'})
    
    # Only staff and superuser can access chart data
    if not (request.user.is_superuser or request.user.is_staff):
        return JsonResponse({'error': 'Permission denied'})
    
    try:
        import calendar
        from datetime import datetime, timedelta
        from django.db.models import Count, Q
        from collections import defaultdict
        
        now = timezone.now()
        current_date = now.date()
        period = request.GET.get('period', 'month')
        
        # Determine date range based on period
        if period == 'month':
            start_date = current_date.replace(day=1)
            if current_date.month == 12:
                end_date = current_date.replace(year=current_date.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                end_date = current_date.replace(month=current_date.month + 1, day=1) - timedelta(days=1)
            period_label = f"{current_date.strftime('%B %Y')}"
        elif period == 'quarter':
            quarter = ((current_date.month - 1) // 3) + 1
            start_month = (quarter - 1) * 3 + 1
            start_date = current_date.replace(month=start_month, day=1)
            end_month = quarter * 3
            if end_month == 12:
                end_date = current_date.replace(month=12, day=31)
            else:
                end_date = current_date.replace(month=end_month + 1, day=1) - timedelta(days=1)
            period_label = f"Q{quarter} {current_date.year}"
        else:  # year
            start_date = current_date.replace(month=1, day=1)
            end_date = current_date.replace(month=12, day=31)
            period_label = f"{current_date.year}"
        
        # Get all leave requests in the period
        leave_requests = LeaveRequest.objects.filter(
            date_prepared__date__gte=start_date,
            date_prepared__date__lte=end_date
        ).select_related('leave_type')
        
        # Prepare time series data
        time_data = defaultdict(lambda: defaultdict(int))
        time_labels = []
        
        if period == 'month':
            # Group by days
            current_day = start_date
            while current_day <= end_date:
                time_labels.append(current_day.day)
                current_day += timedelta(days=1)
            
            for request in leave_requests:
                day_key = request.date_prepared.date().day
                status = request.status
                time_data[day_key][status] += 1
                
        elif period == 'quarter':
            # Group by months
            current_month = start_date.replace(day=1)
            while current_month <= end_date:
                time_labels.append(current_month.strftime('%b'))
                month_key = current_month.month
                for request in leave_requests.filter(date_prepared__month=month_key):
                    status = request.status
                    time_data[month_key][status] += 1
                if current_month.month == 12:
                    current_month = current_month.replace(year=current_month.year + 1, month=1)
                else:
                    current_month = current_month.replace(month=current_month.month + 1)
        else:
            # Group by months for the year
            for month in range(1, 13):
                time_labels.append(calendar.month_abbr[month])
                for request in leave_requests.filter(date_prepared__month=month):
                    status = request.status
                    time_data[month][status] += 1
        
        # Prepare datasets
        statuses = ['routing', 'approved', 'disapproved', 'cancelled']
        status_colors = {
            'routing': '#f59e0b',
            'approved': '#10b981', 
            'disapproved': '#ef4444',
            'cancelled': '#6b7280'
        }
        
        datasets = []
        for status in statuses:
            if period == 'month':
                data = [time_data[day].get(status, 0) for day in range(1, len(time_labels) + 1)]
            elif period == 'quarter':
                quarter_start = ((current_date.month - 1) // 3) * 3 + 1
                data = [time_data[month].get(status, 0) for month in range(quarter_start, quarter_start + 3)]
            else:
                data = [time_data[month].get(status, 0) for month in range(1, 13)]
            
            color = status_colors[status]
            
            datasets.append({
                'label': status.title(),
                'data': data,
                'borderColor': color,
                'backgroundColor': f"{color}40",  # 25% opacity
                'tension': 0.4,
                'fill': False,
                'pointRadius': 4,
                'pointHoverRadius': 6,
                'pointBackgroundColor': color,
                'pointBorderColor': '#ffffff',
                'pointBorderWidth': 2
            })
        
        return JsonResponse({
            'success': True,
            'period_label': period_label,
            'chart_data': {
                'labels': time_labels,
                'datasets': datasets
            }
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url="user-login")
def get_leave_balance(request):
    """AJAX endpoint to get leave balance for validation"""
    if request.method == 'GET':
        leave_type_id = request.GET.get('leave_type_id')
        
        if not leave_type_id:
            return JsonResponse({'error': 'Leave type ID required'})
        
        # Only for regular employees
        if (hasattr(request.user, 'employment_info') and 
            getattr(request.user.employment_info, 'employment_status', None) == 'Regular'):
            
            today = timezone.now().date()
            balance = LeaveBalance.objects.filter(
                employee=request.user,
                leave_type_id=leave_type_id,
                valid_from__lte=today,
                valid_to__gte=today,
            ).first()
            
            if balance:
                return JsonResponse({
                    'remaining': float(balance.remaining),
                    'entitled': float(balance.entitled),
                    'used': float(balance.used),
                    'valid_period': f"{balance.valid_from.strftime('%b %d, %Y')} - {balance.valid_to.strftime('%b %d, %Y')}"
                })
        
        return JsonResponse({'remaining': None, 'message': 'No active balance found'})
    
    return JsonResponse({'error': 'Invalid request method'})

@login_required(login_url="user-login")
def leave_chart_data(request):
    if request.method != 'GET':
        return JsonResponse({'error': 'Invalid request method'})
    
    try:
        from datetime import datetime, timedelta
        from django.db.models import Count, Q
        from collections import defaultdict
        import calendar
        
        now = timezone.now()
        current_date = now.date()
        
        # Calculate fiscal year (May to April)
        if current_date.month >= 5:  # May onwards = current fiscal year
            fiscal_start = datetime(current_date.year, 5, 1).date()
            fiscal_end = datetime(current_date.year + 1, 4, 30).date()
            fiscal_year = f"{current_date.year}-{current_date.year + 1}"
        else:  # January to April = previous fiscal year
            fiscal_start = datetime(current_date.year - 1, 5, 1).date()
            fiscal_end = datetime(current_date.year, 4, 30).date()
            fiscal_year = f"{current_date.year - 1}-{current_date.year}"
        
        # Status distribution data (for pie chart)
        leave_requests = LeaveRequest.objects.filter(
            employee=request.user,
            date_prepared__date__gte=fiscal_start,
            date_prepared__date__lte=fiscal_end
        )
        
        status_counts = leave_requests.aggregate(
            approved=Count('id', filter=Q(status='approved')),
            routing=Count('id', filter=Q(status='routing')),
            disapproved=Count('id', filter=Q(status='disapproved')),
            cancelled=Count('id', filter=Q(status='cancelled'))
        )
        
        total_requests = sum(status_counts.values())
        
        status_percentages = {}
        if total_requests > 0:
            for status, count in status_counts.items():
                status_percentages[status] = round((count / total_requests) * 100, 1)
        else:
            status_percentages = {'approved': 0, 'routing': 0, 'disapproved': 0, 'cancelled': 0}
        
        leave_types = LeaveType.objects.filter(is_active=True)
        monthly_data = defaultdict(lambda: defaultdict(int))
        months = []
        current_month = fiscal_start.replace(day=1)
        while current_month <= fiscal_end:
            months.append(current_month)
            if current_month.month == 12:
                current_month = current_month.replace(year=current_month.year + 1, month=1)
            else:
                current_month = current_month.replace(month=current_month.month + 1)
        
        for leave_request in leave_requests.filter(status='approved'):
            month_key = leave_request.date_prepared.date().replace(day=1)
            if month_key in months:
                monthly_data[month_key][leave_request.leave_type.name] += 1
        
        line_chart_labels = [month.strftime('%b') for month in months]
        line_chart_datasets = []
        
        colors = ['#6366f1', '#10b981', '#f59e0b', '#ef4444', '#06b6d4', '#8b5cf6']
        
        for i, leave_type in enumerate(leave_types):
            data = [monthly_data[month].get(leave_type.name, 0) for month in months]
            color = colors[i % len(colors)]
            
            line_chart_datasets.append({
                'label': leave_type.name,
                'data': data,
                'borderColor': color,
                'backgroundColor': f"{color}20",
                'tension': 0.4,
                'fill': True,
                'pointRadius': 4,
                'pointHoverRadius': 6,
                'pointBackgroundColor': color,
                'pointBorderColor': '#ffffff',
                'pointBorderWidth': 2
            })
        
        return JsonResponse({
            'success': True,
            'fiscal_year': fiscal_year,
            'status_chart': {
                'labels': ['Approved', 'Routing', 'Disapproved', 'Cancelled'],
                'data': [
                    status_counts['approved'],
                    status_counts['routing'], 
                    status_counts['disapproved'],
                    status_counts['cancelled']
                ],
                'backgroundColor': ['#10b981', '#f59e0b', '#ef4444', '#6b7280'],
                'total_requests': total_requests
            },
            'line_chart': {
                'labels': line_chart_labels,
                'datasets': line_chart_datasets
            }
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url="user-login")
def approval_chart_data(request):
    if request.method != 'GET':
        return JsonResponse({'error': 'Invalid request method'})
    
    try:        
        now = timezone.now()
        current_date = now.date()
        period = request.GET.get('period', 'month')
        
        if period == 'month':
            start_date = current_date.replace(day=1)
            if current_date.month == 12:
                end_date = current_date.replace(year=current_date.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                end_date = current_date.replace(month=current_date.month + 1, day=1) - timedelta(days=1)
            period_label = f"{current_date.strftime('%B %Y')}"
        elif period == 'quarter':
            quarter = ((current_date.month - 1) // 3) + 1
            start_month = (quarter - 1) * 3 + 1
            start_date = current_date.replace(month=start_month, day=1)
            end_month = quarter * 3
            if end_month == 12:
                end_date = current_date.replace(month=12, day=31)
            else:
                end_date = current_date.replace(month=end_month + 1, day=1) - timedelta(days=1)
            period_label = f"Q{quarter} {current_date.year}"
        else:  # year
            start_date = current_date.replace(month=1, day=1)
            end_date = current_date.replace(month=12, day=31)
            period_label = f"{current_date.year}"
        
        approval_actions = LeaveApprovalAction.objects.filter(
            approver=request.user,
            action_at__date__gte=start_date,
            action_at__date__lte=end_date
        ).select_related('leave_request__leave_type').exclude(status='cancelled')
        
        leave_types_all = LeaveType.objects.filter(is_active=True).order_by('created_at')
        labels = []
        type_data = []
        colors = ['#6366f1', '#10b981', '#f59e0b', '#ef4444', '#06b6d4', '#8b5cf6', '#f472b6', '#facc15']
        background_colors = []
        for idx, lt in enumerate(leave_types_all):
            labels.append(lt.name)
            cnt = approval_actions.filter(leave_request__leave_type=lt).count()
            type_data.append(cnt)
            background_colors.append(colors[idx % len(colors)])
        
        leave_types = LeaveType.objects.filter(is_active=True)
        
        time_data = defaultdict(lambda: defaultdict(int))
        time_labels = []
        
        if period == 'month':
            current_day = start_date
            while current_day <= end_date:
                time_labels.append(current_day.day)
                current_day += timedelta(days=1)
            
            for action in approval_actions:
                day_key = action.action_at.date().day
                leave_type_name = action.leave_request.leave_type.name
                time_data[day_key][leave_type_name] += 1
                
        elif period == 'quarter':
            current_month = start_date.replace(day=1)
            while current_month <= end_date:
                time_labels.append(current_month.strftime('%b'))
                month_key = current_month.month
                for action in approval_actions.filter(action_at__month=month_key):
                    leave_type_name = action.leave_request.leave_type.name
                    time_data[month_key][leave_type_name] += 1
                if current_month.month == 12:
                    current_month = current_month.replace(year=current_month.year + 1, month=1)
                else:
                    current_month = current_month.replace(month=current_month.month + 1)
        else:
            for month in range(1, 13):
                time_labels.append(calendar.month_abbr[month])
                for action in approval_actions.filter(action_at__month=month):
                    leave_type_name = action.leave_request.leave_type.name
                    time_data[month][leave_type_name] += 1
        
        line_chart_datasets = []
        colors = ['#6366f1', '#10b981', '#f59e0b', '#ef4444', '#06b6d4', '#8b5cf6']
        
        for i, leave_type in enumerate(leave_types):
            if period == 'month':
                data = [time_data[day].get(leave_type.name, 0) for day in range(1, len(time_labels) + 1)]
            elif period == 'quarter':
                quarter_start = ((current_date.month - 1) // 3) * 3 + 1
                data = [time_data[month].get(leave_type.name, 0) for month in range(quarter_start, quarter_start + 3)]
            else:
                data = [time_data[month].get(leave_type.name, 0) for month in range(1, 13)]
            
            color = colors[i % len(colors)]
            
            line_chart_datasets.append({
                'label': leave_type.name,
                'data': data,
                'borderColor': color,
                'backgroundColor': f"{color}20",
                'tension': 0.4,
                'fill': True,
                'pointRadius': 4,
                'pointHoverRadius': 6,
                'pointBackgroundColor': color,
                'pointBorderColor': '#ffffff',
                'pointBorderWidth': 2
            })
        
        return JsonResponse({
            'success': True,
            'period_label': period_label,
            'status_chart': {
                'labels': labels,
                'data': type_data,
                'backgroundColor': background_colors
            },
            'line_chart': {
                'labels': time_labels,
                'datasets': line_chart_datasets
            }
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

def leave_reasons_api(request, leave_type_id):
    reasons = LeaveReason.objects.filter(leave_type_id=leave_type_id, is_active=True)
    data = [
        {
            'id': reason.id,
            'reason_text': reason.reason_text
        }
        for reason in reasons
    ]
    return JsonResponse(data, safe=False)

@login_required(login_url="user-login")
def holidays_and_exceptions_api(request):
    from usercalendar.models import Holiday
    from .models import SundayException
    
    holidays = Holiday.objects.all().values('date', 'name', 'holiday_type')
    holidays_list = [{'date': holiday['date'].strftime('%Y-%m-%d'), 'name': holiday['name'], 'type': holiday['holiday_type']} for holiday in holidays]
    
    sunday_exceptions = SundayException.objects.all().values('date', 'description')
    exceptions_list = [{'date': exception['date'].strftime('%Y-%m-%d'), 'description': exception['description']} for exception in sunday_exceptions]
    
    return JsonResponse({
        'holidays': holidays_list,
        'sunday_exceptions': exceptions_list
    })

@login_required(login_url="user-login")
def check_approver_api(request):
    try:
        employment_info = request.user.employment_info
        has_approver = employment_info.approver is not None and employment_info.approver.is_active
        
        return JsonResponse({
            'has_approver': has_approver,
            'approver_name': f"{employment_info.approver.firstname} {employment_info.approver.lastname}" if has_approver else None
        })
    except:
        return JsonResponse({
            'has_approver': False,
            'approver_name': None,
            'error': 'Employment information not found'
        })

@login_required(login_url="user-login")
def search_approvals_ajax(request):
    """AJAX endpoint for searching approvals"""
    user = request.user
    search_query = request.GET.get('search', '').strip()
    page = request.GET.get('page', 1)
    
    # Get pending approvals with search
    for_leave_approval = (
        LeaveApprovalAction.objects
        .filter(approver=user)
        .annotate(routing_priority=Case(When(status='routing', then=0), default=1, output_field=IntegerField()))
        .exclude(Q(status__iexact='cancelled') | Q(comments='Updated my leave request'))
        .select_related('leave_request__employee', 'leave_request__leave_type', 'leave_request__leave_reason')
        .order_by('routing_priority', '-created_at')
    )

    # Apply search filter
    if search_query:
        for_leave_approval = for_leave_approval.filter(
            Q(leave_request__control_number__icontains=search_query) |
            Q(leave_request__employee__firstname__icontains=search_query) |
            Q(leave_request__employee__lastname__icontains=search_query) |
            Q(leave_request__employee__idnumber__icontains=search_query) |
            Q(leave_request__leave_type__name__icontains=search_query) |
            Q(leave_request__leave_reason__reason_text__icontains=search_query)
        )

    # Pagination
    paginator = Paginator(for_leave_approval, 10)
    pending_approvals_page_obj = paginator.get_page(page)
    
    # Prepare data for JSON response
    approvals_data = []
    for approval in pending_approvals_page_obj:
        approvals_data.append({
            'control_number': approval.leave_request.control_number,
            'employee_name': approval.leave_request.employee.full_name,
            'employee_id': approval.leave_request.employee.idnumber,
            'leave_type': approval.leave_request.leave_type.name,
            'leave_reason': approval.leave_request.leave_reason.reason_text if approval.leave_request.leave_reason else '',
            'duration_display': approval.leave_request.duration_display,
            'status': approval.status,
            'status_display': 'Waiting for your approval' if approval.status == 'routing' else approval.get_status_display(),
            'date_prepared': approval.leave_request.date_prepared.strftime('%b %d, %Y'),
            'is_routing': approval.status == 'routing'
        })
    
    return JsonResponse({
        'approvals': approvals_data,
        'pagination': {
            'has_previous': pending_approvals_page_obj.has_previous(),
            'has_next': pending_approvals_page_obj.has_next(),
            'previous_page_number': pending_approvals_page_obj.previous_page_number() if pending_approvals_page_obj.has_previous() else None,
            'next_page_number': pending_approvals_page_obj.next_page_number() if pending_approvals_page_obj.has_next() else None,
            'number': pending_approvals_page_obj.number,
            'num_pages': pending_approvals_page_obj.paginator.num_pages,
            'start_index': pending_approvals_page_obj.start_index(),
            'end_index': pending_approvals_page_obj.end_index(),
            'count': pending_approvals_page_obj.paginator.count,
            'page_range': list(pending_approvals_page_obj.paginator.page_range)
        },
        'search_query': search_query
    })

@login_required(login_url="user-login")
def admin_leave_detail(request, control_number):
    if not (request.user.hr_admin or request.user.hr_manager or 
            request.user.is_superuser or request.user.is_staff):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        leave_request = get_object_or_404(
            LeaveRequest.objects.select_related(
                'employee', 'leave_type', 'leave_reason', 
                'employee__employment_info__department'
            ).prefetch_related('approval_actions__approver'),
            control_number=control_number
        )
        
        hr_approval_action = leave_request.approval_actions.filter(
            approver__hr_admin=True,
            action_at__isnull=True
        ).first()
        
        can_approve = (
            hr_approval_action and 
            hr_approval_action.approver == request.user and 
            leave_request.status == 'routing'
        )
        
        approval_actions = leave_request.approval_actions.all().order_by('action_at')
        from django.template.loader import render_to_string
        content = render_to_string('leaverequest/detail_content.html', {
            'leave_request': leave_request,
            'can_approve': can_approve,
            'is_admin_view': True,
            'approval_actions': approval_actions
        }, request=request)
        return JsonResponse({
            'success': True,
            'content': content,
            'can_approve': can_approve
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required(login_url="user-login")  
@require_POST
def admin_approve_leave(request, control_number):
    if not (request.user.hr_admin or request.user.hr_manager or 
            request.user.is_superuser or request.user.is_staff):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        leave_request = get_object_or_404(LeaveRequest, control_number=control_number)
        
        approval_action = leave_request.approval_actions.filter(
            approver=request.user,
            action_at__isnull=True
        ).first()
        
        if not approval_action:
            return JsonResponse({'error': 'No pending approval found for this user'}, status=400)
        
        comments = request.POST.get('comments', '').strip()
        
        approval_action.action = 'approved'
        approval_action.status = 'approved'
        approval_action.comments = comments
        approval_action.action_at = timezone.now()
        approval_action.save()
        
        if leave_request.status != "disapproved":
            leave_request.status = 'approved'
            leave_request.updated_at = timezone.now()
            leave_request.save()

            start_date_fmt = leave_request.date_from.strftime("%b %d, %Y")
            end_date_fmt = leave_request.date_to.strftime("%b %d, %Y")

            Notification.objects.create(
                title="Leave Approved",
                message=f"Your leave request for {start_date_fmt} to {end_date_fmt} has been approved.",
                notification_type="approval",
                sender=request.user,
                recipient=leave_request.employee,
                for_all=False,
                module="leave"
            )
        
        return JsonResponse({
            'success': True,
            'message': 'Leave request approved successfully'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required(login_url="user-login")
@require_POST  
def admin_disapprove_leave(request, control_number):
    if not (request.user.hr_admin or request.user.hr_manager or 
            request.user.is_superuser or request.user.is_staff):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        leave_request = get_object_or_404(LeaveRequest, control_number=control_number)
        
        approval_action = leave_request.approval_actions.filter(
            approver=request.user,
            action_at__isnull=True
        ).first()
        
        if not approval_action:
            return JsonResponse({'error': 'No pending approval found for this user'}, status=400)
        
        comments = request.POST.get('comments', '').strip()
        
        approval_action.action = 'disapproved'
        approval_action.status = 'disapproved'
        approval_action.comments = comments
        approval_action.action_at = timezone.now()
        approval_action.save()
        
        leave_request.status = 'disapproved'
        leave_request.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Leave request disapproved successfully'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required(login_url="user-login")
@require_POST
def hr_admin_process_approval(request, control_number):
    if not (request.user.hr_admin or request.user.hr_manager or 
            request.user.is_superuser or request.user.is_staff):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        import json
        data = json.loads(request.body)
        action = data.get('action')
        comments = data.get('comments', '').strip()
        
        if action not in ['approve', 'disapprove']:
            return JsonResponse({'error': 'Invalid action'}, status=400)
        
        leave_request = get_object_or_404(LeaveRequest, control_number=control_number)

        approval_action = leave_request.approval_actions.filter(
            approver=request.user,
        ).last()
        
        if not approval_action:
            return JsonResponse({'error': 'No pending approval found for this user'}, status=400)
        
        # Convert action to proper model format
        action_value = 'approved' if action == 'approve' else 'disapproved'
        status_value = 'approved' if action == 'approve' else 'disapproved'
        
        approval_action.action = action_value
        approval_action.status = status_value
        approval_action.action_at = timezone.now()

        if action == 'approve' and leave_request.status != 'approved':
            updated_comments = process_leave_approval_with_balance_check(leave_request, comments)
            approval_action.comments = updated_comments

            leave_request.status = 'approved'
            leave_request.save()
        elif action == 'approve' and leave_request.status == 'approved':
            # Additional logic: when action == 'approve' and leave_request.status == 'approved' 
            # then update selected leave_request.status to approved (already approved, but ensure consistency)
            approval_action.comments = comments
            leave_request.status = 'approved'
            leave_request.save()
        else:
            approval_action.comments = comments
            
        approval_action.save()
        
        # leave_request.status = action
        # leave_request.save()
        
        send_leave_decision_email(leave_request, action == 'approve', approval_action.comments)
        
        return JsonResponse({
            'success': True,
            'message': f'Leave request {action} successfully'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def process_leave_approval_with_balance_check(leave_request, original_comments):
    employee = leave_request.employee
    
    employment_type = getattr(employee.employment_info, 'employment_type', '').lower()
    
    if employment_type in ['probationary', 'ojt']:
        additional_comment = f"Leave approved without balance deduction (Employee type: {employment_type.title()})."
        if original_comments:
            return f"{original_comments}\n\n{additional_comment}"
        else:
            return additional_comment
    
    leave_type = leave_request.leave_type
    
    # Check if this leave type should be deducted from balance
    if not leave_type.is_deducted:
        additional_comment = f"Leave approved without balance deduction (Leave type '{leave_type.name}' is configured to not deduct from balance)."
        if original_comments:
            return f"{original_comments}\n\n{additional_comment}"
        else:
            return additional_comment
    
    request_start = leave_request.date_from
    request_end = leave_request.date_to
    hrs_requested = leave_request.hrs_requested / 8
    
    balances = LeaveBalance.objects.filter(
        employee=employee,
        leave_type=leave_type,
        valid_from__lte=request_end,
        valid_to__gte=request_start,
    ).order_by('valid_from')
    
    for bal in balances:
        print(f"  Balance {bal.id}: {bal.valid_from} to {bal.valid_to}, entitled: {bal.entitled}, used: {bal.used}, remaining: {bal.remaining}")
    
    if not balances.exists():
        additional_comment = f"No active leave balance found for this leave type ({leave_type.name})."
        if original_comments:
            return f"{original_comments}\n\n{additional_comment}"
        else:
            return additional_comment
    
    total_remaining = sum(balance.remaining for balance in balances)

    if total_remaining < hrs_requested:
        additional_comment = f"Insufficient leave balance. Available: {total_remaining} hours, Requested: {hrs_requested} days. Leave approved but balance may go negative."
        if original_comments:
            comments_with_warning = f"{original_comments}\n\n{additional_comment}"
        else:
            comments_with_warning = additional_comment
        
        try:
            deduct_leave_balance(leave_request)
        except Exception as e:
            error_comment = f"Balance deduction failed: {str(e)}"
            comments_with_warning += f"\n{error_comment}"
        
        return comments_with_warning
    
    try:
        deduct_leave_balance(leave_request)
        success_comment = f"Leave balance deducted successfully. Remaining balance: {total_remaining - hrs_requested} days."
        if original_comments:
            return f"{original_comments}\n\n{success_comment}"
        else:
            return success_comment
    except Exception as e:
        error_comment = f"Leave approved but balance deduction failed: {str(e)}"
        if original_comments:
            return f"{original_comments}\n\n{error_comment}"
        else:
            return error_comment

def deduct_leave_balance(leave_request):
    """
    Deduct leave balances when approving a leave request.
    Split the leave request across matching leave balance periods and deduct balances.
    Uses the same logic as restore_leave_balances but in reverse.
    """
    from datetime import timedelta
    
    employee = leave_request.employee
    leave_type = leave_request.leave_type
    
    # Get all leave balances for this employee and leave type
    leave_balances = LeaveBalance.objects.filter(
        employee=employee,
        leave_type=leave_type
    ).order_by('valid_from')
    
    if not leave_balances.exists():
        raise Exception(f"No valid leave balance found for {employee.full_name} - {leave_type.name}")
    
    current_date = leave_request.date_from
    
    while current_date <= leave_request.date_to:
        # Find the leave balance that covers this date
        matching_balance = None
        for balance in leave_balances:
            if balance.valid_from <= current_date <= balance.valid_to:
                matching_balance = balance
                break
        
        if not matching_balance:
            current_date += timedelta(days=1)
            continue
        
        # Find the end date for this balance period
        period_end = min(leave_request.date_to, matching_balance.valid_to)
        
        # Calculate hours for this period
        period_hours = 0
        temp_date = current_date
        
        # Get holidays and sunday exceptions for the period
        from usercalendar.models import Holiday
        holidays = set(Holiday.objects.filter(
            date__range=[current_date, period_end]
        ).values_list('date', flat=True))
        
        sunday_exceptions = set(SundayException.objects.filter(
            date__range=[current_date, period_end]
        ).values_list('date', flat=True))
        
        while temp_date <= period_end:
            is_holiday = temp_date in holidays
            is_sunday = temp_date.weekday() == 6
            is_sunday_exempted = temp_date in sunday_exceptions
            
            # Count as working day if not holiday and (not sunday or sunday exempted)
            if not is_holiday and (not is_sunday or is_sunday_exempted):
                period_hours += 8  # 8 hours per working day
            
            temp_date += timedelta(days=1)
        
        # Convert hours to days (8 hours = 1 day)
        days_to_deduct = period_hours / 8
        
        if days_to_deduct > 0:
            # Deduct the balance
            from decimal import Decimal
            days_decimal = Decimal(str(days_to_deduct))
            
            # Deduct from balance (opposite of restore)
            matching_balance.used += days_decimal
            matching_balance.remaining -= days_decimal
            
            # Ensure remaining doesn't go below 0 (though we allow negative for approved leaves)
            if matching_balance.remaining < 0:
                print(f"Warning: Balance {matching_balance.id} remaining went negative: {matching_balance.remaining}")
            
            matching_balance.save()
            print(f"Deducted {days_to_deduct} days from balance {matching_balance.id} for period {current_date} to {period_end}. New remaining: {matching_balance.remaining}")
        
        # Move to the next period
        current_date = period_end + timedelta(days=1)
    
    print(f"Successfully processed leave deduction for {employee.full_name} - {leave_type.name}")

def send_leave_decision_email(leave_request, is_approved, comments):
    try:
        employee = leave_request.employee
        # Check the actual leave request status to determine the final decision
        final_decision_approved = leave_request.status == 'approved'
        subject = f"Leave Request {leave_request.control_number} - {'Approved' if final_decision_approved else 'Disapproved'}"
        
        context = {
            'employee_name': employee.full_name,
            'control_number': leave_request.control_number,
            'leave_type': leave_request.leave_type.name,
            'date_from': leave_request.date_from,
            'date_to': leave_request.date_to,
            'days_requested': leave_request.hrs_requested / 8,
            'reason': leave_request.reason,
            'is_approved': final_decision_approved,
            'comments': comments,
            'decision_date': timezone.now().date(),
        }
        
        if final_decision_approved:
            message = f"""
Dear {employee.full_name},

Your leave request has been APPROVED.

Request Details:
- Control Number: {leave_request.control_number}
- Leave Type: {leave_request.leave_type.name}
- Duration: {leave_request.date_from} to {leave_request.date_to} ({leave_request.days_requested} days)
- Reason: {leave_request.reason}

{f'HR Comments: {comments}' if comments else ''}

Please ensure to coordinate with your team and complete any necessary handover procedures before your leave period.

Best regards,
HR Department
REPConnect
"""
        else:
            message = f"""
Dear {employee.full_name},

Your leave request has been DISAPPROVED.

Request Details:
- Control Number: {leave_request.control_number}
- Leave Type: {leave_request.leave_type.name}
- Duration: {leave_request.date_from} to {leave_request.date_to} ({leave_request.days_requested} days)
- Reason: {leave_request.reason}

{f'HR Comments: {comments}' if comments else ''}

If you have any questions or would like to discuss this decision, please contact the HR department.

Best regards,
HR Department
REPConnect
"""
        send_mail(
            subject=subject,
            message=message,
            from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@repconnect.com'),
            recipient_list=[employee.email] if hasattr(employee, 'email') and employee.email else [],
            fail_silently=True,
        )
        
    except Exception as e:
        print(f"Email notification failed: {str(e)}")

@login_required(login_url="user-login")
def hr_admin_approval_detail(request, control_number):
    if not (request.user.hr_admin or request.user.hr_manager or 
            request.user.is_superuser or request.user.is_staff):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        leave_request = get_object_or_404(
            LeaveRequest.objects.select_related(
                'employee', 'leave_type', 'leave_reason', 
                'employee__employment_info__department'
            ).prefetch_related('approval_actions__approver'),
            control_number=control_number
        )
        
        approval_action = leave_request.approval_actions.filter(
            approver=request.user,
            action_at__isnull=True
        ).first()
        
        can_approve = approval_action is not None and leave_request.status == 'routing'
    
        
        leave_balances = LeaveBalance.objects.filter(
            employee=leave_request.employee,
            leave_type=leave_request.leave_type,
            valid_from__lte=leave_request.date_to,
            valid_to__gte=leave_request.date_from,
        ).order_by('valid_from')
        
        for bal in leave_balances:
            overlap_check = bal.valid_from <= leave_request.date_to and bal.valid_to >= leave_request.date_from
            print(f"  Balance {bal.id}: {bal.valid_from} to {bal.valid_to}, remaining: {bal.remaining}, overlap: {overlap_check}")
        
        total_remaining = sum(balance.remaining for balance in leave_balances)
        remaining_after_approval = total_remaining - leave_request.hrs_requested / 8
        
        all_active_balances = LeaveBalance.objects.filter(
            employee=leave_request.employee,
            leave_type=leave_request.leave_type,
        ).order_by('valid_from')
        
        for bal in all_active_balances:
            print(f"  Balance {bal.id}: {bal.valid_from} to {bal.valid_to}, remaining: {bal.remaining}")
        
        from django.template.loader import render_to_string
        content = render_to_string('leaverequest/approval_detail_content.html', {
            'leave_request': leave_request,
            'can_approve': can_approve,
            'is_admin_view': True,
            'leave_balances': leave_balances,
            'all_active_balances': all_active_balances,
            'total_remaining': total_remaining,
            'remaining_after_approval': remaining_after_approval,
            'sufficient_balance': total_remaining >= leave_request.hrs_requested / 8
        }, request=request)
        
        return JsonResponse({
            'success': True,
            'content': content,
            'can_approve': can_approve
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required(login_url="user-login")
def export_leave_report(request):
    # Get date range from request
    date_from_str = request.GET.get('date_from')
    date_to_str = request.GET.get('date_to')
    
    if not date_from_str or not date_to_str:
        return JsonResponse({'error': 'Date range is required'}, status=400)
    
    try:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Invalid date format'}, status=400)
    
    # Get filtered leave requests
    leave_requests = LeaveRequest.objects.filter(
        Q(date_from__gte=date_from, date_from__lte=date_to) |
        Q(date_to__gte=date_from, date_to__lte=date_to) |
        Q(date_from__lte=date_from, date_to__gte=date_to)
    ).select_related('employee', 'leave_type', 'leave_reason', 'employee__employment_info__department')
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheets
    create_overview_sheet(wb, leave_requests, date_from, date_to)
    create_ranking_sheet(wb, date_from, date_to)
    create_summary_sheet(wb, leave_requests, date_from, date_to)
    create_routing_sheet(wb, leave_requests, date_from, date_to)
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="leave_report_{date_from_str}_to_{date_to_str}.xlsx"'
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    
    return response

def apply_sheet_formatting(ws, start_row=1):
    """Apply consistent formatting to all sheets"""
    # Define styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply borders to all cells with data
    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            if cell.value is not None:
                cell.border = thin_border
    
    # Apply header formatting only to header row (start_row + 2)
    header_row_num = start_row
    for cell in ws[header_row_num]:
        if cell.value is not None:
            cell.font = header_font
            cell.fill = header_fill

def create_overview_sheet(wb, leave_requests, date_from, date_to):
    ws = wb.create_sheet("Overview")
    
    ws['A1'] = "RYONAN ELECTRIC PHILIPPINES CORPORATION"
    ws['A2'] = f"Leave Overview ({date_from} to {date_to})"
    
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'].font = Font(italic=True, size=12)
    ws.merge_cells('A1:F1')
    ws.merge_cells('A2:F2')
    
    headers = ['Leave Type', 'Leave Count', 'Leave Percentage']
    
    departments = Department.objects.values_list('department_name', flat=True).distinct()
    headers.extend(departments)
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    
    leave_type_stats = leave_requests.values('leave_type__name').annotate(
        count=Count('id')
    ).order_by('-count')
    
    total_requests = leave_requests.count()
    
    row = 5
    for stat in leave_type_stats:
        leave_type = stat['leave_type__name']
        count = stat['count']
        percentage = (count / total_requests * 100) if total_requests > 0 else 0
        
        ws.cell(row=row, column=1, value=leave_type)
        ws.cell(row=row, column=2, value=count)
        ws.cell(row=row, column=3, value=f"{percentage:.1f}%")
        
        dept_stats = leave_requests.filter(leave_type__name=leave_type).values(
            'employee__employment_info__department__department_name'
        ).annotate(count=Count('id'))
        
        dept_dict = {dept['employee__employment_info__department__department_name']: dept['count'] for dept in dept_stats}
        
        for col, dept_name in enumerate(departments, 4):
            dept_count = dept_dict.get(dept_name, 0)
            ws.cell(row=row, column=col, value=dept_count)
        
        row += 1
    
    chart_start_row = row + 2
    chart = LineChart()
    chart.title = "Leave Type Distribution"
    chart.style = 13
    chart.y_axis.title = 'Number of Requests'
    chart.x_axis.title = 'Leave Types'
    
    # Data for chart (Leave Type and Count)
    data = Reference(ws, min_col=2, min_row=4, max_row=row-1, max_col=2)  # Count column
    cats = Reference(ws, min_col=1, min_row=5, max_row=row-1)  # Leave type names
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    ws.add_chart(chart, f"A{chart_start_row}")
    
    apply_sheet_formatting(ws, 4)

def create_ranking_sheet(wb, date_from, date_to):
    ws = wb.create_sheet("Leave Ranking")
    
    ws['A1'] = "RYONAN ELECTRIC PHILIPPINES CORPORATION"
    
    current_date = datetime.now()
    if current_date.month >= 5:
        fiscal_year_start = current_date.year
        fiscal_year_end = current_date.year + 1
    else:
        fiscal_year_start = current_date.year - 1
        fiscal_year_end = current_date.year
    
    ws['A2'] = f"Leave Reason Ranking for FY {fiscal_year_start}-{fiscal_year_end}"
    
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'].font = Font(italic=True, size=12)
    ws.merge_cells('A1:N1')
    ws.merge_cells('A2:N2')
    
    months = ['May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar', 'Apr']
    headers = ['Leave Reason'] + [f"{month} {fiscal_year_start if i < 8 else fiscal_year_end}" for i, month in enumerate(months)]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    
    fiscal_start_date = datetime(fiscal_year_start, 5, 1).date()
    fiscal_end_date = datetime(fiscal_year_end, 4, 30).date()
    
    fiscal_requests = LeaveRequest.objects.filter(
        date_prepared__date__gte=fiscal_start_date,
        date_prepared__date__lte=fiscal_end_date
    )
    
    leave_reason_stats = fiscal_requests.values('leave_reason__reason_text').annotate(
        total_count=Count('id')
    ).order_by('-total_count')
    
    row = 5
    for stat in leave_reason_stats:
        reason = stat['leave_reason__reason_text'] or 'No Reason'
        ws.cell(row=row, column=1, value=reason)
        
        for month_idx, month in enumerate(months):
            if month_idx < 8:
                year = fiscal_year_start
                month_num = month_idx + 5
            else:
                year = fiscal_year_end
                month_num = month_idx - 7
            
            monthly_count = fiscal_requests.filter(
                leave_reason__reason_text=stat['leave_reason__reason_text'],
                date_prepared__year=year,
                date_prepared__month=month_num
            ).count()
            
            ws.cell(row=row, column=month_idx + 2, value=monthly_count)
        
        row += 1
    
    apply_sheet_formatting(ws, 4)

def create_summary_sheet(wb, leave_requests, date_from, date_to):
    """Create Leave Summary sheet with all leave requests"""
    ws = wb.create_sheet("Leave Summary")
    
    # Headers
    ws['A1'] = "RYONAN ELECTRIC PHILIPPINES CORPORATION"
    ws['A2'] = f"Leave Request Summary ({date_from} to {date_to})"
    
    # Make headers bold and italic for subheader
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'].font = Font(italic=True, size=12)
    ws.merge_cells('A1:G1')
    ws.merge_cells('A2:G2')
    
    # Column headers
    headers = ['Control Number', 'Date Prepared', 'Employee Name', 'Department', 'Leave Type', 'Leave Reason', 'Status']
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    
    # Status colors
    status_colors = {
        'approved': PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),  # Light green
        'disapproved': PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"),  # Light red
        'routing': PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid"),  # Light yellow
        'cancelled': PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),  # Light gray
    }
    
    # Add data
    row = 5
    for leave in leave_requests.order_by('-date_prepared'):
        ws.cell(row=row, column=1, value=leave.control_number)
        ws.cell(row=row, column=2, value=leave.date_prepared.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=3, value=leave.employee.full_name)
        ws.cell(row=row, column=4, value=leave.employee.employment_info.department.department_name if leave.employee.employment_info else 'N/A')
        ws.cell(row=row, column=5, value=leave.leave_type.name)
        ws.cell(row=row, column=6, value=leave.leave_reason.reason_text if leave.leave_reason else 'N/A')
        
        # Status cell with color
        status_cell = ws.cell(row=row, column=7, value=leave.get_status_display())
        if leave.status in status_colors:
            status_cell.fill = status_colors[leave.status]
        
        row += 1
    
    apply_sheet_formatting(ws, 4)

def create_routing_sheet(wb, leave_requests, date_from, date_to):
    """Create Routing Leave sheet with pending requests"""
    ws = wb.create_sheet("Routing Leave")
    
    # Headers
    ws['A1'] = "RYONAN ELECTRIC PHILIPPINES CORPORATION"
    ws['A2'] = f"Leave For Routing ({date_from} to {date_to})"
    
    # Make headers bold and italic for subheader
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'].font = Font(italic=True, size=12)
    ws.merge_cells('A1:G1')
    ws.merge_cells('A2:G2')
    
    # Column headers
    headers = ['Control Number', 'Employee Name', 'Department', 'Leave Type', 'Leave Reason', 'Duration', 'Current Approver']
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    
    # Filter for routing status only
    routing_requests = leave_requests.filter(status='routing').order_by('-date_prepared')
    
    # Add data
    row = 5
    for leave in routing_requests:
        # Get current approver
        current_action = leave.approval_actions.filter(status='routing').first()
        current_approver = current_action.approver.full_name if current_action else 'N/A'
        
        ws.cell(row=row, column=1, value=leave.control_number)
        ws.cell(row=row, column=2, value=leave.employee.full_name)
        ws.cell(row=row, column=3, value=leave.employee.employment_info.department.department_name if leave.employee.employment_info else 'N/A')
        ws.cell(row=row, column=4, value=leave.leave_type.name)
        ws.cell(row=row, column=5, value=leave.leave_reason.reason_text if leave.leave_reason else 'N/A')
        ws.cell(row=row, column=6, value=leave.duration_display)
        ws.cell(row=row, column=7, value=current_approver)
        row += 1
    
    apply_sheet_formatting(ws, 4)

@login_required(login_url="user-login")
def get_balance_details(request, balance_id):
    """Get balance details for modal display"""
    try:
        balance = LeaveBalance.objects.select_related(
            'employee', 'leave_type', 'employee__employment_info__department'
        ).get(id=balance_id)
        
        return JsonResponse({
            'success': True,
            'data': {
                'id': balance.id,
                'employee_id': balance.employee.idnumber,
                'employee_name': balance.employee.full_name,
                'leave_type': balance.leave_type.name,
                'entitled': str(balance.entitled),
                'used': str(balance.used),
                'remaining': str(balance.remaining),
                'valid_from': balance.valid_from.strftime('%Y-%m-%d'),
                'valid_to': balance.valid_to.strftime('%Y-%m-%d'),
                'can_edit': hasattr(request.user, 'accounting_admin') and request.user.accounting_admin
            }
        })
    except LeaveBalance.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Balance not found'}, status=404)

@login_required(login_url="user-login") 
def update_balance(request, balance_id):
    """Update balance details"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)
    
    # Check if user has permission
    if not (hasattr(request.user, 'accounting_admin') and request.user.accounting_admin):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    try:
        balance = LeaveBalance.objects.get(id=balance_id)
        
        # Update editable fields with proper type conversion
        entitled_str = request.POST.get('entitled', '').strip()
        valid_from_str = request.POST.get('valid_from', '').strip()
        valid_to_str = request.POST.get('valid_to', '').strip()
        
        # Validate and convert entitled to Decimal
        if entitled_str:
            try:
                balance.entitled = Decimal(entitled_str)
            except (ValueError, TypeError):
                return JsonResponse({'success': False, 'error': 'Invalid entitled value'}, status=400)
        
        # Validate and convert date fields
        if valid_from_str:
            try:
                balance.valid_from = datetime.strptime(valid_from_str, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({'success': False, 'error': 'Invalid valid_from date format'}, status=400)
        
        if valid_to_str:
            try:
                balance.valid_to = datetime.strptime(valid_to_str, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({'success': False, 'error': 'Invalid valid_to date format'}, status=400)
        
        # Validate business rules
        if balance.valid_from and balance.valid_to and balance.valid_to <= balance.valid_from:
            return JsonResponse({'success': False, 'error': 'Valid to date must be after valid from date'}, status=400)
        
        if balance.entitled and balance.entitled <= 0:
            return JsonResponse({'success': False, 'error': 'Entitled days must be greater than 0'}, status=400)
        
        balance.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Balance updated successfully'
        })
    except LeaveBalance.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Balance not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required(login_url="user-login")
def delete_balance(request, balance_id):
    """Delete balance"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)
    
    try:
        balance = LeaveBalance.objects.get(id=balance_id)
        balance.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Balance deleted successfully'
        })
    except LeaveBalance.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Balance not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required(login_url="user-login")
def download_balance_template(request):
    """Download Excel template for balance import"""
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create Upload Balance sheet
    upload_sheet = wb.create_sheet("Upload Balance")
    upload_headers = ['Id Number', 'Name', 'Leave Type', 'Entitled', 'Valid From', 'Valid To']
    for col, header in enumerate(upload_headers, 1):
        upload_sheet.cell(row=1, column=col, value=header)
        upload_sheet.cell(row=1, column=col).font = Font(bold=True)
    
    # Add sample data
    sample_data = [
        ['EMP001', 'John Doe', 'Vacation Leave', 15, '2024-01-01', '2024-12-31'],
        ['EMP002', 'Jane Smith', 'Sick Leave', 12, '2024-01-01', '2024-12-31']
    ]
    
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            upload_sheet.cell(row=row_num, column=col_num, value=value)
    
    # Add instructions
    instructions = [
        "",
        "INSTRUCTIONS:",
        "1. Replace sample data with actual employee information",
        "2. Id Number: Use employee ID exactly as in system",
        "3. Name: Employee's full name",
        "4. Leave Type: Select from dropdown (see Leave Type sheet)",
        "5. Entitled: Number of leave days allocated",
        "6. Valid From/To: Date range format YYYY-MM-DD",
        "7. Delete these instruction rows before uploading",
    ]
    
    start_row = len(sample_data) + 3
    for i, instruction in enumerate(instructions):
        upload_sheet.cell(row=start_row + i, column=1, value=instruction)
        if instruction.startswith("INSTRUCTIONS:"):
            upload_sheet.cell(row=start_row + i, column=1).font = Font(bold=True)
    
    # Create Leave Type sheet
    leave_type_sheet = wb.create_sheet("Leave Type")
    leave_type_sheet.cell(row=1, column=1, value="Leave Types")
    leave_type_sheet.cell(row=1, column=1).font = Font(bold=True)
    
    leave_types = LeaveType.objects.filter(is_active=True).values_list('name', flat=True)
    for row, leave_type in enumerate(leave_types, 2):
        leave_type_sheet.cell(row=row, column=1, value=leave_type)
    
    # Add data validation for Leave Type column in Upload Balance sheet
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(
        type="list",
        formula1=f"'Leave Type'!$A$2:$A${len(leave_types) + 1}",
        showDropDown=True
    )
    upload_sheet.add_data_validation(dv)
    dv.add(f'C2:C1000')  # Apply to Leave Type column
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="leave_balance_template.xlsx"'
    
    return response

@login_required(login_url="user-login")
def import_balance(request):
    """Import balance from Excel file"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)
    
    try:
        uploaded_file = request.FILES.get('file')
        delete_all = request.POST.get('delete_all') == 'true'
        
        if not uploaded_file:
            return JsonResponse({'success': False, 'error': 'No file uploaded'}, status=400)
        
        # Delete all existing balances if requested
        if delete_all:
            LeaveBalance.objects.all().delete()
        
        # Process Excel file
        import pandas as pd
        df = pd.read_excel(uploaded_file, sheet_name='Upload Balance')
        
        created_count = 0
        skipped_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                # Get employee by ID number
                employee = EmployeeLogin.objects.get(idnumber=row['Id Number'])

                # Get leave type
                leave_type = LeaveType.objects.get(name=row['Leave Type'])

                # Convert pandas Timestamp to date if necessary
                valid_from = row['Valid From']
                valid_to = row['Valid To']
                if hasattr(valid_from, 'to_pydatetime'):
                    valid_from = valid_from.to_pydatetime().date()
                elif hasattr(valid_from, 'date'):
                    valid_from = valid_from.date()
                if hasattr(valid_to, 'to_pydatetime'):
                    valid_to = valid_to.to_pydatetime().date()
                elif hasattr(valid_to, 'date'):
                    valid_to = valid_to.date()

                # Check for duplicates
                existing = LeaveBalance.objects.filter(
                    employee=employee,
                    leave_type=leave_type,
                    valid_from=valid_from,
                    valid_to=valid_to
                ).exists()

                if existing:
                    skipped_count += 1
                    continue

                # Create balance
                LeaveBalance.objects.create(
                    employee=employee,
                    leave_type=leave_type,
                    entitled=row['Entitled'],
                    valid_from=valid_from,
                    valid_to=valid_to
                )
                created_count += 1
                
            except EmployeeLogin.DoesNotExist:
                errors.append(f"Row {index + 2}: Employee with ID {row['Id Number']} not found")
            except LeaveType.DoesNotExist:
                errors.append(f"Row {index + 2}: Leave type '{row['Leave Type']}' not found")
            except Exception as e:
                errors.append(f"Row {index + 2}: {str(e)}")
        
        return JsonResponse({
            'success': True,
            'message': f'Import completed. {created_count} records created, {skipped_count} skipped.',
            'errors': errors if errors else None
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)