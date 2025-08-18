from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.contrib import messages
from django.db.models import Count, Q
from django.utils import timezone
from datetime import timedelta, datetime
from userlogin.models import EmployeeLogin
from notification.models import Notification
from .models import Device, Ticket, DeviceType, TicketCategory
from .forms import DeviceForm, TicketForm, TicketUpdateForm, TicketReviewForm, DeviceUpdateForm
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from django.utils.dateparse import parse_date
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.units import inch
from io import BytesIO


@login_required
def mis_user_dashboard(request):
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    now = timezone.now()
    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    start_of_last_month = (start_of_month - timedelta(days=1)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    end_of_last_month = start_of_month - timedelta(seconds=1)
    
    current_month_tickets = Ticket.objects.filter(
        requestor=request.user,
        created_at__gte=start_of_month
    )
    
    last_month_tickets = Ticket.objects.filter(
        requestor=request.user,
        created_at__gte=start_of_last_month,
        created_at__lte=end_of_last_month
    )
    
    total_current = current_month_tickets.count()
    total_last = last_month_tickets.count()
    
    if total_last > 0:
        percentage_change = ((total_current - total_last) / total_last) * 100
    else:
        percentage_change = 100 if total_current > 0 else 0
    
    # Calculate individual status percentages
    approved_current = current_month_tickets.filter(status='Approved').count()
    approved_last = last_month_tickets.filter(status='Approved').count()
    approved_percentage = ((approved_current - approved_last) / approved_last * 100) if approved_last > 0 else (100 if approved_current > 0 else 0)
    
    disapproved_current = current_month_tickets.filter(status='Disapproved').count()
    disapproved_last = last_month_tickets.filter(status='Disapproved').count()
    disapproved_percentage = ((disapproved_current - disapproved_last) / disapproved_last * 100) if disapproved_last > 0 else (100 if disapproved_current > 0 else 0)
    
    processing_current = current_month_tickets.filter(status='Processing').count()
    processing_last = last_month_tickets.filter(status='Processing').count()
    processing_percentage = ((processing_current - processing_last) / processing_last * 100) if processing_last > 0 else (100 if processing_current > 0 else 0)
    
    stats = {
        'total_requests': total_current,
        'approved': approved_current,
        'disapproved': disapproved_current,
        'processing': processing_current,
        'percentage_change': round(percentage_change, 1),
        'is_increase': percentage_change >= 0,
        'approved_percentage': round(approved_percentage, 1),
        'approved_is_increase': approved_percentage >= 0,
        'disapproved_percentage': round(disapproved_percentage, 1),
        'disapproved_is_increase': disapproved_percentage >= 0,
        'processing_percentage': round(processing_percentage, 1),
        'processing_is_increase': processing_percentage >= 0,
    }
    
    # Handle AJAX requests for search and pagination
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        search_query = request.GET.get('search', '').strip()
        tab = request.GET.get('tab', 'requests')
        
        if tab == 'requests':
            tickets_qs = Ticket.objects.filter(requestor=request.user)
            if search_query:
                tickets_qs = tickets_qs.filter(
                    Q(ticket_number__icontains=search_query) |
                    Q(device__device_name__icontains=search_query) |
                    Q(category__name__icontains=search_query) |
                    Q(priority_level__icontains=search_query) |
                    Q(status__icontains=search_query)
                )
            tickets_qs = tickets_qs.order_by('-created_at')
            
            # Pagination
            page = request.GET.get('page', 1)
            paginator = Paginator(tickets_qs, 10)
            tickets = paginator.get_page(page)
            
            return JsonResponse({
                'success': True,
                'html': render(request, 'ticketing/partials/tickets_table.html', {'tickets': tickets}).content.decode(),
                'pagination': {
                    'has_previous': tickets.has_previous(),
                    'has_next': tickets.has_next(),
                    'number': tickets.number,
                    'num_pages': tickets.paginator.num_pages,
                    'start_index': tickets.start_index(),
                    'end_index': tickets.end_index(),
                    'count': tickets.paginator.count,
                }
            })
        elif tab == 'devices':
            devices_qs = Device.objects.filter(user=request.user, is_active=True)
            if search_query:
                devices_qs = devices_qs.filter(
                    Q(device_name__icontains=search_query) |
                    Q(device_code__icontains=search_query) |
                    Q(device_type__name__icontains=search_query) |
                    Q(device_location__icontains=search_query)
                )
            devices_qs = devices_qs.order_by('-created_at')
            
            # Pagination
            page = request.GET.get('page', 1)
            paginator = Paginator(devices_qs, 10)
            devices = paginator.get_page(page)
            
            return JsonResponse({
                'success': True,
                'html': render(request, 'ticketing/partials/devices_table.html', {'devices': devices}).content.decode(),
                'pagination': {
                    'has_previous': devices.has_previous(),
                    'has_next': devices.has_next(),
                    'number': devices.number,
                    'num_pages': devices.paginator.num_pages,
                    'start_index': devices.start_index(),
                    'end_index': devices.end_index(),
                    'count': devices.paginator.count,
                }
            })
    
    # Regular page load
    tickets_qs = Ticket.objects.filter(requestor=request.user).order_by('-created_at')
    devices_qs = Device.objects.filter(user=request.user, is_active=True).order_by('-created_at')
    
    # Pagination for initial load
    tickets_paginator = Paginator(tickets_qs, 10)
    devices_paginator = Paginator(devices_qs, 10)
    tickets_page = request.GET.get('tickets_page', 1)
    devices_page = request.GET.get('devices_page', 1)
    tickets = tickets_paginator.get_page(tickets_page)
    devices = devices_paginator.get_page(devices_page)
    device_types = DeviceType.objects.all()
    categories = TicketCategory.objects.all()
    
    device_form = DeviceForm()
    ticket_form = TicketForm(user=request.user)
    
    context = {
        'stats': stats,
        'tickets': tickets,
        'device_types': device_types,
        'devices': devices,
        'categories': categories,
        'device_form': device_form,
        'ticket_form': ticket_form,
    }
    
    return render(request, 'ticketing/user-ticket.html', context)


@login_required
def mis_admin_dashboard(request):
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    if not request.user.mis_admin:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('mis_user_dashboard')
    
    # Get current month and last month dates
    now = timezone.now()
    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    start_of_last_month = (start_of_month - timedelta(days=1)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    end_of_last_month = start_of_month - timedelta(seconds=1)
    
    # Current month statistics for all users
    current_month_tickets = Ticket.objects.filter(created_at__gte=start_of_month)
    
    # Last month statistics for all users
    last_month_tickets = Ticket.objects.filter(
        created_at__gte=start_of_last_month,
        created_at__lte=end_of_last_month
    )
    
    # Calculate detailed statistics with individual percentages
    total_current = current_month_tickets.count()
    total_last = last_month_tickets.count()
    
    approved_current = current_month_tickets.filter(status='Approved').count()
    approved_last = last_month_tickets.filter(status='Approved').count()
    
    disapproved_current = current_month_tickets.filter(status='Disapproved').count()
    disapproved_last = last_month_tickets.filter(status='Disapproved').count()
    
    processing_current = current_month_tickets.filter(status='Processing').count()
    processing_last = last_month_tickets.filter(status='Processing').count()
    
    # Calculate percentage changes
    def calculate_percentage(current, last):
        if last > 0:
            return ((current - last) / last) * 100
        else:
            return 100 if current > 0 else 0
    
    total_percentage = calculate_percentage(total_current, total_last)
    approved_percentage = calculate_percentage(approved_current, approved_last)
    disapproved_percentage = calculate_percentage(disapproved_current, disapproved_last)
    processing_percentage = calculate_percentage(processing_current, processing_last)
    
    stats = {
        'total_requests': total_current,
        'approved': approved_current,
        'disapproved': disapproved_current,
        'processing': processing_current,
        'percentage_change': round(total_percentage, 1),
        'is_increase': total_percentage >= 0,
        'approved_percentage': round(approved_percentage, 1),
        'approved_is_increase': approved_percentage >= 0,
        'disapproved_percentage': round(disapproved_percentage, 1),
        'disapproved_is_increase': disapproved_percentage >= 0,
        'processing_percentage': round(processing_percentage, 1),
        'processing_is_increase': processing_percentage >= 0,
    }
    
    # Handle AJAX requests for search and pagination
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        search_query = request.GET.get('search', '').strip()
        tab = request.GET.get('tab', 'tickets')
        
        if tab == 'tickets':
            tickets_qs = Ticket.objects.all()
            if search_query:
                tickets_qs = tickets_qs.filter(
                    Q(ticket_number__icontains=search_query) |
                    Q(requestor_name__icontains=search_query) |
                    Q(device__device_name__icontains=search_query) |
                    Q(category__name__icontains=search_query) |
                    Q(priority_level__icontains=search_query) |
                    Q(status__icontains=search_query)
                )
            tickets_qs = tickets_qs.order_by('-created_at')
            
            # Pagination
            page = request.GET.get('page', 1)
            paginator = Paginator(tickets_qs, 10)
            tickets = paginator.get_page(page)
            
            return JsonResponse({
                'success': True,
                'html': render(request, 'ticketing/partials/admin_tickets_table.html', {'tickets': tickets}).content.decode(),
                'pagination': {
                    'has_previous': tickets.has_previous(),
                    'has_next': tickets.has_next(),
                    'number': tickets.number,
                    'num_pages': tickets.paginator.num_pages,
                    'start_index': tickets.start_index(),
                    'end_index': tickets.end_index(),
                    'count': tickets.paginator.count,
                }
            })
        elif tab == 'devices':
            devices_qs = Device.objects.all()
            if search_query:
                devices_qs = devices_qs.filter(
                    Q(device_name__icontains=search_query) |
                    Q(device_code__icontains=search_query) |
                    Q(device_type__name__icontains=search_query) |
                    Q(device_location__icontains=search_query) |
                    Q(user__firstname__icontains=search_query) |
                    Q(user__lastname__icontains=search_query)
                )
            devices_qs = devices_qs.order_by('-created_at')
            
            # Pagination
            page = request.GET.get('page', 1)
            paginator = Paginator(devices_qs, 10)
            devices = paginator.get_page(page)
            
            return JsonResponse({
                'success': True,
                'html': render(request, 'ticketing/partials/admin_devices_table.html', {'devices': devices}).content.decode(),
                'pagination': {
                    'has_previous': devices.has_previous(),
                    'has_next': devices.has_next(),
                    'number': devices.number,
                    'num_pages': devices.paginator.num_pages,
                    'start_index': devices.start_index(),
                    'end_index': devices.end_index(),
                    'count': devices.paginator.count,
                }
            })
    
    # Regular page load with pagination
    tickets_qs = Ticket.objects.all().order_by('-created_at')
    devices_qs = Device.objects.all().order_by('-created_at')
    
    # Pagination for initial load
    tickets_paginator = Paginator(tickets_qs, 10)
    devices_paginator = Paginator(devices_qs, 10)
    tickets_page = request.GET.get('tickets_page', 1)
    devices_page = request.GET.get('devices_page', 1)
    tickets = tickets_paginator.get_page(tickets_page)
    devices = devices_paginator.get_page(devices_page)
    
    context = {
        'stats': stats,
        'tickets': tickets,
        'devices': devices,
        'categories': TicketCategory.objects.all(),
    }
    
    return render(request, 'ticketing/admin-ticket.html', context)


@login_required
@require_http_methods(["POST"])
def add_device(request):
    try:
        form = DeviceForm(request.POST)
        if form.is_valid():
            device = form.save(commit=False)
            device.user = request.user
            device.save()
            return JsonResponse({
                'status': 'success',
                'message': 'Device added successfully',
                'device': {
                    'id': device.id,
                    'device_name': device.device_name,
                    'device_code': device.device_code,
                    'device_location': device.device_location,
                    'device_type': device.device_type.name if device.device_type else ''
                }
            })
        return JsonResponse({
            'status': 'error',
            'errors': form.errors
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def create_ticket(request):
    mis_personnel = EmployeeLogin.objects.filter(mis_admin=True).first()
    form = TicketForm(request.POST, user=request.user)
    if form.is_valid():
        ticket = form.save(commit=False)
        ticket.requestor = request.user
        ticket.save()

        Notification.objects.create(
            title="New Ticket Created",
            message=f"{request.user.firstname} {request.user.lastname} has created a new ticket.",
            notification_type="approval",
            sender=request.user,
            recipient=mis_personnel,
            for_all=True,
            module="ticketing/admin"
        )
        return JsonResponse({
            'status': 'success',
            'message': 'Ticket created successfully',
            'ticket_number': ticket.ticket_number
        })
    return JsonResponse({
        'status': 'error',
        'errors': form.errors
    }, status=400)


@login_required
def get_ticket_details(request, ticket_id):
    ticket = get_object_or_404(Ticket, id=ticket_id)
    
    # Check permission
    is_owner = ticket.requestor == request.user
    is_admin = request.user.mis_admin
    
    if not (is_owner or is_admin):
        return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)
    
    data = {
        'id': ticket.id,
        'ticket_number': ticket.ticket_number,
        'requestor_name': ticket.requestor_name,
        'device': ticket.device.device_name if ticket.device else '',
        'device_id': ticket.device.id if ticket.device else '',
        'priority_level': ticket.priority_level,
        'category': ticket.category.name if ticket.category else '',
        'category_id': ticket.category.id if ticket.category else '',
        'problem_details': ticket.problem_details,
        'status': ticket.status,
        'technician_name': ticket.technician_name,
        'diagnosis': ticket.diagnosis,
        'action_taken': ticket.action_taken,
        'possible_reason': ticket.possible_reason,
        'recommendation': ticket.recommendation,
        'created_at': ticket.created_at.strftime('%Y-%m-%d %H:%M'),
        'is_owner': is_owner,
        'is_admin': is_admin,
        'can_edit': is_owner and ticket.status == 'Processing',
        'can_review': is_admin and ticket.status == 'Processing'
    }
    
    return JsonResponse(data)


@login_required
@require_http_methods(["POST"])
def update_ticket(request, ticket_id):
    ticket = get_object_or_404(Ticket, id=ticket_id, requestor=request.user)
    
    if ticket.status != 'Processing':
        return JsonResponse({'status': 'error', 'message': 'Cannot edit this ticket'}, status=400)
    
    form = TicketUpdateForm(request.POST, instance=ticket)
    if form.is_valid():
        form.save()
        return JsonResponse({
            'status': 'success',
            'message': 'Ticket updated successfully'
        })
    return JsonResponse({
        'status': 'error',
        'errors': form.errors
    }, status=400)


@login_required
@require_http_methods(["POST"])
def cancel_ticket(request, ticket_id):
    ticket = get_object_or_404(Ticket, id=ticket_id, requestor=request.user)
    
    if ticket.status != 'Processing':
        return JsonResponse({'status': 'error', 'message': 'Cannot cancel this ticket'}, status=400)
    
    ticket.status = 'Cancelled'
    ticket.save()
    
    return JsonResponse({
        'status': 'success',
        'message': 'Ticket cancelled successfully'
    })


@login_required
@require_http_methods(["POST"])
def review_ticket(request, ticket_id):
    if not request.user.mis_admin:
        return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)
    
    ticket = get_object_or_404(Ticket, id=ticket_id)
    
    if ticket.status != 'Processing':
        return JsonResponse({'status': 'error', 'message': 'This ticket has already been reviewed'}, status=400)
    
    form = TicketReviewForm(request.POST, instance=ticket, user=request.user)
    if form.is_valid():
        ticket = form.save(commit=False)
        ticket.reviewed_by = request.user
        ticket.reviewed_at = timezone.now()
        ticket.save()

        if ticket.status == 'Approved':
            notification_title = "Ticket Approved"
            notification_message = f"Your Ticket {ticket.ticket_number} has been approved by the MIS."
            notification_type = "approved"
        elif ticket.status == 'Disapproved':
            notification_title = "Ticket Disapproved"
            notification_message = f"Your Ticket {ticket.ticket_number} has been disapproved by the MIS."
            notification_type = "disapproved"
        else:
            notification_title = "Ticket Updated"
            notification_message = f"Your Ticket {ticket.ticket_number} has been updated by the MIS."
            notification_type = "general"

        Notification.objects.create(
            title=notification_title,
            message=notification_message,
            notification_type=notification_type,
            sender=request.user,
            recipient=ticket.requestor,
            for_all=False,
            module="ticketing"
        )
        
        return JsonResponse({
            'status': 'success',
            'message': 'Ticket reviewed successfully'
        })
    return JsonResponse({
        'status': 'error',
        'errors': form.errors
    }, status=400)


@login_required
def get_device_details(request, device_id):
    device = get_object_or_404(Device, id=device_id)
    
    # Check permission
    is_owner = device.user == request.user
    is_admin = request.user.mis_admin
    
    if not (is_owner or is_admin):
        return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)
    
    data = {
        'id': device.id,
        'device_name': device.device_name,
        'device_code': device.device_code,
        'device_location': device.device_location,
        'device_type': device.device_type.name if device.device_type else '',
        'device_type_id': device.device_type.id if device.device_type else '',
        'user': device.user.full_name,
        'created_at': device.created_at.strftime('%Y-%m-%d %H:%M'),
        'is_owner': is_owner,
        'can_edit': is_owner
    }
    
    return JsonResponse(data)


@login_required
@require_http_methods(["POST"])
def update_device(request, device_id):
    device = get_object_or_404(Device, id=device_id, user=request.user)
    
    form = DeviceUpdateForm(request.POST, instance=device)
    if form.is_valid():
        form.save()
        return JsonResponse({
            'status': 'success',
            'message': 'Device updated successfully'
        })
    return JsonResponse({
        'status': 'error',
        'errors': form.errors
    }, status=400)


@login_required
@require_http_methods(["POST"])
def delete_device(request, device_id):
    device = get_object_or_404(Device, id=device_id, user=request.user)
    device.delete()
    
    return JsonResponse({
        'status': 'success',
        'message': 'Device deleted successfully'
    })


@login_required
def get_form_options(request):
    # Get device types and categories for forms
    device_types = list(DeviceType.objects.values('id', 'name'))
    categories = list(TicketCategory.objects.values('id', 'name'))
    
    # Get user's devices if user is authenticated
    devices = []
    if request.user.is_authenticated:
        devices = list(Device.objects.filter(user=request.user, is_active=True).values('id', 'device_name'))
    
    return JsonResponse({
        'device_types': device_types,
        'categories': categories,
        'devices': devices
    })


@login_required
def export_tickets_report(request):
    """Export tickets report based on filters"""
    if not request.user.mis_admin:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    # Get filter parameters
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    category_id = request.GET.get('category')
    status = request.GET.get('status')
    
    # Parse dates
    try:
        date_from = parse_date(date_from) if date_from else None
        date_to = parse_date(date_to) if date_to else None
    except:
        return JsonResponse({'error': 'Invalid date format'}, status=400)
    
    # Build query
    tickets_qs = Ticket.objects.all()
    
    if date_from and date_to:
        tickets_qs = tickets_qs.filter(created_at__date__gte=date_from, created_at__date__lte=date_to)
    elif date_from:
        tickets_qs = tickets_qs.filter(created_at__date__gte=date_from)
    elif date_to:
        tickets_qs = tickets_qs.filter(created_at__date__lte=date_to)
    
    if category_id:
        tickets_qs = tickets_qs.filter(category_id=category_id)
    
    if status:
        tickets_qs = tickets_qs.filter(status=status)
    
    tickets_qs = tickets_qs.order_by('-created_at')
    
    # Create Excel workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Ticket Report"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=14)
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    subheader_font = Font(bold=True, size=12)
    table_header_font = Font(bold=True, size=11)
    table_header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Company header
    worksheet.merge_cells('A1:M1')  # Updated to M1 for 13 columns
    worksheet['A1'] = "RYONAN ELECTRIC PHILIPPINES"
    worksheet['A1'].font = header_font
    worksheet['A1'].fill = header_fill
    worksheet['A1'].alignment = center_alignment
    worksheet['A1'].border = border
    
    # Subheader
    date_range = ""
    if date_from and date_to:
        date_range = f" for {date_from.strftime('%B %d, %Y')} - {date_to.strftime('%B %d, %Y')}"
    elif date_from:
        date_range = f" from {date_from.strftime('%B %d, %Y')}"
    elif date_to:
        date_range = f" up to {date_to.strftime('%B %d, %Y')}"
    
    worksheet.merge_cells('A2:M2')  # Updated to M2 for 13 columns
    worksheet['A2'] = f"Ticket Output Report{date_range}"
    worksheet['A2'].font = subheader_font
    worksheet['A2'].alignment = center_alignment
    worksheet['A2'].border = border
    
    # Table headers
    headers = ['Submitted At', 'Ticket #', 'Requestor', 'Device', 'Category', 'Priority', 'Status', 'Problem Details', 'Technician', 'Diagnosis', 'Action Taken', 'Possible Reason', 'Recommendation']
    row = 4
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=row, column=col, value=header)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Data rows
    row = 5
    for ticket in tickets_qs:
        data = [
            ticket.created_at.strftime('%m/%d/%Y %H:%M'),
            ticket.ticket_number,
            ticket.requestor_name,
            ticket.device.device_name if ticket.device else '-',
            ticket.category.name if ticket.category else '-',
            ticket.priority_level,
            ticket.status,
            ticket.problem_details,
            ticket.technician_name or '-',
            ticket.diagnosis or '-',
            ticket.action_taken or '-',
            ticket.possible_reason or '-',
            ticket.recommendation or '-'
        ]
        
        for col, value in enumerate(data, 1):
            cell = worksheet.cell(row=row, column=col, value=value)
            cell.border = border
            if col in [6, 7]:  # Priority and Status columns
                cell.alignment = center_alignment
        row += 1
    
    # Adjust column widths
    for col in range(1, 14):  # Updated to 13 columns
        if col == 1:  # Submitted At
            worksheet.column_dimensions[get_column_letter(col)].width = 18
        elif col == 8:  # Problem Details
            worksheet.column_dimensions[get_column_letter(col)].width = 30
        elif col in [10, 11, 12, 13]:  # Diagnosis, Action Taken, Possible Reason, Recommendation
            worksheet.column_dimensions[get_column_letter(col)].width = 25
        else:
            worksheet.column_dimensions[get_column_letter(col)].width = 15
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"ticket_report_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    workbook.save(response)
    return response


@login_required
def device_ticket_history(request, device_id):
    """Get ticket history for a specific device"""
    if not request.user.mis_admin:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    device = get_object_or_404(Device, id=device_id)
    tickets = Ticket.objects.filter(device=device).order_by('-created_at')
    
    # Group tickets by month/year
    history_data = {}
    for ticket in tickets:
        month_year = ticket.created_at.strftime('%b %Y')
        if month_year not in history_data:
            history_data[month_year] = []
        
        ticket_data = {
            'id': ticket.id,
            'ticket_number': ticket.ticket_number,
            'requestor_name': ticket.requestor_name,
            'priority_level': ticket.priority_level,
            'category': ticket.category.name if ticket.category else '-',
            'status': ticket.status,
            'problem_details': ticket.problem_details,
            'created_at': ticket.created_at.strftime('%m/%d/%Y %H:%M'),
            'technician_name': ticket.technician_name or '-',
            'diagnosis': ticket.diagnosis or '-',
            'action_taken': ticket.action_taken or '-',
            'possible_reason': ticket.possible_reason or '-',
            'recommendation': ticket.recommendation or '-'
        }
        history_data[month_year].append(ticket_data)
    
    return JsonResponse({
        'success': True,
        'device': {
            'id': device.id,
            'device_name': device.device_name,
            'device_code': device.device_code,
            'owner': device.user.full_name
        },
        'history': history_data
    })


@login_required
def export_device_history(request, device_id):
    """Export device ticket history to Excel"""
    if not request.user.mis_admin:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    device = get_object_or_404(Device, id=device_id)
    tickets = Ticket.objects.filter(device=device).order_by('-created_at')
    
    # Create Excel workbook
    workbook = openpyxl.Workbook()
    
    # Remove default sheet
    workbook.remove(workbook.active)
    
    # Group tickets by month/year and create sheets
    history_data = {}
    for ticket in tickets:
        month_year = ticket.created_at.strftime('%b %Y')
        if month_year not in history_data:
            history_data[month_year] = []
        history_data[month_year].append(ticket)
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_font = Font(bold=True, size=10)
    table_header_font = Font(bold=True, size=9)
    table_header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    for month_year, month_tickets in history_data.items():
        # Create sheet for each month
        worksheet = workbook.create_sheet(title=month_year)
        
        # Device info header
        worksheet.merge_cells('A1:F1')
        worksheet['A1'] = f"Device History - {device.device_name} ({device.device_code})"
        worksheet['A1'].font = header_font
        worksheet['A1'].fill = header_fill
        worksheet['A1'].alignment = center_alignment
        worksheet['A1'].border = border
        
        worksheet.merge_cells('A2:F2')
        worksheet['A2'] = f"Owner: {device.user.full_name} | Period: {month_year}"
        worksheet['A2'].font = subheader_font
        worksheet['A2'].alignment = center_alignment
        worksheet['A2'].border = border
        
        row = 4
        for ticket in month_tickets:
            # Ticket header
            worksheet.merge_cells(f'A{row}:F{row}')
            worksheet[f'A{row}'] = f"Ticket #{ticket.ticket_number} - {ticket.created_at.strftime('%m/%d/%Y %H:%M')}"
            worksheet[f'A{row}'].font = table_header_font
            worksheet[f'A{row}'].fill = table_header_fill
            worksheet[f'A{row}'].alignment = center_alignment
            worksheet[f'A{row}'].border = border
            row += 1
            
            # Ticket details
            details = [
                ['Status:', ticket.status],
                ['Priority:', ticket.priority_level],
                ['Category:', ticket.category.name if ticket.category else '-'],
                ['Problem:', ticket.problem_details],
                ['Technician:', ticket.technician_name or '-'],
                ['Diagnosis:', ticket.diagnosis or '-'],
                ['Action Taken:', ticket.action_taken or '-'],
                ['Possible Reason:', ticket.possible_reason or '-'],
                ['Recommendation:', ticket.recommendation or '-']
            ]
            
            for label, value in details:
                worksheet[f'A{row}'] = label
                worksheet[f'A{row}'].font = Font(bold=True)
                worksheet[f'A{row}'].border = border
                
                worksheet.merge_cells(f'B{row}:F{row}')
                worksheet[f'B{row}'] = value
                worksheet[f'B{row}'].border = border
                row += 1
            
            row += 1  # Add space between tickets
        
        # Adjust column widths
        worksheet.column_dimensions['A'].width = 20
        for col in range(2, 7):
            worksheet.column_dimensions[get_column_letter(col)].width = 15
    
    # If no tickets, create a summary sheet
    if not history_data:
        worksheet = workbook.create_sheet(title="Summary")
        worksheet['A1'] = f"No ticket history found for {device.device_name}"
        worksheet['A1'].font = header_font
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"device_history_{device.device_code}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    workbook.save(response)
    return response


@login_required
def print_ticket_pdf(request, ticket_id):
    """Generate PDF receipt for approved ticket"""
    ticket = get_object_or_404(Ticket, id=ticket_id)
    
    # Check if user owns the ticket or is admin
    if not (ticket.requestor == request.user or request.user.mis_admin):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    # Create PDF buffer with reduced margins
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    
    # Container for the 'Flowable' objects
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        spaceAfter=20,
        alignment=1,  # Center alignment
        textColor=colors.HexColor('#2563eb')
    )
    
    header_style = ParagraphStyle(
        'CustomHeader',
        parent=styles['Heading2'],
        fontSize=13,
        spaceAfter=12,
        textColor=colors.HexColor('#1f2937')
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=13,
        spaceAfter=6,
        textColor=colors.HexColor('#374151')
    )
    
    # Company header
    company_name = Paragraph("RYONAN ELECTRIC PHILIPPINES", title_style)
    elements.append(company_name)
    elements.append(Spacer(1, 10))
    
    # Ticket receipt title
    receipt_title = Paragraph("SUPPORT TICKET RECEIPT", header_style)
    elements.append(receipt_title)
    elements.append(Spacer(1, 15))
    
    # Ticket Information Table with Problem Details included
    ticket_data = [
        ['Ticket Number:', ticket.ticket_number],
        ['Status:', ticket.status],
        ['Requestor:', ticket.requestor_name],
        ['Device:', ticket.device.device_name if ticket.device else '-'],
        ['Priority Level:', ticket.priority_level],
        ['Category:', ticket.category.name if ticket.category else '-'],
        ['Submitted At:', ticket.created_at.strftime('%B %d, %Y at %I:%M %p')],
        ['Problem Details:', Paragraph(ticket.problem_details, normal_style)],
    ]
    
    ticket_table = Table(ticket_data, colWidths=[2.2*inch, 4.8*inch])
    ticket_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#374151')),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#111827')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    elements.append(ticket_table)
    elements.append(Spacer(1, 20))
    
    # MIS Diagnosis Section (only if ticket is approved/disapproved and has diagnosis)
    if ticket.status in ['Approved', 'Disapproved'] and ticket.diagnosis:
        diagnosis_header = Paragraph("MIS DIAGNOSIS & RESOLUTION", header_style)
        elements.append(diagnosis_header)
        elements.append(Spacer(1, 10))
        
        # Create paragraph styles for wrapping text in table cells
        cell_style = ParagraphStyle(
            'CellStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#111827'),
            wordWrap='LTR'
        )
        
        diagnosis_data = [
            ['Technician:', Paragraph(ticket.technician_name or '-', cell_style)],
            ['Diagnosis:', Paragraph(ticket.diagnosis or '-', cell_style)],
            ['Action Taken:', Paragraph(ticket.action_taken or '-', cell_style)],
            ['Possible Reason:', Paragraph(ticket.possible_reason or '-', cell_style)],
            ['Recommendation:', Paragraph(ticket.recommendation or '-', cell_style)],
        ]
        
        diagnosis_table = Table(diagnosis_data, colWidths=[2.2*inch, 4.8*inch])
        diagnosis_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#ecfdf5')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#065f46')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#a7f3d0')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        elements.append(diagnosis_table)
        elements.append(Spacer(1, 20))
    
    # Footer
    footer_text = Paragraph(
        f"Generated on {timezone.now().strftime('%B %d, %Y at %I:%M %p')}<br/>REPConnect MIS Ticketing System",
        ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            alignment=1,
            textColor=colors.HexColor('#6b7280')
        )
    )
    elements.append(footer_text)
    
    # Build PDF
    doc.build(elements)
    
    # FileResponse
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"ticket_{ticket.ticket_number}_{timezone.now().strftime('%Y%m%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response